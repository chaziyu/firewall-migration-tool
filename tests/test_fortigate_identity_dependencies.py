from io import BytesIO

from openpyxl import load_workbook

import fwmigrate.generators  # noqa: F401 - register built-in target generators
from fwmigrate.core.registry import PluginRegistry
from fwmigrate.parsers.fortigate.extractor import extract_fortigate_config
from fwmigrate.report.excel_exporter import IRExcelExporter


IDENTITY_CONFIG = r'''
config vpn certificate remote
    edit "SAML_CERT"
    next
    edit "AuthCert"
    next
    edit "AuthCA"
    next
end
config user ldap
    edit "LDAP"
        set server "192.0.2.10"
        set ca-cert "AuthCA"
        set client-cert "MissingClientCert"
        set account-key-cert-field othername
        set account-key-processing strip
        set group-filter "(&(objectClass=group)(member={0}))"
        set group-search-base "ou=Groups,dc=example,dc=com"
        set search-type recursive nested
    next
end
config user saml
    edit "SAML"
        set idp-cert "SAML_CERT"
        set cert "AuthCert"
    next
    edit "BrokenSAML"
        set idp-cert "MissingCert"
    next
end
config user fsso
    edit "FSSO"
        set server "192.0.2.20"
    next
end
config user radius
    edit "RADIUS"
        set server "192.0.2.30"
        set secret RADIUS_SECRET
    next
end
config user tacacs+
    edit "TACACS"
        set server "192.0.2.40"
        set key TACACS_SECRET
    next
end
config user adgrp
    edit "DOMAIN/USERS"
        set server-name "FSSO"
    next
end
config user local
    edit "alice"
        set status enable
        set type password
        set passwd ENC SENTINEL_SECRET
    next
end
config user group
    edit "LocalGroup"
        set member "alice"
    next
    edit "FSSOGroup"
        set group-type fsso-service
        set member "DOMAIN/USERS"
    next
    edit "LDAPGroup"
        set member "LDAP"
        config match
            edit 1
                set server-name "LDAP"
                set group-name "CN=Allowed,DC=example,DC=com"
            next
        end
    next
    edit "SAMLGroup"
        set member "SAML"
    next
    edit "RemoteProviders"
        set member "RADIUS" "TACACS"
    next
    edit "Broken"
        set member "DeletedUser"
        config match
            edit 1
                set server-name "DeletedLDAP"
                set group-name "CN=ExternalGroup,DC=example,DC=com"
            next
        end
    next
end
config authentication scheme
    edit "LDAP-Scheme"
        set user-database "LDAP"
    next
    edit "Broken-Scheme"
        set user-database "MissingLDAP"
    next
end
config authentication rule
    edit "GoodRule"
        set active-auth-method "LDAP-Scheme"
    next
    edit "BadRule"
        set active-auth-method "MissingScheme"
    next
end
config user fortitoken
    edit "TOKEN1"
        set seed "TOKEN_SEED_SECRET"
        set activation-code "ACTIVATION_SECRET"
    next
end
config system admin
    edit "admin-good"
        set accprofile "super_admin"
        set two-factor fortitoken
        set fortitoken "TOKEN1"
    next
    edit "admin-bad"
        set accprofile "super_admin"
        set two-factor fortitoken
        set fortitoken "MISSING_TOKEN"
    next
end
config firewall addrgrp
    edit "QuarantinedDevices"
    next
end
config user setting
    set auth-cert "AuthCert"
    set auth-ca-cert "AuthCA"
    set auth-timeout 300
end
config user quarantine
    set firewall-groups "QuarantinedDevices"
end
config firewall policy
    edit 1
        set action accept
        set groups "LocalGroup"
        set service "ALL"
    next
    edit 2
        set action accept
        set users "alice"
        set service "ALL"
    next
    edit 3
        set action accept
        set groups "MissingGroup"
        set service "ALL"
    next
end
'''


def _ir():
    return extract_fortigate_config(IDENTITY_CONFIG).canonical_ir


def test_identity_dependencies_resolve_by_source_type_without_external_dn_validation():
    ir = _ir()
    groups = {group.name: group for group in ir.user_groups}

    assert groups["LocalGroup"].member_dependencies[0].dependency_type == "local-user"
    assert groups["FSSOGroup"].member_dependencies[0].dependency_type == "fsso-ad-group"
    assert groups["LDAPGroup"].member_dependencies[0].dependency_type == "ldap-server"
    assert groups["SAMLGroup"].member_dependencies[0].dependency_type == "saml-server"
    assert [item.dependency_type for item in groups["RemoteProviders"].member_dependencies] == [
        "radius-server", "tacacs-server"
    ]
    assert groups["LDAPGroup"].unresolved_match_servers == []
    assert "CN=Allowed,DC=example,DC=com" not in groups["LDAPGroup"].unresolved_members
    assert ir.fsso_ad_groups[0].provider_resolved is True

    broken = groups["Broken"]
    assert broken.members == ["DeletedUser"]
    assert broken.unresolved_members == ["DeletedUser"]
    assert broken.unresolved_match_servers == ["DeletedLDAP"]
    assert "CN=ExternalGroup,DC=example,DC=com" not in broken.unresolved_members


def test_certificate_authentication_admin_and_singleton_dependencies_are_explicit():
    ir = _ir()
    saml = {item.name: item for item in ir.user_saml_servers}
    assert saml["SAML"].idp_certificate_resolved is True
    assert saml["SAML"].cert_certificate_resolved is True
    assert saml["BrokenSAML"].idp_certificate_resolved is False
    assert saml["BrokenSAML"].unresolved_certificate_references == ["MissingCert"]
    ldap = {item.name: item for item in ir.user_ldap_servers}
    assert ldap["LDAP"].client_certificate_resolved is False
    assert ldap["LDAP"].unresolved_certificate_references == ["MissingClientCert"]
    assert ldap["LDAP"].search_type == ["recursive", "nested"]
    assert ldap["LDAP"].account_key_processing == "strip"

    schemes = {item.name: item for item in ir.authentication_schemes}
    assert schemes["LDAP-Scheme"].resolved_user_databases == ["LDAP"]
    assert schemes["Broken-Scheme"].unresolved_user_databases == ["MissingLDAP"]
    rules = {item.name: item for item in ir.authentication_rules}
    assert rules["GoodRule"].active_auth_method_resolved is True
    assert rules["BadRule"].unresolved_auth_methods == ["MissingScheme"]

    admins = {item.name: item for item in ir.administrators}
    assert admins["admin-good"].fortitoken_resolved is True
    assert admins["admin-good"].access_profile_resolved is True
    assert admins["admin-bad"].fortitoken_resolved is False
    assert admins["admin-bad"].unresolved_references == ["MISSING_TOKEN"]

    assert ir.user_authentication_settings.auth_certificate_resolved is True
    assert ir.user_authentication_settings.auth_ca_certificate_resolved is True
    assert ir.user_quarantine_settings.resolved_firewall_groups == ["QuarantinedDevices"]


def test_identity_policies_are_partial_and_withheld_by_every_target_generator():
    ir = _ir()
    assert ir.policies[0].unresolved_user_groups == []
    assert ir.policies[0].identity_dependency_review is True
    assert ir.policies[1].unresolved_users == []
    assert ir.policies[2].unresolved_user_groups == ["MissingGroup"]
    assert all(policy.migration_status == "PARTIALLY_NORMALIZED" for policy in ir.policies)
    assert all(policy.requires_manual_review for policy in ir.policies)

    forbidden = {
        "cisco_asa": "access-list ACL_",
        "checkpoint": "mgmt_cli add access-rule",
        "juniper_srx": "then permit",
        "fortigate": "set action accept",
        "palo_alto": "<source-user>",
    }
    formats = {
        "cisco_asa": "cli", "checkpoint": "cli", "juniper_srx": "cli",
        "fortigate": "cli", "palo_alto": "xml",
    }
    for vendor, marker in forbidden.items():
        artifacts = PluginRegistry.get_generator(vendor).generate(ir, format=formats[vendor])
        combined = "\n".join(artifact.content for artifact in artifacts)
        assert marker not in combined
        assert "withheld" in combined.lower() or vendor == "palo_alto"


def test_excel_exposes_dependency_results_without_credentials():
    ir = _ir()
    workbook_bytes = IRExcelExporter(ir).generate()
    workbook = load_workbook(BytesIO(workbook_bytes), read_only=True)
    assert "Security Identity Dependencies" in workbook.sheetnames
    assert "User Authentication Settings" in workbook.sheetnames
    assert "User Quarantine" in workbook.sheetnames
    serialized = ir.model_dump_json() + "\n" + "\n".join(
        str(cell.value)
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if cell.value is not None
    )
    for secret in ("SENTINEL_SECRET", "TOKEN_SEED_SECRET", "ACTIVATION_SECRET"):
        assert secret not in serialized


def test_user_setting_and_quarantine_have_typed_extract_only_coverage():
    extraction = extract_fortigate_config(IDENTITY_CONFIG)
    coverage = {section.path: section for section in extraction.source_sections}
    unsupported_paths = {item.source_path for item in extraction.unsupported_items}

    for path in ("user setting", "user quarantine"):
        assert coverage[path].object_count_source == 1
        assert coverage[path].object_count_parsed == 1
        assert coverage[path].object_count_normalized == 1
        assert coverage[path].status.value == "EXTRACT_ONLY"
        assert path not in unsupported_paths


def test_security_profile_existence_is_distinct_from_semantic_migration():
    config = r'''
config antivirus profile
    edit "AV1"
    next
end
config ips sensor
    edit "IPS1"
    next
end
config webfilter profile
    edit "WF1"
    next
end
config application list
    edit "APP1"
    next
end
config firewall ssl-ssh-profile
    edit "certificate-inspection"
    next
end
config firewall policy
    edit 10
        set action accept
        set utm-status enable
        set av-profile "AV1"
        set ips-sensor "IPS1"
        set webfilter-profile "WF1"
        set application-list "APP1"
        set ssl-ssh-profile "certificate-inspection"
        set service "ALL"
    next
    edit 11
        set action accept
        set utm-status enable
        set ips-sensor "MissingIPS"
        set service "ALL"
    next
end
'''
    ir = extract_fortigate_config(config).canonical_ir
    good, broken = ir.policies

    assert good.unresolved_security_profiles == []
    assert good.security_profile_semantics_review is True
    assert good.migration_status == "PARTIALLY_NORMALIZED"
    assert good.requires_manual_review is True
    assert broken.unresolved_security_profiles == ["ips:MissingIPS"]
    assert any(
        entry.id == "policy:11:security-profiles"
        for entry in ir.audit_entries
    )
    correlation = next(
        entry for entry in ir.audit_entries
        if "correlated into Security Profile Group" in entry.message
    )
    assert correlation.confidence.value == "manual"
    assert all(group.requires_manual_review for group in ir.security_profile_groups)
