import io

from openpyxl import load_workbook

from fwmigrate.parsers.fortigate.parser import parse_fortigate_config
from fwmigrate.parsers.fortigate.transformer import FGToIRTransformer
from fwmigrate.report.excel_exporter import IRExcelExporter


def _transform(config: str):
    source = parse_fortigate_config(config)
    return source, FGToIRTransformer(source).transform()


def _policy_config(ztna_settings: str = "") -> str:
    return f'''
config firewall policy
    edit 313
        set name "ZTNA policy"
        set srcintf "LAN"
        set dstintf "WAN"
        set srcaddr "all"
        set dstaddr "all"
        set action accept
        set schedule "always"
        set service "ALL"
{ztna_settings}
    next
end
'''


def test_ztna_policy_fields_are_typed_preserved_and_reviewed():
    source, ir = _transform(_policy_config('''
        set ztna-status enable
        set ztna-device-ownership corporate
        set ztna-ems-tag "PRIMARY_A" "PRIMARY_B"
        set ztna-ems-tag-secondary "SECONDARY_A" "SECONDARY_B"
        set ztna-geo-tag "MY" "SG"
        set ztna-policy-redirect enable
        set ztna-tags-match-logic all
        set ztna-future-setting "retain-me"
'''))

    source_policy = source.policies[0]
    assert source_policy.ztna_status == "enable"
    assert source_policy.ztna_device_ownership == "corporate"
    assert source_policy.ztna_ems_tag == ["PRIMARY_A", "PRIMARY_B"]
    assert source_policy.ztna_ems_tag_secondary == ["SECONDARY_A", "SECONDARY_B"]
    assert source_policy.ztna_geo_tag == ["MY", "SG"]
    assert source_policy.ztna_policy_redirect == "enable"
    assert source_policy.ztna_tags_match_logic == "all"
    assert source_policy.extra_settings == {
        "ztna_future_setting": "retain-me",
    }

    policy = ir.policies[0]
    assert policy.source_ztna_status == "enable"
    assert policy.source_ztna_device_ownership == "corporate"
    assert policy.source_ztna_ems_tags == ["PRIMARY_A", "PRIMARY_B"]
    assert policy.source_ztna_ems_tags_secondary == ["SECONDARY_A", "SECONDARY_B"]
    assert policy.source_ztna_geo_tags == ["MY", "SG"]
    assert policy.source_ztna_policy_redirect == "enable"
    assert policy.source_ztna_tags_match_logic == "all"
    assert policy.source_extra_settings == {
        "ztna_future_setting": "retain-me",
    }
    assert policy.migration_status == "PARTIALLY_NORMALIZED"
    assert policy.requires_manual_review is True
    assert (
        "FortiGate ZTNA policy semantics require target-platform review"
        in policy.review_reasons
    )


def test_each_ztna_field_triggers_review_without_forcing_other_policies():
    cases = (
        ("set ztna-status enable", "source_ztna_status", "enable"),
        (
            "set ztna-device-ownership personal",
            "source_ztna_device_ownership",
            "personal",
        ),
        (
            'set ztna-ems-tag "PRIMARY"',
            "source_ztna_ems_tags",
            ["PRIMARY"],
        ),
        (
            'set ztna-ems-tag-secondary "SECONDARY"',
            "source_ztna_ems_tags_secondary",
            ["SECONDARY"],
        ),
        (
            'set ztna-geo-tag "MY"',
            "source_ztna_geo_tags",
            ["MY"],
        ),
        (
            "set ztna-policy-redirect disable",
            "source_ztna_policy_redirect",
            "disable",
        ),
        (
            "set ztna-tags-match-logic any",
            "source_ztna_tags_match_logic",
            "any",
        ),
    )

    for setting, field, expected in cases:
        _, ir = _transform(_policy_config(f"        {setting}\n"))
        policy = ir.policies[0]
        assert getattr(policy, field) == expected
        assert policy.migration_status == "PARTIALLY_NORMALIZED"
        assert policy.requires_manual_review is True

    _, ordinary_ir = _transform(_policy_config())
    ordinary = ordinary_ir.policies[0]
    assert ordinary.migration_status == "NORMALIZED"
    assert ordinary.requires_manual_review is False
    assert not any("ZTNA" in reason for reason in ordinary.review_reasons)


def test_disabled_ztna_status_alone_is_not_treated_as_active_ztna():
    _, ir = _transform(_policy_config("        set ztna-status disable\n"))

    policy = ir.policies[0]
    assert policy.source_ztna_status == "disable"
    assert policy.migration_status == "NORMALIZED"
    assert policy.requires_manual_review is False
    assert not any("ZTNA" in reason for reason in policy.review_reasons)


def test_typed_ztna_values_are_visible_in_policy_audit_settings():
    _, ir = _transform(_policy_config('''
        set ztna-device-ownership corporate
        set ztna-ems-tag-secondary "SECONDARY_A" "SECONDARY_B"
        set ztna-geo-tag "MY" "SG"
        set ztna-policy-redirect enable
        set ztna-tags-match-logic all
'''))

    workbook = load_workbook(io.BytesIO(IRExcelExporter(ir).generate()))
    sheet = workbook["Policies"]
    headers = {cell.value: cell.column for cell in sheet[3]}
    additional = sheet.cell(4, headers["Additional Settings"]).value

    assert "ztna-device-ownership=corporate" in additional
    assert "ztna-ems-tag-secondary=SECONDARY_A SECONDARY_B" in additional
    assert "ztna-geo-tag=MY SG" in additional
    assert "ztna-policy-redirect=enable" in additional
    assert "ztna-tags-match-logic=all" in additional
