import io

from openpyxl import load_workbook

from fwmigrate.extraction.models import DependencyRecord, ExtractionResult
from fwmigrate.ir.core import IRAuditEntry, IRConfig, IRMetadata
from fwmigrate.ir.enums import MigrationConfidence
from fwmigrate.parsers.fortigate.extractor import extract_fortigate_config
from fwmigrate.report.excel_exporter import IRExcelExporter


def test_dependency_registry_and_unresolved_summary_use_dependency_statuses():
    extraction = ExtractionResult(
        canonical_ir=IRConfig(metadata=IRMetadata()),
        dependencies=[
            DependencyRecord(
                source_path="firewall policy",
                source_field="srcaddr",
                reference="resolved",
                expected_type="address",
                result="RESOLVED",
            ),
            DependencyRecord(
                source_path="firewall policy",
                source_field="dstaddr",
                reference="missing",
                expected_type="address",
                result="UNRESOLVED",
            ),
        ],
    )
    extraction.canonical_ir.audit_entries.append(
        IRAuditEntry(
            id="warning",
            category="test",
            message="unresolved wording only",
            confidence=MigrationConfidence.MANUAL,
        )
    )

    workbook = load_workbook(
        io.BytesIO(IRExcelExporter(extraction.canonical_ir, extraction_result=extraction).generate())
    )

    summary = workbook["Summary"]
    summary_rows = {summary.cell(row, 1).value: summary.cell(row, 2).value for row in range(1, summary.max_row + 1)}
    assert summary_rows["Unresolved References"] == 1

    registry = workbook["Dependency Registry"]
    unresolved = workbook["Unresolved References"]
    assert registry.max_row == 5
    assert unresolved.max_row == 4
    assert registry.cell(4, 7).value == "RESOLVED"
    assert unresolved.cell(4, 7).value == "UNRESOLVED"


def test_ems_address_stays_resolvable_but_is_exported_as_an_address():
    extraction = extract_fortigate_config(
        '''
config firewall address
    edit "ems-address"
        set type dynamic
        set sub-type ems-tag
        set obj-tag "Test_Tag"
    next
end
config firewall addrgrp
    edit "real-group"
        set member "ems-address"
    next
end
config firewall policy
    edit 1
        set srcaddr "ems-address"
    next
end
'''
    )

    dependency = next(
        item for item in extraction.dependencies
        if item.reference == "ems-address"
    )
    assert dependency.result == "RESOLVED"

    workbook = load_workbook(
        io.BytesIO(IRExcelExporter(extraction.canonical_ir, extraction_result=extraction).generate())
    )
    addresses = workbook["Addresses"]
    address_names = {addresses.cell(row, 1).value for row in range(4, addresses.max_row + 1)}
    groups = workbook["Address Groups"]
    group_names = {groups.cell(row, 1).value for row in range(4, groups.max_row + 1)}
    assert "ems-address" in address_names
    assert "real-group" in group_names
    assert "ems-address" not in group_names
