import io

from openpyxl import load_workbook

from fwmigrate.parsers.fortigate.parser import parse_fortigate_config
from fwmigrate.parsers.fortigate.transformer import FGToIRTransformer
from fwmigrate.report.excel_exporter import IRExcelExporter


def test_timezone_and_dns_survive_parser_ir_and_excel() -> None:
    parsed = parse_fortigate_config('''config system global
    set hostname edge-fw
    set timezone 28
    set admin-sport 8443
end
config system dns
    set primary 192.0.2.53
    set secondary 198.51.100.53
    set protocol dot
end
''')
    assert parsed.system_global.timezone == "28"
    assert parsed.dns.primary == "192.0.2.53"
    assert parsed.dns.secondary == "198.51.100.53"
    assert parsed.dns.extra_settings == {"protocol": "dot"}

    ir = FGToIRTransformer(parsed).transform()
    assert ir.system_settings.timezone == "28"
    assert ir.system_settings.admin_https_port == 8443
    assert ir.dns_settings.primary == "192.0.2.53"

    workbook = load_workbook(io.BytesIO(IRExcelExporter(ir).generate()))
    assert workbook["System Settings"][4][0].value == "edge-fw"
    assert workbook["System Settings"][4][1].value == "28"
    assert workbook["DNS Settings"][4][0].value == "192.0.2.53"
    assert workbook["DNS Settings"][4][1].value == "198.51.100.53"


def test_system_global_preserves_safe_fallback_settings_across_layers() -> None:
    parsed = parse_fortigate_config('''config system global
    set hostname FGT01
    set timezone 28
    set admin-sport 8443
    set gui-theme mariner
    set admin-lockout-duration 60
    set admin-console-timeout 0
    set admin-password credential-sentinel
end
''')

    assert parsed.system_global.hostname == "FGT01"
    assert parsed.system_global.admin_sport == 8443
    assert parsed.system_global.timezone == "28"
    assert parsed.system_global.extra_settings == {
        "gui_theme": "mariner",
        "admin_lockout_duration": "60",
        "admin_console_timeout": "0",
        "admin_password": "[REDACTED]",
    }

    ir = FGToIRTransformer(parsed).transform()
    assert ir.system_settings.source_attributes == parsed.system_global.extra_settings

    workbook = load_workbook(io.BytesIO(IRExcelExporter(ir).generate()))
    additional_settings = workbook["System Settings"][4][3].value
    assert "gui-theme=mariner" in additional_settings
    assert "admin-lockout-duration=60" in additional_settings
    assert "credential-sentinel" not in additional_settings


def test_absent_or_partial_dns_does_not_create_fake_values() -> None:
    absent_ir = FGToIRTransformer(parse_fortigate_config("config system global\nend\n")).transform()
    assert absent_ir.dns_settings is None

    partial = parse_fortigate_config("config system dns\nset primary 203.0.113.53\nend\n")
    assert partial.dns.primary == "203.0.113.53"
    assert partial.dns.secondary is None
