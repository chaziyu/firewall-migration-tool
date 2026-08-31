import io

from openpyxl import load_workbook

from fwmigrate.parsers.fortigate.parser import parse_fortigate_config
from fwmigrate.parsers.fortigate.transformer import FGToIRTransformer
from fwmigrate.report.excel_exporter import IRExcelExporter


def _transform(config: str):
    return FGToIRTransformer(parse_fortigate_config(config)).transform()


def test_system_zone_has_system_type_and_source_path():
    ir = _transform(
        '''
config system interface
    edit "port1"
    next
end
config system zone
    edit "INTERNET"
        set interface "port1"
    next
end
'''
    )

    zone = ir.zones[0]
    assert zone.name == "INTERNET"
    assert zone.interfaces == ["port1"]
    assert zone.zone_type == "system"
    assert zone.source_path == "system zone"
    assert zone.source_context == "root"


def test_sdwan_zone_has_sdwan_type_and_member_relationship_membership():
    ir = _transform(
        '''
config system interface
    edit "wan1"
    next
end
config system sdwan
    config zone
        edit "WAN"
        next
    end
    config members
        edit 1
            set interface "wan1"
            set zone "WAN"
        next
    end
end
'''
    )

    zone = ir.zones[0]
    assert zone.name == "WAN"
    assert zone.interfaces == ["wan1"]
    assert zone.zone_type == "sdwan"
    assert zone.source_path == "system sdwan zone"
    assert zone.source_context == "root"


def test_same_zone_name_keeps_system_and_sdwan_records_distinct():
    ir = _transform(
        '''
config system interface
    edit "lan1"
    next
    edit "wan1"
    next
end
config system zone
    edit "SHARED"
        set interface "lan1"
    next
end
config system sdwan
    config members
        edit 1
            set interface "wan1"
            set zone "SHARED"
        next
    end
end
'''
    )

    records = {
        (zone.source_context, zone.zone_type, zone.name): zone
        for zone in ir.zones
    }
    assert set(records) == {
        ("root", "system", "SHARED"),
        ("root", "sdwan", "SHARED"),
    }
    assert records["root", "system", "SHARED"].interfaces == ["lan1"]
    assert records["root", "sdwan", "SHARED"].interfaces == ["wan1"]


def test_same_zone_name_in_different_vdoms_keeps_separate_rows():
    ir = _transform(
        '''
config system interface
    edit "root-port"
    next
end
config system zone
    edit "INTERNET"
        set interface "root-port"
    next
end
config vdom
    edit "tenant-a"
        config system interface
            edit "tenant-port"
            next
        end
        config system zone
            edit "INTERNET"
                set interface "tenant-port"
            next
        end
    next
end
'''
    )

    records = {
        (zone.source_context, zone.zone_type, zone.name): zone
        for zone in ir.zones
    }
    assert set(records) == {
        ("root", "system", "INTERNET"),
        ("tenant-a", "system", "INTERNET"),
    }
    assert records["root", "system", "INTERNET"].interfaces == ["root-port"]
    assert records["tenant-a", "system", "INTERNET"].interfaces == ["tenant-port"]


def test_zones_excel_sheet_exposes_vdom_type_source_path_and_no_duplicates():
    ir = _transform(
        '''
config system interface
    edit "lan1"
    next
    edit "wan1"
    next
end
config system zone
    edit "INTERNET"
        set interface "lan1"
    next
end
config system sdwan
    config zone
        edit "virtual-wan-link"
        next
    end
    config members
        edit 1
            set interface "wan1"
        next
    end
end
'''
    )
    workbook = load_workbook(io.BytesIO(IRExcelExporter(ir).generate()))
    sheet = workbook["Zones"]

    headers = [cell.value for cell in sheet[3]]
    assert headers == [
        "VDOM",
        "Name",
        "Zone Type",
        "Members",
        "Description",
        "Source Path",
        "Manual Review",
        "Additional Settings",
    ]

    rows = list(sheet.iter_rows(min_row=4, values_only=True))
    assert len(rows) == len(ir.zones) == 2
    by_name = {row[1]: row for row in rows}
    assert by_name["virtual-wan-link"][0] == "root"
    assert by_name["virtual-wan-link"][2] == "sdwan"
    assert by_name["virtual-wan-link"][3] == "wan1"
    assert by_name["virtual-wan-link"][5] == "system sdwan zone"
    assert by_name["INTERNET"][2] == "system"
    assert by_name["INTERNET"][5] == "system zone"
