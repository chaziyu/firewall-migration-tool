import io

from openpyxl import load_workbook

from fwmigrate.extraction.models import ExtractionStatus
from fwmigrate.parsers.fortigate.extractor import extract_fortigate_config
from fwmigrate.parsers.fortigate.parser import FortiGateParser
from fwmigrate.parsers.fortigate.tokenizer import FortiGateTokenizer
from fwmigrate.report.excel_exporter import IRExcelExporter


SECURITY_CONFIG = """
config application list
    edit "high-security"
        set comment "Preserve tree"
        unset other-application-action
        append options allow-dns
        config entries
            edit 1
                set application 100 200
                config parameters
                    edit "risk"
                        set value high
                    next
                end
            next
        end
    next
end
config antivirus settings
    set machine-learning-detection enable
end
"""


def test_structured_security_tree_preserves_operations_and_hierarchy():
    parser = FortiGateParser(FortiGateTokenizer(SECURITY_CONFIG))
    parser.parse()
    profiles = parser.structured_source_objects

    application = next(item for item in profiles if item.source_path == "application list")
    assert application.name == "high-security"
    assert [(item.operation, item.key, item.values) for item in application.root.commands] == [
        ("set", "comment", ["Preserve tree"]),
        ("unset", "other-application-action", []),
        ("append", "options", ["allow-dns"]),
    ]
    entries = application.root.children[0]
    assert (entries.node_type, entries.name) == ("config", "entries")
    entry = entries.children[0]
    assert (entry.node_type, entry.name) == ("edit", "1")
    assert entry.commands[0].values == ["100", "200"]
    parameters = entry.children[0]
    assert parameters.name == "parameters"
    assert parameters.children[0].name == "risk"

    result = extract_fortigate_config(SECURITY_CONFIG)
    assert not hasattr(result.canonical_ir, "structured_security_profiles")
    structured = [
        item for item in result.inventory_items
        if "structured-security-profile" in item.notes
    ]
    assert len(structured) == 2
    assert next(item for item in structured if item.source_path == "application list").status == ExtractionStatus.NORMALIZED
    assert next(item for item in structured if item.source_path == "antivirus settings").status == ExtractionStatus.EXTRACT_ONLY
    statuses = {item.path: item.status for item in result.source_sections}
    assert statuses["application list entries"] == ExtractionStatus.EXTRACT_ONLY
    assert statuses["application list entries parameters"] == ExtractionStatus.EXTRACT_ONLY

    workbook = load_workbook(
        io.BytesIO(IRExcelExporter(result.canonical_ir, result).generate())
    )
    assert "Source Security Profiles" in workbook.sheetnames
    # Excel worksheet titles are limited to 31 characters.
    assert "Source Security Profile Setting" in workbook.sheetnames
    settings = workbook["Source Security Profile Setting"]
    headers = {cell.value: cell.column for cell in settings[3]}
    operations = {
        settings.cell(row, headers["Operation"]).value
        for row in range(4, settings.max_row + 1)
    }
    assert {"set", "unset", "append"} <= operations
    subsection_values = {
        settings.cell(row, headers["Subsection"]).value
        for row in range(4, settings.max_row + 1)
    }
    assert "entries / parameters" in subsection_values
