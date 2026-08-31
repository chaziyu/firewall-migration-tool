import io

from openpyxl import load_workbook

from fwmigrate.parsers.fortigate.parser import parse_fortigate_config
from fwmigrate.parsers.fortigate.transformer import FGToIRTransformer
from fwmigrate.report.excel_exporter import IRExcelExporter


INTERFACE_MULTIVALUE_CONFIG = """
config system interface
    edit "port1"
        set ip 10.0.0.1 255.255.255.0
        set secondary-IP enable
        set allowaccess ping https ssh
        set security-groups "group1" "group2"
        set fail-alert-interfaces "port2" "port3"
        set fail-detect-option "link-down" "link-up"
        set dns-server-protocol cleartext dot
        set member "member-one" "member-two"
        config secondaryip
            edit 1
                set ip 10.0.0.2 255.255.255.0
                set detectprotocol ping https
            next
        end
    next
end
"""


def _interface(config, name="port1"):
    return next(item for item in config.interfaces if item.name == name)


def test_system_interface_multivalue_fields_preserve_lists_and_order():
    interface = _interface(parse_fortigate_config(INTERFACE_MULTIVALUE_CONFIG))

    assert interface.security_groups == ["group1", "group2"]
    assert interface.fail_alert_interfaces == ["port2", "port3"]
    assert interface.fail_detect_option == ["link-down", "link-up"]
    assert interface.dns_server_protocol == ["cleartext", "dot"]
    assert interface.allowaccess == ["ping", "https", "ssh"]
    assert interface.members == ["member-one", "member-two"]


def test_secondaryip_detectprotocol_preserves_list_in_extra_settings():
    interface = _interface(parse_fortigate_config(INTERFACE_MULTIVALUE_CONFIG))

    assert interface.secondary_ips[0].extra_settings["detectprotocol"] == [
        "ping",
        "https",
    ]


def test_single_interface_values_remain_one_item_lists():
    config = """
config system interface
    edit "port1"
        set security-groups "group1"
        set fail-alert-interfaces "port2"
        set fail-detect-option "link-down"
        set dns-server-protocol cleartext
        config secondaryip
            edit 1
                set ip 10.0.0.2 255.255.255.0
                set detectprotocol ping
            next
        end
    next
end
"""
    interface = _interface(parse_fortigate_config(config))

    assert interface.security_groups == ["group1"]
    assert interface.fail_alert_interfaces == ["port2"]
    assert interface.fail_detect_option == ["link-down"]
    assert interface.dns_server_protocol == ["cleartext"]
    assert interface.secondary_ips[0].extra_settings["detectprotocol"] == ["ping"]


def test_quoted_interface_value_with_spaces_is_one_list_item():
    config = """
config system interface
    edit "port1"
        set security-groups "Group One"
        config secondaryip
            edit 1
                set ip 10.0.0.2 255.255.255.0
                set detectprotocol "Protocol One"
            next
        end
    next
end
"""
    interface = _interface(parse_fortigate_config(config))

    assert interface.security_groups == ["Group One"]
    assert interface.secondary_ips[0].extra_settings["detectprotocol"] == [
        "Protocol One"
    ]


def test_omitted_interface_multivalue_fields_use_empty_lists():
    config = """
config system interface
    edit "port1"
        set ip 10.0.0.1 255.255.255.0
    next
end
"""
    interface = _interface(parse_fortigate_config(config))

    assert interface.security_groups == []
    assert interface.fail_alert_interfaces == []
    assert interface.fail_detect_option == []
    assert interface.dns_server_protocol == []
    assert interface.secondary_ips == []


def test_interface_multivalue_fields_stay_lists_through_ir():
    ir_interface = FGToIRTransformer(
        parse_fortigate_config(INTERFACE_MULTIVALUE_CONFIG)
    ).transform().interfaces[0]

    assert ir_interface.source_attributes["security_groups"] == ["group1", "group2"]
    assert ir_interface.source_attributes["fail_alert_interfaces"] == ["port2", "port3"]
    assert ir_interface.source_attributes["fail_detect_option"] == ["link-down", "link-up"]
    assert ir_interface.source_attributes["dns_server_protocol"] == ["cleartext", "dot"]
    assert ir_interface.members == ["member-one", "member-two"]
    assert ir_interface.management_access == ["ping", "https", "ssh"]
    assert ir_interface.secondary_ips[0].source_attributes["detectprotocol"] == [
        "ping",
        "https",
    ]


def test_interface_multivalue_fields_render_as_ordered_json_arrays_in_excel():
    ir = FGToIRTransformer(parse_fortigate_config(INTERFACE_MULTIVALUE_CONFIG)).transform()
    workbook = load_workbook(
        io.BytesIO(IRExcelExporter(ir).generate()),
        data_only=False,
    )

    sheet = workbook["Interface Source Settings"]
    rows = list(sheet.iter_rows(min_row=4, values_only=True))
    values_by_setting = {row[2]: row[3] for row in rows}

    assert values_by_setting["security-groups"] == '["group1", "group2"]'
    assert values_by_setting["fail-alert-interfaces"] == '["port2", "port3"]'
    assert values_by_setting["fail-detect-option"] == '["link-down", "link-up"]'
    assert values_by_setting["dns-server-protocol"] == '["cleartext", "dot"]'

    secondary_sheet = workbook["Interface Secondary IPs"]
    secondary_rows = list(secondary_sheet.iter_rows(min_row=4, values_only=True))
    assert secondary_rows[0][9] == "detectprotocol=ping https"
