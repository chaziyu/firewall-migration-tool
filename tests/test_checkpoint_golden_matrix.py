import io
import pytest
from openpyxl import load_workbook

import fwmigrate.generators
import fwmigrate.parsers
from fwmigrate.core.registry import PluginRegistry
from fwmigrate.ir.enums import AddressType, NATTranslationMode, NATType, PolicyAction
from fwmigrate.report.excel_exporter import IRExcelExporter
from tests.fixture_paths import CHECKPOINT_GOLDEN_FIXTURE


def test_checkpoint_r81_golden_matrix_extraction():
    content = CHECKPOINT_GOLDEN_FIXTURE.read_text(encoding="utf-8")
    parser = PluginRegistry.get_parser("checkpoint")
    extraction = parser.extract(content)
    ir = extraction.canonical_ir

    # 1. Metadata and Hostname
    assert ir.metadata.hostname == "CP-Cluster-GW"
    assert ir.metadata.source_vendor == "checkpoint"

    # 2. Topology
    assert len(ir.interfaces) == 3
    assert len(ir.zones) == 3
    zone_names = {z.name for z in ir.zones}
    assert zone_names == {"Trust", "Untrust", "DMZ"}
    assert len(ir.routes) == 2

    # 3. Addresses and Groups
    assert len(ir.addresses) == 6  # 3 hosts, 2 networks, 1 range
    assert any(a.name == "Web_Server_01" and a.type == AddressType.HOST for a in ir.addresses)
    assert any(a.name == "Corp_LAN_Net" and a.type == AddressType.NETWORK for a in ir.addresses)
    assert any(a.name == "Egress_NAT_Pool" and a.type == AddressType.RANGE for a in ir.addresses)
    assert len(ir.address_groups) == 2
    internal_group = next(g for g in ir.address_groups if g.name == "Internal_Servers")
    assert internal_group.members == ["Web_Server_01", "DB_Server_01"]
    exclusion_group = next(g for g in ir.address_groups if g.name == "Corp_Except_DMZ")
    assert exclusion_group.exclusion_enabled
    assert exclusion_group.requires_manual_review

    # 4. Services and Schedules
    assert len(ir.services) == 4  # http, https, dns, ping
    assert len(ir.service_groups) == 1
    assert ir.service_groups[0].name == "Web_Services"
    assert len(ir.schedules) == 1
    assert ir.schedules[0].name == "Business_Hours"

    # 5. Access Policies
    assert len(ir.policies) == 3
    p_https = next(p for p in ir.policies if p.name == "Inbound_HTTPS")
    assert p_https.action == PolicyAction.ALLOW
    assert not p_https.safe_for_target_generation
    assert any("Web_Services" in reason for reason in p_https.review_reasons)

    p_lan = next(p for p in ir.policies if p.name == "LAN_To_DMZ")
    assert p_lan.action == PolicyAction.ALLOW
    assert p_lan.schedule == "Business_Hours"
    assert not p_lan.safe_for_target_generation

    assert all(p.name != "Interactive_Auth_Prompt" for p in ir.policies)
    ask_item = next(i for i in extraction.inventory_items if i.name == "Interactive_Auth_Prompt")
    assert ask_item.status.value == "PARTIALLY_NORMALIZED"
    assert ask_item.requires_manual_review

    p_cleanup = next(p for p in ir.policies if p.name == "Cleanup_Drop")
    assert p_cleanup.action == PolicyAction.DROP
    assert p_cleanup.safe_for_target_generation
    assert all(policy.source_extra_settings["vpn"] == "Any" for policy in ir.policies)

    # 6. NAT Rules
    assert len(ir.nat_rules) == 2
    dnat = next(n for n in ir.nat_rules if n.name == "DNAT_Web_Server")
    assert dnat.type == NATType.DESTINATION
    assert dnat.translated_destinations == ["Web_Server_01"]
    assert not dnat.safe_for_target_generation
    assert any("Web_Services" in reason for reason in dnat.review_reasons)

    snat = next(n for n in ir.nat_rules if n.name == "LAN_Hide_NAT")
    assert snat.type == NATType.SOURCE
    assert snat.translated_sources == ["Egress_NAT_Pool"]
    assert snat.source_translation_mode == NATTranslationMode.DYNAMIC_IP_AND_PORT
    assert snat.safe_for_target_generation

    # 7. Unsupported Items and Accounting
    unsupp_names = {u.source_name for u in extraction.unsupported_items}
    assert "Corp_Except_DMZ" in unsupp_names
    assert "Interactive_Auth_Prompt" in unsupp_names


def test_checkpoint_golden_matrix_cross_vendor_generation():
    content = CHECKPOINT_GOLDEN_FIXTURE.read_text(encoding="utf-8")
    parser = PluginRegistry.get_parser("checkpoint")
    extraction = parser.extract(content)
    ir = extraction.canonical_ir

    # Palo Alto generation
    pa_gen = PluginRegistry.get_generator("palo_alto")
    pa_artifacts = pa_gen.generate(ir, format="xml")
    assert len(pa_artifacts) >= 1
    pa_xml = "\n".join(art.content for art in pa_artifacts)
    assert "Inbound_HTTPS" not in pa_xml
    assert "LAN_To_DMZ" not in pa_xml
    assert "Interactive_Auth_Prompt" not in pa_xml

    # FortiGate CLI generation
    fg_gen = PluginRegistry.get_generator("fortigate")
    fg_artifacts = fg_gen.generate(ir, format="cli")
    assert len(fg_artifacts) >= 1
    fg_cli = "\n".join(art.content for art in fg_artifacts)
    assert "Policy Inbound_HTTPS withheld" in fg_cli
    assert "Policy LAN_To_DMZ withheld" in fg_cli
    assert 'set name "Interactive_Auth_Prompt"' not in fg_cli

    # Cisco ASA CLI generation
    asa_gen = PluginRegistry.get_generator("cisco_asa")
    asa_artifacts = asa_gen.generate(ir, format="cli")
    assert len(asa_artifacts) >= 1
    asa_cli = "\n".join(art.content for art in asa_artifacts)
    assert "access-list" in asa_cli

    # Excel export
    exporter = IRExcelExporter(ir, extraction_result=extraction)
    excel_bytes = exporter.generate()
    wb = load_workbook(io.BytesIO(excel_bytes), read_only=True)
    assert "Source Inventory" in wb.sheetnames
    assert "Extraction Coverage" in wb.sheetnames
