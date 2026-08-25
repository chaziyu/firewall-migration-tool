import io

from openpyxl import load_workbook

from fwmigrate.extraction.models import ExtractionStatus
from fwmigrate.parsers.fortigate.extractor import extract_fortigate_config
from fwmigrate.parsers.fortigate.parser import FortiGateParser
from fwmigrate.parsers.fortigate.tokenizer import FortiGateTokenizer
from fwmigrate.report.excel_exporter import IRExcelExporter


ROUTING_CONFIG = '''config router bgp
    set as 65001
    unset router-id
    append network-import-check enable
    config neighbor
        edit "192.0.2.2"
            set remote-as 65002
            config conditional-advertise
                edit "ADV-1"
                    set condition-route-map "CHECK"
                next
            end
        next
    end
end
config router ospf
    set router-id 192.0.2.1
    config area
        edit "0.0.0.0"
            set type regular
        next
    end
end
'''


def test_dynamic_routing_uses_recursive_extract_only_source_tree() -> None:
    parser = FortiGateParser(FortiGateTokenizer(ROUTING_CONFIG))
    parsed = parser.parse()
    routing = [
        item for item in parser.structured_source_objects
        if item.source_path.startswith("router ")
    ]
    assert [item.source_path for item in routing] == ["router bgp", "router ospf"]
    bgp = routing[0].root
    assert [(item.operation, item.key) for item in bgp.commands] == [
        ("set", "as"),
        ("unset", "router-id"),
        ("append", "network-import-check"),
    ]
    neighbor = bgp.children[0]
    assert neighbor.name == "neighbor"
    assert neighbor.children[0].name == "192.0.2.2"
    assert neighbor.children[0].children[0].children[0].name == "ADV-1"
    assert parsed.static_routes == []

    result = extract_fortigate_config(ROUTING_CONFIG)
    assert result.canonical_ir.routes == []
    assert all(
        section.status == ExtractionStatus.EXTRACT_ONLY
        for section in result.source_sections
    )
    items = [
        item for item in result.inventory_items
        if "structured-routing-protocol" in item.notes
    ]
    assert len(items) == 2
    assert all(item.requires_manual_review for item in items)

    workbook = load_workbook(
        io.BytesIO(IRExcelExporter(result.canonical_ir, result).generate())
    )
    assert "Routing Protocols" in workbook.sheetnames
    assert "Routing Protocol Settings" in workbook.sheetnames
    protocols = {workbook["Routing Protocols"].cell(row, 1).value for row in range(4, 6)}
    assert protocols == {"BGP", "OSPF"}
    operations = {
        workbook["Routing Protocol Settings"].cell(row, 5).value
        for row in range(4, workbook["Routing Protocol Settings"].max_row + 1)
    }
    assert {"set", "unset", "append"} <= operations
