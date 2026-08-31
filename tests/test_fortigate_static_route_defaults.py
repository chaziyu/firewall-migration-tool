import io

import pytest
from openpyxl import load_workbook

from fwmigrate.parsers.fortigate.extractor import extract_fortigate_config
from fwmigrate.parsers.fortigate.parser import parse_fortigate_config
from fwmigrate.parsers.fortigate.transformer import FGToIRTransformer
from fwmigrate.report.excel_exporter import IRExcelExporter


def _config(section: str, body: str) -> str:
    destination = (
        "2001:db8:1::/64"
        if section == "router static6"
        else "198.51.100.0 255.255.255.0"
    )
    return f"""
config {section}
    edit 22
        set dst {destination}
{body}
    next
end
"""


@pytest.mark.parametrize("section", ["router static", "router static6"])
def test_static_route_effective_defaults_apply_to_ipv4_and_ipv6(section):
    fg = parse_fortigate_config(_config(section, ""))
    route = fg.static_routes[0]

    assert (route.distance, route.priority, route.weight, route.status) == (
        10,
        1,
        0,
        "enable",
    )
    assert not {"distance", "priority", "weight", "status"}.intersection(
        route.source_explicit_fields
    )
    assert route.extra_settings == {}

    ir_route = FGToIRTransformer(fg).transform().routes[0]
    assert (
        ir_route.administrative_distance,
        ir_route.priority,
        ir_route.weight,
        ir_route.enabled,
    ) == (10, 1, 0, True)
    assert not {"distance", "priority", "weight", "status"}.intersection(
        ir_route.source_explicit_fields
    )


@pytest.mark.parametrize("section", ["router static", "router static6"])
def test_static_route_explicit_values_override_defaults(section):
    fg = parse_fortigate_config(
        _config(
            section,
            """        set distance 20
        set priority 5
        set weight 10
        set status disable""",
        )
    )
    route = fg.static_routes[0]

    assert (route.distance, route.priority, route.weight, route.status) == (
        20,
        5,
        10,
        "disable",
    )
    assert {"distance", "priority", "weight", "status"} <= set(
        route.source_explicit_fields
    )

    ir_route = FGToIRTransformer(fg).transform().routes[0]
    assert (
        ir_route.administrative_distance,
        ir_route.priority,
        ir_route.weight,
        ir_route.enabled,
    ) == (20, 5, 10, False)
    assert {"distance", "priority", "weight", "status"} <= set(
        ir_route.source_explicit_fields
    )
    assert ir_route.source_attributes == {}


def test_explicit_default_distance_is_distinguished_from_omitted_distance():
    fg = parse_fortigate_config(
        _config("router static", "        set distance 10")
    )
    route = fg.static_routes[0]

    assert route.distance == 10
    assert "distance" in route.source_explicit_fields

    ir_route = FGToIRTransformer(fg).transform().routes[0]
    assert ir_route.administrative_distance == 10
    assert "distance" in ir_route.source_explicit_fields


@pytest.mark.parametrize(
    ("status_line", "expected_enabled"),
    [("", True), ("set status enable", True), ("set status disable", False)],
)
def test_static_route_status_maps_to_enabled(status_line, expected_enabled):
    fg = parse_fortigate_config(_config("router static", f"        {status_line}"))
    route = FGToIRTransformer(fg).transform().routes[0]

    assert route.enabled is expected_enabled


def test_invalid_explicit_numeric_route_value_is_not_replaced_by_default():
    fg = parse_fortigate_config(
        _config("router static", "        set distance invalid")
    )
    route = fg.static_routes[0]

    assert route.distance is None
    assert "distance" in route.source_explicit_fields
    assert route.extra_settings["unparsed_distance"] == "invalid"


def test_static_route_effective_values_are_exported_without_synthetic_source_fields():
    result = extract_fortigate_config(_config("router static", ""))
    workbook = load_workbook(
        io.BytesIO(IRExcelExporter(result.canonical_ir, result).generate())
    )
    sheet = workbook["Routes"]
    headers = {cell.value: cell.column for cell in sheet[3]}

    assert sheet.cell(4, headers["Administrative Distance"]).value == 10
    assert sheet.cell(4, headers["Priority"]).value == 1
    assert sheet.cell(4, headers["Weight"]).value == 0
    assert sheet.cell(4, headers["Enabled"]).value == "Yes"
    assert sheet.cell(4, headers["Additional Settings"]).value is None
