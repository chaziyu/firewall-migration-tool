import io

from openpyxl import load_workbook

from fwmigrate.parsers.palo_alto.parser import PANOSSourceParser
from fwmigrate.report.excel_exporter import IRExcelExporter


def _extract(body: str):
    return PANOSSourceParser().extract(f"<config version='11.1.0'>{body}</config>")


def _nat_rule(name: str, translated: str, to_interface: str) -> str:
    return f"""
      <entry name='{name}'>
        <from><member>trust</member></from><to><member>untrust</member></to>
        <source><member>any</member></source><destination><member>any</member></destination>
        <service>any</service><to-interface>{to_interface}</to-interface>
        <destination-translation><translated-address>{translated}</translated-address></destination-translation>
      </entry>
    """


def test_pan_nat_translation_order_and_to_interface_are_preserved_in_ir():
    result = _extract(f"""
      <devices><entry name='fw-a'><vsys><entry name='vsys1'>
        <address><entry name='AFC-TPW2_10.128.62.10'>
          <ip-netmask>10.128.62.10/32</ip-netmask>
        </entry></address>
        <rulebase><nat><rules>
          {_nat_rule('object', 'AFC-TPW2_10.128.62.10', 'ethernet1/3')}
          {_nat_rule('unresolved', 'missing-object-name', 'any')}
          {_nat_rule('host', '10.1.1.10', 'ethernet1/4')}
          {_nat_rule('range', '10.1.1.10-10.1.1.20', 'ethernet1/5')}
        </rules></nat></rulebase>
      </entry></vsys></entry></devices>
    """)

    rules = {rule.name: rule for rule in result.canonical_ir.nat_rules}
    assert rules['object'].translated_destinations == ['AFC-TPW2_10.128.62.10']
    assert rules['object'].source_to_interfaces == ['ethernet1/3']
    assert rules['object'].source_attributes['pan_translated_destination_values'][0]['classification'] == 'object-reference'

    assert rules['unresolved'].translated_destinations == ['missing-object-name']
    assert 'unresolved-translated-destination' in rules['unresolved'].review_reasons
    assert rules['unresolved'].source_attributes['pan_translated_destination_values'][0]['classification'] == 'unresolved-reference'
    assert rules['unresolved'].source_to_interfaces == ['any']

    assert rules['host'].translated_destinations == ['10.1.1.10']
    assert rules['range'].translated_destinations == ['10.1.1.10-10.1.1.20']


def test_pan_nat_to_interface_is_exported_to_destination_interface():
    result = _extract(f"""
      <devices><entry name='fw-a'><vsys><entry name='vsys1'><rulebase><nat><rules>
        {_nat_rule('excel-nat', '10.1.1.10', 'ethernet1/3')}
      </rules></nat></rulebase></entry></vsys></entry></devices>
    """)

    workbook = load_workbook(io.BytesIO(IRExcelExporter(result.canonical_ir).generate()))
    sheet = workbook['NAT Rules']
    headers = {cell.value: cell.column for cell in sheet[3]}
    assert sheet.cell(4, headers['Destination Interface']).value == 'ethernet1/3'


def test_pan_nat_effective_order_is_separate_from_source_order():
    result = _extract("""<devices><entry name='fw-a'><vsys><entry name='vsys1'><pre-rulebase><nat><rules>
      <entry name='pre'/></rules></nat></pre-rulebase><rulebase><nat><rules>
      <entry name='local'/></rules></nat></rulebase><post-rulebase><nat><rules>
      <entry name='post'/></rules></nat></post-rulebase></entry></vsys></entry></devices>""")
    # Invalid match fields keep these as inventory records, which is enough to
    # verify ordering evidence without inventing canonical NAT semantics.
    items = {item.name: item for item in result.inventory_items if item.domain == "nat"}
    assert items["pre"].source_attributes["effective_policy_rank"] == 0
    assert items["local"].source_attributes["effective_policy_rank"] == 1
    assert items["post"].source_attributes["effective_policy_rank"] == 2
