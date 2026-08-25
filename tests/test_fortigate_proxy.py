import io

from openpyxl import load_workbook

from fwmigrate.extraction.models import ExtractionStatus
from fwmigrate.parsers.fortigate.extractor import extract_fortigate_config
from fwmigrate.parsers.fortigate.parser import parse_fortigate_config
from fwmigrate.parsers.fortigate.transformer import FGToIRTransformer
from fwmigrate.report.excel_exporter import IRExcelExporter


PROXY_CONFIG = r"""
config firewall proxy-address
    edit "proxy-regex"
        set uuid "proxy-uuid-1"
        set type host-regex
        set host "www.example.com"
        set host-regex "^([a-z]+\.)?example\.com$"
        set path "/downloads/.*"
        set query "version=[0-9]+"
        set case-sensitivity disable
    next
end
config web-proxy global
    set proxy-fqdn "proxy.example.com"
    set strict-web-check enable
end
"""


def test_proxy_addresses_remain_exact_extract_only_inventory():
    result = extract_fortigate_config(PROXY_CONFIG)
    parsed = parse_fortigate_config(PROXY_CONFIG).proxy_addresses[0]
    proxy = result.canonical_ir.proxy_addresses[0]

    assert parsed.host_regex == r"^([a-z]+\.)?example\.com$"
    assert proxy.name == "proxy-regex"
    assert proxy.source_uuid == "proxy-uuid-1"
    assert proxy.proxy_address_type == "host-regex"
    assert proxy.host == "www.example.com"
    assert proxy.host_regex == r"^([a-z]+\.)?example\.com$"
    assert proxy.path == "/downloads/.*"
    assert proxy.query == "version=[0-9]+"
    assert proxy.source_attributes == {"case_sensitivity": "disable"}
    assert proxy.migration_status == "EXTRACT_ONLY"
    assert proxy.requires_manual_review is True
    assert all(address.name != proxy.name for address in result.canonical_ir.addresses)

    section = next(
        item for item in result.source_sections
        if item.path == "firewall proxy-address"
    )
    assert section.status == ExtractionStatus.EXTRACT_ONLY


def test_web_proxy_global_is_set_based_and_absent_section_creates_no_object():
    result = extract_fortigate_config(PROXY_CONFIG)
    source = parse_fortigate_config(PROXY_CONFIG).web_proxy_global
    settings = result.canonical_ir.web_proxy_settings

    assert source is not None
    assert source.proxy_fqdn == "proxy.example.com"
    assert source.extra_settings == {"strict_web_check": "enable"}
    assert settings is not None
    assert settings.proxy_fqdn == "proxy.example.com"
    assert settings.source_attributes == {"strict_web_check": "enable"}
    assert settings.migration_status == "EXTRACT_ONLY"
    assert settings.requires_manual_review is True
    global_section = next(
        item for item in result.source_sections
        if item.path == "web-proxy global"
    )
    assert global_section.status == ExtractionStatus.EXTRACT_ONLY

    absent_ir = FGToIRTransformer(parse_fortigate_config("")).transform()
    assert absent_ir.web_proxy_settings is None

    empty_result = extract_fortigate_config("config web-proxy global\nend\n")
    assert empty_result.canonical_ir.web_proxy_settings is None
    empty_section = next(
        item for item in empty_result.source_sections
        if item.path == "web-proxy global"
    )
    assert empty_section.status == ExtractionStatus.EXTRACT_ONLY


def test_proxy_excel_preserves_regex_and_does_not_add_blank_global_row():
    result = extract_fortigate_config(PROXY_CONFIG)
    workbook = load_workbook(
        io.BytesIO(
            IRExcelExporter(
                result.canonical_ir,
                extraction_result=result,
            ).generate()
        )
    )
    proxy_sheet = workbook["Proxy Addresses"]
    proxy_headers = {cell.value: cell.column for cell in proxy_sheet[3]}
    assert proxy_sheet.cell(4, proxy_headers["Host Regex"]).value == r"^([a-z]+\.)?example\.com$"
    assert proxy_sheet.cell(4, proxy_headers["Additional Settings"]).value == "case-sensitivity=disable"

    global_sheet = workbook["Web Proxy Settings"]
    assert global_sheet.max_row == 4
    assert global_sheet["A4"].value == "proxy.example.com"

    empty_ir = FGToIRTransformer(parse_fortigate_config("")).transform()
    empty_workbook = load_workbook(io.BytesIO(IRExcelExporter(empty_ir).generate()))
    assert empty_workbook["Web Proxy Settings"].max_row == 3
