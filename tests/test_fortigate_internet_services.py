from io import BytesIO

from openpyxl import load_workbook

from fwmigrate.parsers.fortigate.parser import (
    parse_fortigate_config,
)
from fwmigrate.parsers.fortigate.extractor import (
    extract_fortigate_config,
)
from fwmigrate.parsers.fortigate.transformer import (
    FGToIRTransformer,
)
from fwmigrate.report.excel_exporter import (
    IRExcelExporter,
)


FORTIGATE_CONFIG = """
config system global
    set hostname "FG-TEST"
end

config firewall internet-service-name
    edit "Google-Other"
        set internet-service-id 65536
    next
    edit "Microsoft-Office365"
        set internet-service-id 327782
    next
end
"""


PRESERVATION_CONFIG = """
config firewall internet-service-name
    edit "Custom-Service"
        set internet-service-id 12345
        set comment "Preserved description"
        set custom-setting enable
        set vendor-flag test
        set vendor-password do-not-retain
    next
end
"""


def test_fortigate_internet_service_source_id_is_parsed():
    fg = parse_fortigate_config(
        FORTIGATE_CONFIG
    )

    assert len(fg.internet_services) == 2

    google = fg.internet_services[0]

    assert google.name == "Google-Other"
    assert google.id == 65536
    assert google.extra_settings == {}

    microsoft = fg.internet_services[1]

    assert (
        microsoft.name
        == "Microsoft-Office365"
    )
    assert microsoft.id == 327782
    assert microsoft.extra_settings == {}


def test_fortigate_internet_service_preserves_unknown_safe_settings():
    service = parse_fortigate_config(
        PRESERVATION_CONFIG
    ).internet_services[0]

    assert service.name == "Custom-Service"
    assert service.id == 12345
    assert service.comment == "Preserved description"
    assert service.extra_settings == {
        "custom_setting": "enable",
        "vendor_flag": "test",
        "vendor_password": "[REDACTED]",
    }
    assert service.model_dump()["extra_settings"] == service.extra_settings
    assert "do-not-retain" not in service.model_dump_json()


def test_fortigate_internet_service_preserves_malformed_source_id():
    service = parse_fortigate_config("""
config firewall internet-service-name
    edit "Broken-Service"
        set internet-service-id abc
        set custom-setting enable
    next
end
""").internet_services[0]

    assert service.name == "Broken-Service"
    assert service.id is None
    assert service.extra_settings == {
        "custom_setting": "enable",
        "unparsed_internet_service_id": "abc",
    }


def test_fortigate_internet_service_source_id_reaches_ir():
    fg = parse_fortigate_config(
        FORTIGATE_CONFIG
    )

    ir = FGToIRTransformer(
        fg
    ).transform()

    assert len(ir.internet_services) == 2

    google = ir.internet_services[0]

    assert google.name == "Google-Other"
    assert google.source_id == 65536

    microsoft = ir.internet_services[1]

    assert (
        microsoft.name
        == "Microsoft-Office365"
    )
    assert microsoft.source_id == 327782


def test_fortigate_internet_service_source_attributes_reach_ir():
    ir = FGToIRTransformer(
        parse_fortigate_config(PRESERVATION_CONFIG)
    ).transform()

    service = ir.internet_services[0]

    assert service.name == "Custom-Service"
    assert service.source_id == 12345
    assert service.description == "Preserved description"
    assert service.source_attributes == {
        "custom_setting": "enable",
        "vendor_flag": "test",
        "vendor_password": "[REDACTED]",
    }


def test_fortigate_internet_service_extraction_accounts_for_all_values():
    result = extract_fortigate_config(PRESERVATION_CONFIG)
    section = result.source_sections[0]
    service = result.canonical_ir.internet_services[0]

    assert section.path == "firewall internet-service-name"
    assert section.object_count_source == 1
    assert section.object_count_parsed == 1
    assert section.object_count_normalized == 1
    assert service.source_attributes == {
        "custom_setting": "enable",
        "vendor_flag": "test",
        "vendor_password": "[REDACTED]",
    }


def test_fortigate_internet_service_source_id_is_exported_to_excel():
    fg = parse_fortigate_config(
        FORTIGATE_CONFIG
    )

    ir = FGToIRTransformer(
        fg
    ).transform()

    workbook_bytes = IRExcelExporter(
        ir
    ).generate()

    workbook = load_workbook(
        BytesIO(workbook_bytes)
    )

    assert "Internet Services" in workbook.sheetnames

    sheet = workbook["Internet Services"]

    headers = [
        cell.value
        for cell in sheet[3]
    ]

    assert "Name" in headers
    assert "Source Vendor" in headers
    assert "Source ID" in headers
    assert "Description" in headers
    assert "Additional Settings" in headers

    header_index = {
        name: index + 1
        for index, name in enumerate(
            headers
        )
    }

    rows = {}

    for row_number in range(
        4,
        sheet.max_row + 1,
    ):
        name = sheet.cell(
            row=row_number,
            column=header_index["Name"],
        ).value

        if name:
            rows[name] = {
                "vendor": sheet.cell(
                    row=row_number,
                    column=header_index[
                        "Source Vendor"
                    ],
                ).value,
                "source_id": sheet.cell(
                    row=row_number,
                    column=header_index[
                        "Source ID"
                    ],
                ).value,
            }

    assert "Google-Other" in rows

    assert (
        rows["Google-Other"]["vendor"]
        == "fortigate"
    )

    assert (
        rows["Google-Other"]["source_id"]
        == 65536
    )

    assert "Microsoft-Office365" in rows

    assert (
        rows["Microsoft-Office365"][
            "source_id"
        ]
        == 327782
    )


def test_fortigate_internet_service_additional_settings_reach_excel():
    ir = FGToIRTransformer(
        parse_fortigate_config(PRESERVATION_CONFIG)
    ).transform()
    workbook = load_workbook(
        BytesIO(IRExcelExporter(ir).generate())
    )
    sheet = workbook["Internet Services"]
    headers = {
        cell.value: cell.column
        for cell in sheet[3]
    }

    assert sheet.cell(4, headers["Name"]).value == "Custom-Service"
    assert sheet.cell(4, headers["Source ID"]).value == 12345
    assert sheet.cell(4, headers["Description"]).value == "Preserved description"

    additional_settings = sheet.cell(
        4,
        headers["Additional Settings"],
    ).value
    assert "custom-setting=enable" in additional_settings
    assert "vendor-flag=test" in additional_settings
    assert "vendor-password=[REDACTED]" in additional_settings
    assert "do-not-retain" not in additional_settings
