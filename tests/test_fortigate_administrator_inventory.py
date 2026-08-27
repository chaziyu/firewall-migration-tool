import io

from openpyxl import load_workbook

from fwmigrate.parsers.fortigate.extractor import extract_fortigate_config
from fwmigrate.parsers.fortigate.parser import parse_fortigate_config
from fwmigrate.parsers.fortigate.transformer import FGToIRTransformer
from fwmigrate.report.excel_exporter import IRExcelExporter


ADMIN_SECRET = "ADMIN_PASSWORD_SENTINEL"
TOKEN_SEED = "FORTITOKEN_SEED_SENTINEL"
ACTIVATION_SECRET = "FORTITOKEN_ACTIVATION_SENTINEL"

ADMIN_CONFIG = f'''config system admin
    edit "review-admin"
        set accprofile "auditor"
        set vdom "root" "customer-a"
        set trusthost1 192.0.2.0 255.255.255.0
        set trusthost2 198.51.100.10 255.255.255.255
        set two-factor fortitoken
        set fortitoken "FTKMOB000000001"
        set email-to "reviewer@example.test"
        set remote-auth enable
        set remote-group "remote-admins"
        set passwd "{ADMIN_SECRET}"
        set api-key "{ADMIN_SECRET}"
        set custom-admin-setting preserve-me
    next
end
config system accprofile
    edit "auditor"
        set secfabgrp read
        set fwgrp read
        set custom-permission enabled
    next
end
config user fortitoken
    edit "FTKMOB000000001"
        set status active
        set comments "Assigned review token"
        set user "review-admin"
        set seed "{TOKEN_SEED}"
        set activation-code "{ACTIVATION_SECRET}"
        set vendor-sku "mobile"
    next
end
'''


def test_administrator_inventory_survives_parser_ir_coverage_and_excel() -> None:
    fg = parse_fortigate_config(ADMIN_CONFIG)

    administrator = fg.administrators[0]
    assert administrator.name == "review-admin"
    assert administrator.accprofile == "auditor"
    assert administrator.vdom == ["root", "customer-a"]
    assert administrator.trusthost1 == "192.0.2.0 255.255.255.0"
    assert administrator.two_factor == "fortitoken"
    assert administrator.fortitoken == "FTKMOB000000001"
    assert administrator.remote_auth == "enable"
    assert administrator.remote_group == "remote-admins"
    assert administrator.credential_configured is True
    assert administrator.extra_settings == {
        "custom_admin_setting": "preserve-me",
    }

    profile = fg.admin_profiles[0]
    assert profile.name == "auditor"
    assert profile.extra_settings == {
        "secfabgrp": "read",
        "fwgrp": "read",
        "custom_permission": "enabled",
    }

    token = fg.fortitokens[0]
    assert token.serial == "FTKMOB000000001"
    assert token.status == "active"
    assert token.comments == "Assigned review token"
    assert token.assigned_user == "review-admin"
    assert token.extra_settings == {"vendor_sku": "mobile"}

    result = extract_fortigate_config(ADMIN_CONFIG)
    ir = result.canonical_ir
    ir_admin = ir.administrators[0]
    assert ir_admin.access_profile == "auditor"
    assert ir_admin.vdoms == ["root", "customer-a"]
    assert ir_admin.token_reference == "FTKMOB000000001"
    assert ir_admin.credential_configured is True
    assert ir_admin.migration_status == "EXTRACT_ONLY"
    assert ir_admin.requires_manual_review is True
    assert ir.admin_profiles[0].source_attributes["custom_permission"] == "enabled"
    assert ir.fortitokens[0].assigned_user == "review-admin"

    coverage = {item.path: item for item in result.source_sections}
    for path in ("system admin", "system accprofile", "user fortitoken"):
        assert coverage[path].status.value == "EXTRACT_ONLY"
        assert coverage[path].object_count_source == 1
        assert coverage[path].object_count_parsed == 1
        assert coverage[path].object_count_normalized == 1

    workbook = load_workbook(
        io.BytesIO(IRExcelExporter(ir, extraction_result=result).generate())
    )
    for sheet_name in ("Administrators", "Admin Profiles", "FortiTokens"):
        assert sheet_name in workbook.sheetnames
        assert workbook[sheet_name].max_row == 4

    administrators = workbook["Administrators"]
    admin_headers = {cell.value: cell.column for cell in administrators[3]}
    assert administrators.cell(4, admin_headers["Name"]).value == "review-admin"
    assert administrators.cell(4, admin_headers["VDOMs"]).value == "root\ncustomer-a"
    assert administrators.cell(4, admin_headers["Credential Configured"]).value == "Yes"
    assert administrators.cell(4, admin_headers["Migration Status"]).value == "EXTRACT_ONLY"
    assert administrators.cell(4, admin_headers["Manual Review"]).value == "Yes"

    profiles = workbook["Admin Profiles"]
    profile_headers = {cell.value: cell.column for cell in profiles[3]}
    assert "custom-permission=enabled" in profiles.cell(
        4, profile_headers["Additional Settings"]
    ).value

    tokens = workbook["FortiTokens"]
    token_headers = {cell.value: cell.column for cell in tokens[3]}
    assert tokens.cell(4, token_headers["Serial / Name"]).value == "FTKMOB000000001"
    assert tokens.cell(4, token_headers["Assigned User"]).value == "review-admin"
    assert tokens.cell(4, token_headers["Migration Status"]).value == "EXTRACT_ONLY"

    serialized_layers = (
        fg.model_dump_json(),
        ir.model_dump_json(),
        result.model_dump_json(),
        "\n".join(
            str(cell.value)
            for sheet in workbook.worksheets
            for row in sheet.iter_rows()
            for cell in row
            if cell.value is not None
        ),
    )
    for serialized in serialized_layers:
        assert ADMIN_SECRET not in serialized
        assert TOKEN_SEED not in serialized
        assert ACTIVATION_SECRET not in serialized


def test_absent_administrator_inventory_stays_empty() -> None:
    ir = FGToIRTransformer(parse_fortigate_config("config system global\nend\n")).transform()
    assert ir.administrators == []
    assert ir.admin_profiles == []
    assert ir.fortitokens == []


def test_administrator_security_fields_and_accprofile_permissions_survive_excel() -> None:
    config = '''config system admin
    edit "guest-admin"
        set guest-usergroups "Guest Group A" "Guest Group B"
        set trusthost1 192.0.2.1 255.255.255.255
        set trusthost3 192.0.2.3 255.255.255.255
        set trusthost10 192.0.2.10 255.255.255.255
        set ip6-trusthost1 2001:db8::1/128
        set ip6-trusthost5 2001:db8::5/128
        set ip6-trusthost10 2001:db8::10/128
    next
end
config system accprofile
    edit "custom-admin"
        config fwgrp-permission
            set policy read
            set address read-write
            set future-permission foo
        end
        config loggrp-permission
            set data-access read-write
        end
    next
end
'''
    result = extract_fortigate_config(config)
    admin = result.canonical_ir.administrators[0]
    assert admin.guest_user_groups == ["Guest Group A", "Guest Group B"]
    assert admin.trusted_hosts_ipv4 == [
        "192.0.2.1 255.255.255.255", "192.0.2.3 255.255.255.255",
        "192.0.2.10 255.255.255.255",
    ]
    assert admin.trusted_hosts_ipv6 == ["2001:db8::1/128", "2001:db8::5/128", "2001:db8::10/128"]
    profile = result.canonical_ir.admin_profiles[0]
    assert profile.permission_blocks[0].name == "fwgrp-permission"
    assert profile.permission_blocks[0].source_attributes["future_permission"] == "foo"

    workbook = load_workbook(io.BytesIO(IRExcelExporter(
        result.canonical_ir, extraction_result=result
    ).generate()))
    sheet = workbook["Admin Profile Permissions"]
    rows = list(sheet.iter_rows(min_row=4, values_only=True))
    assert ("custom-admin", "fwgrp-permission", "policy", "read", "EXTRACT_ONLY", "future-permission=foo") in rows
    assert ("custom-admin", "fwgrp-permission", "future_permission", "foo", "EXTRACT_ONLY", "future-permission=foo") in rows
