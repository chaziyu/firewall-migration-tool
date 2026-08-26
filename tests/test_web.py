import io
import json
import pytest
import zipfile
from openpyxl import load_workbook
from pathlib import Path
from unittest.mock import patch, MagicMock

from fwmigrate.web import create_app, ACTIVE_SESSIONS
from fwmigrate.engine.diagnostics import DiagnosticResult


@pytest.fixture
def client():
    app = create_app({'TESTING': True})
    return app.test_client()


def test_index_page(client):
    response = client.get('/')
    assert response.status_code == 200
    assert b"Firewall Migration" in response.data
    assert b"Live Migration" in response.data
    assert b"Extract Data to Excel" in response.data
    assert response.data.index(b"Extract Data to Excel") < response.data.index(b"Convert Config File")
    assert response.data.index(b"Convert Config File") < response.data.index(b"Live Migration")
    assert b'id="tab-extract" class="tab-btn active"' in response.data
    assert b'id="tab-download" class="tab-btn"' in response.data
    assert b'id="mode-download-form" class="hidden"' in response.data
    assert b'id="mode-extract-form"' in response.data
    assert b'id="mode-extract-form" class="hidden"' not in response.data
    assert b'class="vendor-select-group hidden" id="target-vendor-group"' in response.data
    assert b'id="source-vendor-select"' in response.data
    assert b'id="target-vendor-select"' in response.data
    assert b'source-vendor-pills' not in response.data
    assert b'rule-preview-section' not in response.data
    assert b'v2.5' not in response.data


def test_api_migrate_offline_zip(client):
    conf_content = """config system global
    set hostname "fw-test"
end
config system interface
    edit "port1"
        set ip 10.0.0.1 255.255.255.0
    next
end"""
    data = {
        'file': (io.BytesIO(conf_content.encode('utf-8')), 'test.conf')
    }
    response = client.post('/api/migrate', data=data, content_type='multipart/form-data')
    assert response.status_code == 200
    assert response.headers['Content-Type'] == 'application/zip'
    with zipfile.ZipFile(io.BytesIO(response.data)) as archive:
        assert 'source_inventory_fortigate.xlsx' in archive.namelist()
        inventory = load_workbook(io.BytesIO(archive.read('source_inventory_fortigate.xlsx')))
        assert inventory['Summary']['B7'].value == 'Configuration File'


def test_api_extract_excel_from_file(client):
    conf_content = """config system global
    set hostname "excel-fw"
end
config firewall address
    edit "Inventory-Only"
        set subnet 10.20.30.40 255.255.255.255
    next
end"""
    response = client.post(
        '/api/extract/excel',
        data={
            'source_vendor': 'fortigate',
            'file': (io.BytesIO(conf_content.encode('utf-8')), 'test.conf'),
        },
        content_type='multipart/form-data',
    )

    assert response.status_code == 200
    assert response.headers['Content-Type'].startswith(
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    workbook = load_workbook(io.BytesIO(response.data))
    assert workbook['Summary']['B6'].value == 'excel-fw'
    assert workbook['Summary']['B7'].value == 'Configuration File'
    assert workbook['Addresses']['A4'].value == 'Inventory-Only'


def test_migration_zip_inventory_is_pre_optimization(client):
    conf_content = """config system global
    set hostname "pre-opt-fw"
end
config firewall address
    edit "Unused-But-Extracted"
        set subnet 198.51.100.10 255.255.255.255
    next
end"""
    response = client.post(
        '/api/migrate',
        data={
            'source_vendor': 'fortigate',
            'target_vendor': 'palo_alto',
            'optimize': 'true',
            'file': (io.BytesIO(conf_content.encode('utf-8')), 'test.conf'),
        },
        content_type='multipart/form-data',
    )

    assert response.status_code == 200
    with zipfile.ZipFile(io.BytesIO(response.data)) as archive:
        workbook = load_workbook(io.BytesIO(archive.read('source_inventory_fortigate.xlsx')))
        assert workbook['Addresses']['A4'].value == 'Unused-But-Extracted'


def test_api_diagnostics_endpoint(client):
    mock_results = [
        DiagnosticResult(name="terraform_cli", status="ok", message="CLI ready"),
        DiagnosticResult(name="registry_access", status="ok", message="Registry ok"),
        DiagnosticResult(name="palo_alto_line_of_sight", status="ok", message="Connected"),
        DiagnosticResult(name="palo_alto_auth", status="ok", message="Authenticated")
    ]

    with patch("fwmigrate.web.PaloAltoDiagnostics.run_all", return_value=mock_results):
        response = client.post('/api/diagnostics', json={
            'host': '192.168.1.1',
            'port': 443,
            'api_key': 'test_key'
        })
        assert response.status_code == 200
        data = response.get_json()
        assert data['success'] is True
        assert len(data['results']) == 4


def test_api_terraform_prepare(client, tmp_path):
    conf_content = """config system global
    set hostname "fw-test"
end
config firewall address
    edit "Server_1"
        set subnet 10.1.1.10 255.255.255.255
    next
end"""
    data = {
        'file': (io.BytesIO(conf_content.encode('utf-8')), 'test.conf'),
        'host': '192.168.1.1',
        'api_key': 'secret_key',
        'vsys': 'vsys1'
    }

    with patch("fwmigrate.engine.runner.TerraformSandbox.__init__", return_value=None), \
         patch("fwmigrate.engine.runner.TerraformSandbox.create", return_value=tmp_path):

        prep_resp = client.post('/api/terraform/prepare', data=data, content_type='multipart/form-data')
        assert prep_resp.status_code == 200
        prep_data = prep_resp.get_json()
        assert prep_data['success'] is True
        assert 'session_id' in prep_data


def test_api_terraform_plan(client, tmp_path):
    session_id = "test_plan_sess"
    ACTIVE_SESSIONS[session_id] = {
        'sandbox_dir': tmp_path,
        'secrets': [],
        'host': '10.0.0.1',
        'stats': {}
    }

    with patch("fwmigrate.engine.runner.TerraformRunner.run_init", return_value=(True, "init ok")), \
         patch("fwmigrate.engine.runner.TerraformRunner.run_plan", return_value=(True, "Plan: 5 to add", {"add": 5, "change": 0, "destroy": 0})):

        plan_resp = client.post('/api/terraform/plan', json={'session_id': session_id})
        assert plan_resp.status_code == 200
        plan_data = plan_resp.get_json()
        assert plan_data['success'] is True
        assert plan_data['summary']['add'] == 5


def test_api_terraform_apply_stream(client, tmp_path):
    session_id = "test_stream_sess"
    ACTIVE_SESSIONS[session_id] = {
        'status': 'APPROVED',
        'sandbox_dir': tmp_path,
        'secrets': [],
        'host': '10.0.0.1',
        'stats': {}
    }

    with patch("fwmigrate.engine.runner.TerraformRunner.run_apply_stream", return_value=iter([
        {'event': 'log', 'line': 'panos_address_object created'},
        {'event': 'complete', 'success': True, 'exit_code': 0, 'message': 'Success'}
    ])):
        stream_resp = client.get(f'/api/terraform/apply/stream?session_id={session_id}')
        assert stream_resp.status_code == 200
        assert "text/event-stream" in stream_resp.headers['Content-Type']
        # Fully consume the stream
        data = stream_resp.get_data(as_text=True)
        assert "panos_address_object created" in data
        assert "complete" in data


def test_api_terraform_destroy_stream(client, tmp_path):
    session_id = "test_destroy_sess"
    ACTIVE_SESSIONS[session_id] = {
        'status': 'APPROVED',
        'sandbox_dir': tmp_path,
        'secrets': [],
        'host': '10.0.0.1',
        'stats': {}
    }

    with patch("fwmigrate.engine.runner.TerraformRunner.run_destroy_stream", return_value=iter([
        {'event': 'log', 'line': 'panos_address_object destroyed'},
        {'event': 'complete', 'success': True, 'exit_code': 0, 'message': 'Destroy completed'}
    ])):
        stream_resp = client.get(f'/api/terraform/destroy/stream?session_id={session_id}')
        assert stream_resp.status_code == 200
        assert "text/event-stream" in stream_resp.headers['Content-Type']
        data = stream_resp.get_data(as_text=True)
        assert "panos_address_object destroyed" in data
        assert "complete" in data


def test_api_download_state_and_package(client, tmp_path):
    session_id = "test_dl_sess"
    (tmp_path / "main.tf").write_text("# main tf")
    (tmp_path / "terraform.tfstate").write_text('{"version": 4}')

    ACTIVE_SESSIONS[session_id] = {
        'sandbox_dir': tmp_path,
        'secrets': [],
        'host': '10.0.0.1',
        'stats': {}
    }

    # State download
    state_resp = client.get(f'/api/download/state?session_id={session_id}')
    assert state_resp.status_code == 200
    assert '{"version": 4}' in state_resp.get_data(as_text=True)

    # Package download
    pkg_resp = client.get(f'/api/download/package?session_id={session_id}')
    assert pkg_resp.status_code == 200
    assert pkg_resp.headers['Content-Type'] == 'application/zip'


def test_api_vendors_list(client):
    resp = client.get('/api/vendors')
    assert resp.status_code == 200
    data = resp.get_json()
    assert data['success'] is True
    assert len(data['sources']) >= 5
    assert len(data['targets']) >= 5
    # Check that api_fields metadata is attached for supported vendors
    fg_vendor = next((s for s in data['sources'] if s['vendor_id'] == 'fortigate'), None)
    assert fg_vendor is not None
    assert 'api_fields' in fg_vendor


def test_api_ingest_multivendor(client):
    # Missing host returns 400
    resp = client.post('/api/ingest/palo_alto', json={})
    assert resp.status_code == 400
    data = resp.get_json()
    assert data['success'] is False
    assert 'host is required' in data['error']

