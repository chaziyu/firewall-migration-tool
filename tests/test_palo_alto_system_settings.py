import io
import json

from openpyxl import load_workbook

from fwmigrate.parsers.palo_alto.parser import PANOSSourceParser
from fwmigrate.report.excel_exporter import IRExcelExporter


def _extract(system: str):
    return PANOSSourceParser().extract(f"""
    <config><devices><entry name="fw1"><deviceconfig><system>
      {system}
    </system></deviceconfig></entry></devices></config>
    """)


def test_system_settings_projects_hostname_timezone_and_management_plane():
    result = _extract("""
      <hostname>SBG-PA-FW01</hostname><timezone>Asia/Kuala_Lumpur</timezone>
      <ip-address>10.3.200.211</ip-address><netmask>255.255.255.0</netmask>
      <default-gateway>10.3.200.1</default-gateway><type><static/></type>
      <ipv6-enable>no</ipv6-enable>
      <service><disable-http>yes</disable-http><disable-https>no</disable-https>
        <disable-telnet>yes</disable-telnet><disable-snmp>no</disable-snmp></service>
    """)
    settings = result.canonical_ir.system_settings
    management = settings.management_plane

    assert result.canonical_ir.metadata.hostname == "SBG-PA-FW01"
    assert settings.hostname == "SBG-PA-FW01"
    assert settings.timezone == "Asia/Kuala_Lumpur"
    assert settings.admin_https_port is None
    assert management.ipv4_address == "10.3.200.211"
    assert management.netmask == "255.255.255.0"
    assert management.default_gateway == "10.3.200.1"
    assert management.address_type == "static"
    assert management.ipv6_enabled is False
    assert management.services == {"http": False, "https": True, "telnet": False, "snmp": True}


def test_system_settings_does_not_fabricate_absent_values():
    result = _extract("<hostname>fw1</hostname><service><disable-http>yes</disable-http></service>")
    management = result.canonical_ir.system_settings.management_plane

    assert management.ipv4_address is None
    assert management.default_gateway is None
    assert management.services == {"http": False}
    assert result.canonical_ir.dns_settings is None
    assert result.canonical_ir.ntp_settings is None
    assert result.canonical_ir.management_service_routes == []


def test_dns_and_ntp_are_typed_without_defaults_or_secrets():
    result = _extract("""
      <dns-setting><servers><primary>10.1.4.1</primary><secondary>203.121.16.85</secondary></servers></dns-setting>
      <ntp-servers>
        <primary-ntp-server><ntp-server-address>my.pool.ntp.org</ntp-server-address>
          <authentication-type><none/></authentication-type><authentication-key>FAKE_NTP_SECRET</authentication-key>
        </primary-ntp-server>
        <secondary-ntp-server><ntp-server-address>pool.ntp.org</ntp-server-address>
          <authentication-type><none/></authentication-type>
        </secondary-ntp-server>
      </ntp-servers>
    """)
    ir = result.canonical_ir

    assert (ir.dns_settings.primary, ir.dns_settings.secondary) == ("10.1.4.1", "203.121.16.85")
    assert [(server.role, server.address, server.authentication_type) for server in ir.ntp_settings.servers] == [
        ("primary", "my.pool.ntp.org", "none"), ("secondary", "pool.ntp.org", "none")
    ]
    assert "FAKE_NTP_SECRET" not in json.dumps(ir.model_dump(), default=str)


def test_dns_primary_only_and_malformed_service_preserve_source_evidence():
    result = _extract("""
      <dns-setting><servers><primary>not-an-ip</primary></servers><future-dns><value>keep</value></future-dns></dns-setting>
      <service><disable-http>maybe</disable-http></service>
    """)

    assert result.canonical_ir.dns_settings.primary == "not-an-ip"
    assert result.canonical_ir.dns_settings.secondary is None
    assert "future-dns" in result.canonical_ir.dns_settings.source_attributes["pan_source_entry"]["dns-setting"]
    assert result.canonical_ir.system_settings.management_plane.services == {}
    service = next(item for item in result.inventory_items if item.source_path == "deviceconfig/system/service")
    assert service.source_attributes["pan_system_management_service_disable"]["disable-http"] == "maybe"


def test_permitted_ips_and_single_ntp_server_keep_source_order_without_fallbacks():
    result = _extract("""
      <permitted-ip><entry name="198.51.100.0/24"/><entry name="bad"/></permitted-ip>
      <ntp-servers><primary-ntp-server><ntp-server-address>ntp.example</ntp-server-address></primary-ntp-server></ntp-servers>
    """)
    management = result.canonical_ir.system_settings.management_plane

    assert management.permitted_ips == ["198.51.100.0/24"]
    assert result.canonical_ir.ntp_settings.servers[0].role == "primary"
    assert len(result.canonical_ir.ntp_settings.servers) == 1
    assert any("bad" in item.source_attributes["pan_system_management_invalid_permitted_ips"] for item in result.inventory_items if item.source_path == "deviceconfig/system/permitted-ip")


def test_ntp_malformed_entries_remain_auditable():
    result = _extract("""
      <ntp-servers><primary-ntp-server><authentication-type><none/><password/></authentication-type></primary-ntp-server>
      </ntp-servers>
    """)
    server = result.canonical_ir.ntp_settings.servers[0]

    assert server.address is None
    assert server.authentication_type is None
    assert result.canonical_ir.ntp_settings.requires_manual_review is True
    assert any("missing ntp-server-address" in reason for reason in result.canonical_ir.ntp_settings.source_attributes["review_reasons"])


def test_management_service_routes_are_separate_and_ordered():
    result = _extract("""
      <route><service>
        <entry name="dns"><source><address>211.25.233.75/28</address><interface>ethernet1/3</interface></source></entry>
        <entry name="snmp"><source><address>10.3.200.211</address><interface>management</interface></source></entry>
        <entry name="dns"><source><address>bad-address</address></source></entry>
      </service></route>
    """)
    routes = result.canonical_ir.management_service_routes

    assert [route.name for route in routes] == ["dns", "snmp", "dns"]
    assert routes[0].source_address == "211.25.233.75/28"
    assert routes[0].source_interface == "ethernet1/3"
    assert routes[1].source_interface == "management"
    assert any("malformed source address" in reason for reason in routes[2].review_reasons)
    assert result.canonical_ir.routes == []
    route_records = [item for item in result.inventory_items if item.domain == "system_settings" and "route/service/entry" in item.source_path]
    assert len({item.source_record_id for item in route_records}) == 3


def test_system_residual_boundary_keeps_unimplemented_siblings():
    result = _extract("""
      <timezone>Asia/Kuala_Lumpur</timezone><dns-setting><servers><primary>1.1.1.1</primary></servers></dns-setting>
      <ntp-servers/><route><service/></route><update-server>updates.example</update-server>
      <authentication-profile>TACACS MRL</authentication-profile>
    """)
    residual_paths = {item.source_path for item in result.inventory_items if item.domain == "deviceconfig"}

    assert "deviceconfig/system/timezone" not in residual_paths
    assert "deviceconfig/system/dns-setting" not in residual_paths
    assert "deviceconfig/system/ntp-servers" not in residual_paths
    assert "deviceconfig/system/route" not in residual_paths
    assert "deviceconfig/system/update-server" in residual_paths
    assert "deviceconfig/system/authentication-profile" in residual_paths


def test_multiple_devices_do_not_use_last_system_value():
    result = PANOSSourceParser().extract("""
      <config><devices>
        <entry name="fw-a"><deviceconfig><system><hostname>FW-A</hostname><ip-address>10.0.0.1</ip-address></system></deviceconfig></entry>
        <entry name="fw-b"><deviceconfig><system><hostname>FW-B</hostname><ip-address>10.0.0.2</ip-address></system></deviceconfig></entry>
      </devices></config>
    """)
    settings = result.canonical_ir.system_settings

    assert settings.hostname is None
    assert settings.management_plane.ipv4_address is None
    assert "hostname" in settings.source_attributes["pan_multiple_device_conflicts"]


def test_excel_contains_phase6_sheets_and_no_secret():
    result = _extract("""
      <hostname>fw1</hostname><timezone>Asia/Kuala_Lumpur</timezone>
      <dns-setting><servers><primary>10.1.4.1</primary><secondary>203.121.16.85</secondary></servers></dns-setting>
      <ntp-servers><primary-ntp-server><ntp-server-address>my.pool.ntp.org</ntp-server-address><authentication-type><none/></authentication-type></primary-ntp-server>
      <secondary-ntp-server><ntp-server-address>pool.ntp.org</ntp-server-address><authentication-type><none/></authentication-type></secondary-ntp-server></ntp-servers>
      <route><service><entry name="dns"><source><address>211.25.233.75/28</address><interface>ethernet1/3</interface></source></entry><entry name="ntp"><source><address>211.25.233.75/28</address><interface>ethernet1/3</interface></source></entry></service></route>
    """)
    workbook = load_workbook(io.BytesIO(IRExcelExporter(result.canonical_ir, result).generate()))

    assert workbook["DNS Settings"]["A4"].value == "10.1.4.1"
    assert workbook["NTP Settings"].max_row == 5
    assert workbook["Management Service Routes"].max_row == 5
    assert workbook["Routes"].max_row == 3
    assert "FAKE_NTP_SECRET" not in "\n".join(
        str(cell.value) for sheet in workbook.worksheets for row in sheet.iter_rows() for cell in row
        if cell.value is not None
    )
