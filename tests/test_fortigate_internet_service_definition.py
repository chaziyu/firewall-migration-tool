from io import BytesIO

from openpyxl import load_workbook

from fwmigrate.parsers.fortigate.extractor import extract_fortigate_config
from fwmigrate.parsers.fortigate.parser import parse_fortigate_config
from fwmigrate.parsers.fortigate.transformer import FGToIRTransformer
from fwmigrate.report.excel_exporter import IRExcelExporter


CONFIG = '''
config system global
    set hostname "FG-TEST"
end
config firewall internet-service-definition
    edit 100
        set future-definition enable
        config entry
            edit 1
                set category-id 5
                set name "Custom-Web"
                set protocol 6
                set future-entry enabled
                config port-range
                    edit 1
                        set start-port 443
                        set end-port 443
                        set future-range value
                    next
                    edit 2
                        set start-port 8443
                        set end-port 8444
                    next
                end
            next
            edit 2
                set protocol 17
            next
        end
    next
    edit 101
        config entry
            edit 1
                set protocol 1
            next
        end
    next
end
'''


def test_parses_complete_hierarchy_without_defaults_or_loss():
    definitions = parse_fortigate_config(CONFIG).internet_service_definitions

    assert [definition.id for definition in definitions] == [100, 101]
    first = definitions[0]
    assert first.extra_settings == {"future_definition": "enable"}
    assert [entry.seq_num for entry in first.entries] == [1, 2]
    entry = first.entries[0]
    assert (entry.category_id, entry.name, entry.protocol) == (5, "Custom-Web", 6)
    assert entry.extra_settings == {"future_entry": "enabled"}
    assert [(item.id, item.start_port, item.end_port) for item in entry.port_ranges] == [
        (1, 443, 443), (2, 8443, 8444)
    ]
    assert entry.port_ranges[0].extra_settings == {"future_range": "value"}
    assert first.entries[1].category_id is None
    assert first.entries[1].name is None
    assert first.entries[1].port_ranges == []


def test_malformed_numeric_values_are_preserved_as_source_attributes():
    fg = parse_fortigate_config('''
config firewall internet-service-definition
    edit 1
        config entry
            edit 1
                set protocol invalid-protocol
                config port-range
                    edit 1
                        set start-port invalid-port
                    next
                end
            next
        end
    next
end
''')
    entry = fg.internet_service_definitions[0].entries[0]
    assert entry.protocol is None
    assert entry.extra_settings["unparsed_protocol"] == "invalid-protocol"
    assert entry.port_ranges[0].start_port is None
    assert entry.port_ranges[0].extra_settings["unparsed_start_port"] == "invalid-port"


def test_transformer_preserves_dedicated_extract_only_hierarchy():
    ir = FGToIRTransformer(parse_fortigate_config(CONFIG)).transform()

    definition = ir.internet_service_definitions[0]
    entry = definition.entries[0]
    assert (definition.source_id, definition.migration_status, definition.requires_manual_review) == (
        100, "EXTRACT_ONLY", True
    )
    assert (entry.source_sequence, entry.category_id, entry.name, entry.protocol_number) == (
        1, 5, "Custom-Web", 6
    )
    assert [(item.source_id, item.start_port, item.end_port) for item in entry.port_ranges] == [
        (1, 443, 443), (2, 8443, 8444)
    ]
    assert definition.source_attributes == {"future_definition": "enable"}
    assert entry.source_attributes == {"future_entry": "enabled"}


def test_extraction_coverage_and_excel_inventory_are_relational():
    extraction = extract_fortigate_config(CONFIG)
    coverage = {
        section.path: section
        for section in extraction.source_sections
        if section.path.startswith("firewall internet-service-definition")
    }
    assert coverage["firewall internet-service-definition"].status.value == "EXTRACT_ONLY"
    assert coverage["firewall internet-service-definition"].object_count_parsed == 2
    assert coverage["firewall internet-service-definition entry"].object_count_parsed == 3
    assert coverage["firewall internet-service-definition entry port-range"].object_count_parsed == 2
    source_items = [
        item for item in extraction.inventory_items
        if item.source_path.startswith("firewall internet-service-definition")
    ]
    assert any(
        item.source_path == "firewall internet-service-definition entry port-range"
        and item.name == "1"
        and any(command.key == "future-range" for command in item.commands)
        for item in source_items
    )

    workbook = load_workbook(BytesIO(IRExcelExporter(extraction.canonical_ir).generate()))
    assert {
        "Internet Service Definitions", "Internet Service Def Entries", "Internet Service Def Ports"
    }.issubset(workbook.sheetnames)
    definitions = workbook["Internet Service Definitions"]
    entries = workbook["Internet Service Def Entries"]
    ports = workbook["Internet Service Def Ports"]
    definition_headers = {cell.value: cell.column for cell in definitions[3]}
    entry_headers = {cell.value: cell.column for cell in entries[3]}
    port_headers = {cell.value: cell.column for cell in ports[3]}
    assert definitions.cell(4, definition_headers["Definition ID"]).value == 100
    assert entries.cell(4, entry_headers["Protocol #"]).value == 6
    assert entries.cell(4, entry_headers["Protocol Name"]).value == "TCP"
    assert ports.cell(4, port_headers["Definition ID"]).value == 100
    assert ports.cell(4, port_headers["Entry Sequence #"]).value == 1
    assert ports.cell(4, port_headers["Range ID"]).value == 1


def test_empty_section_is_valid_without_phantom_objects():
    extraction = extract_fortigate_config("config firewall internet-service-definition\nend\n")
    assert extraction.canonical_ir.internet_service_definitions == []
    section = next(item for item in extraction.source_sections if item.path == "firewall internet-service-definition")
    assert section.status.value == "EXTRACT_ONLY"
    assert section.object_count_parsed == 0
