import io

import pytest
from openpyxl import load_workbook

from fwmigrate.extraction.models import ExtractionStatus
from fwmigrate.parsers.fortigate.extractor import extract_fortigate_config
from fwmigrate.parsers.fortigate.net_utils import normalize_ipv6_network
from fwmigrate.report.excel_exporter import IRExcelExporter


def _headers(sheet):
    return {cell.value: cell.column for cell in sheet[3]}


def test_static_route_destination_fidelity_and_ipv6_coverage() -> None:
    result = extract_fortigate_config('''
config router static
    edit 1
        set dstaddr "REMOTE_NET"
    next
    edit 2
        set gateway 192.0.2.1
    next
    edit 3
        set dst 198.51.100.0 255.255.255.0
        set sdwan-zone "zone-a" "zone-b"
        set dynamic-gateway enable
        set link-monitor-exempt enable
        set src 10.0.0.0 255.255.255.0
        set weight 20
        set bfd enable
        set vrf 7
        set tag 42
        set internet-service 12345
        set internet-service-custom "custom-isdb"
        set future-route-option retained
    next
end
config router static6
    edit 10
        set dst 2001:db8:10::1/64
        set gateway 2001:db8::1
    next
    edit 11
        set dst not-an-ipv6-prefix
    next
end
''')

    object_route, default_route, advanced_route, static6, invalid6 = (
        result.canonical_ir.routes
    )
    assert object_route.destination is None
    assert object_route.source_destination_reference == "REMOTE_NET"
    assert object_route.requires_manual_review is True
    assert object_route.safe_for_target_generation is False

    assert default_route.destination == "0.0.0.0/0"
    assert default_route.safe_for_target_generation is True

    assert advanced_route.sdwan_zones == ["zone-a", "zone-b"]
    assert advanced_route.sdwan_zone is None
    assert advanced_route.dynamic_gateway == "enable"
    assert advanced_route.link_monitor_exempt == "enable"
    assert advanced_route.source_prefix == "10.0.0.0 255.255.255.0"
    assert advanced_route.weight == 20
    assert advanced_route.bfd == "enable"
    assert advanced_route.vrf == 7
    assert advanced_route.route_tag == 42
    assert advanced_route.internet_service == 12345
    assert advanced_route.internet_service_custom == "custom-isdb"
    assert advanced_route.source_attributes["future_route_option"] == "retained"
    assert advanced_route.safe_for_target_generation is False

    assert static6.address_family == "ipv6"
    assert static6.destination == "2001:db8:10::/64"
    assert invalid6.address_family == "ipv6"
    assert invalid6.destination is None
    assert invalid6.parse_error
    assert invalid6.safe_for_target_generation is False

    coverage = {item.path: item for item in result.source_sections}
    assert coverage["router static"].object_count_parsed == 3
    assert coverage["router static6"].object_count_parsed == 2
    assert coverage["router static6"].status == ExtractionStatus.PARTIALLY_NORMALIZED

    workbook = load_workbook(io.BytesIO(IRExcelExporter(
        result.canonical_ir, result
    ).generate()))
    routes = workbook["Routes"]
    headers = _headers(routes)
    assert routes.cell(4, headers["Destination Prefix (Normalized)"]).value is None
    assert routes.cell(4, headers["Destination Object / Group"]).value == "REMOTE_NET"
    assert routes.cell(4, headers["Manual Review"]).value == "Yes"


def test_invalid_ipv6_network_is_rejected_without_substitution() -> None:
    with pytest.raises(ValueError):
        normalize_ipv6_network("not-an-ipv6-prefix")


def test_sdwan_cardinality_hierarchy_and_future_child_fallback() -> None:
    result = extract_fortigate_config('''
config system sdwan
    set status enable
    config members
        edit 1
            set interface "wan1"
            set zone "Internet"
            set gateway 192.0.2.1
            set source 192.0.2.2
            set gateway6 2001:db8::1
            set source6 2001:db8::2
            set cost 10
            set weight 20
            set priority 5
            set priority6 6
            set spillover-threshold 100
            set ingress-spillover-threshold 200
            set volume-ratio 3
            set status enable
            set comment "Primary link"
        next
    end
    config health-check
        edit "internet"
            set server "1.1.1.1"
            set members 1
            set protocol ping
            set port 53
            set interval 500
            set probe-timeout 250
            set failtime 3
            set recoverytime 4
            set update-static-route enable
            set vrf 7
            set source 192.0.2.2
            set future-health-option retained
        next
    end
    config service
        edit 10
            set status enable
            set src "LAN"
            set dst "all"
            set health-check "google" "office365"
            set priority-members 1
            set priority-zone "Internet" "Backup"
            set sla-compare-method number
            set tie-break fib-best-match
            set use-shortcut-sla enable
            config sla
                edit "internet"
                    set id 1
                    set future-sla-option retained
                next
            end
        next
    end
    config duplication
        edit 1
            set service-id 10
            set srcaddr "LAN"
            set dstaddr "all"
            set srcaddr6 "LAN6"
            set dstaddr6 "all6"
            set srcintf "lan"
            set dstintf "wan1" "wan2"
            set service "HTTPS" "DNS"
            set packet-duplication force
            set sla-match-service "internet"
            set packet-de-duplication enable
        next
    end
    config neighbor
        edit "peer-a"
            set member 1
            set role primary
        next
    end
    config future-feature
        edit "future-a"
            set future-option retained
        next
    end
end
''')

    sdwan = result.canonical_ir.sdwan
    assert sdwan is not None
    assert sdwan.migration_status == "EXTRACT_ONLY"
    assert sdwan.requires_manual_review is True

    member = sdwan.members[0]
    assert (member.gateway, member.source) == ("192.0.2.1", "192.0.2.2")
    assert (member.gateway6, member.source6) == ("2001:db8::1", "2001:db8::2")
    assert (member.cost, member.weight, member.priority, member.priority6) == (10, 20, 5, 6)
    assert (member.spillover_threshold, member.ingress_spillover_threshold) == (100, 200)
    assert member.volume_ratio == 3

    check = sdwan.health_checks[0]
    assert (check.protocol, check.port, check.interval) == ("ping", 53, 500)
    assert (check.probe_timeout, check.failtime, check.recoverytime) == (250, 3, 4)
    assert (check.update_static_route, check.vrf, check.source) == (
        "enable", 7, "192.0.2.2"
    )
    assert check.source_attributes["future_health_option"] == "retained"

    rule = sdwan.rules[0]
    assert rule.health_checks == ["google", "office365"]
    assert rule.health_check is None
    assert rule.priority_zones == ["Internet", "Backup"]
    assert rule.sla[0].name == "internet"
    assert rule.sla[0].source_id == 1
    assert rule.sla[0].source_attributes["future_sla_option"] == "retained"

    duplication = sdwan.duplication_rules[0]
    assert duplication.service_id == 10
    assert duplication.destination_interfaces == ["wan1", "wan2"]
    assert duplication.services == ["HTTPS", "DNS"]
    assert sdwan.neighbors[0].name == "peer-a"
    assert sdwan.neighbors[0].source_attributes == {
        "member": ["1"],
        "role": "primary",
    }

    coverage = {item.path: item.status for item in result.source_sections}
    assert coverage["system sdwan service sla"] == ExtractionStatus.EXTRACT_ONLY
    assert coverage["system sdwan duplication"] == ExtractionStatus.EXTRACT_ONLY
    assert coverage["system sdwan neighbor"] == ExtractionStatus.EXTRACT_ONLY
    assert coverage["system sdwan future-feature"] == ExtractionStatus.EXTRACT_ONLY

    workbook = load_workbook(io.BytesIO(IRExcelExporter(
        result.canonical_ir, result
    ).generate()))
    assert workbook["SD-WAN Duplication"].max_row == 4
    assert workbook["SD-WAN Neighbors"].max_row == 4
    assert workbook["SD-WAN Rule SLAs"].max_row == 4
    source_rows = list(
        workbook["FortiGate Source Configuration"].iter_rows(
            min_row=4, values_only=True
        )
    )
    assert any(
        row[1] == "system sdwan future-feature"
        and row[6] == "future-option"
        and row[7] == "retained"
        for row in source_rows
    )


def test_routing_dependencies_and_empty_protocol_blocks_are_distinct() -> None:
    result = extract_fortigate_config('''
config router bgp
end
config router route-map
    edit "SET-PREFERENCE"
        config rule
            edit 1
                set match-ip-address "PREFIXES"
                set set-local-preference 200
            next
        end
    next
end
''')
    workbook = load_workbook(io.BytesIO(IRExcelExporter(
        result.canonical_ir, result
    ).generate()))

    protocols = workbook["Routing Protocols"]
    protocol_headers = _headers(protocols)
    assert protocols.cell(4, protocol_headers["Source Block Present"]).value == "Yes"
    assert protocols.cell(4, protocol_headers["Configured"]).value == "No"

    dependencies = workbook["Routing Dependencies"]
    dependency_headers = _headers(dependencies)
    assert dependencies.cell(4, dependency_headers["Source Path"]).value == "router route-map"
    assert dependencies.cell(4, dependency_headers["Configured"]).value == "Yes"
    assert workbook["Routing Dependency Settings"].max_row > 3
