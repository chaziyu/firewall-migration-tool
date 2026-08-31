import io

from openpyxl import load_workbook

from fwmigrate.parsers.fortigate.extractor import extract_fortigate_config
from fwmigrate.report.excel_exporter import IRExcelExporter


def _policy_config(inspection_mode=None):
    inspection_setting = (
        ""
        if inspection_mode is None
        else f"        set inspection-mode {inspection_mode}\n"
    )
    return f"""config firewall policy
    edit 32
        set name "Policy_32"
        set srcintf "any"
        set dstintf "any"
        set srcaddr "all"
        set dstaddr "all"
        set service "ALL"
        set action accept
{inspection_setting}    next
end
"""


def _policy(inspection_mode=None):
    result = extract_fortigate_config(_policy_config(inspection_mode))
    return result, result.canonical_ir.policies[0]


def test_proxy_inspection_mode_requires_manual_review_and_is_preserved():
    _, policy = _policy("proxy")

    reason = "FortiGate proxy inspection mode requires target-platform review"
    assert policy.source_inspection_mode == "proxy"
    assert policy.migration_status == "PARTIALLY_NORMALIZED"
    assert policy.requires_manual_review is True
    assert reason in policy.review_reasons
    assert policy.review_reasons.count(reason) == 1


def test_flow_inspection_mode_does_not_require_review_by_itself():
    _, policy = _policy("flow")

    assert policy.source_inspection_mode == "flow"
    assert policy.migration_status == "NORMALIZED"
    assert policy.requires_manual_review is False
    assert not any("inspection mode" in reason.lower() for reason in policy.review_reasons)


def test_missing_inspection_mode_uses_default_without_manual_review():
    _, policy = _policy()

    assert policy.source_inspection_mode is None
    assert policy.migration_status == "NORMALIZED"
    assert policy.requires_manual_review is False
    assert not any("inspection mode" in reason.lower() for reason in policy.review_reasons)


def test_unknown_inspection_mode_requires_manual_review():
    _, policy = _policy("vendor-future-mode")

    assert policy.source_inspection_mode == "vendor-future-mode"
    assert policy.migration_status == "PARTIALLY_NORMALIZED"
    assert policy.requires_manual_review is True
    assert "Unknown FortiGate inspection mode requires manual review" in policy.review_reasons


def test_policy_inspection_fields_are_visible_in_excel():
    result, _ = _policy("proxy")
    workbook = load_workbook(
        io.BytesIO(IRExcelExporter(result.canonical_ir, result).generate())
    )

    sheet = workbook["Policies"]
    headers = {cell.value: cell.column for cell in sheet[3]}
    row = next(
        row_number
        for row_number in range(4, sheet.max_row + 1)
        if sheet.cell(row_number, headers["Source Policy ID"]).value == "32"
    )

    assert sheet.cell(row, headers["Inspection Mode"]).value == "proxy"
    assert sheet.cell(row, headers["Extraction Status"]).value == "PARTIALLY_NORMALIZED"
    assert sheet.cell(row, headers["Manual Review"]).value == "TRUE"
    assert "FortiGate proxy inspection mode requires target-platform review" in sheet.cell(
        row, headers["Review Reasons"]
    ).value
