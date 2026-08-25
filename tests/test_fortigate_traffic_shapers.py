import io

from openpyxl import load_workbook

from fwmigrate.extraction.models import ExtractionStatus
from fwmigrate.parsers.fortigate.extractor import extract_fortigate_config
from fwmigrate.parsers.fortigate.parser import parse_fortigate_config
from fwmigrate.report.excel_exporter import IRExcelExporter


TRAFFIC_SHAPERS_CONFIG = """
config firewall shaper traffic-shaper
    edit "high-priority"
        set maximum-bandwidth 500
        set bandwidth-unit kbps
        set priority high
        set per-policy enable
        set overhead enable
    next
    edit "guarantee-100kbps"
        set guaranteed-bandwidth 100
        set priority medium
    next
    edit "shared-1M-pipe"
        set maximum-bandwidth 1
        set bandwidth-unit mbps
        set per-policy disable
    next
    edit "low-priority"
        set priority low
    next
    edit "unqualified-bandwidth"
        set maximum-bandwidth 750
    next
end
"""


def test_traffic_shapers_preserve_typed_and_unknown_source_values():
    result = extract_fortigate_config(TRAFFIC_SHAPERS_CONFIG)
    parsed = parse_fortigate_config(TRAFFIC_SHAPERS_CONFIG).traffic_shapers
    shapers = result.canonical_ir.traffic_shapers

    assert len(parsed) == 5
    assert len(shapers) == 5

    high = next(item for item in shapers if item.name == "high-priority")
    assert high.maximum_bandwidth == 500
    assert high.source_bandwidth_unit == "kbps"
    assert high.priority == "high"
    assert high.per_policy is True
    assert high.source_attributes == {"overhead": "enable"}
    assert high.migration_status == "PARTIALLY_NORMALIZED"
    assert high.requires_manual_review is True

    guarantee = next(item for item in shapers if item.name == "guarantee-100kbps")
    assert guarantee.guaranteed_bandwidth == 100
    assert guarantee.source_bandwidth_unit is None
    assert guarantee.per_policy is None

    shared = next(item for item in shapers if item.name == "shared-1M-pipe")
    assert shared.maximum_bandwidth == 1
    assert shared.source_bandwidth_unit == "mbps"
    assert shared.per_policy is False

    section = next(
        item for item in result.source_sections
        if item.path == "firewall shaper traffic-shaper"
    )
    assert section.status == ExtractionStatus.PARTIALLY_NORMALIZED
    assert section.object_count_parsed == 5
    assert section.object_count_normalized == 5


def test_traffic_shapers_excel_rows_match_source_objects():
    result = extract_fortigate_config(TRAFFIC_SHAPERS_CONFIG)
    workbook = load_workbook(
        io.BytesIO(
            IRExcelExporter(
                result.canonical_ir,
                extraction_result=result,
            ).generate()
        )
    )
    sheet = workbook["Traffic Shapers"]
    headers = {cell.value: cell.column for cell in sheet[3]}

    assert sheet.max_row == 8
    assert sheet.cell(4, headers["Name"]).value == "high-priority"
    assert sheet.cell(4, headers["Additional Settings"]).value == "overhead=enable"
    assert sheet.cell(5, headers["Source Bandwidth Unit"]).value is None
    assert sheet.cell(5, headers["Per Policy"]).value is None
