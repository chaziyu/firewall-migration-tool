from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
from fwmigrate.extraction.models import ExtractionStatus
from fwmigrate.ir.enums import ServiceProtocol
from fwmigrate.parsers.palo_alto.parser import PANOSSourceParser
from fwmigrate.report.excel_exporter import IRExcelExporter
from fwmigrate.parsers.palo_alto.source_model import PANScope


FIXTURE = Path(__file__).parent / "fixtures" / "palo_alto" / "services.xml"


def _extract():
    parser = PANOSSourceParser()
    result = parser.extract(FIXTURE.read_text(encoding="utf-8"))
    return parser, result


def _service(result, name):
    return next(service for service in result.canonical_ir.services if service.name == name)


def _group(result, name):
    return next(group for group in result.canonical_ir.service_groups if group.name == name)


def _records(result, domain, name):
    return [item for item in result.inventory_items if item.domain == domain and item.name == name]


def test_tcp_service_destination_port():
    _, result = _extract()
    port = _service(result, "TCP-443").ports[0]
    assert port.protocol == ServiceProtocol.TCP
    assert port.port == "443"


def test_udp_service_destination_port():
    _, result = _extract()
    port = _service(result, "UDP-53").ports[0]
    assert port.protocol == ServiceProtocol.UDP
    assert port.port == "53"


def test_tcp_service_source_port():
    _, result = _extract()
    assert _service(result, "TCP-Source").ports[0].source_port == "1024-65535"


def test_udp_service_source_port():
    _, result = _extract()
    assert _service(result, "UDP-Source").ports[0].source_port == "500,4500"


def test_service_single_port():
    _, result = _extract()
    assert _service(result, "TCP-443").ports[0].port == "443"


def test_service_port_range():
    _, result = _extract()
    assert _service(result, "Port-Range").ports[0].port == "80-90"


def test_service_comma_separated_ports():
    _, result = _extract()
    assert _service(result, "Port-List").ports[0].port == "80,443"


def test_service_multiple_ranges():
    _, result = _extract()
    assert _service(result, "Multi-Range").ports[0].port == "80,443,8000-8100,9000-9010"


def test_invalid_service_port_parse_error():
    parser, result = _extract()
    assert all(service.name != "Bad-Port" for service in result.canonical_ir.services)
    assert _records(result, "services", "Bad-Port")[0].status == ExtractionStatus.PARSE_ERROR
    assert parser.resolver.resolve("Bad-Port", "service-reference", PANScope(kind="vsys", name="vsys1")) is None


def test_reversed_service_range_parse_error():
    _, result = _extract()
    assert all(service.name != "Reversed-Port" for service in result.canonical_ir.services)
    assert _records(result, "services", "Reversed-Port")[0].status == ExtractionStatus.PARSE_ERROR


def test_missing_destination_port_parse_error():
    _, result = _extract()
    assert all(service.name != "Missing-Port" for service in result.canonical_ir.services)
    assert _records(result, "services", "Missing-Port")[0].status == ExtractionStatus.PARSE_ERROR


def test_service_with_tcp_and_udp_parse_error():
    _, result = _extract()
    assert all(service.name != "Both-Protocols" for service in result.canonical_ir.services)
    assert _records(result, "services", "Both-Protocols")[0].status == ExtractionStatus.PARSE_ERROR


def test_service_description_preserved():
    _, result = _extract()
    assert _service(result, "TCP-443").description == "HTTPS service"


def test_service_tags_retained():
    _, result = _extract()
    service = _service(result, "Tagged-Service")
    assert service.source_attributes["pan_tags"] == ["production"]
    assert _records(result, "services", "Tagged-Service")[0].status == ExtractionStatus.PARTIALLY_NORMALIZED


def test_service_timeout_retained():
    _, result = _extract()
    attributes = _service(result, "Timeout-Service").source_attributes
    assert attributes["pan_timeout"] == "3600"
    assert attributes["pan_halfclose_timeout"] == "120"
    assert attributes["pan_timewait_timeout"] == "15"
    assert "override" in str(attributes["pan_timeout_override"])


def test_service_timeout_causes_partial_status():
    _, result = _extract()
    assert _records(result, "services", "Timeout-Service")[0].status == ExtractionStatus.PARTIALLY_NORMALIZED


def test_service_unknown_field_causes_partial():
    _, result = _extract()
    service = _service(result, "Unknown-Service")
    assert service.source_attributes["pan_unknown_fields"] == {"future-setting": "retain-me"}
    assert service.source_attributes["pan_unknown_protocol_fields"] == {"future-protocol": "value"}
    assert _records(result, "services", "Unknown-Service")[0].status == ExtractionStatus.PARTIALLY_NORMALIZED


def test_service_exactly_one_terminal_status():
    _, result = _extract()
    assert len(_records(result, "services", "TCP-443")) == 1
    assert _records(result, "services", "TCP-443")[0].status == ExtractionStatus.NORMALIZED


def test_invalid_service_not_registered():
    parser, _ = _extract()
    scope = PANScope(kind="vsys", name="vsys1")
    assert parser.resolver.resolve("Bad-Port", "service", scope) is None


def test_service_scope_collision_gets_canonical_names():
    _, result = _extract()
    by_port = {service.ports[0].port: service.name for service in result.canonical_ir.services if service.name.endswith("Scoped-Service")}
    assert by_port == {"9443": "Scoped-Service", "10443": "vsys1::Scoped-Service", "1053": "vsys2::Scoped-Service"}


def test_service_group_members_extracted():
    _, result = _extract()
    assert _group(result, "Basic-Services").members == ["TCP-443", "UDP-53"]


def test_service_group_description_preserved():
    _, result = _extract()
    assert _group(result, "Basic-Services").description == "Basic group"


def test_service_group_tags_retained():
    _, result = _extract()
    assert _group(result, "Basic-Services").source_attributes["pan_tags"] == ["group-tag"]


def test_service_group_member_canonicalized():
    _, result = _extract()
    assert _group(result, "Nested-Services").members == ["Basic-Services", "vsys1::Scoped-Service"]


def test_unresolved_service_group_member_requires_review():
    _, result = _extract()
    group = _group(result, "Unresolved-Services")
    assert group.unsafe_members == ["Missing-Service"]
    assert group.requires_manual_review is True
    assert _records(result, "service_groups", "Unresolved-Services")[0].status == ExtractionStatus.PARTIALLY_NORMALIZED


def test_service_and_service_group_collision_not_order_dependent():
    parser, result = _extract()
    scope = PANScope(kind="vsys", name="vsys1")
    assert all(group.name != "Collision" for group in result.canonical_ir.service_groups)
    assert _records(result, "service_groups", "Collision")[0].status == ExtractionStatus.PARSE_ERROR
    assert parser.resolver.resolve("Collision", "service-reference", scope).kind == "service"


def test_service_group_exactly_one_terminal_status():
    _, result = _extract()
    assert len(_records(result, "service_groups", "Basic-Services")) == 1


def test_service_group_scope_collision_gets_canonical_names():
    _, result = _extract()
    names = sorted(group.name for group in result.canonical_ir.service_groups if group.name.endswith("Scoped-Group"))
    assert names == ["vsys1::Scoped-Group", "vsys2::Scoped-Group"]


def _group_xml(members, services="", groups="", rules="", nat=""):
    return f"""<config version="11.1.0"><devices><entry name="localhost.localdomain"><vsys>
    <entry name="vsys1"><service>{services}</service><service-group>{groups}
    </service-group><rulebase><security><rules>{rules}</rules></security><nat><rules>{nat}</rules></nat></rulebase>
    </entry></vsys></entry></devices></config>"""


def _service_group_case(members, services="", groups=""):
    groups = groups or (
        f'<entry name="WEB"><members>{"".join(f"<member>{m}</member>" for m in members)}</members></entry>'
        if members else ""
    )
    xml = _group_xml(
        members,
        services=services,
        groups=groups,
    )
    return PANOSSourceParser().extract(xml)


def test_predefined_services_are_safe_service_group_members():
    result = _service_group_case(["service-http", "service-https"])
    group = _group(result, "WEB")
    assert group.members == ["service-http", "service-https"]
    assert group.unsafe_members == []
    assert group.requires_manual_review is False
    assert group.migration_status == "NORMALIZED"
    assert group.source_attributes["pan_recognized_predefined_services"] == ["service-http", "service-https"]
    assert "pan_unresolved_members" not in group.source_attributes


def test_predefined_and_custom_service_group_members_are_safe():
    services = '<entry name="TCP8443"><protocol><tcp><port>8443</port></tcp></protocol></entry>'
    result = _service_group_case(["service-http", "service-https", "TCP8443"], services=services)
    group = _group(result, "WEB")
    assert group.members == ["service-http", "service-https", "TCP8443"]
    assert group.unsafe_members == []
    assert group.migration_status == "NORMALIZED"


def test_nested_predefined_service_group_is_safe():
    groups = (
        '<entry name="WEB-INNER"><members><member>service-http</member><member>service-https</member></members></entry>'
        '<entry name="WEB-OUTER"><members><member>WEB-INNER</member></members></entry>'
    )
    result = _service_group_case([], groups=groups)
    assert _group(result, "WEB-INNER").unsafe_members == []
    assert _group(result, "WEB-OUTER").unsafe_members == []


def test_predefined_service_does_not_hide_unresolved_or_partial_custom_members():
    unresolved = _service_group_case(["service-http", "MISSING-SERVICE"])
    group = _group(unresolved, "WEB")
    assert group.unsafe_members == ["MISSING-SERVICE"]
    assert group.source_attributes["pan_unresolved_members"] == ["MISSING-SERVICE"]
    assert "service-http" not in group.unsafe_members

    partial_service = '<entry name="PARTIAL"><protocol><tcp><port>8443</port><future>value</future></tcp></protocol></entry>'
    partial = _service_group_case(["service-https", "PARTIAL"], services=partial_service)
    group = _group(partial, "WEB")
    assert group.unsafe_members == ["PARTIAL"]
    assert "service-https" not in group.unsafe_members
    assert group.migration_status == "PARTIALLY_NORMALIZED"


def test_application_default_is_not_a_service_group_predefined_service():
    result = _service_group_case(["application-default"])
    group = _group(result, "WEB")
    assert group.unsafe_members == ["application-default"]
    assert "pan_recognized_predefined_services" not in group.source_attributes


def test_predefined_service_group_is_exported_without_review_flags():
    result = _service_group_case(["service-http", "service-https"])
    workbook = load_workbook(BytesIO(IRExcelExporter(result.canonical_ir, result).generate()))
    sheet = workbook["Service Groups"]
    headers = {cell.value: cell.column for cell in sheet[3]}
    assert sheet.cell(4, headers["Members"]).value == "service-http\nservice-https"
    assert sheet.cell(4, headers["Unsafe Members"]).value in (None, "")
    assert sheet.cell(4, headers["Migration Status"]).value == "NORMALIZED"
