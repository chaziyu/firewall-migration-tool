import io

from openpyxl import load_workbook

from fwmigrate.parsers.fortigate.extractor import extract_fortigate_config
from fwmigrate.parsers.fortigate.parser import parse_fortigate_config
from fwmigrate.report.excel_exporter import IRExcelExporter


FORBIDDEN = {
    "PASSWORD_SENTINEL",
    "PRIVATE_KEY_SENTINEL",
    "PSK_SENTINEL",
    "FORTITOKEN_SEED_SENTINEL",
    "BEGIN PRIVATE KEY",
    "BEGIN ENCRYPTED PRIVATE KEY",
    "BEGIN RSA PRIVATE KEY",
    "BEGIN EC PRIVATE KEY",
}


SECRET_CONFIG = '''config user ldap
    edit "ldap"
        set password "PASSWORD_SENTINEL"
    next
end
config user local
    edit "local"
        set passwd "PASSWORD_SENTINEL"
        set passwd-time 2025-07-09 23:11:09
        set password-expiry-warning 7
        set password-renewal enable
        set seed "FORTITOKEN_SEED_SENTINEL"
        set activation-code "PASSWORD_SENTINEL"
    next
end
config user fsso
    edit "fsso"
        set password "PASSWORD_SENTINEL"
    next
end
config user fortitoken
    edit "token"
        set seed "FORTITOKEN_SEED_SENTINEL"
        set activation-code "PASSWORD_SENTINEL"
    next
end
config vpn certificate local
    edit "local-cert"
        set password "PASSWORD_SENTINEL"
        set private-key "-----BEGIN PRIVATE KEY-----PRIVATE_KEY_SENTINEL-----END PRIVATE KEY-----"
    next
end
config firewall ssh local-key
    edit "ssh-key"
        set password "PASSWORD_SENTINEL"
        set private-key "-----BEGIN RSA PRIVATE KEY-----PRIVATE_KEY_SENTINEL-----END RSA PRIVATE KEY-----"
    next
end
config firewall ssh local-ca
    edit "ssh-ca"
        set passwd "PASSWORD_SENTINEL"
        set private-key "PRIVATE_KEY_SENTINEL"
    next
end
config vpn ipsec phase1-interface
    edit "tunnel"
        set interface "wan1"
        set psksecret "PSK_SENTINEL"
    next
end
'''


def test_secret_material_never_crosses_parser_extraction_ir_or_excel_layers() -> None:
    fg = parse_fortigate_config(SECRET_CONFIG)
    result = extract_fortigate_config(SECRET_CONFIG)
    workbook = load_workbook(
        io.BytesIO(IRExcelExporter(result.canonical_ir, result).generate())
    )

    assert fg.user_ldap_servers[0].has_password is True
    assert fg.local_users[0].has_password is True
    assert fg.local_users[0].passwd_time == "2025-07-09 23:11:09"
    assert fg.local_users[0].extra_settings["password_expiry_warning"] == "7"
    assert fg.local_users[0].extra_settings["password_renewal"] == "enable"
    assert fg.fsso_servers[0].has_password is True
    assert result.canonical_ir.fsso_providers[0].has_password is True
    assert result.canonical_ir.local_users[0].source_passwd_time == (
        "2025-07-09 23:11:09"
    )
    assert fg.certificates[0].has_private_key is True
    assert fg.ssh_keys[0].has_private_key is True
    assert fg.ssh_keys[0].has_password is True
    assert fg.phase1_interfaces[0].has_psk is True
    assert result.canonical_ir.vpn_tunnels[0].has_psk is True

    serialized = "\n".join((
        fg.model_dump_json(),
        result.model_dump_json(),
        result.canonical_ir.model_dump_json(),
    ))
    assert "psksecret" not in serialized.lower()
    vpn_sheet = workbook["VPN Tunnels"]
    vpn_headers = {cell.value: cell.column for cell in vpn_sheet[3]}
    assert vpn_sheet.cell(4, vpn_headers["PSK"]).value == "Configured / Redacted"
    local_sheet = workbook["Local Users"]
    local_headers = {cell.value: cell.column for cell in local_sheet[3]}
    assert local_sheet.cell(4, local_headers["Password Configured"]).value == "Yes"
    assert local_sheet.cell(4, local_headers["Password Time"]).value == (
        "2025-07-09 23:11:09"
    )
    excel_cells = "\n".join(
        str(cell.value)
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if cell.value is not None
    )
    for forbidden in FORBIDDEN:
        assert forbidden not in serialized
        assert forbidden not in excel_cells
        for sheet_name in (
            "Warnings", "Unsupported", "Extraction Coverage",
            "FortiGate Source Configuration",
        ):
            assert forbidden not in "\n".join(
                str(cell.value)
                for row in workbook[sheet_name].iter_rows()
                for cell in row
                if cell.value is not None
            )


def test_fortigate_secret_sanitizer_keeps_password_metadata() -> None:
    from fwmigrate.parsers.fortigate.extraction import sanitize_source_attributes

    sanitized = sanitize_source_attributes({
        "passwd": "PASSWORD_SENTINEL",
        "password2": "PASSWORD2_SENTINEL",
        "passwd-time": "2025-07-09 23:11:09",
        "password-policy": "policy-name",
        "password-expiry-warning": "7",
        "password-renewal": "enable",
    })

    assert sanitized == {
        "passwd": "[REDACTED]",
        "password2": "[REDACTED]",
        "passwd_time": "2025-07-09 23:11:09",
        "password_policy": "policy-name",
        "password_expiry_warning": "7",
        "password_renewal": "enable",
    }
