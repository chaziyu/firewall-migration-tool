import io
import json
import pytest
from openpyxl import load_workbook

from fwmigrate.core.registry import PluginRegistry
from fwmigrate.report.excel_exporter import IRExcelExporter
from tests.fixture_paths import CHECKPOINT_FIXTURE


def test_checkpoint_excel_export_includes_source_inventory():
    content = CHECKPOINT_FIXTURE.read_text(encoding="utf-8")
    parser = PluginRegistry.get_parser("checkpoint")
    extraction = parser.extract(content)

    exporter = IRExcelExporter(extraction.canonical_ir, extraction_result=extraction)
    excel_bytes = exporter.generate()

    assert excel_bytes is not None
    assert len(excel_bytes) > 0

    wb = load_workbook(io.BytesIO(excel_bytes), read_only=True)
    sheet_names = wb.sheetnames

    assert "Source Inventory" in sheet_names
    assert "Extraction Coverage" in sheet_names
    assert "Policies" in sheet_names
    assert "Addresses" in sheet_names
    assert "Services" in sheet_names


def test_checkpoint_excel_export_defense_in_depth_secrets():
    bundle_with_secrets = json.dumps({
        "format": "checkpoint-export-v1",
        "responses": [
            {
                "command": "show-gateways-and-servers",
                "data": {
                    "objects": [
                        {
                            "name": "SecretGW",
                            "type": "simple-gateway",
                            "one-time-password": "TopSecretPassword999!",
                            "shared-secret": "MySuperSecretKey"
                        }
                    ]
                }
            }
        ]
    })

    parser = PluginRegistry.get_parser("checkpoint")
    extraction = parser.extract(bundle_with_secrets)

    exporter = IRExcelExporter(extraction.canonical_ir, extraction_result=extraction)
    excel_bytes = exporter.generate()

    wb = load_workbook(io.BytesIO(excel_bytes), data_only=True)
    inv_sheet = wb["Source Inventory"]

    # Search all cells for raw sensitive strings
    for row in inv_sheet.iter_rows(values_only=True):
        for cell_val in row:
            if cell_val is not None:
                assert "TopSecretPassword999!" not in str(cell_val)
                assert "MySuperSecretKey" not in str(cell_val)
