import io

from openpyxl import load_workbook

from fwmigrate.extraction.models import ExtractionStatus
from fwmigrate.parsers.fortigate.extractor import extract_fortigate_config
from fwmigrate.parsers.fortigate.parser import parse_fortigate_config
from fwmigrate.report.excel_exporter import IRExcelExporter


SCHEDULE_CONFIG = """
config firewall schedule recurring
    edit "business-hours"
        set start 08:00
        set end 18:00
        set day monday tuesday wednesday thursday friday
        set color 3
        set fabric-object enable
    next
    edit "weekend"
        set start 10:00
        set end 14:00
        set day saturday sunday
    next
    edit "nightly"
        set start 22:00
        set end 02:00
        set day sunday monday tuesday wednesday thursday
    next
end
config firewall schedule onetime
    edit "maintenance-window"
        set start "23:00 2026/08/25"
        set end "01:00 2026/08/26"
        set color 7
        set expiration-days 2
        set visibility disable
    next
    edit "one-time-no-expiry"
        set start "08:30 2026/09/01"
        set end "09:00 2026/09/01"
    next
end
"""


def test_recurring_and_one_time_schedules_preserve_source_semantics():
    result = extract_fortigate_config(SCHEDULE_CONFIG)
    parsed = parse_fortigate_config(SCHEDULE_CONFIG).schedules
    schedules = result.canonical_ir.schedules

    assert len(parsed) == 5
    assert len(schedules) == 5

    recurring = next(item for item in schedules if item.name == "business-hours")
    assert recurring.schedule_type == "recurring"
    assert recurring.start == "08:00"
    assert recurring.end == "18:00"
    assert recurring.days == ["monday", "tuesday", "wednesday", "thursday", "friday"]
    assert recurring.source_color == 3
    assert recurring.expiration_days is None
    assert recurring.source_attributes == {"fabric_object": "enable"}

    onetime = next(item for item in schedules if item.name == "maintenance-window")
    assert onetime.schedule_type == "onetime"
    assert onetime.start == "23:00 2026/08/25"
    assert onetime.end == "01:00 2026/08/26"
    assert onetime.days == []
    assert onetime.source_color == 7
    assert onetime.expiration_days == 2
    assert onetime.source_attributes == {"visibility": "disable"}

    no_expiry = next(item for item in schedules if item.name == "one-time-no-expiry")
    assert no_expiry.expiration_days is None

    coverage = {item.path: item for item in result.source_sections}
    assert coverage["firewall schedule recurring"].status == ExtractionStatus.NORMALIZED
    assert coverage["firewall schedule recurring"].object_count_parsed == 3
    assert coverage["firewall schedule onetime"].status == ExtractionStatus.NORMALIZED
    assert coverage["firewall schedule onetime"].object_count_parsed == 2


def test_schedules_excel_contains_recurring_and_one_time_rows():
    result = extract_fortigate_config(SCHEDULE_CONFIG)
    workbook = load_workbook(
        io.BytesIO(
            IRExcelExporter(
                result.canonical_ir,
                extraction_result=result,
            ).generate()
        )
    )
    sheet = workbook["Schedules"]
    headers = {cell.value: cell.column for cell in sheet[3]}
    rows = {
        sheet.cell(row, headers["Name"]).value: row
        for row in range(4, sheet.max_row + 1)
    }

    assert len(rows) == 5
    assert sheet.cell(rows["business-hours"], headers["Type"]).value == "recurring"
    assert sheet.cell(rows["maintenance-window"], headers["Type"]).value == "onetime"
    assert sheet.cell(rows["maintenance-window"], headers["Expiration Days"]).value == 2
    assert sheet.cell(rows["one-time-no-expiry"], headers["Expiration Days"]).value is None
