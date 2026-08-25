import io

from openpyxl import load_workbook

from fwmigrate.parsers.fortigate.extractor import extract_fortigate_config
from fwmigrate.parsers.fortigate.parser import parse_fortigate_config
from fwmigrate.parsers.fortigate.transformer import FGToIRTransformer
from fwmigrate.report.excel_exporter import IRExcelExporter


SSH_CONFIG = '''config firewall ssh local-key
    edit "admin-key"
        set public-key "ssh-rsa AAAASANITIZED"
        set private-key "PRIVATE_KEY_SENTINEL"
        set password "PASSWORD_SENTINEL"
        set source imported
        set comment "Safe metadata"
    next
end
config firewall ssh local-ca
    edit "admin-ca"
        set public-key "ssh-ed25519 AAAACA"
        set private-key "PRIVATE_KEY_SENTINEL"
        set passwd "PASSWORD_SENTINEL"
        set source generated
    next
end
'''


def test_ssh_keys_preserve_public_metadata_and_discard_secrets() -> None:
    parsed = parse_fortigate_config(SSH_CONFIG)
    assert [(item.name, item.key_type) for item in parsed.ssh_keys] == [
        ("admin-key", "local-key"),
        ("admin-ca", "local-ca"),
    ]
    key = parsed.ssh_keys[0]
    assert key.public_key == "ssh-rsa AAAASANITIZED"
    assert key.source == "imported"
    assert key.has_private_key is True
    assert key.has_password is True
    assert key.extra_settings == {"comment": "Safe metadata"}

    ir = FGToIRTransformer(parsed).transform()
    assert ir.ssh_keys[1].public_key == "ssh-ed25519 AAAACA"
    assert ir.ssh_keys[1].has_private_key is True
    assert ir.ssh_keys[1].has_password is True

    result = extract_fortigate_config(SSH_CONFIG)
    workbook = load_workbook(
        io.BytesIO(IRExcelExporter(result.canonical_ir, result).generate())
    )
    assert "SSH Keys" in workbook.sheetnames
    assert workbook["SSH Keys"][4][3].value == "Yes"

    serialized = parsed.model_dump_json() + ir.model_dump_json() + result.model_dump_json()
    all_cells = "\n".join(
        str(cell.value)
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if cell.value is not None
    )
    for forbidden in ("PRIVATE_KEY_SENTINEL", "PASSWORD_SENTINEL"):
        assert forbidden not in serialized
        assert forbidden not in all_cells
