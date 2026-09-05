import io

from openpyxl import load_workbook

from fwmigrate.extraction.models import ExtractionStatus
from fwmigrate.parsers.fortigate.extractor import extract_fortigate_config
from fwmigrate.parsers.fortigate.parser import FortiGateParser, parse_fortigate_config
from fwmigrate.parsers.fortigate.tokenizer import FortiGateTokenizer
from fwmigrate.report.excel_exporter import IRExcelExporter


def test_dhcp_v4_coverage_is_typed_extract_only_with_exact_child_counts() -> None:
    result = extract_fortigate_config("""
config system dhcp server
    edit 1
        config ip-range
            edit 3
            next
            edit 1
            next
        end
        config exclude-range
            edit 2
            next
        end
        config reserved-address
            edit 4
            next
        end
        config options
            edit 5
            next
            edit 6
            next
        end
    next
end
""")
    sections = {section.path: section for section in result.source_sections}
    for path, count in {
        "system dhcp server": 1,
        "system dhcp server ip-range": 2,
        "system dhcp server exclude-range": 1,
        "system dhcp server reserved-address": 1,
        "system dhcp server options": 2,
    }.items():
        assert sections[path].status == ExtractionStatus.EXTRACT_ONLY
        assert sections[path].object_count_source == count
        assert sections[path].object_count_parsed == count
        assert sections[path].object_count_normalized == count


def test_typed_operational_parents_keep_context_and_redact_credentials() -> None:
    content = """config vdom
edit "root"
    config system sdn-connector
        edit "sdn1"
            set type aws
            set server "10.0.0.10"
            set api-key "SDN_SECRET"
        next
    end
    config firewall network-service-dynamic
        edit "dynamic1"
            set filter "cloud"
            set sdn "sdn1"
        next
    end
    config user radius
        edit "radius1"
            set server "10.0.0.20"
            set secondary-server "10.0.0.21"
            set tertiary-server "10.0.0.22"
            set auth-type ms_chap_v2
            set transport-protocol tcp
            set tls-min-proto-version tls1-2
            set ca-cert "RADIUS_CA"
            set client-cert "RADIUS_CLIENT"
            set nas-ip "192.0.2.20"
            set nas-ip6 "2001:db8::20"
            set source-ip "192.0.2.21"
            set source-ip6 "2001:db8::21"
            set radius-port 1812
            set auth-port 1812
            set acct-port 1813
            set coa-port 3799
            set timeout 10
            set retries 3
            set acct-interim-interval 300
            set secret "RADIUS_SECRET"
            set tertiary-secret "TERTIARY_SECRET"
            set class "staff" "vpn"
            set switch-controller-service-type login authenticate
            set account-key-cert-field rfc822name
            set account-key-processing strip
            set use-management-vdom enable
            set vrf-select 7
            set username-case-sensitive enable
            config accounting-server
                edit "1"
                    set status enable
                    set server "10.0.0.23"
                    set port 1813
                    set source-ip "192.0.2.22"
                    set source-ip6 "2001:db8::22"
                    set interface-select-method specify
                    set interface "mgmt"
                    set vrf-select 8
                    set secret "ACCOUNTING_SECRET"
                next
            end
        next
    end
    config user tacacs+
        edit "tacacs1"
            set server "10.0.0.30"
            set secondary-server "10.0.0.31"
            set tertiary-server "10.0.0.32"
            set port 49
            set authen-type pap
            set authorization enable
            set source-ip "10.0.0.5"
            set interface-select-method specify
            set interface "mgmt"
            set vrf-select 9
            set status-ttl 300
            set secondary-key "TACACS_SECONDARY_SECRET"
            set tertiary-key "TACACS_TERTIARY_SECRET"
            set key-string "TACACS_SECRET"
        next
    end
    config system link-monitor
        edit "wan-monitor"
            set srcintf "wan1" "wan2"
            set server "1.1.1.1" "1.0.0.1"
            set protocol ping http
            set gateway-ip 192.0.2.1
            set source-ip 192.0.2.2
            set port 443
            set timeout 500
            set update-static-route enable
            set update-policy-route disable
            set update-cascade-interface enable
        next
    end
    config vpn ipsec manualkey-interface
        edit "manual1"
            set interface "wan1"
            set encryption-key "ENC_SECRET"
            set authentication-key "AUTH_SECRET"
        next
    end
next
end
"""

    parser = FortiGateParser(FortiGateTokenizer(content))
    parsed = parser.parse()

    assert parsed.sdn_connectors[0].source_context == "root"
    assert parsed.sdn_connectors[0].server == "10.0.0.10"
    assert parsed.sdn_connectors[0].has_secret is True
    assert parsed.radius_servers[0].has_secret is True
    assert parsed.radius_servers[0].transport_protocol == "tcp"
    assert parsed.radius_servers[0].class_ == ["staff", "vpn"]
    assert parsed.radius_servers[0].switch_controller_service_type == [
        "login", "authenticate"
    ]
    assert parsed.radius_servers[0].account_key_processing == "strip"
    assert parsed.radius_servers[0].vrf_select == 7
    assert parsed.radius_servers[0].accounting_servers[0].has_secret is True
    assert parsed.radius_servers[0].accounting_servers[0].source_ip6 == "2001:db8::22"
    assert parsed.radius_servers[0].accounting_servers[0].vrf_select == 8
    assert parsed.tacacs_servers[0].has_secret is True
    assert parsed.link_monitors[0].server == ["1.1.1.1", "1.0.0.1"]
    assert parsed.link_monitors[0].srcintf == ["wan1", "wan2"]
    assert parsed.link_monitors[0].protocol == ["ping", "http"]
    assert (
        parsed.link_monitors[0].gateway_ip,
        parsed.link_monitors[0].source_ip,
        parsed.link_monitors[0].port,
        parsed.link_monitors[0].timeout,
    ) == ("192.0.2.1", "192.0.2.2", 443, 500)
    assert (
        parsed.link_monitors[0].update_static_route,
        parsed.link_monitors[0].update_policy_route,
        parsed.link_monitors[0].update_cascade_interface,
    ) == ("enable", "disable", "enable")
    assert parsed.manualkey_interfaces[0].has_encryption_key is True
    assert parsed.manualkey_interfaces[0].has_authentication_key is True

    serialized = parsed.model_dump_json()
    for secret in (
        "SDN_SECRET", "RADIUS_SECRET", "TACACS_SECRET",
        "TERTIARY_SECRET",
        "TACACS_SECONDARY_SECRET", "TACACS_TERTIARY_SECRET",
        "ENC_SECRET", "AUTH_SECRET",
    ):
        assert secret not in serialized
    result = extract_fortigate_config(content)
    radius = result.canonical_ir.user_radius_servers[0]
    assert (
        radius.server, radius.secondary_server, radius.tertiary_server,
        radius.auth_type, radius.port, radius.acct_interim_interval,
        radius.nas_ip, radius.source_ip,
    ) == (
        "10.0.0.20", "10.0.0.21", "10.0.0.22", "ms_chap_v2", 1812,
        300, "192.0.2.20", "192.0.2.21",
    )
    assert radius.accounting_servers[0].server == "10.0.0.23"
    assert radius.accounting_servers[0].port == 1813
    assert radius.accounting_servers[0].has_secret is True
    radius_section = next(section for section in result.source_sections if section.path == "user radius")
    assert (
        radius_section.parser_handler,
        radius_section.object_count_parsed,
        radius_section.object_count_normalized,
    ) == ("FortiGateParser.build_model", 1, 1)
    tacacs = result.canonical_ir.user_tacacs_servers[0]
    assert (
        tacacs.server, tacacs.secondary_server, tacacs.tertiary_server,
        tacacs.port, tacacs.authentication_type, tacacs.authorization,
        tacacs.interface, tacacs.status_ttl,
    ) == ("10.0.0.30", "10.0.0.31", "10.0.0.32", 49, "pap", "enable", "mgmt", 300)
    assert parsed.tacacs_servers[0].vrf_select == 9
    tacacs_section = next(section for section in result.source_sections if section.path == "user tacacs+")
    assert (
        tacacs_section.parser_handler,
        tacacs_section.object_count_parsed,
        tacacs_section.object_count_normalized,
    ) == ("FortiGateParser.build_model", 1, 1)
    workbook = load_workbook(
        io.BytesIO(IRExcelExporter(result.canonical_ir, result).generate())
    )
    tacacs_sheet = workbook["TACACS+ Servers"]
    tacacs_headers = {cell.value: cell.column for cell in tacacs_sheet[3]}
    assert tacacs_sheet.cell(4, tacacs_headers["Secret Configured"]).value == "Yes"
    assert tacacs_sheet.cell(4, tacacs_headers["Status TTL"]).value == 300
    assert all(secret not in "\n".join(
        str(cell.value)
        for row in tacacs_sheet.iter_rows()
        for cell in row
        if cell.value is not None
    ) for secret in ("TACACS_SECRET", "TACACS_SECONDARY_SECRET", "TACACS_TERTIARY_SECRET"))
    radius_sheet = workbook["RADIUS Servers"]
    radius_headers = {cell.value: cell.column for cell in radius_sheet[3]}
    assert radius_sheet.cell(4, radius_headers["Accounting Interim Interval"]).value == 300
    accounting_sheet = workbook["RADIUS Accounting Servers"]
    accounting_headers = {cell.value: cell.column for cell in accounting_sheet[3]}
    assert accounting_sheet.cell(4, accounting_headers["Server"]).value == "10.0.0.23"
    assert accounting_sheet.cell(4, accounting_headers["Secret Configured"]).value == "Yes"
    assert "ACCOUNTING_SECRET" not in "\n".join(
        str(cell.value)
        for row in accounting_sheet.iter_rows()
        for cell in row
        if cell.value is not None
    )


def test_dynamic_service_dependencies_are_resolved_within_source_context() -> None:
    content = """config firewall network-service-dynamic
    edit "dynamic1"
        set sdn "sdn1"
    next
end
config system sdn-connector
    edit "sdn1"
        set type aws
    next
end
config firewall policy
    edit 1
        set srcintf "wan1"
        set dstintf "lan"
        set srcaddr "all"
        set dstaddr "all"
        set service "ALL"
        set action accept
        set network-service-dynamic "dynamic1"
    next
end
"""

    result = extract_fortigate_config(content)
    dependency_pairs = {
        (dependency.source_path, dependency.source_field, dependency.reference, dependency.result)
        for dependency in result.dependencies
    }
    assert (
        "firewall network-service-dynamic",
        "sdn",
        "sdn1",
        "RESOLVED",
    ) in dependency_pairs
    assert (
        "firewall policy",
        "network-service-dynamic",
        "dynamic1",
        "RESOLVED",
    ) in dependency_pairs


def test_identity_dependency_chain_reaches_radius_source_inventory() -> None:
    content = """config user radius
    edit "radius1"
        set server "10.0.0.20"
        set secret "RADIUS_SECRET"
    next
end
config user group
    edit "vpn-users"
        set member "radius1"
    next
end
config firewall policy
    edit 1
        set srcintf "any"
        set dstintf "any"
        set srcaddr "all"
        set dstaddr "all"
        set service "ALL"
        set action accept
        set groups "vpn-users"
    next
end
"""

    result = extract_fortigate_config(content)
    chain = [
        dependency for dependency in result.dependencies
        if dependency.reference in {"vpn-users", "radius1"}
    ]
    assert {(dependency.source_path, dependency.reference, dependency.result) for dependency in chain} == {
        ("firewall policy", "vpn-users", "RESOLVED"),
        ("user group", "radius1", "RESOLVED"),
    }


def test_profile_group_nested_security_dependencies_are_inventory_resolvable() -> None:
    content = """config firewall profile-group
    edit "secure"
        set ssh-filter-profile "ssh-prod"
        config ssh-filter
            edit "ssh-prod"
                set status enable
            next
        end
    next
end
config firewall policy
    edit 1
        set srcintf "any"
        set dstintf "any"
        set srcaddr "all"
        set dstaddr "all"
        set service "ALL"
        set action accept
        set profile-group "secure"
    next
end
"""

    result = extract_fortigate_config(content)
    assert any(
        dependency.source_path == "firewall profile-group"
        and dependency.reference == "ssh-prod"
        and dependency.result == "RESOLVED"
        and dependency.target_path == "firewall profile-group ssh-filter"
        for dependency in result.dependencies
    )


def test_unresolved_reference_is_partial_and_reported_without_broadening_policy() -> None:
    content = """config firewall policy
    edit 1
        set srcintf "missing-interface"
        set dstintf "any"
        set srcaddr "all"
        set dstaddr "all"
        set service "ALL"
        set action accept
    next
end
"""

    result = extract_fortigate_config(content)
    unresolved = [dependency for dependency in result.dependencies if dependency.result == "UNRESOLVED"]
    assert len(unresolved) == 1
    assert unresolved[0].reference == "missing-interface"
    assert result.generation_safe is False
    policy_section = next(section for section in result.source_sections if section.path == "firewall policy")
    assert policy_section.status == ExtractionStatus.PARTIALLY_NORMALIZED
    assert policy_section.unresolved_dependencies == 1
    assert any("missing-interface" in entry.message for entry in result.canonical_ir.audit_entries)

    workbook = load_workbook(
        io.BytesIO(IRExcelExporter(result.canonical_ir, result).generate())
    )
    assert "Unresolved References" in workbook.sheetnames
    unresolved_sheet = workbook["Unresolved References"]
    values = "\n".join(
        str(cell.value)
        for row in unresolved_sheet.iter_rows()
        for cell in row
        if cell.value is not None
    )
    assert "missing-interface" in values


def test_dns_source_semantics_are_partial_even_when_typed_vendor_fields_exist() -> None:
    content = """config system dns
    set primary 1.1.1.1
    set secondary 8.8.8.8
    set protocol dot
    set server-select-method failover
    set domain example.com
end
"""

    result = extract_fortigate_config(content)
    dns_section = next(section for section in result.source_sections if section.path == "system dns")
    assert dns_section.status == ExtractionStatus.PARTIALLY_NORMALIZED
    assert {"protocol", "server_select_method", "domain"} <= set(dns_section.semantic_unknowns)
    assert result.canonical_ir.dns_settings.primary == "1.1.1.1"
    assert result.canonical_ir.dns_settings.secondary == "8.8.8.8"


def test_coverage_reports_semantic_tiers_for_typed_and_structured_sections() -> None:
    result = extract_fortigate_config('''
config user local
    edit "alice"
        set status enable
        set passwd-time 123
    next
end
config user radius
    edit "radius"
        set server "192.0.2.1"
        set obscure-setting value
    next
end
config ips sensor
    edit "IPS1"
        config entries
            edit 1
                set action block
            next
        end
    next
end
config webfilter profile
    edit "WF1"
        config web
            set feature enable
        end
    next
end
''')
    sections = {section.path: section for section in result.source_sections}
    assert any("Semantic support level: TYPED_EXTRACT_ONLY" in note for note in sections["user local"].notes)
    assert any("Semantic support level: TYPED_EXTRACT_ONLY" in note for note in sections["user radius"].notes)
    assert any("Semantic support level: TYPED_EXTRACT_ONLY" in note for note in sections["ips sensor"].notes)
    assert any("Dedicated typed profile semantics" in note for note in sections["webfilter profile"].notes)


def test_dos_policy_identity_includes_address_family() -> None:
    content = """config firewall DoS-policy
    edit 1
        set status enable
        set interface "wan4"
        set srcaddr "all"
        set dstaddr "all"
        set service "ALL"
    next
end
config firewall DoS-policy6
    edit 1
        set status enable
        set interface "wan6"
        set srcaddr "all"
        set dstaddr "all"
        set service "ALL"
    next
end
"""

    result = extract_fortigate_config(content)
    assert [(policy.source_id, policy.address_family, policy.interface) for policy in result.canonical_ir.dos_policies] == [
        (1, "ipv4", "wan4"),
        (1, "ipv6", "wan6"),
    ]


def test_unknown_migration_relevant_sections_are_recursively_captured() -> None:
    content = """config firewall future-policy
    edit "future1"
        set action accept
        set api-key "FUTURE_SECRET"
        config nested-setting
            edit "child1"
                set value preserved
            next
        end
    next
end
"""

    result = extract_fortigate_config(content)
    section = result.source_sections[0]
    assert section.status == ExtractionStatus.EXTRACT_ONLY_UNKNOWN
    assert result.generation_safe is False
    root = next(item for item in result.inventory_items if item.source_path == "firewall future-policy")
    assert root.name == "future1"
    assert root.children[0].children[0].name == "child1"
    serialized = result.model_dump_json()
    assert "FUTURE_SECRET" not in serialized


def test_webfilter_url_sections_are_structured_extract_only_including_empty() -> None:
    content = """config webfilter search-engine
    edit "search"
        set safe-search enable
    next
end
config webfilter ips-urlfilter-setting
end
config webfilter ips-urlfilter-setting6
    edit "ipv6"
        set status enable
    next
end
"""

    result = extract_fortigate_config(content)
    sections = {section.path: section for section in result.source_sections}
    assert {
        path: sections[path].status
        for path in (
            "webfilter search-engine",
            "webfilter ips-urlfilter-setting",
            "webfilter ips-urlfilter-setting6",
        )
    } == {
        path: ExtractionStatus.EXTRACT_ONLY
        for path in (
            "webfilter search-engine",
            "webfilter ips-urlfilter-setting",
            "webfilter ips-urlfilter-setting6",
        )
    }
    assert any(
        item.source_path == "webfilter ips-urlfilter-setting"
        for item in result.inventory_items
    )


def test_system_interface_static_and_secondary_ip_746_fields() -> None:
    content = """config system interface
    edit "port1"
        set mode static
        set ip 192.0.2.1 255.255.255.0
        set allowaccess ping https ssh
        set distance 10
        set priority 5
        set gwdetect enable
        set ha-priority 50
        set ping-serv "192.0.2.254"
        set secondary-IP enable
        config secondaryip
            edit 1
                set ip 192.0.2.2 255.255.255.0
                set allowaccess ping
                set gwdetect enable
                set ha-priority 60
                set ping-serv "192.0.2.253"
            next
        end
    next
end
"""
    intf = parse_fortigate_config(content).interfaces[0]
    assert intf.mode == "static"
    assert intf.ip == "192.0.2.1 255.255.255.0"
    assert intf.distance == 10
    assert intf.priority == 5
    assert intf.gwdetect == "enable"
    assert intf.ha_priority == 50
    assert intf.ping_serv == "192.0.2.254"
    assert intf.secondary_ip == "enable"
    assert len(intf.secondary_ips) == 1
    sec = intf.secondary_ips[0]
    assert sec.ip == "192.0.2.2 255.255.255.0"
    assert sec.gwdetect == "enable"
    assert sec.ha_priority == 60
    assert sec.ping_serv == "192.0.2.253"
    assert {"distance", "priority", "gwdetect", "ha_priority", "ping_serv"} <= intf.source_explicit_fields


def test_system_interface_dhcp_and_relay_746_fields() -> None:
    content = """config system interface
    edit "dhcp-wan"
        set mode dhcp
        set distance 20
        set priority 15
        set defaultgw enable
        set dhcp-renew-time 7200
        set dhcp-client-identifier "client-42"
        set dhcp-relay-service enable
        set dhcp-relay-ip 198.51.100.1 198.51.100.2
        append dhcp-relay-ip 198.51.100.3
        set dhcp-relay-type regular
        set dhcp-relay-link-selection enable
        set dhcp-relay-interface-select-method specify
        set dhcp-relay-interface "lan1"
        set dhcp-snooping enable
        set dhcp-snooping-option82 enable
        set dhcp-snooping-trust enable
        config client-options
            edit 1
                set code 12
                set type string
                set value "my-hostname"
            next
        end
        config dhcp-snooping-server-list
            edit "srv1"
                set server-ip 198.51.100.10
            next
        end
    next
end
"""
    intf = parse_fortigate_config(content).interfaces[0]
    assert intf.mode == "dhcp"
    assert intf.distance == 20
    assert intf.priority == 15
    assert intf.defaultgw == "enable"
    assert intf.dhcp_renew_time == 7200
    assert intf.dhcp_client_identifier == "client-42"
    assert intf.dhcp_relay_service == "enable"
    assert intf.dhcp_relay_ip == ["198.51.100.1", "198.51.100.2", "198.51.100.3"]
    assert intf.dhcp_relay_type == "regular"
    assert intf.dhcp_relay_link_selection == "enable"
    assert intf.dhcp_relay_interface_select_method == "specify"
    assert intf.dhcp_relay_interface == "lan1"
    assert intf.dhcp_snooping == "enable"
    assert intf.dhcp_snooping_option82 == "enable"
    assert intf.dhcp_snooping_trust == "enable"
    assert len(intf.client_options) == 1
    assert intf.client_options[0].code == 12
    assert intf.client_options[0].value == "my-hostname"
    assert len(intf.dhcp_snooping_server_list) == 1
    assert intf.dhcp_snooping_server_list[0].server_ip == "198.51.100.10"
    assert {
        "distance", "priority", "defaultgw", "dhcp_renew_time",
        "dhcp_client_identifier", "dhcp_relay_service", "dhcp_relay_ip",
        "dhcp_relay_type", "dhcp_relay_link_selection",
        "dhcp_relay_interface_select_method", "dhcp_relay_interface",
        "dhcp_snooping", "dhcp_snooping_option82", "dhcp_snooping_trust",
    } <= intf.source_explicit_fields


def test_system_interface_vlan_746_fields() -> None:
    content = """config system interface
    edit "vlan100"
        set type vlan
        set vlanid 100
        set interface "port1"
        set vlan-protocol 8021q
        set switch "switch1"
        set vlanforward enable
    next
end
"""
    intf = parse_fortigate_config(content).interfaces[0]
    assert intf.type == "vlan"
    assert intf.vlanid == 100
    assert intf.interface == "port1"
    assert intf.vlan_protocol == "8021q"
    assert intf.switch == "switch1"
    assert intf.vlanforward == "enable"
    assert {"vlan_protocol", "switch"} <= intf.source_explicit_fields


def test_system_interface_aggregate_and_redundant_746_fields() -> None:
    content = """config system interface
    edit "agg1"
        set type aggregate
        set member "port1" "port2"
        set lacp-select-timeout 30
        set bandwidth 20000
        set link-up-delay 100
        set link-down-delay 50
    next
    edit "agg2"
        set type aggregate
        set lacp-select-timeout 30
        set bandwidth 20000
        unset lacp-select-timeout
        unset bandwidth
    next
end
"""
    parsed = parse_fortigate_config(content)
    agg1 = parsed.interfaces[0]
    assert agg1.type == "aggregate"
    assert agg1.members == ["port1", "port2"]
    assert agg1.lacp_select_timeout == 30
    assert agg1.bandwidth == 20000
    assert agg1.link_up_delay == 100
    assert agg1.link_down_delay == 50
    assert {"lacp_select_timeout", "bandwidth"} <= agg1.source_explicit_fields

    agg2 = parsed.interfaces[1]
    assert agg2.lacp_select_timeout is None
    assert agg2.bandwidth is None
    assert "lacp_select_timeout" not in agg2.source_explicit_fields
    assert "bandwidth" not in agg2.source_explicit_fields


def test_system_interface_type_dependent_and_vrrp6_746_fields() -> None:
    content = """config system interface
    edit "port10"
        set type physical
        set speed 1000auto
        set mediatype sr-lr
        set fec cl74-fc-fec
        set flowcontrol both
        set fortilink enable
        set fortilink-neighbor-detect enable
        set auto-auth-extension enable
        set security-mode 802.1X
        set security-mac-auth enable
        set security-exempt-list "exempt-profile"
        set security-redirect-url "https://auth.corp.test"
        set management-ip 192.0.2.99 255.255.255.0
        set ip-managed-by-fortiipam enable
        config ipv6
            config vrrp6
                edit 1
                    set vrip6 2001:db8::fe
                    set priority 180
                    set start-time 45
                    set vrdst6 2001:db8::1
                    set vrdst6-priority 25
                next
            end
        end
    next
end
"""
    intf = parse_fortigate_config(content).interfaces[0]
    assert intf.type == "physical"
    assert intf.speed == "1000auto"
    assert intf.mediatype == "sr-lr"
    assert intf.fec == "cl74-fc-fec"
    assert intf.flowcontrol == "both"
    assert intf.fortilink == "enable"
    assert intf.fortilink_neighbor_detect == "enable"
    assert intf.auto_auth_extension == "enable"
    assert intf.security_mode == "802.1X"
    assert intf.security_mac_auth == "enable"
    assert intf.security_exempt_list == "exempt-profile"
    assert intf.security_redirect_url == "https://auth.corp.test"
    assert intf.management_ip == "192.0.2.99 255.255.255.0"
    assert intf.ip_managed_by_fortiipam == "enable"
    assert {
        "fec", "flowcontrol", "fortilink", "fortilink_neighbor_detect",
        "auto_auth_extension", "security_mode", "security_mac_auth",
        "security_exempt_list", "security_redirect_url", "management_ip",
        "ip_managed_by_fortiipam",
    } <= intf.source_explicit_fields

    assert len(intf.vrrp6) == 1
    vrrp6 = intf.vrrp6[0]
    assert vrrp6.vrip6 == "2001:db8::fe"
    assert vrrp6.priority == 180
    assert vrrp6.start_time == 45
    assert vrrp6.vrdst6 == "2001:db8::1"
    assert vrrp6.vrdst6_priority == 25


def test_router_static6_devindex_handling_and_ipv4_unaffected() -> None:
    content = """config router static6
    edit 1
        set dst 2001:db8:1::/64
        set gateway 2001:db8::1
        set device "wan1"
        set devindex 12
    next
    edit 2
        set dst 2001:db8:2::/64
        set gateway 2001:db8::2
        set device "wan1"
        set devindex 14
        unset devindex
    next
    edit 3
        set dst 2001:db8:3::/64
        set gateway 2001:db8::3
        set device "wan1"
    next
    edit 4
        set dst 2001:db8:4::/64
        set gateway 2001:db8::4
        set device "wan1"
        set devindex invalid-index
    next
end
config router static
    edit 10
        set dst 10.0.0.0 255.255.255.0
        set gateway 192.0.2.1
        set device "wan1"
    next
end
"""
    routes = parse_fortigate_config(content).static_routes

    # Configured static6 devindex
    r1 = routes[0]
    assert r1.address_family == "ipv6"
    assert r1.dst == "2001:db8:1::/64"
    assert r1.gateway == "2001:db8::1"
    assert r1.device == "wan1"
    assert r1.devindex == 12
    assert "devindex" in r1.source_explicit_fields

    # Unset static6 devindex
    r2 = routes[1]
    assert r2.address_family == "ipv6"
    assert r2.dst == "2001:db8:2::/64"
    assert r2.gateway == "2001:db8::2"
    assert r2.device == "wan1"
    assert r2.devindex is None
    assert "devindex" not in r2.source_explicit_fields

    # Omitted static6 devindex
    r3 = routes[2]
    assert r3.address_family == "ipv6"
    assert r3.dst == "2001:db8:3::/64"
    assert r3.gateway == "2001:db8::3"
    assert r3.device == "wan1"
    assert r3.devindex is None
    assert "devindex" not in r3.source_explicit_fields

    # Malformed static6 devindex
    r4 = routes[3]
    assert r4.address_family == "ipv6"
    assert r4.dst == "2001:db8:4::/64"
    assert r4.gateway == "2001:db8::4"
    assert r4.device == "wan1"
    assert r4.devindex is None
    assert r4.extra_settings.get("unparsed_devindex") == "invalid-index"
    assert "devindex" in r4.source_explicit_fields

    # IPv4 route is completely unaffected
    r_v4 = routes[4]
    assert r_v4.address_family == "ipv4"
    assert r_v4.dst == "10.0.0.0 255.255.255.0"
    assert r_v4.gateway == "192.0.2.1"
    assert r_v4.device == "wan1"
    assert r_v4.devindex is None
    assert "devindex" not in r_v4.source_explicit_fields


def test_system_dns_746_typed_fields_and_explicit_omitted_values() -> None:
    content_explicit = """config system dns
    set primary 198.51.100.1
    set secondary 198.51.100.2
    set alt-primary 198.51.100.3
    set alt-secondary 198.51.100.4
    set ip6-primary 2001:db8::1
    set ip6-secondary 2001:db8::2
    set protocol cleartext dot doh
    set server-select-method failover
    set domain "example.corp" "internal.local"
    set server-hostname "dns.example.corp"
    set interface-select-method specify
    set interface "wan1"
    set source-ip 198.51.100.50
    set source-ip6 2001:db8::50
    set ssl-certificate "Custom_DNS_Cert"
    set timeout 4
    set retry 3
    set dns-cache-limit 10000
    set dns-cache-ttl 3600
    set cache-notfound-responses enable
    set fqdn-cache-ttl 300
    set fqdn-max-refresh 7200
    set fqdn-min-refresh 120
    set log error
end
"""
    cfg_explicit = parse_fortigate_config(content_explicit)
    dns = cfg_explicit.dns
    assert dns is not None
    assert dns.primary == "198.51.100.1"
    assert dns.secondary == "198.51.100.2"
    assert dns.alt_primary == "198.51.100.3"
    assert dns.alt_secondary == "198.51.100.4"
    assert dns.ip6_primary == "2001:db8::1"
    assert dns.ip6_secondary == "2001:db8::2"
    assert dns.protocol == ["cleartext", "dot", "doh"]
    assert dns.server_select_method == "failover"
    assert dns.domain == ["example.corp", "internal.local"]
    assert dns.server_hostname == "dns.example.corp"
    assert dns.interface_select_method == "specify"
    assert dns.interface == "wan1"
    assert dns.source_ip == "198.51.100.50"
    assert dns.source_ip6 == "2001:db8::50"
    assert dns.ssl_certificate == "Custom_DNS_Cert"
    assert dns.timeout == 4
    assert dns.retry == 3
    assert dns.dns_cache_limit == 10000
    assert dns.dns_cache_ttl == 3600
    assert dns.cache_notfound_responses == "enable"
    assert dns.fqdn_cache_ttl == 300
    assert dns.fqdn_max_refresh == 7200
    assert dns.fqdn_min_refresh == 120
    assert dns.log == "error"

    expected_explicit = {
        "primary", "secondary", "alt_primary", "alt_secondary",
        "ip6_primary", "ip6_secondary", "protocol", "server_select_method",
        "domain", "server_hostname", "interface_select_method", "interface",
        "source_ip", "source_ip6", "ssl_certificate", "timeout", "retry",
        "dns_cache_limit", "dns_cache_ttl", "cache_notfound_responses",
        "fqdn_cache_ttl", "fqdn_max_refresh", "fqdn_min_refresh", "log",
    }
    assert expected_explicit <= dns.source_explicit_fields

    content_minimal = """config system dns
    set primary 8.8.8.8
end
"""
    cfg_minimal = parse_fortigate_config(content_minimal)
    min_dns = cfg_minimal.dns
    assert min_dns is not None
    assert min_dns.primary == "8.8.8.8"
    assert min_dns.secondary is None
    assert min_dns.alt_primary is None
    assert min_dns.alt_secondary is None
    assert min_dns.protocol == []
    assert min_dns.interface is None
    assert min_dns.interface_select_method is None
    assert min_dns.source_ip is None
    assert min_dns.source_ip6 is None
    assert min_dns.ssl_certificate is None
    assert min_dns.timeout is None
    assert min_dns.retry is None
    assert min_dns.dns_cache_limit is None
    assert min_dns.dns_cache_ttl is None
    assert min_dns.cache_notfound_responses is None
    assert min_dns.fqdn_cache_ttl is None
    assert min_dns.fqdn_max_refresh is None
    assert min_dns.fqdn_min_refresh is None
    assert min_dns.log is None
    assert min_dns.source_explicit_fields == {"primary"}

    content_unset = """config system dns
    set timeout 4
    set retry 3
    set interface "wan1"
    set protocol dot
    unset timeout
    unset retry
    unset interface
    unset protocol
end
"""
    cfg_unset = parse_fortigate_config(content_unset)
    unset_dns = cfg_unset.dns
    assert unset_dns.timeout is None
    assert unset_dns.retry is None
    assert unset_dns.interface is None
    assert unset_dns.protocol == []
    assert not {"timeout", "retry", "interface", "protocol"}.intersection(unset_dns.source_explicit_fields)
    assert not {"timeout", "retry", "interface", "protocol"}.intersection(unset_dns.extra_settings.keys())


def test_system_link_monitor_server_list_hierarchy_and_ordering() -> None:
    content = """config system link-monitor
    edit "wan-health"
        set srcintf "wan1" "wan2"
        set protocol ping
        set gateway-ip 192.0.2.1
        set source-ip 192.0.2.100
        set port 80
        set interval 1000
        set timeout 2000
        set failtime 3
        set recoverytime 5
        set update-static-route enable
        set update-policy-route disable
        set update-cascade-interface enable
        config server-list
            edit 1
                set dst 198.51.100.10
                set protocol ping
                set port 0
                set weight 10
            next
            edit 2
                set dst 198.51.100.20
                set protocol tcp-echo http
                set port 8080
                set weight 20
            next
            edit 3
                set server 198.51.100.30
                set protocol twamp
                set port 862
            next
        end
    next
    edit "secondary-monitor"
        set srcintf "port1"
        config server-list
            edit 1
                set dst 8.8.8.8
            next
        end
    next
end
"""
    cfg = parse_fortigate_config(content)
    assert len(cfg.link_monitors) == 2

    mon1 = cfg.link_monitors[0]
    assert mon1.name == "wan-health"
    assert mon1.srcintf == ["wan1", "wan2"]
    assert mon1.protocol == ["ping"]
    assert mon1.gateway_ip == "192.0.2.1"
    assert mon1.source_ip == "192.0.2.100"
    assert mon1.port == 80
    assert mon1.interval == 1000
    assert mon1.timeout == 2000
    assert mon1.failtime == 3
    assert mon1.recoverytime == 5
    assert mon1.update_static_route == "enable"
    assert mon1.update_policy_route == "disable"
    assert mon1.update_cascade_interface == "enable"

    assert len(mon1.nested_configs) > 0
    assert mon1.nested_configs[0].name == "server-list"
    assert len(mon1.nested_configs[0].children) == 3

    assert len(mon1.server_list) == 3
    assert mon1.servers == mon1.server_list

    s1 = mon1.server_list[0]
    assert s1.id == 1
    assert s1.dst == "198.51.100.10"
    assert s1.server == "198.51.100.10"
    assert s1.protocol == ["ping"]
    assert s1.port == 0
    assert s1.weight == 10
    assert {"dst", "protocol", "port", "weight"} <= s1.source_explicit_fields

    s2 = mon1.server_list[1]
    assert s2.id == 2
    assert s2.dst == "198.51.100.20"
    assert s2.server == "198.51.100.20"
    assert s2.protocol == ["tcp-echo", "http"]
    assert s2.port == 8080
    assert s2.weight == 20
    assert {"dst", "protocol", "port", "weight"} <= s2.source_explicit_fields

    s3 = mon1.server_list[2]
    assert s3.id == 3
    assert s3.dst == "198.51.100.30"
    assert s3.server == "198.51.100.30"
    assert s3.protocol == ["twamp"]
    assert s3.port == 862
    assert s3.weight is None
    assert "weight" not in s3.source_explicit_fields

    assert [s.id for s in mon1.server_list] == [1, 2, 3]

    mon2 = cfg.link_monitors[1]
    assert mon2.name == "secondary-monitor"
    assert len(mon2.server_list) == 1
    assert mon2.server_list[0].dst == "8.8.8.8"


