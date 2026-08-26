import ast
from collections import Counter
import io
from pathlib import Path

from openpyxl import load_workbook

from fwmigrate.ir.core import (
    IRAddress,
    IRAddressGroup,
    IRAuditEntry,
    IRCertificate,
    IRConfig,
    IRFSSOADGroup,
    IRFSSOProvider,
    IRInterface,
    IRInterfaceSecondaryIP,
    IRIPPool,
    IRLocalUser,
    IRMetadata,
    IRNATRule,
    IRPolicy,
    IRRoute,
    IRSecurityProfileGroup,
    IRService,
    IRServiceCategory,
    IRServiceGroup,
    IRServicePort,
    IRSSHKey,
    IRUserGroup,
    IRUserLDAP,
    IRVPNTunnel,
    IRZone,
)
from fwmigrate.ir.enums import (
    AddressType,
    MigrationConfidence,
    NATType,
    PolicyAction,
    ServiceProtocol,
)
from fwmigrate.ir.version import IR_SCHEMA_VERSION
from fwmigrate.report.excel_exporter import IRExcelExporter
from fwmigrate.parsers.fortigate.parser import parse_fortigate_config
from fwmigrate.parsers.fortigate.transformer import FGToIRTransformer
from fwmigrate.parsers.fortigate.extractor import extract_fortigate_config


def _sample_ir() -> IRConfig:
    return IRConfig(
        metadata=IRMetadata(
            hostname="HQ-FW-東京",
            source_vendor="fortigate",
            input_type="Configuration File",
            source_version="7.4.x",
            source_context="root",
        ),
        interfaces=[IRInterface(name="port1", zone="trust", ip="10.0.0.1/24")],
        zones=[IRZone(name="trust", interfaces=["port1"])],
        addresses=[
            IRAddress(name="Users", type=AddressType.NETWORK, subnet="10.0.0.0/24"),
            IRAddress(
                name="=HYPERLINK(\"https://invalid.example\")",
                type=AddressType.FQDN,
                fqdn="app.example.test",
                description="Unicode ✓ " + ("x" * 40000),
            ),
        ],
        address_groups=[
            IRAddressGroup(
                name="User Group",
                members=["Users", "Remote Users"],
                source_uuid="address-group-uuid",
                allow_routing=True,
                source_color=25,
                source_category="ztna-ems-tag",
                source_attributes={"visibility": "enable"},
            )
        ],
        service_categories=[
            IRServiceCategory(
                name="Web Access",
                description="Web access.",
                source_attributes={"color": "1"},
            )
        ],
        services=[
            IRService(
                name="Web",
                source_uuid="service-uuid",
                source_category="Web Access",
                source_protocol="tcp/udp/sctp",
                ports=[
                    IRServicePort(
                        protocol=ServiceProtocol.TCP,
                        port="443",
                        source_port="1024-65535",
                        raw_source_value="443:1024-65535",
                    ),
                    IRServicePort(protocol=ServiceProtocol.UDP, port="443"),
                ],
                source_proxy=False,
                source_attributes={"helper": "https"},
            )
        ],
        service_groups=[
            IRServiceGroup(
                name="Web Group",
                members=["Web"],
                source_uuid="service-group-uuid",
                source_attributes={"color": "4"},
            )
        ],
        policies=[
            IRPolicy(
                name="Allow-Web",
                source_rule_id="25",
                source_uuid="0819b852-ebb4-51eb-210e-517744c1e41b",
                source_from_interfaces=["LAN"],
                source_to_interfaces=["WAN"],
                source_address_references=["Users", "Remote Users"],
                destination_address_references=["all"],
                source_service_references=["Web"],
                source_action="accept",
                source_schedule="always",
                source_user_groups=["SSLVPN Users", "Domain_Users"],
                source_users=["alice", "bob.smith"],
                source_log_setting="all",
                source_inspection_mode="proxy",
                source_ztna_status="enable",
                source_ztna_ems_tags=["TAG_A", "TAG B"],
                source_extra_settings={
                    "timeout_send_rst": "enable",
                    "port_preserve": "disable",
                },
                nat_enabled=True,
                nat_pool_enabled=True,
                nat_pool_names=["PUBLIC_POOL"],
                from_zone=["trust"],
                to_zone=["untrust"],
                source=["Users", "Remote Users"],
                destination=["any"],
                service=["Web"],
                action=PolicyAction.ALLOW,
            )
        ],
        fsso_providers=[
            IRFSSOProvider(
                name="corp-fsso",
                server="10.10.10.10",
                has_password=True,
                source_attributes={"custom_option": "test"},
            )
        ],
        fsso_ad_groups=[
            IRFSSOADGroup(
                name="CORP/DOMAIN USERS",
                provider_name="corp-fsso",
                provider_resolved=True,
            )
        ],
        ip_pools=[
            IRIPPool(
                name="PUBLIC_POOL",
                pool_type="overload",
                start_ip="203.0.113.10",
                end_ip="203.0.113.20",
                associated_interface="wan1",
                arp_reply=True,
                permit_any_host=False,
                excluded_ips=["203.0.113.11", "203.0.113.12"],
                description="Internet SNAT pool",
            )
        ],
        nat_rules=[
            IRNATRule(
                name="Outbound-NAT",
                type=NATType.TWICE,
                source_policy_reference="25",
                source_policy_uuid="nat-policy-uuid",
                enabled=False,
                source_from_interfaces=["LAN"],
                source_to_interfaces=["WAN"],
                from_zone=["trust"],
                to_zone=["untrust"],
                source=["Users"],
                destination=["any"],
                services=["Web", "HTTPS"],
                internet_services=["Microsoft-Office365"],
                source_translation_mode="pool",
                source_pool_references=["PUBLIC_POOL"],
                translated_sources=["203.0.113.10"],
                source_vip_reference="VIP_WEB",
                source_vip_group_reference="VIP_GROUP",
                translated_destinations=["10.0.0.10"],
                original_destination_port="8443",
                translated_port="443",
                requires_manual_review=True,
            )
        ],
        vpn_tunnels=[
            IRVPNTunnel(
                name="VPN-HQ",
                peer_address="192.0.2.1",
                local_interface="port1",
                psk="do-not-export-this-secret",
            )
        ],
        routes=[IRRoute(name="default", destination="0.0.0.0/0", next_hop="10.0.0.254")],
        security_profile_groups=[IRSecurityProfileGroup(name="strict", antivirus="default")],
        audit_entries=[
            IRAuditEntry(
                id="warn-1",
                category="Reference",
                message="Unresolved object requires review",
                confidence=MigrationConfidence.PARTIAL,
            ),
            IRAuditEntry(
                id="unsupported-1",
                category="router bgp",
                message="No IR mapping implemented for bgp router",
                confidence=MigrationConfidence.UNSUPPORTED,
            ),
        ],
    )


def test_excel_exporter_generates_complete_safe_workbook():
    workbook_bytes = IRExcelExporter(_sample_ir()).generate()
    workbook = load_workbook(io.BytesIO(workbook_bytes), data_only=False)

    assert workbook.sheetnames == list(IRExcelExporter.SHEET_ORDER)
    assert workbook["Addresses"].max_row == 5  # title, note, header, and exactly two objects
    assert workbook["Addresses"]["A4"].value == "Users"
    assert workbook["Addresses"]["A5"].value.startswith("'")
    assert workbook["Addresses"]["A5"].data_type == "s"
    address_headers = {
        cell.value: cell.column
        for cell in workbook["Addresses"][3]
    }
    assert len(
        workbook["Addresses"].cell(
            5,
            address_headers["Description"],
        ).value
    ) <= 32767
    assert workbook["Address Groups"]["C4"].value == "Users\nRemote Users"
    assert workbook["Address Groups"]["B4"].value == "address-group-uuid"
    assert workbook["Service Categories"]["A4"].value == "Web Access"
    services = workbook["Services"]
    service_headers = {cell.value: cell.column for cell in services[3]}
    assert services.cell(4, service_headers["Source UUID"]).value == "service-uuid"
    assert services.cell(4, service_headers["Category"]).value == "Web Access"
    assert services.cell(4, service_headers["Source Port Constraint"]).value == "1024-65535"
    assert services.cell(4, service_headers["Additional Settings"]).value == "helper=https"
    service_groups = workbook["Service Groups"]
    group_headers = {cell.value: cell.column for cell in service_groups[3]}
    assert service_groups.cell(4, group_headers["Source UUID"]).value == "service-group-uuid"
    policy_headers = {
        cell.value: cell.column for cell in workbook["Policies"][3]
    }
    assert workbook["Policies"].cell(
        4, policy_headers["Source Address (Normalized)"]
    ).value == "Users\nRemote Users"
    assert workbook["VPN Tunnels"]["E4"].value == "Configured / Redacted"
    assert workbook["FSSO Servers"]["A4"].value == "corp-fsso"
    assert workbook["FSSO Servers"]["C4"].value == "Yes"
    assert workbook["FSSO AD Groups"]["A4"].value == "CORP/DOMAIN USERS"
    assert workbook["FSSO AD Groups"]["C4"].value == "Yes"
    assert workbook["Extraction Coverage"]["A4"].value == "Interfaces"

    all_text = "\n".join(
        str(cell.value)
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if cell.value is not None
    )
    assert "do-not-export-this-secret" not in all_text
    assert "HQ-FW-東京" in all_text


def test_excel_exporter_exposes_source_policy_audit_fields():
    workbook = load_workbook(io.BytesIO(IRExcelExporter(_sample_ir()).generate()))
    policies = workbook["Policies"]

    assert [cell.value for cell in policies[3]] == [
        "Rule #", "Source Policy ID", "Source UUID", "Name", "Source Interface", "From Zone",
        "Destination Interface", "To Zone", "Source Address (FortiGate)",
        "Source Address (Normalized)", "Destination Address (FortiGate)",
        "Destination Address (Normalized)", "User Groups", "Users",
        "Service (FortiGate)", "Service (Normalized)", "Action (FortiGate)",
        "Action (Normalized)", "Schedule (FortiGate)", "Schedule (Normalized)",
        "Disabled", "Log Setting", "Log Start", "Log End",
        "NAT Enabled", "IP Pool Enabled", "NAT Pool", "Applications",
        "Internet Services", "Security Profile Group", "Antivirus", "IPS Sensor",
        "Web Filter", "Application List", "SSL/SSH Profile", "Inspection Mode",
        "ZTNA Status", "ZTNA EMS Tags", "Additional Settings", "Description",
    ]
    headers = {cell.value: cell.column for cell in policies[3]}
    assert policies.cell(4, headers["Rule #"]).value == 1
    assert policies.cell(4, headers["Source Policy ID"]).value == "25"
    assert policies.cell(4, headers["Source UUID"]).value == (
        "0819b852-ebb4-51eb-210e-517744c1e41b"
    )
    assert policies.cell(4, headers["Source Interface"]).value == "LAN"
    assert policies.cell(4, headers["Destination Interface"]).value == "WAN"
    assert policies.cell(4, headers["Source Address (FortiGate)"]).value == (
        "Users\nRemote Users"
    )
    assert policies.cell(4, headers["Source Address (Normalized)"]).value == (
        "Users\nRemote Users"
    )
    assert policies.cell(4, headers["Destination Address (FortiGate)"]).value == "all"
    assert policies.cell(4, headers["Destination Address (Normalized)"]).value == "any"
    assert policies.cell(4, headers["Service (FortiGate)"]).value == "Web"
    assert policies.cell(4, headers["Service (Normalized)"]).value == "Web"
    assert policies.cell(4, headers["Action (FortiGate)"]).value == "accept"
    assert policies.cell(4, headers["Action (Normalized)"]).value == "allow"
    assert policies.cell(4, headers["Schedule (FortiGate)"]).value == "always"
    assert policies.cell(4, headers["Schedule (Normalized)"]).value is None
    assert policies.cell(4, headers["User Groups"]).value == "SSLVPN Users\nDomain_Users"
    assert policies.cell(4, headers["Users"]).value == "alice\nbob.smith"
    assert policies.cell(4, headers["Log Setting"]).value == "all"
    assert policies.cell(4, headers["NAT Enabled"]).value == "TRUE"
    assert policies.cell(4, headers["IP Pool Enabled"]).value == "TRUE"
    assert policies.cell(4, headers["NAT Pool"]).value == "PUBLIC_POOL"
    assert policies.cell(4, headers["Inspection Mode"]).value == "proxy"
    assert policies.cell(4, headers["ZTNA Status"]).value == "enable"
    assert policies.cell(4, headers["ZTNA EMS Tags"]).value == "TAG_A\nTAG B"
    additional = policies.cell(4, headers["Additional Settings"]).value
    assert "timeout-send-rst=enable" in additional
    assert "port-preserve=disable" in additional


def test_excel_exporter_leaves_empty_policy_identity_selectors_blank():
    ir = IRConfig(
        metadata=IRMetadata(hostname="minimal", source_vendor="fortigate"),
        policies=[IRPolicy(name="No_Identity", action=PolicyAction.DENY)],
    )
    workbook = load_workbook(io.BytesIO(IRExcelExporter(ir).generate()))

    headers = {
        cell.value: cell.column for cell in workbook["Policies"][3]
    }
    assert workbook["Policies"].cell(4, headers["User Groups"]).value is None
    assert workbook["Policies"].cell(4, headers["Users"]).value is None


def test_excel_exporter_includes_ip_pool_inventory_and_existing_nat_output():
    workbook = load_workbook(io.BytesIO(IRExcelExporter(_sample_ir()).generate()))
    pools = workbook["IP Pools"]

    assert [cell.value for cell in pools[3]] == [
        "Name", "Type", "Start IP", "End IP", "Source Start IP",
        "Source End IP", "Start Port", "End Port", "Associated Interface",
        "ARP Reply", "ARP Interface", "Permit Any Host", "Excluded IPs",
        "Block Size", "Blocks Per User", "PBA Timeout", "Ports Per User",
        "NAT64", "TCP Session Quota", "UDP Session Quota",
        "ICMP Session Quota", "Description",
    ]
    assert pools["A4"].value == "PUBLIC_POOL"
    assert pools["B4"].value == "overload"
    assert pools["C4"].value == "203.0.113.10"
    assert pools["D4"].value == "203.0.113.20"
    assert pools["I4"].value == "wan1"
    assert pools["J4"].value == "TRUE"
    assert pools["L4"].value == "FALSE"
    assert pools["M4"].value == "203.0.113.11\n203.0.113.12"
    assert pools["V4"].value == "Internet SNAT pool"

    summary_counts = {
        workbook["Summary"].cell(row, 1).value: workbook["Summary"].cell(row, 2).value
        for row in range(1, workbook["Summary"].max_row + 1)
    }
    assert summary_counts["IP Pools"] == 1
    assert summary_counts["IR Schema Version"] == IR_SCHEMA_VERSION

    coverage_rows = {
        workbook["Extraction Coverage"].cell(row, 1).value:
            workbook["Extraction Coverage"].cell(row, 3).value
        for row in range(4, workbook["Extraction Coverage"].max_row + 1)
    }
    assert coverage_rows["IP Pools"] == 1
    nat_rules = workbook["NAT Rules"]
    assert [cell.value for cell in nat_rules[3]] == [
        "Rule #", "Name", "Type", "Source Policy ID", "Source Policy UUID",
        "Enabled", "Source Interface", "From Zone", "Destination Interface",
        "To Zone", "Original Source", "Original Destination", "Services",
        "Internet Services", "Source Translation Mode", "IP Pool",
        "Translated Source", "VIP", "VIP Group", "Translated Destination",
        "Original Destination Port", "Translated Port", "Manual Review", "Description",
    ]
    headers = {cell.value: cell.column for cell in nat_rules[3]}
    assert nat_rules.cell(4, headers["Name"]).value == "Outbound-NAT"
    assert nat_rules.cell(4, headers["Source Policy ID"]).value == "25"
    assert nat_rules.cell(4, headers["Source Policy UUID"]).value == "nat-policy-uuid"
    assert nat_rules.cell(4, headers["Enabled"]).value == "FALSE"
    assert nat_rules.cell(4, headers["Source Interface"]).value == "LAN"
    assert nat_rules.cell(4, headers["Destination Interface"]).value == "WAN"
    assert nat_rules.cell(4, headers["Services"]).value == "Web\nHTTPS"
    assert nat_rules.cell(4, headers["Internet Services"]).value == "Microsoft-Office365"
    assert nat_rules.cell(4, headers["Source Translation Mode"]).value == "pool"
    assert nat_rules.cell(4, headers["IP Pool"]).value == "PUBLIC_POOL"
    assert nat_rules.cell(4, headers["Translated Source"]).value == "203.0.113.10"
    assert nat_rules.cell(4, headers["VIP"]).value == "VIP_WEB"
    assert nat_rules.cell(4, headers["VIP Group"]).value == "VIP_GROUP"
    assert nat_rules.cell(4, headers["Translated Destination"]).value == "10.0.0.10"
    assert nat_rules.cell(4, headers["Translated Port"]).value == "443"
    assert nat_rules.cell(4, headers["Manual Review"]).value == "TRUE"


def test_excel_exporter_marks_missing_parser_coverage_as_unknown():
    ir = IRConfig(metadata=IRMetadata(hostname="minimal", source_vendor="juniper_srx"))
    workbook = load_workbook(io.BytesIO(IRExcelExporter(ir).generate()))
    coverage = workbook["Extraction Coverage"]

    assert coverage["D4"].value == "Not reported"
    assert coverage["E4"].value == "Empty / unknown"
    assert "awaits ExtractionResult" in coverage["F4"].value


def test_excel_exporter_uses_extraction_result_source_evidence():
    config = """config application list
    edit "inventory-only"
        set comment "retained"
    next
end
config switch-controller global
end
config system unknown-feature
    edit "x"
    next
end
"""
    extraction = extract_fortigate_config(config)
    workbook = load_workbook(
        io.BytesIO(
            IRExcelExporter(
                extraction.canonical_ir,
                extraction_result=extraction,
            ).generate()
        )
    )

    coverage = workbook["Extraction Coverage"]
    headers = {cell.value: cell.column for cell in coverage[3]}
    rows = {
        coverage.cell(row, headers["Source Section"]).value: row
        for row in range(4, coverage.max_row + 1)
    }
    assert coverage.cell(
        rows["application list"], headers["Status"]
    ).value == "EXTRACT_ONLY"
    assert coverage.cell(
        rows["switch-controller global"], headers["Status"]
    ).value == "IGNORED_BY_POLICY"
    assert coverage.cell(
        rows["system unknown-feature"], headers["Status"]
    ).value == "UNSUPPORTED"
    assert coverage.cell(
        rows["system unknown-feature"], headers["Line Start"]
    ).value == 8
    assert "unavailable" not in coverage["A2"].value

    unsupported = workbook["Unsupported"]
    assert unsupported["A4"].value == "system unknown-feature"
    assert unsupported["C4"].value == "UNSUPPORTED"
    assert unsupported["E4"].value == "Yes"


def test_fortigate_interface_source_settings_are_exported():
    config = """
config system interface
    edit "x1"
        set vdom "root"
        set mode dhcp
        set allowaccess ping
        set type physical
        set lldp-reception disable
        set role wan
        set snmp-index 3
    next
end
    """
    ir = FGToIRTransformer(parse_fortigate_config(config)).transform()
    workbook = load_workbook(io.BytesIO(IRExcelExporter(ir).generate()))

    interfaces = workbook["Interfaces"]
    headers = {cell.value: cell.column for cell in interfaces[3]}
    assert interfaces.cell(4, headers["Name"]).value == "x1"
    assert interfaces.cell(4, headers["Source VDOM"]).value == "root"
    assert interfaces.cell(4, headers["Interface Type"]).value == "physical"
    assert interfaces.cell(4, headers["Role"]).value == "wan"
    assert interfaces.cell(4, headers["Addressing Mode"]).value == "dhcp"
    assert interfaces.cell(4, headers["DHCP Client"]).value == "Yes"
    assert interfaces.cell(4, headers["Management Access"]).value == "ping"

    settings = workbook["Interface Source Settings"]
    extracted = {
        settings.cell(row, 3).value: settings.cell(row, 4).value
        for row in range(4, settings.max_row + 1)
    }
    assert extracted["lldp-reception"] == "disable"
    assert extracted["snmp-index"] == "3"


def test_fortigate_tunnel_remote_ip_is_exported_with_source_evidence():
    config = """
config system interface
    edit "Tunnel_With_IP"
        set vdom "root"
        set ip 10.255.0.1 255.255.255.255
        set type tunnel
        set remote-ip 10.255.0.2 255.255.255.255
        set interface "port1"
    next
    edit "Tunnel_No_IP"
        set type tunnel
        set interface "port1"
    next
end
    """
    ir = FGToIRTransformer(parse_fortigate_config(config)).transform()
    workbook = load_workbook(io.BytesIO(IRExcelExporter(ir).generate()))

    interfaces = workbook["Interfaces"]
    headers = {cell.value: cell.column for cell in interfaces[3]}
    rows = {
        interfaces.cell(row, headers["Name"]).value: row
        for row in range(4, interfaces.max_row + 1)
    }

    with_ip_row = rows["Tunnel_With_IP"]
    assert interfaces.cell(with_ip_row, headers["IP / Prefix"]).value == "10.255.0.1/32"
    assert interfaces.cell(with_ip_row, headers["Remote IP / Prefix"]).value == "10.255.0.2/32"
    assert interfaces.cell(with_ip_row, headers["Interface Type"]).value == "tunnel"
    assert interfaces.cell(
        with_ip_row, headers["Parent / Underlay Interface"]
    ).value == "port1"

    no_ip_row = rows["Tunnel_No_IP"]
    assert interfaces.cell(no_ip_row, headers["IP / Prefix"]).value is None
    assert interfaces.cell(no_ip_row, headers["Remote IP / Prefix"]).value is None
    assert interfaces.cell(no_ip_row, headers["Interface Type"]).value == "tunnel"
    assert interfaces.cell(
        no_ip_row, headers["Parent / Underlay Interface"]
    ).value == "port1"

    settings = workbook["Interface Source Settings"]
    remote_ip_rows = [
        row for row in range(4, settings.max_row + 1)
        if settings.cell(row, 1).value == "Tunnel_With_IP"
        and settings.cell(row, 3).value == "remote-ip"
    ]
    assert len(remote_ip_rows) == 1
    remote_ip_row = remote_ip_rows[0]
    assert settings.cell(remote_ip_row, 4).value == "10.255.0.2 255.255.255.255"
    assert settings.cell(remote_ip_row, 5).value == "EXTRACT_ONLY"


def test_unresolved_interface_zone_exports_as_blank():
    ir = IRConfig(
        metadata=IRMetadata(hostname="edge-fw", source_vendor="fortigate"),
        interfaces=[IRInterface(name="port1", role="wan", zone=None)],
    )
    workbook = load_workbook(io.BytesIO(IRExcelExporter(ir).generate()))
    interfaces = workbook["Interfaces"]
    headers = {cell.value: cell.column for cell in interfaces[3]}

    assert interfaces.cell(4, headers["Name"]).value == "port1"
    assert interfaces.cell(4, headers["Zone"]).value is None
    assert interfaces.cell(4, headers["Role"]).value == "wan"


def test_invalid_route_source_and_parse_error_are_visible_in_excel():
    ir = IRConfig(
        metadata=IRMetadata(hostname="edge-fw", source_vendor="fortigate"),
        routes=[
            IRRoute(
                name="route_20",
                destination=None,
                source_destination="10.20.30.0 255.0.255.0",
                next_hop="192.0.2.1",
                requires_manual_review=True,
                parse_error="Invalid IPv4 network",
                source_attributes={"priority": "7"},
            )
        ],
    )
    workbook = load_workbook(io.BytesIO(IRExcelExporter(ir).generate()))
    routes = workbook["Routes"]
    headers = {cell.value: cell.column for cell in routes[3]}

    assert routes.cell(4, headers["Destination"]).value is None
    assert routes.cell(4, headers["Source Destination"]).value == (
        "10.20.30.0 255.0.255.0"
    )
    assert routes.cell(4, headers["Manual Review"]).value == "Yes"
    assert routes.cell(4, headers["Parse Error"]).value == "Invalid IPv4 network"


def test_route_excel_keeps_administrative_distance_separate_from_metric():
    ir = IRConfig(
        metadata=IRMetadata(hostname="edge-fw", source_vendor="fortigate"),
        routes=[
            IRRoute(
                name="route_10",
                source_route_id=10,
                destination="10.10.0.0/16",
                source_destination="10.10.0.0 255.255.0.0",
                administrative_distance=5,
                metric=None,
                priority=20,
                blackhole=False,
                enabled=False,
                sdwan_zone="virtual-wan-link",
            )
        ],
    )
    workbook = load_workbook(io.BytesIO(IRExcelExporter(ir).generate()))
    routes = workbook["Routes"]
    headers = {cell.value: cell.column for cell in routes[3]}

    assert routes.cell(4, headers["Source Route ID"]).value == 10
    assert routes.cell(4, headers["Administrative Distance"]).value == 5
    assert routes.cell(4, headers["Metric"]).value is None
    assert routes.cell(4, headers["Priority"]).value == 20
    assert routes.cell(4, headers["Blackhole"]).value == "No"
    assert routes.cell(4, headers["Enabled"]).value == "No"
    assert routes.cell(4, headers["SD-WAN Zone"]).value == "virtual-wan-link"
    assert routes.cell(4, headers["Migration Status"]).value == "NORMALIZED"

def _summary_navigation_rows(workbook):
    summary = workbook["Summary"]

    header_row = None

    for row in range(
        1,
        summary.max_row + 1,
    ):
        if (
            summary.cell(row, 1).value == "Category"
            and summary.cell(row, 2).value == "Sheet"
            and summary.cell(row, 3).value == "Records"
        ):
            header_row = row
            break

    assert header_row is not None

    rows = {}

    for row in range(
        header_row + 1,
        summary.max_row + 1,
    ):
        sheet_name = summary.cell(
            row,
            2,
        ).value

        if not sheet_name:
            break

        rows[sheet_name] = row

    return rows


def _tab_rgb(sheet):
    color = sheet.sheet_properties.tabColor

    assert color is not None

    return str(color.rgb or "").upper()


def test_excel_exporter_uses_logical_sheet_order():
    workbook = load_workbook(
        io.BytesIO(
            IRExcelExporter(
                _sample_ir()
            ).generate()
        )
    )

    assert workbook.sheetnames[0] == "Summary"

    assert workbook.sheetnames[-3:] == [
        "Warnings",
        "Unsupported",
        "Extraction Coverage",
    ]

    assert (
        workbook.sheetnames.index("Policies")
        <
        workbook.sheetnames.index(
            "Interface Source Settings"
        )
    )

    assert (
        workbook.sheetnames.index("VPN Phase 2")
        <
        workbook.sheetnames.index(
            "Source Security Profile Setting"
        )
    )

    assert len(
        IRExcelExporter.SHEET_ORDER
    ) == len(
        set(
            IRExcelExporter.SHEET_ORDER
        )
    )


def test_excel_exporter_summary_contains_navigation_links():
    workbook = load_workbook(
        io.BytesIO(
            IRExcelExporter(
                _sample_ir()
            ).generate()
        )
    )

    navigation = _summary_navigation_rows(
        workbook
    )

    for required_sheet in (
        "Interfaces",
        "Addresses",
        "Policies",
        "IP Pools",
        "VPN Tunnels",
        "Warnings",
        "Unsupported",
        "Extraction Coverage",
    ):
        assert required_sheet in navigation

        row = navigation[
            required_sheet
        ]

        cell = workbook[
            "Summary"
        ].cell(
            row,
            2,
        )

        assert cell.hyperlink is not None

        assert cell.hyperlink.target == (
            f"#'{required_sheet}'!A1"
        )

    policies_row = navigation["Policies"]

    assert workbook[
        "Summary"
    ].cell(
        policies_row,
        3,
    ).value == len(
        _sample_ir().policies
    )


def test_excel_exporter_non_summary_sheets_link_back_to_summary():
    workbook = load_workbook(
        io.BytesIO(
            IRExcelExporter(
                _sample_ir()
            ).generate()
        )
    )

    for sheet_name in (
        "Interfaces",
        "Policies",
        "Interface Source Settings",
        "Warnings",
        "Extraction Coverage",
    ):
        sheet = workbook[
            sheet_name
        ]

        assert sheet["A2"].hyperlink is not None

        assert (
            sheet["A2"].hyperlink.target
            == "#'Summary'!A1"
        )

        assert "Back to Summary" in str(
            sheet["A2"].value
        )


def test_excel_exporter_applies_sheet_group_tab_colors():
    workbook = load_workbook(
        io.BytesIO(
            IRExcelExporter(
                _sample_ir()
            ).generate()
        )
    )

    assert _tab_rgb(
        workbook["Summary"]
    ).endswith(
        IRExcelExporter._NAVY
    )

    assert _tab_rgb(
        workbook["Policies"]
    ).endswith(
        IRExcelExporter._TEAL
    )

    assert _tab_rgb(
        workbook[
            "Interface Source Settings"
        ]
    ).endswith(
        IRExcelExporter._MUTED
    )

    assert _tab_rgb(
        workbook["Unsupported"]
    ).endswith(
        IRExcelExporter._LIGHT_RED
    )


def test_excel_exporter_preserves_table_navigation_features():
    workbook = load_workbook(
        io.BytesIO(
            IRExcelExporter(
                _sample_ir()
            ).generate()
        )
    )

    interfaces = workbook[
        "Interfaces"
    ]

    assert (
        interfaces.sheet_view.showGridLines
        is False
    )

    assert (
        interfaces.freeze_panes
        == "A4"
    )

    assert interfaces.auto_filter.ref

    policies = workbook[
        "Policies"
    ]

    assert (
        policies.freeze_panes
        == "E4"
    )

    assert policies.auto_filter.ref


def test_excel_exporter_warning_highlight_is_limited_to_confidence():
    workbook = load_workbook(
        io.BytesIO(
            IRExcelExporter(
                _sample_ir()
            ).generate()
        )
    )

    warnings = workbook[
        "Warnings"
    ]

    # Sample warning row uses PARTIAL confidence.
    confidence_fill = (
        warnings["C4"]
        .fill
        .fgColor
        .rgb
    )

    message_fill = (
        warnings["D4"]
        .fill
        .fgColor
    )

    assert confidence_fill is not None

    # The warning emphasis should be localized rather than filling the
    # complete warning row.
    assert message_fill != confidence_fill


def test_excel_exporter_preserves_names_with_sensitive_keywords_without_false_positive_redaction():
    names = [
        "DELEUM/KEY ADMINS",
        "DELEUM/ENTERPRISE KEY ADMINS",
        "DELEUM/ALLOWED RODC PASSWORD REPLICATION GROUP",
        "DELEUM/DENIED RODC PASSWORD REPLICATION GROUP",
    ]
    ir = IRConfig(
        metadata=IRMetadata(hostname="HQ-FW", source_vendor="fortigate"),
        fsso_ad_groups=[
            IRFSSOADGroup(name=name, provider_name="fsso-srv", provider_resolved=True)
            for name in names
        ],
        addresses=[
            IRAddress(name="DELEUM/KEY ADMINS", type=AddressType.FQDN, fqdn="key-admins.deleum.com"),
        ],
        address_groups=[
            IRAddressGroup(name="DELEUM/ENTERPRISE KEY ADMINS", members=["DELEUM/KEY ADMINS"]),
        ],
        user_groups=[
            IRUserGroup(name="DELEUM/ALLOWED RODC PASSWORD REPLICATION GROUP", members=["DELEUM/KEY ADMINS"]),
            IRUserGroup(name="DELEUM/DENIED RODC PASSWORD REPLICATION GROUP", members=["DELEUM/KEY ADMINS"]),
        ],
    )
    workbook = load_workbook(io.BytesIO(IRExcelExporter(ir).generate()))

    # Verify FSSO AD Groups sheet
    fsso_sheet = workbook["FSSO AD Groups"]
    fsso_names = [fsso_sheet.cell(r, 1).value for r in range(4, fsso_sheet.max_row + 1)]
    assert fsso_names == names

    # Verify Addresses sheet
    addr_sheet = workbook["Addresses"]
    assert addr_sheet["A4"].value == "DELEUM/KEY ADMINS"

    # Verify Address Groups sheet
    grp_sheet = workbook["Address Groups"]
    assert grp_sheet["A4"].value == "DELEUM/ENTERPRISE KEY ADMINS"
    assert grp_sheet["C4"].value == "DELEUM/KEY ADMINS"

    # Verify User Groups sheet
    ugrp_sheet = workbook["User Groups"]
    ugrp_names = [ugrp_sheet.cell(r, 1).value for r in range(4, ugrp_sheet.max_row + 1)]
    assert "DELEUM/ALLOWED RODC PASSWORD REPLICATION GROUP" in ugrp_names
    assert "DELEUM/DENIED RODC PASSWORD REPLICATION GROUP" in ugrp_names

    # Check across entire workbook that no name was partially replaced with asterisks
    all_text = "\n".join(
        str(cell.value)
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if cell.value is not None
    )
    for name in names:
        assert name in all_text
    assert "******" not in all_text


def test_excel_exporter_excludes_actual_secrets():
    ir = IRConfig(
        metadata=IRMetadata(hostname="HQ-FW", source_vendor="fortigate"),
        vpn_tunnels=[
            IRVPNTunnel(name="vpn1", peer_address="1.1.1.1", local_interface="wan1", psk="super_secret_psk_999"),
        ],
        local_users=[
            IRLocalUser(
                name="admin_user",
                has_password=True,
                source_attributes={"password": "[REDACTED]"},
            ),
        ],
        certificates=[
            IRCertificate(
                name="local_cert",
                certificate_type="local",
                has_private_key=True,
                has_password=True,
                source_attributes={"private_key": "[REDACTED]"},
            ),
        ],
        ssh_keys=[
            IRSSHKey(
                name="ssh1",
                key_type="local",
                has_private_key=True,
                has_password=True,
                source_attributes={"private_key": "[REDACTED]"},
            ),
        ],
    )
    workbook = load_workbook(io.BytesIO(IRExcelExporter(ir).generate()))

    # VPN PSK is redacted to 'Configured / Redacted'
    vpn_sheet = workbook["VPN Tunnels"]
    assert vpn_sheet["E4"].value == "Configured / Redacted"

    # Local user password flag is Yes
    user_sheet = workbook["Local Users"]
    assert user_sheet["D4"].value == "Yes"

    # Certificate private key flag is Yes
    cert_sheet = workbook["Certificates"]
    headers = {cell.value: cell.column for cell in cert_sheet[3]}
    assert cert_sheet.cell(4, headers["Has Private Key"]).value == "Yes"

    all_text = "\n".join(
        str(cell.value)
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if cell.value is not None
    )
    assert "super_secret_psk_999" not in all_text


def test_excel_exporter_leaves_absent_policy_profiles_blank_and_exports_explicit_profiles():
    ir = IRConfig(
        metadata=IRMetadata(hostname="HQ-FW", source_vendor="fortigate"),
        policies=[
            IRPolicy(
                name="Policy_Partial_UTM",
                source_rule_id="1",
                action=PolicyAction.ALLOW,
                security_profile_group="SPG_IPS_default",
                antivirus=None,
                ips_sensor="default",
                webfilter=None,
                application_list=None,
                ssl_ssh_profile="certificate-inspection",
            ),
            IRPolicy(
                name="Policy_Explicit_UTM",
                source_rule_id="2",
                action=PolicyAction.ALLOW,
                security_profile_group="SPG_AV_default_IPS_protect_WF_custom",
                antivirus="default",
                ips_sensor="protect_server",
                webfilter="custom_filter",
                application_list="app_ctrl",
                ssl_ssh_profile=None,
            ),
        ],
        security_profile_groups=[
            IRSecurityProfileGroup(
                name="SPG_IPS_default",
                antivirus=None,
                vulnerability="default",
                anti_spyware=None,
                url_filtering=None,
                file_blocking=None,
                wildfire=None,
                ssl_decryption="certificate-inspection",
            ),
        ],
    )
    workbook = load_workbook(io.BytesIO(IRExcelExporter(ir).generate()))

    pol_sheet = workbook["Policies"]
    headers = {cell.value: cell.column for cell in pol_sheet[3]}

    # Policy 1 (Partial UTM): absent AV/WF/AppCtrl must be blank (None), explicit IPS and SSL/SSH must be present
    assert pol_sheet.cell(4, headers["Antivirus"]).value is None
    assert pol_sheet.cell(4, headers["IPS Sensor"]).value == "default"
    assert pol_sheet.cell(4, headers["Web Filter"]).value is None
    assert pol_sheet.cell(4, headers["Application List"]).value is None
    assert pol_sheet.cell(4, headers["SSL/SSH Profile"]).value == "certificate-inspection"

    # Policy 2 (Explicit UTM): explicit AV="default", IPS="protect_server", WF="custom_filter", AppList="app_ctrl", SSL/SSH=None
    assert pol_sheet.cell(5, headers["Antivirus"]).value == "default"
    assert pol_sheet.cell(5, headers["IPS Sensor"]).value == "protect_server"
    assert pol_sheet.cell(5, headers["Web Filter"]).value == "custom_filter"
    assert pol_sheet.cell(5, headers["Application List"]).value == "app_ctrl"
    assert pol_sheet.cell(5, headers["SSL/SSH Profile"]).value is None

    # Security Profiles sheet: absent profiles must be blank
    spg_sheet = workbook["Security Profiles"]
    spg_headers = {cell.value: cell.column for cell in spg_sheet[3]}
    assert spg_sheet.cell(4, spg_headers["Name"]).value == "SPG_IPS_default"
    assert spg_sheet.cell(4, spg_headers["Antivirus"]).value is None
    assert spg_sheet.cell(4, spg_headers["Vulnerability"]).value == "default"
    assert spg_sheet.cell(4, spg_headers["Anti-Spyware"]).value is None
    assert spg_sheet.cell(4, spg_headers["URL Filtering"]).value is None
    assert spg_sheet.cell(4, spg_headers["File Blocking"]).value is None
    assert spg_sheet.cell(4, spg_headers["WildFire"]).value is None
    assert spg_sheet.cell(4, spg_headers["SSL Decryption"]).value == "certificate-inspection"


def test_excel_exporter_interface_secondary_ips_summary_navigation():
    ir = IRConfig(
        metadata=IRMetadata(hostname="fw-sec-test", source_vendor="fortigate"),
        interfaces=[
            IRInterface(
                name="port1",
                ip="10.0.0.1/24",
                secondary_ips=[
                    IRInterfaceSecondaryIP(
                        source_id="1",
                        source_ip="10.0.0.2 255.255.255.0",
                        ip="10.0.0.2/24",
                        requires_manual_review=True,
                    )
                ],
            )
        ],
    )
    workbook = load_workbook(io.BytesIO(IRExcelExporter(ir).generate()))
    navigation = _summary_navigation_rows(workbook)
    assert "Interface Secondary IPs" in navigation
    row = navigation["Interface Secondary IPs"]
    assert workbook["Summary"].cell(row, 1).value == "Core Inventory"
    assert workbook["Summary"].cell(row, 3).value == 1
    assert workbook["Summary"].cell(row, 5).value == "Yes"


def test_excel_exporter_no_duplicate_methods():
    source_path = (
        Path(__file__).resolve().parent.parent
        / "src"
        / "fwmigrate"
        / "report"
        / "excel_exporter.py"
    )
    tree = ast.parse(source_path.read_text(encoding="utf-8"))
    exporter_class = next(
        node
        for node in tree.body
        if isinstance(node, ast.ClassDef) and node.name == "IRExcelExporter"
    )

    method_names = [
        node.name
        for node in exporter_class.body
        if isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef))
    ]

    duplicates = {
        name: count
        for name, count in Counter(method_names).items()
        if count > 1
    }

    assert duplicates == {}
