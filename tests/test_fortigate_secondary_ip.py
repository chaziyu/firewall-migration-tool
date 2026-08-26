import io
import pytest
from openpyxl import load_workbook

from fwmigrate.extraction.models import ExtractionStatus
from fwmigrate.parsers.fortigate.parser import parse_fortigate_config
from fwmigrate.parsers.fortigate.transformer import FGToIRTransformer
from fwmigrate.parsers.fortigate.extractor import extract_fortigate_config
from fwmigrate.parsers.fortigate.model import FGInterfaceSecondaryIP
from fwmigrate.ir.core import IRInterfaceSecondaryIP, IRInterface, IRConfig, IRMetadata
from fwmigrate.report.excel_exporter import IRExcelExporter
from fwmigrate.core.registry import PluginRegistry
import fwmigrate.generators


def _transform(config: str) -> IRConfig:
    return FGToIRTransformer(parse_fortigate_config(config)).transform()


def test_parse_and_transform_valid_secondary_ips():
    config = """
config system interface
    edit "port1"
        set vdom "root"
        set ip 10.0.0.1 255.255.255.0
        set allowaccess ping https ssh
        set type physical
        set secondary-IP enable
        config secondaryip
            edit 1
                set ip 10.0.0.2 255.255.255.0
                set allowaccess ping https
            next
            edit 2
                set ip 10.0.0.3 255.255.255.0
                set allowaccess ping
            next
        end
    next
end
"""
    fg = parse_fortigate_config(config)
    assert len(fg.interfaces) == 1
    intf = fg.interfaces[0]
    assert intf.name == "port1"
    assert intf.ip == "10.0.0.1 255.255.255.0"
    assert intf.secondary_ip == "enable"
    assert intf.source_attributes.get("secondary_ip") == "enable"
    assert "secondary_ips" not in intf.source_attributes
    assert len(intf.secondary_ips) == 2

    sec1 = intf.secondary_ips[0]
    assert isinstance(sec1, FGInterfaceSecondaryIP)
    assert sec1.id == 1
    assert sec1.ip == "10.0.0.2 255.255.255.0"
    assert sec1.allowaccess == ["ping", "https"]
    assert sec1.extra_settings == {}

    sec2 = intf.secondary_ips[1]
    assert isinstance(sec2, FGInterfaceSecondaryIP)
    assert sec2.id == 2
    assert sec2.ip == "10.0.0.3 255.255.255.0"
    assert sec2.allowaccess == ["ping"]
    assert sec2.extra_settings == {}

    ir = FGToIRTransformer(fg).transform()
    assert len(ir.interfaces) == 1
    ir_intf = ir.interfaces[0]
    assert ir_intf.name == "port1"
    assert ir_intf.ip == "10.0.0.1/24"
    assert len(ir_intf.secondary_ips) == 2

    ir_sec1 = ir_intf.secondary_ips[0]
    assert isinstance(ir_sec1, IRInterfaceSecondaryIP)
    assert ir_sec1.source_id == "1"
    assert ir_sec1.source_ip == "10.0.0.2 255.255.255.0"
    assert ir_sec1.ip == "10.0.0.2/24"
    assert ir_sec1.management_access == ["ping", "https"]
    assert ir_sec1.requires_manual_review is False
    assert ir_sec1.parse_error is None
    assert ir_sec1.source_attributes == {}

    ir_sec2 = ir_intf.secondary_ips[1]
    assert isinstance(ir_sec2, IRInterfaceSecondaryIP)
    assert ir_sec2.source_id == "2"
    assert ir_sec2.source_ip == "10.0.0.3 255.255.255.0"
    assert ir_sec2.ip == "10.0.0.3/24"
    assert ir_sec2.management_access == ["ping"]
    assert ir_sec2.requires_manual_review is False
    assert ir_sec2.parse_error is None
    assert ir_sec2.source_attributes == {}


def test_parse_and_transform_secondary_ip_with_extra_settings():
    config = """
config system interface
    edit "port1"
        set ip 10.0.0.1 255.255.255.0
        set secondary-IP enable
        config secondaryip
            edit 1
                set ip 10.0.0.2 255.255.255.0
                set allowaccess ping
                set custom-future-option "test-val"
            next
        end
    next
end
"""
    fg = parse_fortigate_config(config)
    intf = fg.interfaces[0]
    sec = intf.secondary_ips[0]
    assert sec.extra_settings == {"custom_future_option": "test-val"}

    ir = FGToIRTransformer(fg).transform()
    ir_sec = ir.interfaces[0].secondary_ips[0]
    assert ir_sec.source_id == "1"
    assert ir_sec.source_ip == "10.0.0.2 255.255.255.0"
    assert ir_sec.ip == "10.0.0.2/24"
    assert ir_sec.requires_manual_review is True
    assert ir_sec.parse_error is None
    assert ir_sec.source_attributes == {"custom_future_option": "test-val"}

    # Audit entry verification
    audit = next(
        (
            e for e in ir.audit_entries
            if e.id == "interface:port1:secondaryip:1:source-settings"
        ),
        None,
    )
    assert audit is not None
    assert audit.category == "Interface Secondary IP"
    assert "custom_future_option" in audit.message


def test_parse_and_transform_invalid_secondary_ip_netmask():
    config = """
config system interface
    edit "port1"
        set ip 10.0.0.1 255.255.255.0
        set secondary-IP enable
        config secondaryip
            edit 1
                set ip 10.0.0.2 255.255.999.0
                set allowaccess ping
            next
        end
    next
end
"""
    ir = _transform(config)
    ir_sec = ir.interfaces[0].secondary_ips[0]
    assert ir_sec.source_id == "1"
    assert ir_sec.source_ip == "10.0.0.2 255.255.999.0"
    assert ir_sec.ip is None
    assert ir_sec.parse_error is not None
    assert ir_sec.requires_manual_review is True

    audit = next(
        (
            e for e in ir.audit_entries
            if e.id == "interface:port1:secondaryip:1"
        ),
        None,
    )
    assert audit is not None
    assert "invalid IP/netmask syntax" in audit.message


def test_parse_and_transform_missing_secondary_ip():
    config = """
config system interface
    edit "port1"
        set ip 10.0.0.1 255.255.255.0
        set secondary-IP enable
        config secondaryip
            edit 1
                set allowaccess ping
            next
        end
    next
end
"""
    ir = _transform(config)
    ir_sec = ir.interfaces[0].secondary_ips[0]
    assert ir_sec.source_id == "1"
    assert ir_sec.source_ip is None
    assert ir_sec.ip is None
    assert ir_sec.parse_error == "Missing source secondary IP value."
    assert ir_sec.requires_manual_review is True

    audit = next(
        (
            e for e in ir.audit_entries
            if e.id == "interface:port1:secondaryip:1"
        ),
        None,
    )
    assert audit is not None
    assert "has no configured IP/netmask value" in audit.message


def test_parse_and_transform_unusable_secondary_ip():
    config = """
config system interface
    edit "port1"
        set ip 10.0.0.1 255.255.255.0
        set secondary-IP enable
        config secondaryip
            edit 1
                set ip 0.0.0.0 0.0.0.0
                set allowaccess ping
            next
        end
    next
end
"""
    ir = _transform(config)
    ir_sec = ir.interfaces[0].secondary_ips[0]
    assert ir_sec.source_id == "1"
    assert ir_sec.source_ip == "0.0.0.0 0.0.0.0"
    assert ir_sec.ip is None
    assert ir_sec.parse_error is None
    assert ir_sec.requires_manual_review is True

    audit = next(
        (
            e for e in ir.audit_entries
            if e.id == "interface:port1:secondaryip:1"
        ),
        None,
    )
    assert audit is not None
    assert "does not represent a usable configured secondary address" in audit.message


def test_extraction_coverage_secondary_ips():
    valid_config = """
config system interface
    edit "port1"
        set ip 10.0.0.1 255.255.255.0
        set secondary-IP enable
        config secondaryip
            edit 1
                set ip 10.0.0.2 255.255.255.0
            next
        end
    next
end
"""
    result = extract_fortigate_config(valid_config)
    sec_section = next(
        (s for s in result.source_sections if s.path == "system interface secondaryip"),
        None,
    )
    assert sec_section is not None
    assert sec_section.status == ExtractionStatus.NORMALIZED
    assert sec_section.object_count_source == 1
    assert sec_section.object_count_parsed == 1
    assert sec_section.object_count_normalized == 1

    partial_config = """
config system interface
    edit "port1"
        set ip 10.0.0.1 255.255.255.0
        set secondary-IP enable
        config secondaryip
            edit 1
                set ip 10.0.0.2 255.255.999.0
            next
        end
    next
end
"""
    result_partial = extract_fortigate_config(partial_config)
    sec_section_partial = next(
        (s for s in result_partial.source_sections if s.path == "system interface secondaryip"),
        None,
    )
    assert sec_section_partial is not None
    assert sec_section_partial.status == ExtractionStatus.PARTIALLY_NORMALIZED
    assert any("missing or invalid IP/netmask values" in n for n in sec_section_partial.notes)


def test_excel_export_interface_secondary_ips():
    ir = IRConfig(
        metadata=IRMetadata(hostname="FW-1", source_vendor="fortigate"),
        interfaces=[
            IRInterface(
                name="port1",
                ip="10.0.0.1/24",
                secondary_ips=[
                    IRInterfaceSecondaryIP(
                        source_id="1",
                        source_ip="10.0.0.2 255.255.255.0",
                        ip="10.0.0.2/24",
                        management_access=["ping", "https"],
                        requires_manual_review=False,
                        parse_error=None,
                        source_attributes={},
                    ),
                    IRInterfaceSecondaryIP(
                        source_id="2",
                        source_ip="10.0.0.3 255.255.255.0",
                        ip="10.0.0.3/24",
                        management_access=["ping"],
                        requires_manual_review=True,
                        parse_error=None,
                        source_attributes={"setting_x": "val_y"},
                    ),
                    IRInterfaceSecondaryIP(
                        source_id="3",
                        source_ip="10.0.0.4 255.255.999.0",
                        ip=None,
                        management_access=[],
                        requires_manual_review=True,
                        parse_error="Invalid IPv4 mask: 255.255.999.0",
                        source_attributes={},
                    ),
                ],
            )
        ],
    )

    exporter = IRExcelExporter(ir)
    wb_bytes = exporter.generate()
    wb = load_workbook(io.BytesIO(wb_bytes), data_only=False)

    assert "Interface Secondary IPs" in wb.sheetnames
    assert (
        wb.sheetnames.index("Interface Secondary IPs")
        == wb.sheetnames.index("Interfaces") + 1
    )

    sheet = wb["Interface Secondary IPs"]
    assert sheet.max_row == 6  # Title, Subtitle, Headers + 3 rows

    # Check header row (row 3)
    headers = [cell.value for cell in sheet[3]]
    assert headers == [
        "Interface",
        "Source ID",
        "Source IP",
        "IP / Prefix",
        "Management Access",
        "Extraction Status",
        "Manual Review",
        "Parse Error",
        "Additional Settings",
    ]

    # Row 1 (sec 1: normalized)
    row4 = [sheet.cell(4, c).value for c in range(1, 10)]
    assert row4[0] == "port1"
    assert row4[1] == "1"
    assert row4[2] == "10.0.0.2 255.255.255.0"
    assert row4[3] == "10.0.0.2/24"
    assert "ping" in row4[4]
    assert row4[5] == "NORMALIZED"
    assert row4[6] == "FALSE"
    assert not row4[7]
    assert not row4[8]

    # Row 2 (sec 2: partially normalized)
    row5 = [sheet.cell(5, c).value for c in range(1, 10)]
    assert row5[0] == "port1"
    assert row5[1] == "2"
    assert row5[2] == "10.0.0.3 255.255.255.0"
    assert row5[3] == "10.0.0.3/24"
    assert row5[5] == "PARTIALLY_NORMALIZED"
    assert row5[6] == "TRUE"
    assert not row5[7]
    assert "setting-x=val_y" in row5[8]

    # Row 3 (sec 3: parse error)
    row6 = [sheet.cell(6, c).value for c in range(1, 10)]
    assert row6[0] == "port1"
    assert row6[1] == "3"
    assert row6[2] == "10.0.0.4 255.255.999.0"
    assert not row6[3]
    assert row6[5] == "PARSE_ERROR"
    assert row6[6] == "TRUE"
    assert "Invalid IPv4 mask" in row6[7]

    # Summary navigation check
    summary = wb["Summary"]
    nav_row = None
    for r in range(1, summary.max_row + 1):
        if summary.cell(r, 2).value == "Interface Secondary IPs":
            nav_row = r
            break
    assert nav_row is not None
    assert summary.cell(nav_row, 1).value == "Core Inventory"
    assert summary.cell(nav_row, 5).value == "Yes"


def test_parent_enabled_zero_children():
    config = """
config system interface
    edit "port1"
        set ip 10.0.0.1 255.255.255.0
        set secondary-IP enable
    next
end
"""
    fg = parse_fortigate_config(config)
    intf = fg.interfaces[0]
    assert intf.secondary_ip == "enable"
    assert intf.source_attributes["secondary_ip"] == "enable"
    assert intf.secondary_ips == []

    ir = FGToIRTransformer(fg).transform()
    assert ir.interfaces[0].secondary_ips == []


def test_parent_disabled_zero_children():
    config = """
config system interface
    edit "port1"
        set ip 10.0.0.1 255.255.255.0
        set secondary-IP disable
    next
end
"""
    fg = parse_fortigate_config(config)
    intf = fg.interfaces[0]
    assert intf.secondary_ip == "disable"
    assert intf.source_attributes["secondary_ip"] == "disable"
    assert intf.secondary_ips == []

    ir = FGToIRTransformer(fg).transform()
    assert ir.interfaces[0].secondary_ips == []


def test_nested_child_without_parent_enable_setting():
    config = """
config system interface
    edit "port1"
        set ip 10.0.0.1 255.255.255.0
        config secondaryip
            edit 1
                set ip 10.0.0.2 255.255.255.0
                set allowaccess ping
            next
        end
    next
end
"""
    fg = parse_fortigate_config(config)
    intf = fg.interfaces[0]
    assert intf.secondary_ip is None
    assert len(intf.secondary_ips) == 1

    ir = FGToIRTransformer(fg).transform()
    ir_intf = ir.interfaces[0]
    assert len(ir_intf.secondary_ips) == 1
    sec = ir_intf.secondary_ips[0]
    assert sec.source_id == "1"
    assert sec.source_ip == "10.0.0.2 255.255.255.0"
    assert sec.ip == "10.0.0.2/24"
    assert sec.management_access == ["ping"]


def test_primary_and_remote_ip_remain_unchanged():
    config = """
config system interface
    edit "port1"
        set ip 10.0.0.1 255.255.255.0
        set remote-ip 10.10.10.1 255.255.255.255
        set secondary-IP enable
        config secondaryip
            edit 1
                set ip 10.0.0.2 255.255.255.0
            next
        end
    next
end
"""
    ir = _transform(config)
    intf = ir.interfaces[0]
    assert intf.ip == "10.0.0.1/24"
    assert intf.remote_ip == "10.10.10.1/32"
    assert len(intf.secondary_ips) == 1
    assert intf.secondary_ips[0].ip == "10.0.0.2/24"


def test_unset_secondary_ip_normalization():
    config = """
config system interface
    edit "port1"
        set secondary-IP enable
        unset secondary-IP
    next
end
"""
    fg = parse_fortigate_config(config)
    intf = fg.interfaces[0]
    assert intf.secondary_ip is None
    assert "secondary_ip" not in intf.source_attributes


def test_target_generators_unaffected_by_secondary_ips():
    ir = _transform("""
config system interface
    edit "port1"
        set ip 10.0.0.1 255.255.255.0
        set secondary-IP enable
        config secondaryip
            edit 1
                set ip 10.0.0.2 255.255.255.0
            next
        end
    next
end
""")
    for vendor in ["palo_alto", "cisco_asa", "checkpoint", "juniper_srx", "fortigate"]:
        generator = PluginRegistry.get_generator(vendor)
        target_config = generator.generate(ir)
        assert target_config is not None
