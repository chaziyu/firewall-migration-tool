import io

from openpyxl import load_workbook
import pytest

from fwmigrate.extraction.models import ExtractionStatus
from fwmigrate.parsers.fortigate.dependencies import build_dependency_registry
from fwmigrate.parsers.fortigate.extractor import extract_fortigate_config
from fwmigrate.parsers.fortigate.model import FGInterface
from fwmigrate.parsers.fortigate.parser import parse_fortigate_config
from fwmigrate.parsers.fortigate.transformer import FGToIRTransformer
from fwmigrate.report.excel_exporter import IRExcelExporter


def _interface(config: str, name: str) -> FGInterface:
    parsed = parse_fortigate_config(config)
    return next(interface for interface in parsed.interfaces if interface.name == name)


def _aggregate_config(interface_type: str = "aggregate", members: str = '"port2" "port1"') -> str:
    return f'''
config system interface
    edit "{interface_type}0"
        set type {interface_type}
        set member {members}
    next
    edit "port1"
    next
    edit "port2"
    next
end
'''


def test_aggregate_interface_members_are_typed_in_source_order():
    config = _aggregate_config(members='"port2" "port1"')

    interface = _interface(config, "aggregate0")

    assert interface.type == "aggregate"
    assert interface.members == ["port2", "port1"]
    assert "member" not in interface.source_attributes
    assert "members" not in interface.source_attributes


def test_redundant_interface_members_are_typed_in_source_order():
    interface = _interface(
        _aggregate_config("redundant", '"port4" "port3"'),
        "redundant0",
    )

    assert interface.type == "redundant"
    assert interface.members == ["port4", "port3"]


def test_interface_member_with_one_value_remains_a_one_item_list():
    interface = _interface(
        _aggregate_config("aggregate", '"port1"'),
        "aggregate0",
    )

    assert interface.members == ["port1"]


def test_aggregate_interface_without_member_has_an_empty_list():
    interface = _interface(
        '''
config system interface
    edit "aggregate0"
        set type aggregate
    next
end
''',
        "aggregate0",
    )

    assert interface.members == []


def test_aggregate_settings_are_typed_with_explicit_provenance():
    interface = _interface(
        '''
config system interface
    edit "aggregate0"
        set type aggregate
        set member "port1" "port2"
        set min-links 2
        set min-links-down operational
        set algorithm L4
        set lacp-mode active
        set lacp-speed fast
        set lacp-ha-secondary enable
        set system-id-type user
        set system-id 00:11:22:33:44:55
        set link-up-delay 100
        set aggregate-type physical
        set priority-override disable
    next
end
''',
        "aggregate0",
    )

    assert interface.members == ["port1", "port2"]
    assert interface.source_attributes == {
        "type": "aggregate",
        "source_context": "root",
        "vdom": "root",
        "link_up_delay": 100,
    }
    assert {
        "member",
        "lacp_mode",
        "lacp_ha_secondary",
        "min_links",
        "min_links_down",
        "algorithm",
        "lacp_speed",
        "system_id_type",
        "system_id",
        "link_up_delay",
        "aggregate_type",
        "priority_override",
    } <= interface.source_explicit_fields
    assert (
        interface.lacp_mode,
        interface.lacp_ha_secondary,
        interface.system_id_type,
        interface.system_id,
        interface.lacp_speed,
        interface.min_links,
        interface.min_links_down,
        interface.algorithm,
        interface.link_up_delay,
        interface.aggregate_type,
        interface.priority_override,
    ) == (
        "active", "enable", "user", "00:11:22:33:44:55", "fast", 2,
        "operational", "L4", 100, "physical", "disable",
    )


def test_legacy_lacp_ha_slave_is_preserved_as_unknown_source_setting():
    interface = _interface(
        '''
config system interface
    edit "aggregate0"
        set type aggregate
        set lacp-ha-slave enable
    next
end
''',
        "aggregate0",
    )

    assert interface.lacp_ha_secondary == "enable"
    assert interface.source_attributes["lacp_ha_slave"] == "enable"
    assert "lacp_ha_slave" not in interface.source_explicit_fields


def test_aggregate_unset_restores_effective_defaults_and_clears_provenance():
    interface = _interface(
        '''
config system interface
    edit "aggregate0"
        set type aggregate
        set lacp-mode passive
        set lacp-ha-secondary disable
        set system-id-type user
        set system-id 00:11:22:33:44:55
        set lacp-speed fast
        set min-links 2
        set min-links-down operational
        set algorithm L2
        set aggregate-type vxlan
        set priority-override disable
        unset lacp-mode
        unset lacp-ha-secondary
        unset system-id-type
        unset system-id
        unset lacp-speed
        unset min-links
        unset min-links-down
        unset algorithm
        unset aggregate-type
        unset priority-override
    next
end
''',
        "aggregate0",
    )

    assert (
        interface.lacp_mode,
        interface.lacp_ha_secondary,
        interface.system_id_type,
        interface.system_id,
        interface.lacp_speed,
        interface.min_links,
        interface.min_links_down,
        interface.algorithm,
        interface.aggregate_type,
        interface.priority_override,
    ) == (
        "active", "enable", "auto", None, "slow", 1, "operational", "L4",
        "physical", "enable",
    )
    assert not interface.source_explicit_fields
    assert "source_unset_settings" in interface.source_attributes


def test_member_set_append_and_unset_preserve_order_and_clear_members():
    interface = _interface(
        '''
config system interface
    edit "aggregate0"
        set type aggregate
        set member "port3" "port1"
        append member "port2"
    next
end
''',
        "aggregate0",
    )
    assert interface.members == ["port3", "port1", "port2"]

    cleared = _interface(
        '''
config system interface
    edit "aggregate0"
        set type aggregate
        set member "port3"
        unset member
    next
end
''',
        "aggregate0",
    )
    assert cleared.members == []


def test_aggregate_defaults_are_effective_but_not_source_explicit():
    interface = _interface(
        '''
config system interface
    edit "aggregate0"
        set type aggregate
    next
end
''',
        "aggregate0",
    )

    assert (
        interface.lacp_mode,
        interface.lacp_ha_secondary,
        interface.system_id_type,
        interface.system_id,
        interface.lacp_speed,
        interface.min_links,
        interface.min_links_down,
        interface.algorithm,
        interface.aggregate_type,
        interface.priority_override,
    ) == (
        "active", "enable", "auto", None, "slow", 1, "operational", "L4",
        "physical", "enable",
    )
    assert not interface.source_explicit_fields


def test_aggregate_relationships_use_unambiguous_source_model_fields():
    interface = _interface(
        '''
config system interface
    edit "aggregate0"
        set type aggregate
        set aggregate "parent0"
        set redundant-interface "redundant0"
    next
end
''',
        "aggregate0",
    )

    assert interface.aggregate_parent == "parent0"
    assert interface.redundant_interface_parent == "redundant0"
    assert {"aggregate", "redundant_interface"} <= interface.source_explicit_fields


def test_malformed_min_links_preserves_evidence_and_effective_default():
    interface = _interface(
        '''
config system interface
    edit "aggregate0"
        set type aggregate
        set min-links invalid
    next
end
''',
        "aggregate0",
    )

    assert interface.min_links == 1
    assert interface.source_attributes["unparsed_min_links"] == "invalid"
    assert "min_links" in interface.source_explicit_fields


def test_aggregate_transform_preserves_topology_and_requires_review():
    ir = FGToIRTransformer(
        parse_fortigate_config(_aggregate_config())
    ).transform()
    interface = next(item for item in ir.interfaces if item.name == "aggregate0")

    assert interface.interface_type == "aggregate"
    assert interface.members == ["port2", "port1"]
    assert interface.migration_status == "PARTIALLY_NORMALIZED"
    assert interface.requires_manual_review is True
    assert (
        "FortiGate aggregate or redundant interface topology requires "
        "target-platform review"
    ) in interface.review_reasons
    assert interface.zone is None
    assert interface.parent is None


def test_aggregate_review_propagates_to_coverage_and_generation_safety():
    result = extract_fortigate_config(_aggregate_config())
    section = next(item for item in result.source_sections if item.path == "system interface")

    assert section.status == ExtractionStatus.PARTIALLY_NORMALIZED
    assert result.generation_safe is False
    assert any("canonical objects require manual review" in reason for reason in result.blocking_reasons)


def test_aggregate_transform_projects_typed_source_semantics():
    ir = FGToIRTransformer(parse_fortigate_config('''
config system interface
    edit "aggregate0"
        set type aggregate
        set member "port1" "port2"
        set lacp-mode passive
        set lacp-ha-secondary disable
        set system-id-type user
        set system-id 00:11:22:33:44:55
        set lacp-speed fast
        set min-links 2
        set min-links-down administrative
        set algorithm L3
        set link-up-delay 100
        set aggregate-type physical
        set priority-override disable
    next
end
''')).transform()
    interface = ir.interfaces[0]

    assert interface.source_lacp_mode == "passive"
    assert interface.source_lacp_ha_secondary == "disable"
    assert interface.source_lacp_system_id_type == "user"
    assert interface.source_lacp_system_id == "00:11:22:33:44:55"
    assert interface.source_lacp_speed == "fast"
    assert interface.source_min_links == 2
    assert interface.source_min_links_down == "administrative"
    assert interface.source_aggregate_algorithm == "L3"
    assert interface.source_aggregate_type == "physical"
    assert interface.source_priority_override == "disable"
    assert interface.source_explicit_aggregate_fields == [
        "aggregate_type",
        "algorithm",
        "lacp_ha_secondary",
        "lacp_mode",
        "lacp_speed",
        "link_up_delay",
        "member",
        "min_links",
        "min_links_down",
        "priority_override",
        "system_id",
        "system_id_type",
    ]


def test_redundant_transform_preserves_topology_and_requires_review():
    ir = FGToIRTransformer(
        parse_fortigate_config(_aggregate_config("redundant", '"port1" "port2"'))
    ).transform()
    interface = next(item for item in ir.interfaces if item.name == "redundant0")

    assert interface.interface_type == "redundant"
    assert interface.members == ["port1", "port2"]
    assert interface.requires_manual_review is True
    assert interface.migration_status == "PARTIALLY_NORMALIZED"


def test_normal_interface_has_no_topology_review_reason():
    ir = FGToIRTransformer(parse_fortigate_config('''
config system interface
    edit "port1"
        set type physical
    next
end
''')).transform()
    interface = ir.interfaces[0]

    assert interface.interface_type == "physical"
    assert interface.members == []
    assert interface.migration_status == "NORMALIZED"
    assert interface.requires_manual_review is False
    assert not any("topology" in reason.lower() for reason in interface.review_reasons)
    assert interface.source_lacp_mode is None
    assert interface.source_min_links is None


def test_aggregate_validation_preserves_invalid_values_and_requires_review():
    ir = FGToIRTransformer(parse_fortigate_config('''
config system interface
    edit "aggregate0"
        set type aggregate
        set lacp-mode invalid
        set lacp-ha-secondary maybe
        set lacp-speed medium
        set min-links 33
        set link-up-delay 49
    next
end
''')).transform()
    interface = ir.interfaces[0]

    assert interface.source_lacp_mode == "invalid"
    assert interface.source_lacp_ha_secondary == "maybe"
    assert interface.source_lacp_speed == "medium"
    assert interface.source_min_links == 33
    assert interface.requires_manual_review is True
    assert interface.migration_status == "PARTIALLY_NORMALIZED"
    assert any("lacp-mode" in reason for reason in interface.review_reasons)
    assert any("min-links" in reason for reason in interface.review_reasons)
    assert any("link-up-delay" in reason for reason in interface.review_reasons)


def test_aggregate_validation_covers_system_id_members_and_source_tokens():
    ir = FGToIRTransformer(parse_fortigate_config('''
config system interface
    edit "aggregate0"
        set type aggregate
        set member "port1"
        set system-id-type user
        set system-id not-a-mac
        set min-links 2
        set min-links-down disabled
        set algorithm L5
        set aggregate-type vxlan
        set priority-override maybe
        set link-up-delay 3600001
    next
end
''')).transform()
    interface = ir.interfaces[0]

    assert interface.source_lacp_system_id == "not-a-mac"
    assert interface.source_min_links == 2
    assert interface.source_aggregate_algorithm == "L5"
    assert interface.source_aggregate_type == "vxlan"
    assert interface.source_priority_override == "maybe"
    assert any("system-id" in reason for reason in interface.review_reasons)
    assert any("member count" in reason for reason in interface.review_reasons)
    assert any("min-links-down" in reason for reason in interface.review_reasons)
    assert any("algorithm" in reason for reason in interface.review_reasons)
    assert any("VXLAN" in reason for reason in interface.review_reasons)
    assert any("priority-override" in reason for reason in interface.review_reasons)
    assert any("link-up-delay" in reason for reason in interface.review_reasons)


def test_explicit_aggregate_setting_on_physical_interface_is_preserved():
    ir = FGToIRTransformer(parse_fortigate_config('''
config system interface
    edit "port1"
        set type physical
        set lacp-mode passive
    next
end
''')).transform()
    interface = ir.interfaces[0]

    assert interface.source_lacp_mode == "passive"
    assert interface.source_explicit_aggregate_fields == ["lacp_mode"]
    assert interface.requires_manual_review is True
    assert any("aggregate-specific" in reason for reason in interface.review_reasons)


def test_redundant_lacp_settings_are_reviewed_without_inference():
    ir = FGToIRTransformer(parse_fortigate_config('''
config system interface
    edit "redundant0"
        set type redundant
        set lacp-mode passive
    next
end
''')).transform()
    interface = ir.interfaces[0]

    assert interface.source_lacp_mode == "passive"
    assert any("LACP-specific" in reason for reason in interface.review_reasons)


def test_aggregate_topology_reviews_duplicates_cycles_conflicts_and_nested_members():
    ir = FGToIRTransformer(parse_fortigate_config('''
config system interface
    edit "agg1"
        set type aggregate
        set member "agg2" "agg2"
    next
    edit "agg2"
        set type aggregate
        set member "agg1"
    next
    edit "agg3"
        set type aggregate
        set member "port1"
    next
    edit "red3"
        set type redundant
        set member "port1"
    next
    edit "port1"
    next
end
''')).transform()
    agg1 = next(interface for interface in ir.interfaces if interface.name == "agg1")
    agg2 = next(interface for interface in ir.interfaces if interface.name == "agg2")
    agg3 = next(interface for interface in ir.interfaces if interface.name == "agg3")
    red3 = next(interface for interface in ir.interfaces if interface.name == "red3")
    port1 = next(interface for interface in ir.interfaces if interface.name == "port1")

    assert agg1.members == ["agg2", "agg2"]
    assert any("duplicate" in reason for reason in agg1.review_reasons)
    assert any("cycle" in reason for reason in agg1.review_reasons)
    assert any("nested" in reason for reason in agg1.review_reasons)
    assert any("cycle" in reason for reason in agg2.review_reasons)
    assert any("multiple" in reason for reason in port1.review_reasons)


def test_read_only_parent_relationship_mismatch_is_preserved_for_review():
    ir = FGToIRTransformer(parse_fortigate_config('''
config system interface
    edit "agg0"
        set type aggregate
        set member "port2"
    next
    edit "port1"
        set aggregate "agg0"
    next
end
''')).transform()
    interface = next(interface for interface in ir.interfaces if interface.name == "port1")

    assert interface.source_aggregate_parent == "agg0"
    assert any("disagrees" in reason for reason in interface.review_reasons)


def test_explicit_read_only_parent_relationships_are_dependency_checked():
    result = extract_fortigate_config('''
config system interface
    edit "port1"
        set aggregate "agg0"
        set redundant-interface "red0"
    next
    edit "agg0"
        set type aggregate
    next
    edit "red0"
        set type redundant
    next
end
''')
    dependencies = {
        dependency.source_field: dependency
        for dependency in result.dependencies
        if dependency.source_object == "port1"
    }

    assert dependencies["aggregate"].result == "RESOLVED"
    assert dependencies["redundant-interface"].result == "RESOLVED"


def test_empty_aggregate_members_require_manual_review():
    ir = FGToIRTransformer(parse_fortigate_config('''
config system interface
    edit "aggregate0"
        set type aggregate
    next
end
''')).transform()
    interface = ir.interfaces[0]

    assert interface.members == []
    assert interface.requires_manual_review is True
    assert interface.migration_status == "PARTIALLY_NORMALIZED"
    assert (
        "FortiGate aggregate or redundant interface has no configured members"
        in interface.review_reasons
    )


def test_self_referencing_aggregate_is_preserved_and_reviewed():
    config = '''
config system interface
    edit "aggregate0"
        set type aggregate
        set member "aggregate0"
    next
end
'''

    result = extract_fortigate_config(config)
    interface = result.canonical_ir.interfaces[0]
    dependency = next(
        item
        for item in result.dependencies
        if item.source_path == "system interface"
    )

    assert interface.members == ["aggregate0"]
    assert interface.requires_manual_review is True
    assert any("itself" in reason for reason in interface.review_reasons)
    assert dependency.result == "UNRESOLVED"
    assert dependency.target_path is None


def test_interface_member_dependencies_resolve_in_same_vdom():
    result = extract_fortigate_config(_aggregate_config())
    dependencies = [
        item
        for item in result.dependencies
        if item.source_path == "system interface"
        and item.source_object == "aggregate0"
    ]

    assert [(item.reference, item.result, item.target_path) for item in dependencies] == [
        ("port2", "RESOLVED", "system interface"),
        ("port1", "RESOLVED", "system interface"),
    ]


def test_missing_interface_member_is_unresolved():
    result = extract_fortigate_config(_aggregate_config(members='"missing-port"'))
    dependency = next(
        item
        for item in result.dependencies
        if item.source_path == "system interface"
    )

    assert dependency.reference == "missing-port"
    assert dependency.result == "UNRESOLVED"
    assert dependency.expected_type == "system interface"
    assert dependency.target_path is None


def test_interface_members_do_not_resolve_across_vdoms():
    config = '''
config vdom
    edit "tenant-a"
        config system interface
            edit "aggregate0"
                set type aggregate
                set member "port1"
            next
        end
    next
    edit "tenant-b"
        config system interface
            edit "port1"
            next
        end
    next
end
'''

    result = extract_fortigate_config(config)
    dependency = next(
        item
        for item in result.dependencies
        if item.source_path == "system interface"
    )

    assert dependency.source_context == "tenant-a"
    assert dependency.result == "UNRESOLVED"
    assert dependency.target_path is None


def test_interface_member_dependency_registry_rejects_self_reference():
    from fwmigrate.extraction.models import SourceCommand, SourceInventoryItem

    dependencies = build_dependency_registry([
        SourceInventoryItem(
            domain="system",
            source_path="system interface",
            name="aggregate0",
            source_context="root",
            commands=[SourceCommand(operation="set", key="member", values=["aggregate0"])],
        ),
    ])

    assert len(dependencies) == 1
    assert dependencies[0].result == "UNRESOLVED"
    assert dependencies[0].notes == "Interface member cannot reference its own interface."


def test_interface_topology_is_exported_with_review_metadata():
    ir = FGToIRTransformer(parse_fortigate_config(_aggregate_config())).transform()
    workbook = load_workbook(io.BytesIO(IRExcelExporter(ir).generate()))
    sheet = workbook["Interfaces"]
    headers = {cell.value: cell.column for cell in sheet[3]}
    values = {
        header: sheet.cell(4, column).value
        for header, column in headers.items()
    }

    assert values["Interface Type"] == "aggregate"
    assert values["Members"] == "port2, port1"
    assert values["Extraction Status"] == "PARTIALLY_NORMALIZED"
    assert values["Manual Review"] == "TRUE"
    assert "target-platform review" in values["Review Reasons"]
    assert "member" not in values["Additional Settings"]


def test_interface_coverage_is_partial_for_aggregate_topology():
    result = extract_fortigate_config(_aggregate_config())
    section = next(
        item for item in result.source_sections if item.path == "system interface"
    )

    assert section.status == ExtractionStatus.PARTIALLY_NORMALIZED
    assert result.requires_manual_review is True


def test_redundant_settings_preserve_topology_without_lacp_inference():
    ir = FGToIRTransformer(parse_fortigate_config('''
config system interface
    edit "redundant0"
        set type redundant
        set member "port2" "port1"
        set link-up-delay 100
        set priority-override disable
    next
end
''')).transform()
    interface = ir.interfaces[0]

    assert interface.members == ["port2", "port1"]
    assert interface.source_attributes["link_up_delay"] == 100
    assert interface.source_priority_override == "disable"
    assert "lacp_mode" not in interface.source_explicit_aggregate_fields
    assert interface.migration_status == "PARTIALLY_NORMALIZED"
    assert interface.requires_manual_review is True
    assert any("target-platform review" in reason for reason in interface.review_reasons)


def test_explicit_effective_aggregate_defaults_keep_provenance():
    ir = FGToIRTransformer(parse_fortigate_config('''
config system interface
    edit "aggregate0"
        set type aggregate
        set lacp-mode active
        set lacp-speed slow
        set min-links 1
        set algorithm L4
        set priority-override enable
    next
end
''')).transform()
    interface = ir.interfaces[0]

    assert interface.source_lacp_mode == "active"
    assert interface.source_lacp_speed == "slow"
    assert interface.source_min_links == 1
    assert interface.source_aggregate_algorithm == "L4"
    assert interface.source_priority_override == "enable"
    assert {
        "lacp_mode", "lacp_speed", "min_links", "algorithm", "priority_override"
    } <= set(interface.source_explicit_aggregate_fields)


def test_aggregate_unset_history_covers_category9_settings():
    interface = _interface('''
config system interface
    edit "aggregate0"
        set type aggregate
        set member "port1"
        set lacp-mode passive
        set lacp-ha-secondary disable
        set system-id-type user
        set system-id 00:11:22:33:44:55
        set lacp-speed fast
        set min-links 2
        set min-links-down administrative
        set algorithm L2
        set aggregate-type vxlan
        set priority-override disable
        unset member
        unset lacp-mode
        unset lacp-ha-secondary
        unset system-id-type
        unset system-id
        unset lacp-speed
        unset min-links
        unset min-links-down
        unset algorithm
        unset aggregate-type
        unset priority-override
    next
end
''', "aggregate0")

    assert interface.members == []
    assert (interface.lacp_mode, interface.lacp_ha_secondary, interface.system_id_type,
            interface.system_id, interface.lacp_speed, interface.min_links,
            interface.min_links_down, interface.algorithm, interface.aggregate_type,
            interface.priority_override) == (
        "active", "enable", "auto", None, "slow", 1, "operational", "L4",
        "physical", "enable",
    )
    assert interface.source_attributes["source_unset_settings"] == [
        "member", "lacp-mode", "lacp-ha-secondary", "system-id-type", "system-id",
        "lacp-speed", "min-links", "min-links-down", "algorithm", "aggregate-type",
        "priority-override",
    ]
    assert not interface.source_explicit_fields


def test_malformed_aggregate_numbers_preserve_evidence_and_continue():
    interface = _interface('''
config system interface
    edit "aggregate0"
        set type aggregate
        set min-links invalid
        set link-up-delay unknown
        set priority-override disable
    next
end
''', "aggregate0")

    assert interface.min_links == 1
    assert interface.link_up_delay is None
    assert interface.source_attributes["unparsed_min_links"] == "invalid"
    assert interface.source_attributes["unparsed_link_up_delay"] == "unknown"
    assert interface.priority_override == "disable"


def test_aggregate_numeric_ranges_and_member_count_require_review():
    ir = FGToIRTransformer(parse_fortigate_config('''
config system interface
    edit "aggregate0"
        set type aggregate
        set member "port1" "port2"
        set min-links 3
        set link-up-delay 49
    next
    edit "aggregate1"
        set type aggregate
        set member "port1"
        set min-links 0
        set link-up-delay 3600001
    next
    edit "aggregate2"
        set type aggregate
        set member "port1"
        set min-links 33
    next
end
''')).transform()
    by_name = {interface.name: interface for interface in ir.interfaces}

    assert by_name["aggregate0"].source_min_links == 3
    assert by_name["aggregate0"].source_attributes["link_up_delay"] == 49
    assert any("member count" in reason for reason in by_name["aggregate0"].review_reasons)
    assert by_name["aggregate1"].source_min_links == 0
    assert by_name["aggregate1"].source_attributes["link_up_delay"] == 3600001
    assert by_name["aggregate2"].source_min_links == 33
    assert all(
        by_name[name].requires_manual_review
        for name in ("aggregate0", "aggregate1", "aggregate2")
    )


def test_invalid_lacp_options_are_preserved_and_reviewed():
    ir = FGToIRTransformer(parse_fortigate_config('''
config system interface
    edit "aggregate0"
        set type aggregate
        set lacp-mode invalid-mode
        set lacp-ha-secondary maybe
        set lacp-speed medium
        set system-id-type generated
    next
end
''')).transform()
    interface = ir.interfaces[0]

    assert interface.source_lacp_mode == "invalid-mode"
    assert interface.source_lacp_ha_secondary == "maybe"
    assert interface.source_lacp_speed == "medium"
    assert interface.source_lacp_system_id_type == "generated"
    assert all(any(key in reason for reason in interface.review_reasons)
               for key in ("lacp-mode", "lacp-ha-secondary", "lacp-speed", "system-id-type"))


@pytest.mark.parametrize(
    ("system_id_type", "system_id", "reviewed"),
    [
        ("user", "00:11:22:33:44:55", False),
        ("user", "not-a-mac", True),
        ("user", None, True),
    ],
)
def test_user_system_id_validation_preserves_source_and_review(
    system_id_type, system_id, reviewed
):
    system_id_line = f"        set system-id {system_id}\n" if system_id else ""
    ir = FGToIRTransformer(parse_fortigate_config(f'''
config system interface
    edit "aggregate0"
        set type aggregate
        set system-id-type {system_id_type}
{system_id_line}    next
end
''')).transform()
    interface = ir.interfaces[0]

    assert interface.source_lacp_system_id_type == "user"
    assert interface.source_lacp_system_id == system_id
    has_system_id_review = any(
        "system-id" in reason or "system ID" in reason
        for reason in interface.review_reasons
    )
    assert has_system_id_review is reviewed


@pytest.mark.parametrize("algorithm", ["L2", "L3", "L4", "Source-MAC"])
def test_valid_aggregate_algorithms_preserve_source_token(algorithm):
    interface = FGToIRTransformer(parse_fortigate_config(f'''
config system interface
    edit "aggregate0"
        set type aggregate
        set algorithm {algorithm}
    next
end
''')).transform().interfaces[0]

    assert interface.source_aggregate_algorithm == algorithm
    assert not any("algorithm" in reason for reason in interface.review_reasons)


def test_invalid_aggregate_algorithm_preserves_source_token_and_requires_review():
    interface = FGToIRTransformer(parse_fortigate_config('''
config system interface
    edit "aggregate0"
        set type aggregate
        set algorithm L5
    next
end
''')).transform().interfaces[0]

    assert interface.source_aggregate_algorithm == "L5"
    assert any("algorithm" in reason for reason in interface.review_reasons)


def test_aggregate_types_preserve_physical_and_vxlan_semantics():
    physical = _interface(_aggregate_config(), "aggregate0")
    vxlan = _interface('''
config system interface
    edit "aggregate0"
        set type aggregate
        set aggregate-type vxlan
    next
end
''', "aggregate0")

    assert physical.aggregate_type == "physical"
    assert vxlan.aggregate_type == "vxlan"

    ir = FGToIRTransformer(parse_fortigate_config('''
config system interface
    edit "aggregate0"
        set type aggregate
        set aggregate-type vxlan
    next
end
''')).transform()
    assert any("VXLAN" in reason for reason in ir.interfaces[0].review_reasons)


def test_invalid_aggregate_type_preserves_source_token_and_requires_review():
    interface = FGToIRTransformer(parse_fortigate_config('''
config system interface
    edit "aggregate0"
        set type aggregate
        set aggregate-type mystery
    next
end
''')).transform().interfaces[0]

    assert interface.source_aggregate_type == "mystery"
    assert any("aggregate-type" in reason for reason in interface.review_reasons)


def test_duplicate_members_are_preserved_and_reviewed():
    interface = FGToIRTransformer(parse_fortigate_config('''
config system interface
    edit "aggregate0"
        set type aggregate
        set member "port1" "port1"
    next
end
''')).transform().interfaces[0]

    assert interface.members == ["port1", "port1"]
    assert any("duplicate" in reason for reason in interface.review_reasons)


def test_recursive_cycle_completes_extraction_and_reviews_both_interfaces():
    result = extract_fortigate_config('''
config system interface
    edit "agg1"
        set type aggregate
        set member "agg2"
    next
    edit "agg2"
        set type aggregate
        set member "agg1"
    next
end
''')
    by_name = {interface.name: interface for interface in result.canonical_ir.interfaces}

    assert set(by_name) == {"agg1", "agg2"}
    assert by_name["agg1"].members == ["agg2"]
    assert by_name["agg2"].members == ["agg1"]
    assert all(any("cycle" in reason for reason in item.review_reasons)
               for item in by_name.values())


def test_conflicting_aggregate_parents_preserve_both_relationships_and_review():
    ir = FGToIRTransformer(parse_fortigate_config('''
config system interface
    edit "agg1"
        set type aggregate
        set member "port1"
    next
    edit "agg2"
        set type aggregate
        set member "port1"
    next
    edit "port1"
    next
end
''')).transform()
    by_name = {interface.name: interface for interface in ir.interfaces}
    port1 = next(interface for interface in ir.interfaces if interface.name == "port1")

    assert by_name["agg1"].members == ["port1"]
    assert by_name["agg2"].members == ["port1"]
    assert port1.members == []
    assert any("multiple" in reason for reason in port1.review_reasons)


def test_read_only_parent_relationships_are_separate_from_member_topology():
    ir = FGToIRTransformer(parse_fortigate_config('''
config system interface
    edit "agg1"
        set type aggregate
    next
    edit "red1"
        set type redundant
    next
    edit "port1"
        set aggregate "agg1"
        set redundant-interface "red1"
    next
end
''')).transform()
    port1 = next(interface for interface in ir.interfaces if interface.name == "port1")

    assert port1.members == []
    assert port1.source_aggregate_parent == "agg1"
    assert port1.source_redundant_interface_parent == "red1"


def test_matching_read_only_parent_relationship_has_no_inconsistency_reason():
    ir = FGToIRTransformer(parse_fortigate_config('''
config system interface
    edit "agg1"
        set type aggregate
        set member "port1"
    next
    edit "port1"
        set aggregate "agg1"
    next
end
''')).transform()
    port1 = next(interface for interface in ir.interfaces if interface.name == "port1")

    assert not any("disagrees" in reason for reason in port1.review_reasons)
