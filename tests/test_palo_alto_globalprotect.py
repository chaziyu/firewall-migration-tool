import io
import json

from openpyxl import load_workbook

from fwmigrate.ir.migrations import migrate_ir_payload
from fwmigrate.ir.version import IR_SCHEMA_VERSION
from fwmigrate.extraction.models import ExtractionStatus
from fwmigrate.parsers.palo_alto.parser import PANOSSourceParser
from fwmigrate.report.excel_exporter import IRExcelExporter


def _xml(global_protect: str = "", network_gateway: str = "") -> str:
    return f"""<config>
      <shared>
        <certificate><entry name="GP_Cert"><ca>yes</ca></entry></certificate>
        <ssl-tls-service-profile><entry name="GP_SSL_TLS_Profile"><certificate>GP_Cert</certificate></entry></ssl-tls-service-profile>
      </shared>
      <devices><entry name="fw1"><network>
        <interface><loopback><units><entry name="loopback.78"/></units></loopback><tunnel><units><entry name="tunnel.78"/></units></tunnel></interface>
        <tunnel>{network_gateway}</tunnel>
      </network></entry></devices>
      <vsys><entry name="vsys1"><address>
        <entry name="211.25.233.78_SSLVPN"><ip-netmask>211.25.233.78</ip-netmask></entry>
        <entry name="P-LAN_10.0.0.0"><ip-netmask>10.0.0.0/8</ip-netmask></entry>
      </address><authentication-profile>
        <entry name="GPVPN Profile"><method><local-database/></method></entry>
      </authentication-profile><global-protect>{global_protect}</global-protect></entry></vsys>
    </config>"""


def _client_auth() -> str:
    return """<client-auth>
      <entry name="ldap"><os>Windows</os><authentication-profile>GPVPN Profile</authentication-profile></entry>
      <entry name="radius"><os>Android</os><authentication-profile>GPVPN Profile</authentication-profile></entry>
      <entry name="saml"><os>iOS</os><authentication-profile>GPVPN Profile</authentication-profile></entry>
    </client-auth>"""


def _app_settings() -> str:
    return "".join(
        f'<entry name="setting-{index:02d}"><member>value-{index:02d}</member></entry>'
        for index in range(53)
    )


def _portal_and_gateway() -> str:
    auth = _client_auth()
    return f"""
      <global-protect-portal><entry name="GPVPN_Portal"><portal-config>
        <local-address><interface>loopback.78</interface><ip><ipv4>211.25.233.78_SSLVPN</ipv4></ip></local-address>
        <ssl-tls-service-profile>GP_SSL_TLS_Profile</ssl-tls-service-profile>{auth}
      </portal-config><client-config>
        <configs><entry name="default"><source-user><member>any</member></source-user><os><member>any</member></os>
          <gateways><external><list><entry name="External_Gateway"><ipv4>211.25.233.78</ipv4><manual>no</manual><priority-rule><entry name="Any"><priority>1</priority></entry></priority-rule></entry></list><cutoff-time>5</cutoff-time></external></gateways>
          <gp-app-config><config>{_app_settings()}</config></gp-app-config>
          <authentication-override><generate-cookie>yes</generate-cookie></authentication-override><save-user-credentials>1</save-user-credentials>
        </entry></configs>
        <root-ca><entry name="root"><certificate>GP_Cert</certificate><install-in-cert-store>yes</install-in-cert-store></entry></root-ca>
        <agent-user-override-key>PHASE81_AGENT_OVERRIDE_SECRET</agent-user-override-key>
      </client-config></entry></global-protect-portal>
      <global-protect-gateway><entry name="GP_Gateway"><gateway-config><ssl-tls-service-profile>GP_SSL_TLS_Profile</ssl-tls-service-profile><tunnel-mode>yes</tunnel-mode><remote-user-tunnel>tunnel.78</remote-user-tunnel>{auth}
        <roles><entry name="default"><login-lifetime><days>30</days></login-lifetime><inactivity-logout><hours>3</hours></inactivity-logout><disconnect-on-idle><minutes>180</minutes></disconnect-on-idle></entry></roles>
        <remote-user-tunnel-configs><entry name="GP_Test"><source-user><member>any</member></source-user><os><member>any</member></os><ip-pool><member>10.3.38.20-10.3.38.250</member></ip-pool><split-tunnel><access-route><member>P-LAN_10.0.0.0</member></access-route></split-tunnel><retrieve-framed-ip-address>no</retrieve-framed-ip-address><no-direct-access-to-local-network>no</no-direct-access-to-local-network></entry></remote-user-tunnel-configs>
      </gateway-config></entry></global-protect-gateway>
    """


def _network_gateway() -> str:
    return """<global-protect-gateway><entry name="GP_Gateway-N"><client><exclude-video-traffic><enabled>yes</enabled></exclude-video-traffic><dns-server><primary>10.1.4.2</primary><secondary>10.1.4.1</secondary></dns-server><dns-suffix><member>prasarana.com.my</member></dns-suffix><dns-suffix-inherited>no</dns-suffix-inherited></client><ipsec><third-party-client><enable>yes</enable><group-name>SubangVPN</group-name><group-password>PHASE81_GROUP_PASSWORD_SECRET</group-password></third-party-client></ipsec><local-address><interface>loopback.78</interface></local-address><ip-pool><member>10.3.38.20-10.3.38.250</member></ip-pool><tunnel-interface>tunnel.78</tunnel-interface></entry></global-protect-gateway>"""


def test_globalprotect_extracts_separate_portal_gateway_and_dependencies():
    result = PANOSSourceParser().extract(_xml(_portal_and_gateway(), _network_gateway()))
    ir = result.canonical_ir
    assert len(ir.global_protect_portals) == 1
    assert len(ir.global_protect_gateways) == 1
    portal = ir.global_protect_portals[0]
    gateway = ir.global_protect_gateways[0]
    assert len(ir.global_protect_network_gateways) == 1
    assert len(portal.client_authentication) == 3
    assert len(gateway.client_authentication) == 3
    assert all(auth.authentication_profile_resolved for auth in portal.client_authentication)
    assert all(auth.authentication_profile_resolved for auth in gateway.client_authentication)
    assert len(portal.client_configs) == 1
    assert len(portal.client_configs[0].external_gateways) == 1
    assert len(portal.client_configs[0].app_settings) == 53
    assert len(portal.root_ca_certificates) == 1
    assert len(gateway.roles) == 1
    assert len(gateway.remote_user_tunnel_configs) == 1

    assert portal.name == "GPVPN_Portal"
    assert portal.local_interface == "loopback.78"
    assert portal.local_interface_resolved is True
    assert portal.local_ipv4 == "211.25.233.78_SSLVPN"
    assert portal.resolved_local_address == "211.25.233.78_SSLVPN"
    assert portal.ssl_tls_service_profile == "GP_SSL_TLS_Profile"
    assert portal.ssl_tls_service_profile_resolved is True
    assert portal.has_agent_user_override_key is True
    assert portal.root_ca_certificates[0].certificate == "GP_Cert"
    assert portal.root_ca_certificates[0].certificate_resolved is True

    client_config = portal.client_configs[0]
    external_gateway = client_config.external_gateways[0]
    assert external_gateway.name == "External_Gateway"
    assert external_gateway.ipv4 == "211.25.233.78"
    assert external_gateway.manual is False
    assert client_config.external_gateway_cutoff_time == "5"
    assert len(external_gateway.priority_rules) == 1
    assert external_gateway.priority_rules[0].name == "Any"
    assert external_gateway.priority_rules[0].priority == 1
    assert [setting.source_order for setting in client_config.app_settings] == list(range(53))
    assert client_config.app_settings[0].values == ["value-00"]

    assert gateway.name == "GP_Gateway"
    assert gateway.tunnel_mode is True
    assert gateway.remote_user_tunnel == "tunnel.78"
    assert gateway.remote_user_tunnel_resolved is True
    role = gateway.roles[0]
    assert (role.login_lifetime_days, role.inactivity_logout_hours, role.disconnect_on_idle_minutes) == (30, 3, 180)
    tunnel = gateway.remote_user_tunnel_configs[0]
    assert tunnel.name == "GP_Test"
    assert tunnel.ip_pools == ["10.3.38.20-10.3.38.250"]
    assert tunnel.split_include_routes == ["P-LAN_10.0.0.0"]
    assert tunnel.resolved_split_include_routes == ["P-LAN_10.0.0.0"]
    assert tunnel.retrieve_framed_ip_address is False
    assert tunnel.no_direct_access_to_local_network is False

    serialized = json.dumps(result.model_dump(), default=str)
    assert "PHASE81_AGENT_OVERRIDE_SECRET" not in serialized
    assert "PHASE81_GROUP_PASSWORD_SECRET" not in serialized


def test_globalprotect_network_gateway_is_not_vsys_gateway_and_redacts_password():
    result = PANOSSourceParser().extract(_xml(_portal_and_gateway(), _network_gateway()))
    ir = result.canonical_ir
    assert [item.name for item in ir.global_protect_gateways] == ["GP_Gateway"]
    network_gateway = ir.global_protect_network_gateways[0]
    assert network_gateway.name == "GP_Gateway-N"
    assert network_gateway.local_interface_resolved is True
    assert network_gateway.tunnel_interface_resolved is True
    assert network_gateway.client_dns_primary == "10.1.4.2"
    assert network_gateway.client_dns_secondary == "10.1.4.1"
    assert network_gateway.dns_suffixes == ["prasarana.com.my"]
    assert network_gateway.dns_suffix_inherited is False
    assert network_gateway.exclude_video_traffic_enabled is True
    assert network_gateway.third_party_client_enabled is True
    assert network_gateway.third_party_group_name == "SubangVPN"
    assert network_gateway.third_party_group_password_configured is True
    serialized = result.model_dump_json()
    assert "PHASE81_GROUP_PASSWORD_SECRET" not in serialized


def test_globalprotect_production_children_are_typed_and_unknown_children_remain_unsupported():
    xml = _xml(f"""
      <global-protect-portal><entry name="Portal">
        <portal-config><unknown-portal><entry name="portal-unknown"><value>1</value></entry></unknown-portal></portal-config>
        <client-config><unknown-client><entry name="client-unknown"><value>1</value></entry></unknown-client></client-config>
      </entry></global-protect-portal>
      <global-protect-gateway><entry name="Gateway"><gateway-config><unknown-gateway><entry name="gateway-unknown"><value>1</value></entry></unknown-gateway></gateway-config></entry></global-protect-gateway>
    """)
    result = PANOSSourceParser().extract(xml)
    unsupported = {(item.source_path, item.name) for item in result.inventory_items if item.status == ExtractionStatus.UNSUPPORTED}
    assert any("portal-config/unknown-portal" in path and name == "portal-unknown" for path, name in unsupported)
    assert any("client-config/unknown-client" in path and name == "client-unknown" for path, name in unsupported)
    assert any("gateway-config/unknown-gateway" in path and name == "gateway-unknown" for path, name in unsupported)

    production_result = PANOSSourceParser().extract(_xml(_portal_and_gateway(), _network_gateway()))
    typed_names = {"GPVPN_Portal", "GP_Gateway", "GP_Gateway-N", "default", "External_Gateway", "Any", "root", "GP_Test"}
    assert not any(item.status == ExtractionStatus.UNSUPPORTED and item.name in typed_names for item in production_result.inventory_items)


def test_globalprotect_nested_timer_parse_errors_keep_evidence_and_review_reason():
    result = PANOSSourceParser().extract(_xml("""
      <global-protect-gateway><entry name="Gateway"><gateway-config><roles><entry name="role">
        <login-lifetime><days>bad</days></login-lifetime>
      </entry></roles></gateway-config></entry></global-protect-gateway>
    """))
    role = result.canonical_ir.global_protect_gateways[0].roles[0]
    assert role.login_lifetime_days is None
    assert "malformed-login_lifetime_days:bad" in result.canonical_ir.global_protect_gateways[0].review_reasons
    assert "bad" in json.dumps(role.source_attributes)


def test_globalprotect_strict_booleans_and_unresolved_references_require_review():
    xml = _xml("""<global-protect-portal><entry name="p"><portal-config><local-address><interface>missing</interface><ip><ipv4>missing-address</ipv4></ip></local-address><client-config><entry name="c"><external-gateway><entry name="g"><manual>maybe</manual></entry></external-gateway></entry></client-config></portal-config></entry></global-protect-portal>""")
    result = PANOSSourceParser().extract(xml)
    portal = result.canonical_ir.global_protect_portals[0]
    assert portal.local_interface_resolved is False
    assert portal.local_address_resolved is False
    assert portal.requires_manual_review is True
    assert any("unresolved" in reason for reason in portal.review_reasons)


def test_globalprotect_excel_and_schema_migration():
    result = PANOSSourceParser().extract(_xml(_portal_and_gateway(), _network_gateway()))
    workbook = load_workbook(io.BytesIO(IRExcelExporter(result.canonical_ir, result).generate()))
    assert "GlobalProtect Portals" in workbook.sheetnames
    assert "GlobalProtect Gateways" in workbook.sheetnames
    assert "GlobalProtect Network Gateways" in workbook.sheetnames
    assert workbook["GlobalProtect Portal Configs"].max_row == 4
    assert workbook["GlobalProtect External Gateways"].max_row == 4
    assert workbook["GlobalProtect App Settings"].max_row == 56
    assert workbook["GlobalProtect Root CAs"].max_row == 4
    assert workbook["GlobalProtect Gateway Roles"].max_row == 4
    assert workbook["GlobalProtect Tunnel Configs"].max_row == 4
    assert workbook["GlobalProtect Network Gateways"].max_row == 4
    assert "Group Password" not in [cell.value for cell in workbook["GlobalProtect Network Gateways"][3]]
    assert all(
        "PHASE81_AGENT_OVERRIDE_SECRET" not in str(cell.value)
        and "PHASE81_GROUP_PASSWORD_SECRET" not in str(cell.value)
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
    )
    migrated = migrate_ir_payload({"schema_version": "1.45", "metadata": {"source_vendor": "palo_alto"}})
    assert migrated["schema_version"] == IR_SCHEMA_VERSION
    assert migrated["global_protect_portals"] == []


def test_globalprotect_scopes_do_not_deduplicate_and_unknown_site_to_site_is_visible():
    result = PANOSSourceParser().extract("""<config><vsys>
      <entry name="vsys1"><global-protect><global-protect-portal><entry name="same"><portal-config/></entry></global-protect-portal></global-protect></entry>
      <entry name="vsys2"><global-protect><global-protect-portal><entry name="same"><portal-config/></entry></global-protect-portal><global-protect-site-to-site><entry name="future"><peer>configured</peer></entry></global-protect-site-to-site></global-protect></entry>
    </vsys></config>""")
    assert [portal.source_context for portal in result.canonical_ir.global_protect_portals] == ["vsys:vsys1", "vsys:vsys2"]
    assert any(item.status == ExtractionStatus.UNSUPPORTED and item.name == "future" for item in result.inventory_items)
