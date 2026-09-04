import io

from openpyxl import load_workbook

import fwmigrate.parsers  # noqa: F401 - registers Palo Alto plugins
from fwmigrate.parsers.palo_alto.parser import PANOSSourceParser
from fwmigrate.report.excel_exporter import IRExcelExporter


PSK_MARKER = "TEST-SHOULD-NOT-LEAK"


def _extract(body: str):
    return PANOSSourceParser().extract(f"<config version='11.1.0'>{body}</config>")


def _vpn_fixture() -> str:
    return f"""
    <devices><entry name="fw-a"><network>
      <ike><crypto-profiles>
        <ike-crypto-profiles><entry name="IKE-PROFILE">
          <encryption><member>aes-256-cbc</member></encryption>
          <hash><member>sha256</member></hash>
          <dh-group><member>group14</member></dh-group>
        </entry></ike-crypto-profiles>
        <ipsec-crypto-profiles><entry name="IPSEC-PROFILE">
          <protocol><member>esp</member></protocol>
          <esp><encryption><member>aes-256-cbc</member></encryption></esp>
        </entry></ipsec-crypto-profiles>
      </crypto-profiles><gateway>
        <entry name="GW-V1"><protocol><version>ikev1</version><ikev1>
          <ike-crypto-profile>IKE-V1</ike-crypto-profile>
          <dpd><enable>yes</enable></dpd>
        </ikev1></protocol><protocol-common>
          <nat-traversal><enable>no</enable></nat-traversal>
          <passive-mode>no</passive-mode>
          <fragmentation><enable>yes</enable></fragmentation>
        </protocol-common><local-address><interface>ethernet1/1</interface></local-address>
        <peer-address><ip>198.51.100.2</ip></peer-address>
        <authentication><pre-shared-key><key>{PSK_MARKER}</key></pre-shared-key></authentication>
      </entry>
        <entry name="GW-V2"><protocol><version>ikev2</version><ikev2>
          <ike-crypto-profile>IKE-V2</ike-crypto-profile>
          <dpd><enable>yes</enable></dpd>
        </ikev2></protocol><local-address><interface>ethernet1/2</interface></local-address>
      </entry>
      </gateway></ike>
      <tunnel><ipsec><entry name="IPSEC-TUNNEL"><tunnel-interface>tunnel.1</tunnel-interface>
        <auto-key><ike-gateway><entry name="GW-V1"/></ike-gateway>
          <ipsec-crypto-profile>IPSEC-1</ipsec-crypto-profile>
          <proxy-id><entry name="proxy-1"><local>10.0.0.0/24</local></entry></proxy-id>
        </auto-key><tunnel-monitor><enable>no</enable></tunnel-monitor>
      </entry></ipsec></tunnel>
    </network></entry></devices>
    """


def _inventory(result, domain, name):
    return next(item for item in result.inventory_items
                if item.domain == domain and item.name == name)


def test_ike_hash_and_ipsec_profiles_use_pan_ike_crypto_profile_root():
    result = _extract(_vpn_fixture())

    ike = _inventory(result, "vpn:ike_crypto_profile", "IKE-PROFILE")
    ipsec = _inventory(result, "vpn:ipsec_crypto_profile", "IPSEC-PROFILE")

    assert ike.source_attributes["pan_hash"] == ["sha256"]
    assert ike.source_attributes["pan_authentication"] == ["sha256"]
    assert ipsec.source_path.startswith("network/ike/crypto-profiles/")


def test_ike_gateway_version_profile_and_version_specific_dpd_are_preserved():
    result = _extract(_vpn_fixture())
    v1 = next(t for t in result.canonical_ir.vpn_tunnels if t.name == "GW-V1")
    v2 = next(t for t in result.canonical_ir.vpn_tunnels if t.name == "GW-V2")

    assert v1.ike_version == "ikev1"
    assert v1.ike_crypto_profile == "IKE-V1"
    assert v2.ike_version == "ikev2"
    assert v2.ike_crypto_profile == "IKE-V2"
    assert v1.source_attributes["pan_ikev1_dpd"]
    assert v2.source_attributes["pan_ikev2_dpd"]
    assert v1.source_attributes["pan_nat_traversal"]
    assert v1.source_attributes["pan_passive_mode"] == "no"
    assert v1.source_attributes["pan_fragmentation"]


def test_ipsec_tunnel_relationship_monitor_and_proxy_id_are_preserved():
    result = _extract(_vpn_fixture())
    phase2 = result.canonical_ir.vpn_phase2[0]

    assert phase2.phase1_name == "GW-V1"
    assert phase2.proposals == ["IPSEC-1"]
    assert phase2.source_attributes["pan_tunnel_monitor"]
    assert phase2.source_attributes["pan_tunnel_monitoring"]
    assert phase2.source_attributes["pan_proxy_ids"]


def test_multiple_ike_gateway_references_are_preserved_without_selecting_one():
    result = _extract("""
    <devices><entry name="fw-a"><network><tunnel><ipsec><entry name="TUNNEL">
      <auto-key><ike-gateway><entry name="GW-1"/><entry name="GW-2"/></ike-gateway>
    </auto-key></entry></ipsec></tunnel></network></entry></devices>
    """)
    phase2 = result.canonical_ir.vpn_phase2[0]
    item = _inventory(result, "vpn:ipsec_tunnel", "TUNNEL")

    assert phase2.phase1_name == ""
    assert phase2.source_attributes["pan_ike_gateways"] == ["GW-1", "GW-2"]
    assert "multiple IKE gateways" in item.notes[0]


def test_psk_is_presence_only_in_ir_inventory_and_excel():
    result = _extract(_vpn_fixture())
    tunnel = next(t for t in result.canonical_ir.vpn_tunnels if t.name == "GW-V1")

    assert tunnel.has_psk is True
    assert PSK_MARKER not in result.model_dump_json()
    assert PSK_MARKER not in str([item.source_attributes for item in result.inventory_items])

    workbook_bytes = IRExcelExporter(result.canonical_ir, extraction_result=result).generate()
    assert PSK_MARKER.encode() not in workbook_bytes


def test_vpn_fields_are_visible_in_existing_excel_sheets():
    result = _extract(_vpn_fixture())
    workbook = load_workbook(io.BytesIO(
        IRExcelExporter(result.canonical_ir, extraction_result=result).generate()
    ), data_only=True)

    tunnels = workbook["VPN Tunnels"]
    tunnel_headers = {cell.value: cell.column for cell in tunnels[3]}
    tunnel_row = next(row for row in tunnels.iter_rows(min_row=4)
                      if row[tunnel_headers["Name"] - 1].value == "GW-V1")
    assert tunnel_row[tunnel_headers["IKE Version"] - 1].value == "ikev1"
    assert tunnel_row[tunnel_headers["IKE Crypto Profile"] - 1].value == "IKE-V1"
    assert tunnel_row[tunnel_headers["PSK"] - 1].value == "Configured / Redacted"

    phase2 = workbook["VPN Phase 2"]
    phase2_headers = {cell.value: cell.column for cell in phase2[3]}
    phase2_row = next(row for row in phase2.iter_rows(min_row=4)
                      if row[phase2_headers["Name"] - 1].value == "IPSEC-TUNNEL")
    assert phase2_row[phase2_headers["Phase 1"] - 1].value == "GW-V1"
    assert phase2_row[phase2_headers["Proposal"] - 1].value == "IPSEC-1"
    settings = phase2_row[phase2_headers["Additional Settings"] - 1].value
    assert "tunnel-monitor" in settings
    assert "proxy-ids" in settings
