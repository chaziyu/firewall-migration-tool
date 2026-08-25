import io

from openpyxl import load_workbook

from fwmigrate.parsers.fortigate.parser import parse_fortigate_config
from fwmigrate.parsers.fortigate.transformer import FGToIRTransformer
from fwmigrate.report.excel_exporter import IRExcelExporter


PUBLIC_CERTIFICATE = """-----BEGIN CERTIFICATE-----
MIIB0zCCATygAwIBAgIEB1vNFTANBgkqhkiG9w0BAQsFADAlMSMwIQYDVQQDDBpT
YW5pdGl6ZWQgVGVzdCBDZXJ0aWZpY2F0ZTAeFw0yNDAxMDEwMDAwMDBaFw0zNDAx
MDEwMDAwMDBaMCUxIzAhBgNVBAMMGlNhbml0aXplZCBUZXN0IENlcnRpZmljYXRl
MIGfMA0GCSqGSIb3DQEBAQUAA4GNADCBiQKBgQDNFFoDwHIWYOCHZkF+Xv67SU66
e8o0Dyttu2bCXlXPirg9Yj2FIn83t5+V0pgZ+fKOFnFD1HAzLb8yhtB7bxNszwvq
EI4TIeU7pCFBnzAoc38LFeRiAFL/oe8/qonS6U7Q2w3WfKdDyfIUh9t6Eb2Wo4+1
ltLCqeevBgAqwf+/awIDAQABoxAwDjAMBgNVHRMBAf8EAjAAMA0GCSqGSIb3DQEB
CwUAA4GBADm0Fq4t7TIFVnQMG1zvFW35iOvs/eVdbNF4LYu0oRj0cLyM3VeLEQw6
5IlVbPc0GDbdHhqADrIm8r1eFR+nwNcGxcbgE2is74fZpD20mh2wuOeSDlhc8TD9
86tohF1Wd0uI52ZSaQvMvhnBX664DTccPkfLRja8NjoktiQx6paN
-----END CERTIFICATE-----"""

PRIVATE_KEY_MARKER = """-----BEGIN ENCRYPTED PRIVATE KEY-----
TEST_PRIVATE_KEY_SECRET
-----END ENCRYPTED PRIVATE KEY-----"""


def _certificate_config() -> str:
    return f'''config vpn certificate remote
    edit "REMOTE_Cert_1"
        set remote "{PUBLIC_CERTIFICATE}"
        set range global
        set auto-regenerate-days 30
    next
end
config vpn certificate local
    edit "Fortinet_SSL"
        set password ENC TEST_PASSWORD_SECRET
        set private-key "{PRIVATE_KEY_MARKER}"
        set certificate "{PUBLIC_CERTIFICATE}"
        set range global
        set source factory
        set comments "Factory certificate"
        set last-updated 1773822478
    next
    edit "User_Cert"
        set password ENC TEST_PASSWORD_SECRET
        set private-key "-----BEGIN PRIVATE KEY-----
TEST_PRIVATE_KEY_SECRET
-----END PRIVATE KEY-----"
        set certificate "{PUBLIC_CERTIFICATE}"
        set range global
        set source user
    next
end
'''


def test_remote_and_local_certificates_are_extracted_without_secrets() -> None:
    config = parse_fortigate_config(_certificate_config())

    assert len(config.certificates) == 3
    remote, factory, user = config.certificates

    assert remote.name == "REMOTE_Cert_1"
    assert remote.certificate_type == "remote"
    assert remote.range == "global"
    assert remote.has_certificate is True
    assert remote.has_private_key is False
    assert remote.subject == "CN=Sanitized Test Certificate"
    assert remote.issuer == "CN=Sanitized Test Certificate"
    assert remote.serial_number
    assert remote.valid_from is not None
    assert remote.valid_until is not None
    assert remote.sha256_fingerprint
    assert remote.public_key_algorithm == "RSA"
    assert remote.public_key_size == 1024
    assert remote.extra_settings == {"auto_regenerate_days": "30"}

    assert factory.certificate_type == "local"
    assert factory.source == "factory"
    assert factory.has_certificate is True
    assert factory.has_private_key is True
    assert factory.private_key_encrypted is True
    assert factory.has_password is True

    assert user.has_private_key is True
    assert user.private_key_encrypted is False
    assert user.has_password is True

    serialized = config.model_dump_json()
    assert "TEST_PRIVATE_KEY_SECRET" not in serialized
    assert "TEST_PASSWORD_SECRET" not in serialized
    assert "BEGIN ENCRYPTED PRIVATE KEY" not in serialized


def test_certificates_transform_to_extract_only_ir_without_secret_fields() -> None:
    ir = FGToIRTransformer(
        parse_fortigate_config(_certificate_config())
    ).transform()

    assert len(ir.certificates) == 3
    remote, factory, user = ir.certificates
    assert remote.migration_status == "EXTRACT_ONLY"
    assert remote.requires_manual_review is True
    assert factory.migration_status == "EXTRACT_ONLY"
    assert factory.requires_manual_review is False
    assert factory.source_last_updated is not None
    assert factory.source_last_updated.tzinfo is not None
    assert user.requires_manual_review is True
    assert factory.public_certificate_pem == PUBLIC_CERTIFICATE
    assert "private_key" not in factory.__class__.model_fields
    assert "password" not in factory.__class__.model_fields

    serialized = ir.model_dump_json()
    assert "TEST_PRIVATE_KEY_SECRET" not in serialized
    assert "TEST_PASSWORD_SECRET" not in serialized
    assert "BEGIN ENCRYPTED PRIVATE KEY" not in serialized


def test_certificate_excel_inventory_contains_metadata_but_no_secret_material() -> None:
    ir = FGToIRTransformer(
        parse_fortigate_config(_certificate_config())
    ).transform()
    workbook = load_workbook(
        io.BytesIO(IRExcelExporter(ir).generate()),
        data_only=False,
    )

    assert "Certificates" in workbook.sheetnames
    sheet = workbook["Certificates"]
    headers = {cell.value: cell.column for cell in sheet[3]}
    rows = {
        sheet.cell(row, headers["Name"]).value: row
        for row in range(4, sheet.max_row + 1)
    }

    remote_row = rows["REMOTE_Cert_1"]
    assert sheet.cell(remote_row, headers["Certificate Type"]).value == "remote"
    assert sheet.cell(remote_row, headers["Range"]).value == "global"
    assert sheet.cell(remote_row, headers["Subject"]).value
    assert sheet.cell(remote_row, headers["Issuer"]).value
    assert sheet.cell(remote_row, headers["SHA-256 Fingerprint"]).value

    local_row = rows["Fortinet_SSL"]
    assert sheet.cell(local_row, headers["Has Private Key"]).value == "Yes"
    assert sheet.cell(local_row, headers["Private Key Encrypted"]).value == "Yes"
    assert sheet.cell(local_row, headers["Has Password"]).value == "Yes"

    summary = {
        workbook["Summary"].cell(row, 1).value:
            workbook["Summary"].cell(row, 2).value
        for row in range(1, workbook["Summary"].max_row + 1)
    }
    assert summary["Certificates"] == 3

    all_text = "\n".join(
        str(cell.value)
        for worksheet in workbook.worksheets
        for row in worksheet.iter_rows()
        for cell in row
        if cell.value is not None
    )
    assert "TEST_PRIVATE_KEY_SECRET" not in all_text
    assert "TEST_PASSWORD_SECRET" not in all_text
    assert "BEGIN ENCRYPTED PRIVATE KEY" not in all_text
    assert "BEGIN CERTIFICATE" not in all_text


def test_malformed_certificate_is_retained_for_manual_review() -> None:
    config = '''config vpn certificate remote
    edit "Malformed"
        set remote "-----BEGIN CERTIFICATE-----
INVALID
-----END CERTIFICATE-----"
    next
end
'''

    parsed = parse_fortigate_config(config)
    assert len(parsed.certificates) == 1
    assert parsed.certificates[0].has_certificate is True
    assert parsed.certificates[0].parse_error

    ir = FGToIRTransformer(parsed).transform()
    assert len(ir.certificates) == 1
    assert ir.certificates[0].requires_manual_review is True
    assert ir.certificates[0].parse_error
    assert "INVALID" not in ir.certificates[0].parse_error


def test_empty_and_unsupported_certificate_sections_create_no_records() -> None:
    config = '''config vpn certificate ca
end
config vpn certificate remote
end
'''

    assert parse_fortigate_config(config).certificates == []


def test_invalid_last_updated_is_preserved_safely_without_crashing() -> None:
    config = f'''config vpn certificate local
    edit "Bad_Timestamp"
        set certificate "{PUBLIC_CERTIFICATE}"
        set last-updated not-an-epoch
    next
end
'''

    parsed = parse_fortigate_config(config)
    assert parsed.certificates[0].last_updated is None
    assert parsed.certificates[0].extra_settings["last_updated_raw"] == (
        "not-an-epoch"
    )

    ir = FGToIRTransformer(parsed).transform()
    assert ir.certificates[0].source_last_updated is None
    assert ir.certificates[0].source_attributes["last_updated_raw"] == (
        "not-an-epoch"
    )
