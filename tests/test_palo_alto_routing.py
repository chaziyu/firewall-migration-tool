import io

from openpyxl import load_workbook

from fwmigrate.parsers.palo_alto.parser import PANOSSourceParser
from fwmigrate.report.excel_exporter import IRExcelExporter


def _config(destination: str, objects: str = "", name: str = "route") -> str:
    return f"""<config version='11.1.0'><devices><entry name='fw'>
      <network><virtual-router><entry name='vr'><routing-table><ip><static-route>
        <entry name='{name}'><destination>{destination}</destination>
          <nexthop><ip-address>192.0.2.1</ip-address></nexthop><interface>ethernet1/8</interface><metric>10</metric><admin-dist>10</admin-dist>
        </entry>
      </static-route></ip></routing-table></entry></virtual-router></network>
      <vsys><entry name='vsys1'>{objects}</entry></vsys>
    </entry></devices></config>"""


def _route(destination: str, objects: str = "", name: str = "route"):
    result = PANOSSourceParser().extract(_config(destination, objects, name))
    return result, result.canonical_ir.routes[0]


def test_literal_and_host_destinations_are_normalized_without_references():
    result, route = _route("10.10.0.0/16")
    assert len(result.canonical_ir.routes) == 1
    assert route.destination == "10.10.0.0/16"
    assert route.source_destination == "10.10.0.0/16"
    assert route.source_destination_reference is None

    _, host = _route("10.10.10.10", name="host")
    assert host.destination == "10.10.10.10/32"


def test_named_ip_netmask_and_host_objects_keep_source_reference():
    objects = """
      <address><entry name='ROUTE-NET'><ip-netmask>10.20.0.0/16</ip-netmask></entry>
        <entry name='ROUTE-HOST'><ip-netmask>10.20.1.10</ip-netmask></entry></address>
    """
    _, network = _route("ROUTE-NET", objects)
    assert network.destination == "10.20.0.0/16"
    assert network.source_destination == "ROUTE-NET"
    assert network.source_destination_reference == "ROUTE-NET"
    assert network.requires_manual_review is True
    assert network.migration_status == "PARTIALLY_NORMALIZED"
    assert network.metric == 10
    assert network.administrative_distance == 10
    assert network.source_attributes["pan_resolved_destination"] == "10.20.0.0/16"

    _, host = _route("ROUTE-HOST", objects, name="host-object")
    assert host.destination == "10.20.1.10/32"
    assert host.source_destination_reference == "ROUTE-HOST"


def test_unresolved_destination_is_retained_for_review():
    result, route = _route("UNKNOWN_ROUTE_OBJECT")
    assert route in result.canonical_ir.routes
    assert route.destination is None
    assert route.source_destination == "UNKNOWN_ROUTE_OBJECT"
    assert route.source_destination_reference == "UNKNOWN_ROUTE_OBJECT"
    assert route.requires_manual_review is True
    assert "unresolved-destination-reference" in route.review_reasons
    assert all(item.status.value != "PARSE_ERROR" for item in result.inventory_items if item.domain == "routes")
    section = next(section for section in result.source_sections if "static-route" in section.path)
    assert (section.object_count_source, section.object_count_parsed) == (1, 1)


def test_group_fqdn_and_range_destinations_are_not_flattened():
    cases = [
        ("ROUTE-GROUP", "<address-group><entry name='ROUTE-GROUP'><static><member>x</member></static></entry></address-group>", "destination-address-group-reference"),
        ("ROUTE-FQDN", "<address><entry name='ROUTE-FQDN'><fqdn>example.test</fqdn></entry></address>", "unsupported-destination-reference-type"),
        ("ROUTE-RANGE", "<address><entry name='ROUTE-RANGE'><ip-range>10.20.1.1-10.20.1.9</ip-range></entry></address>", "unsupported-destination-reference-type"),
    ]
    for destination, objects, reason in cases:
        _, route = _route(destination, objects)
        assert route.destination is None
        assert route.source_destination_reference == destination
        assert route.requires_manual_review is True
        assert reason in route.review_reasons
        assert route.source_attributes["pan_destination_reference_kind"] in {"address", "address-group"}


def test_address_family_mismatch_is_retained_for_review():
    objects = "<address><entry name='V6-NET'><ip-netmask>2001:db8::/32</ip-netmask></entry></address>"
    _, route = _route("V6-NET", objects)
    assert route.destination is None
    assert route.source_destination_reference == "V6-NET"
    assert "destination-address-family-mismatch" in route.review_reasons


def test_same_reference_name_resolves_only_in_each_device_vsys():
    def device(name: str, prefix: str) -> str:
        return f"""<entry name='{name}'><network><virtual-router><entry name='vr'>
          <routing-table><ip><static-route><entry name='{name}-route'><destination>ROUTE-NET</destination>
            <nexthop><discard/></nexthop></entry></static-route></ip></routing-table>
        </entry></virtual-router></network><vsys><entry name='vsys1'><address>
          <entry name='ROUTE-NET'><ip-netmask>{prefix}</ip-netmask></entry>
        </address></entry></vsys></entry>"""

    xml = "<config><devices>" + device("fw-a", "10.1.0.0/16") + device("fw-b", "10.2.0.0/16") + "</devices></config>"
    result = PANOSSourceParser().extract(xml)
    routes = {route.name: route.destination for route in result.canonical_ir.routes}
    assert routes == {"fw-a-route": "10.1.0.0/16", "fw-b-route": "10.2.0.0/16"}


def test_named_route_is_visible_in_excel():
    result, _ = _route(
        "ROUTE-NET",
        "<address><entry name='ROUTE-NET'><ip-netmask>10.20.0.0/16</ip-netmask></entry></address>",
        name="named-route",
    )
    workbook = load_workbook(io.BytesIO(IRExcelExporter(result.canonical_ir).generate()))
    sheet = workbook["Routes"]
    headers = {cell.value: cell.column for cell in sheet[3]}
    values = {header: sheet.cell(4, column).value for header, column in headers.items()}
    assert values["Name"] == "named-route"
    assert values["Destination"] == "10.20.0.0/16"
    assert values["Source Destination"] == "ROUTE-NET"
    assert values["Destination Object / Group"] == "ROUTE-NET"
    assert values["Next Hop"] == "192.0.2.1"
    assert values["Interface"] == "ethernet1/8"
    assert result.canonical_ir.routes[0].safe_for_target_generation is False
