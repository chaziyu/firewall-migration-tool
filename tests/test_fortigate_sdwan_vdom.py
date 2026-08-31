import io

from openpyxl import load_workbook

from fwmigrate.ir.io import load_ir_payload
from fwmigrate.parsers.fortigate.extractor import extract_fortigate_config
from fwmigrate.parsers.fortigate.parser import parse_fortigate_config
from fwmigrate.parsers.fortigate.transformer import FGToIRTransformer
from fwmigrate.report.excel_exporter import IRExcelExporter


MULTI_VDOM_SDWAN_CONFIG = '''
config system interface
    edit "wan1"
        set ip 192.0.2.1 255.255.255.0
    next
end
config system sdwan
    set status enable
    config zone
        edit "WAN"
        next
    end
    config members
        edit 1
            set interface "wan1"
            set zone "WAN"
        next
    end
    config health-check
        edit "internet"
            set server "1.1.1.1"
            config sla
                edit 1
                    set latency-threshold 100
                next
            end
        next
    end
    config service
        edit 10
            set name "tenant-independent-name"
            set health-check "internet"
            config sla
                edit "service-sla"
                    set latency-threshold 50
                next
            end
        next
    end
    config duplication
        edit 1
            set service-id 10
        next
    end
    config neighbor
        edit "peer-a"
            set member 1
        next
    end
end
config vdom
    edit "tenant-a"
        config system interface
            edit "wan1"
                set ip 198.51.100.1 255.255.255.0
            next
        end
        config system sdwan
            set status enable
            config zone
                edit "WAN"
                next
            end
            config members
                edit 1
                    set interface "wan1"
                    set zone "WAN"
                next
            end
            config health-check
                edit "internet"
                    set server "9.9.9.9"
                next
            end
            config service
                edit 10
                    set name "tenant-a-rule"
                    set health-check "internet"
                next
            end
        end
        config firewall policy
            edit 101
                set srcintf "wan1"
                set dstintf "ROOT_ONLY"
                set srcaddr "all"
                set dstaddr "all"
                set action accept
                set schedule "always"
                set service "ALL"
            next
        end
    next
end
'''


def test_parser_keeps_sdwan_state_and_duplicate_member_ids_per_vdom():
    parsed = parse_fortigate_config(MULTI_VDOM_SDWAN_CONFIG)

    assert [(sdwan.source_context, len(sdwan.zones)) for sdwan in parsed.sdwans] == [
        ("root", 1),
        ("tenant-a", 1),
    ]
    assert [
        (sdwan.source_context, member.id, member.interface)
        for sdwan in parsed.sdwans
        for member in sdwan.members
    ] == [
        ("root", 1, "wan1"),
        ("tenant-a", 1, "wan1"),
    ]

    for sdwan in parsed.sdwans:
        assert all(zone.source_context == sdwan.source_context for zone in sdwan.zones)
        assert all(member.source_context == sdwan.source_context for member in sdwan.members)
        assert all(check.source_context == sdwan.source_context for check in sdwan.health_checks)
        assert all(
            sla.source_context == sdwan.source_context
            for check in sdwan.health_checks
            for sla in check.sla
        )
        assert all(rule.source_context == sdwan.source_context for rule in sdwan.services)
        assert all(
            sla.source_context == sdwan.source_context
            for rule in sdwan.services
            for sla in rule.sla
        )
        assert all(rule.source_context == sdwan.source_context for rule in sdwan.duplication_rules)
        assert all(neighbor.source_context == sdwan.source_context for neighbor in sdwan.neighbors)

    extraction = extract_fortigate_config(MULTI_VDOM_SDWAN_CONFIG)
    sdwan_inventory = [
        item
        for item in extraction.inventory_items
        if item.source_path.startswith("system sdwan")
    ]
    assert {item.source_context for item in sdwan_inventory} == {"root", "tenant-a"}


def test_transformer_preserves_default_zone_ownership_and_interface_context():
    ir = FGToIRTransformer(parse_fortigate_config(MULTI_VDOM_SDWAN_CONFIG)).transform()

    assert [(sdwan.source_context, len(sdwan.zones)) for sdwan in ir.sdwans] == [
        ("root", 1),
        ("tenant-a", 1),
    ]
    assert [
        (sdwan.source_context, zone.name)
        for sdwan in ir.sdwans
        for zone in sdwan.zones
    ] == [("root", "WAN"), ("tenant-a", "WAN")]

    interfaces = {
        (interface.source_context, interface.name): interface
        for interface in ir.interfaces
    }
    assert interfaces["root", "wan1"].zone == "WAN"
    assert interfaces["tenant-a", "wan1"].zone == "WAN"
    assert interfaces["root", "wan1"].ip == "192.0.2.1/24"
    assert interfaces["tenant-a", "wan1"].ip == "198.51.100.1/24"
    assert all(
        child.source_context == sdwan.source_context
        for sdwan in ir.sdwans
        for check in sdwan.health_checks
        for child in check.sla
    )
    assert all(
        child.source_context == sdwan.source_context
        for sdwan in ir.sdwans
        for rule in sdwan.rules
        for child in rule.sla
    )


def test_duplicate_sdwan_names_do_not_cross_context_dependency_resolution():
    result = extract_fortigate_config(MULTI_VDOM_SDWAN_CONFIG)

    tenant_dependencies = [
        dependency
        for dependency in result.dependencies
        if dependency.source_context == "tenant-a"
    ]
    assert any(
        dependency.reference == "ROOT_ONLY"
        and dependency.result == "UNRESOLVED"
        for dependency in tenant_dependencies
    )


def test_excel_sdwan_rows_identify_their_vdom():
    result = extract_fortigate_config(MULTI_VDOM_SDWAN_CONFIG)
    workbook = load_workbook(
        io.BytesIO(IRExcelExporter(result.canonical_ir, result).generate())
    )

    assert "VDOM" in [cell.value for cell in workbook["SD-WAN Zones"][3]]
    zone_rows = list(workbook["SD-WAN Zones"].iter_rows(min_row=4, values_only=True))
    assert {(row[0], row[2]) for row in zone_rows} == {
        ("WAN", "root"),
        ("WAN", "tenant-a"),
    }

    member_sheet = workbook["SD-WAN Members"]
    member_headers = [cell.value for cell in member_sheet[3]]
    vdom_column = member_headers.index("VDOM")
    member_rows = list(member_sheet.iter_rows(min_row=4, values_only=True))
    assert {row[vdom_column] for row in member_rows} == {"root", "tenant-a"}


def test_default_virtual_wan_link_is_scoped_per_vdom():
    config = '''
config system sdwan
    config zone
        edit "virtual-wan-link"
        next
    end
end
config vdom
    edit "tenant-a"
        config system sdwan
            config zone
                edit "virtual-wan-link"
                next
            end
        end
    next
end
'''

    ir = FGToIRTransformer(parse_fortigate_config(config)).transform()

    assert [
        (sdwan.source_context, zone.name)
        for sdwan in ir.sdwans
        for zone in sdwan.zones
    ] == [
        ("root", "virtual-wan-link"),
        ("tenant-a", "virtual-wan-link"),
    ]


def test_legacy_ir_sdwan_payload_migrates_to_context_scoped_collection():
    ir = load_ir_payload({
        "schema_version": "1.17",
        "metadata": {"source_vendor": "fortigate"},
        "sdwan": {
            "rules": [{"source_id": 10}],
        },
    })

    assert len(ir.sdwans) == 1
    assert ir.sdwans[0].source_context == "root"
    assert ir.sdwans[0].rules[0].source_context == "root"
    assert ir.model_dump().get("sdwan") is None
