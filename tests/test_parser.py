import pytest
import io
from openpyxl import load_workbook
from fwmigrate.parsers.fortigate.tokenizer import FortiGateTokenizer
from fwmigrate.parsers.fortigate.parser import FortiGateParser
from fwmigrate.parsers.fortigate.transformer import FGToIRTransformer
from fwmigrate.report.excel_exporter import IRExcelExporter


def test_fortios_header_metadata_survives_parser_ir_and_excel():
    config = """#config-version=FG2H0G-7.4.11-FW-build2878-260126:opmode=0:vdom=0:user=test
#buildno=2878
config system global
    set hostname "FGT"
end
"""

    fg = FortiGateParser(FortiGateTokenizer(config)).parse()
    assert fg.source_version == "7.4.11"
    assert fg.source_build == "2878"

    ir = FGToIRTransformer(fg).transform()
    assert ir.metadata.source_version == "FortiOS 7.4.11 build 2878"

    workbook = load_workbook(io.BytesIO(IRExcelExporter(ir).generate()))
    summary_values = {
        row[0].value: row[1].value
        for row in workbook["Summary"].iter_rows()
        if row[0].value
    }
    assert summary_values["Source Version"] == "FortiOS 7.4.11 build 2878"


def test_missing_fortios_header_does_not_invent_source_version():
    fg = FortiGateParser(FortiGateTokenizer("config system global\nend\n")).parse()
    assert fg.source_version is None
    assert fg.source_build is None
    assert FGToIRTransformer(fg).transform().metadata.source_version is None

def test_parse_system_global():
    config = """
config system global
    set hostname "GENE-FW2"
    set admin-sport 8443
    set timezone 60
end
    """
    tokenizer = FortiGateTokenizer(config)
    parser = FortiGateParser(tokenizer)
    cfg = parser.parse()
    
    assert cfg.system_global is not None
    assert cfg.system_global.hostname == "GENE-FW2"
    assert cfg.system_global.admin_sport == 8443

def test_parse_interface():
    config = """
config system interface
    edit "port1"
        set vdom "root"
        set ip 10.0.0.1 255.255.255.0
        set allowaccess ping https ssh
        set type physical
        set role lan
    next
    edit "port2"
        set vdom "root"
        set ip 10.0.0.2 255.255.255.0
        set allowaccess ping
        set type physical
        set role wan
    next
end
    """
    tokenizer = FortiGateTokenizer(config)
    parser = FortiGateParser(tokenizer)
    cfg = parser.parse()
    
    assert len(cfg.interfaces) == 2
    
    intf1 = cfg.interfaces[0]
    assert intf1.name == "port1"
    assert intf1.ip == "10.0.0.1 255.255.255.0"
    assert "ssh" in intf1.allowaccess
    assert intf1.role == "lan"

def test_parse_interface_preserves_all_explicit_source_settings():
    config = """
config system interface
    edit "x1"
        set vdom "root"
        set mode dhcp
        set allowaccess ping
        set type physical
        set lldp-reception disable
        set role wan
        set snmp-index 3
        set password "must-not-be-retained"
    next
end
    """
    cfg = FortiGateParser(FortiGateTokenizer(config)).parse()

    interface = cfg.interfaces[0]
    assert interface.mode == "dhcp"
    assert interface.allowaccess == ["ping"]
    assert interface.type == "physical"
    assert interface.role == "wan"
    assert interface.source_attributes["lldp_reception"] == "disable"
    assert interface.source_attributes["snmp_index"] == "3"
    assert interface.source_attributes["password"] == "[REDACTED]"


def test_parse_and_transform_tunnel_interface_remote_ip():
    config = """
config system interface
    edit "Tunnel_With_IP"
        set vdom "root"
        set ip 10.255.0.1 255.255.255.255
        set type tunnel
        set remote-ip 10.255.0.2 255.255.255.255
        set interface "port1"
    next
    edit "Tunnel_No_IP"
        set type tunnel
        set interface "port1"
    next
    edit "Tunnel_Remote_Only"
        set type tunnel
        set remote-ip 10.1.1.2 255.255.255.255
        set interface "port1"
    next
end
    """
    cfg = FortiGateParser(FortiGateTokenizer(config)).parse()

    parsed = {interface.name: interface for interface in cfg.interfaces}
    with_ip = parsed["Tunnel_With_IP"]
    assert with_ip.type == "tunnel"
    assert with_ip.interface == "port1"
    assert with_ip.ip == "10.255.0.1 255.255.255.255"
    assert with_ip.remote_ip == "10.255.0.2 255.255.255.255"
    assert with_ip.source_attributes["remote_ip"] == "10.255.0.2 255.255.255.255"

    ir = FGToIRTransformer(cfg).transform()
    transformed = {interface.name: interface for interface in ir.interfaces}

    with_ip_ir = transformed["Tunnel_With_IP"]
    assert with_ip_ir.interface_type == "tunnel"
    assert with_ip_ir.ip == "10.255.0.1/32"
    assert with_ip_ir.remote_ip == "10.255.0.2/32"
    assert with_ip_ir.parent == "port1"
    assert with_ip_ir.source_attributes["remote_ip"] == "10.255.0.2 255.255.255.255"

    no_ip_ir = transformed["Tunnel_No_IP"]
    assert no_ip_ir.ip is None
    assert no_ip_ir.remote_ip is None
    assert no_ip_ir.parent == "port1"
    assert no_ip_ir.interface_type == "tunnel"

    remote_only_ir = transformed["Tunnel_Remote_Only"]
    assert remote_only_ir.ip is None
    assert remote_only_ir.remote_ip == "10.1.1.2/32"
    assert remote_only_ir.parent == "port1"


def test_parse_firewall_address():
    config = """
config firewall address
    edit "local_net"
        set subnet 192.168.1.0 255.255.255.0
    next
    edit "google"
        set type fqdn
        set fqdn "google.com"
    next
end
    """
    tokenizer = FortiGateTokenizer(config)
    parser = FortiGateParser(tokenizer)
    cfg = parser.parse()
    
    assert len(cfg.addresses) == 2
    assert cfg.addresses[0].name == "local_net"
    assert cfg.addresses[0].subnet == "192.168.1.0 255.255.255.0"
    
    assert cfg.addresses[1].name == "google"
    assert cfg.addresses[1].type == "fqdn"
    assert cfg.addresses[1].fqdn == "google.com"

def test_parse_firewall_policy():
    config = """
config firewall policy
    edit 1
        set name "allow_out"
        set uuid 0819b852-ebb4-51eb-210e-517744c1e41b
        set srcintf "lan"
        set dstintf "wan1"
        set srcaddr "all"
        set dstaddr "all"
        set action accept
        set schedule "always"
        set service "ALL"
        set nat enable
    next
    edit 2
        set name "deny_in"
        set srcintf "wan1"
        set dstintf "lan"
        set srcaddr "all"
        set dstaddr "all"
        set action deny
        set schedule "always"
        set service "ALL"
    next
end
    """
    tokenizer = FortiGateTokenizer(config)
    parser = FortiGateParser(tokenizer)
    cfg = parser.parse()
    
    assert len(cfg.policies) == 2
    p = cfg.policies[0]
    assert p.id == 1
    assert p.uuid == "0819b852-ebb4-51eb-210e-517744c1e41b"
    assert p.name == "allow_out"
    assert p.srcintf == ["lan"]
    assert p.action == "accept"
    assert p.nat == "enable"
    assert cfg.policies[1].uuid is None


def test_parse_firewall_policy_preserves_identity_selectors_as_lists():
    config = """
config firewall policy
    edit 100
        set name "Identity_Test"
        set srcintf "LAN"
        set dstintf "WAN"
        set srcaddr "all"
        set dstaddr "all"
        set groups "SSLVPN Users" "Domain_Users"
        set users "alice" "bob.smith"
        set action accept
        set schedule "always"
        set service "ALL"
    next
    edit 101
        set name "No_Identity"
        set srcintf "LAN"
        set dstintf "WAN"
        set srcaddr "all"
        set dstaddr "all"
        set action accept
        set schedule "always"
        set service "ALL"
    next
end
    """

    cfg = FortiGateParser(FortiGateTokenizer(config)).parse()

    assert cfg.policies[0].groups == ["SSLVPN Users", "Domain_Users"]
    assert cfg.policies[0].users == ["alice", "bob.smith"]
    assert cfg.policies[1].groups == []
    assert cfg.policies[1].users == []

    ir = FGToIRTransformer(cfg).transform()
    assert ir.policies[0].source_user_groups == ["SSLVPN Users", "Domain_Users"]
    assert ir.policies[0].source_users == ["alice", "bob.smith"]
    assert ir.policies[1].source_user_groups == []
    assert ir.policies[1].source_users == []


def test_parse_firewall_policy_preserves_advanced_and_unknown_settings():
    config = """
config firewall policy
    edit 100
        set name "advanced_policy"
        set uuid 11111111-2222-3333-4444-555555555555
        set srcintf "LAN"
        set dstintf "WAN"
        set srcaddr "all"
        set dstaddr "all"
        set action accept
        set schedule "always"
        set service "ALL"
        set inspection-mode proxy
        set ztna-status enable
        set ztna-ems-tag "TAG_A" "TAG B"
        set timeout-send-rst enable
        set auto-asic-offload disable
        set np-acceleration disable
        set port-preserve disable
        set future-policy-option some-value
        set custom-secret "must-not-survive"
    next
end
    """

    policy = FortiGateParser(FortiGateTokenizer(config)).parse().policies[0]

    assert policy.inspection_mode == "proxy"
    assert policy.ztna_status == "enable"
    assert policy.ztna_ems_tag == ["TAG_A", "TAG B"]
    assert policy.extra_settings == {
        "timeout_send_rst": "enable",
        "auto_asic_offload": "disable",
        "np_acceleration": "disable",
        "port_preserve": "disable",
        "future_policy_option": "some-value",
        "custom_secret": "[REDACTED]",
    }
    assert "must-not-survive" not in policy.extra_settings.values()
    assert "inspection_mode" not in policy.extra_settings
    assert "ztna_status" not in policy.extra_settings
    assert "ztna_ems_tag" not in policy.extra_settings


def test_parse_firewall_policy_preservation_fields_default_to_absent():
    config = """
config firewall policy
    edit 101
        set name "ordinary_policy"
        set srcintf "LAN"
        set dstintf "WAN"
        set srcaddr "all"
        set dstaddr "all"
        set action accept
        set schedule "always"
        set service "ALL"
    next
end
    """

    policy = FortiGateParser(FortiGateTokenizer(config)).parse().policies[0]

    assert policy.inspection_mode is None
    assert policy.ztna_status is None
    assert policy.ztna_ems_tag == []
    assert policy.extra_settings == {}

def test_parse_nested_config():
    config = """
config system sdwan
    set status enable
    config zone
        edit "virtual-wan-link"
        next
    end
    config members
        edit 1
            set interface "wan1"
            set zone "virtual-wan-link"
        next
    end
end
    """
    tokenizer = FortiGateTokenizer(config)
    parser = FortiGateParser(tokenizer)
    cfg = parser.parse()
    
    # We silently ignored the nested ones for now in mvp parser
    # But it shouldn't crash.
    assert cfg.sdwan is not None
    assert cfg.sdwan.status == "enable"

