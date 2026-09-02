import io

from openpyxl import load_workbook

from fwmigrate.ir.core import IRAddress
from fwmigrate.ir.enums import AddressType
from fwmigrate.parsers.fortigate.parser import parse_fortigate_config
from fwmigrate.parsers.fortigate.transformer import FGToIRTransformer
from fwmigrate.parsers.fortigate.extractor import extract_fortigate_config
from fwmigrate.report.excel_exporter import IRExcelExporter


ADDRESS_CONFIG = """
config firewall address
    edit "normal-net"
        set uuid 11111111-1111-1111-1111-111111111111
        set subnet 192.168.10.0 255.255.255.0
        set allow-routing enable
        set associated-interface "port6"
        set color 9
        set cache-ttl 300
        set password "must-not-be-retained"
    next

    edit "geo"
        set type geography
        set country "MY"
    next

    edit "geo-missing-country"
        set type geography
    next

    edit "ems"
        set type dynamic
        set sub-type ems-tag
        set dirty clean
        set obj-tag "Deleum_ADUser"
        set tag-type "zero_trust"
        set obj-type mac
    next

    edit "mac-source"
        set uuid 33333333-3333-3333-3333-333333333333
        set type mac
        set macaddr "00:11:22:33:44:55"
        set associated-interface "port7"
        set comment "Unsupported MAC inventory"
    next
end

config firewall address6
    edit "SSLVPN_TUNNEL_IPv6_ADDR1"
        set uuid 17523864-65a4-51e9-c45e-65c6367ea4e3
        set ip6 fdff:ffff::/120
    next
    edit "ipv6-test"
        set uuid 22222222-2222-2222-2222-222222222222
        set ip6 fdff:ffff::/120
        set fabric-object enable
    next
    edit "all"
        set uuid aaaaaaaa-aaaa-aaaa-aaaa-aaaaaaaaaaaa
    next
    edit "none"
        set uuid bbbbbbbb-bbbb-bbbb-bbbb-bbbbbbbbbbbb
        set ip6 ::/128
    next
end


config firewall multicast-address
    edit "multicast-test"
        set start-ip 239.1.1.1
        set end-ip 239.1.1.2
        set visibility enable
    next
end

config firewall multicast-address6
    edit "all"
        set ip6 ff00::/8
        set visibility enable
    next
end

config firewall wildcard-fqdn custom
    edit "cdn-apple"
        set uuid cccccccc-cccc-cccc-cccc-cccccccccccc
        set wildcard-fqdn "*.cdn-apple.com"
        set cache-ttl 60
    next
    edit "google-play"
        set uuid dddddddd-dddd-dddd-dddd-dddddddddddd
        set wildcard-fqdn "*play.google.com"
    next
end
"""


SPECIAL_ADDRESS_CONFIG = """
config firewall address
    edit "all"
        set uuid 00000000-0000-0000-0000-000000000001
        set subnet 0.0.0.0 0.0.0.0
    next
    edit "none"
        set uuid 00000000-0000-0000-0000-000000000002
        set subnet 0.0.0.0 255.255.255.255
    next
    edit "FABRIC_DEVICE"
    next
    edit "FIREWALL_AUTH_PORTAL_ADDRESS"
    next
end

config firewall address6
    edit "all"
        set ip6 ::/0
    next
    edit "none"
        set ip6 ::/128
    next
end

config firewall multicast-address
    edit "none"
        set start-ip 239.255.255.255
        set end-ip 239.255.255.255
    next
end

config firewall multicast-address6
    edit "FABRIC_DEVICE"
        set ip6 ff00::/8
    next
end
"""


def _by_name(items):
    return {item.name: item for item in items}


def test_fortigate_address_parser_preserves_typed_and_unknown_settings():
    config = parse_fortigate_config(ADDRESS_CONFIG)
    addresses = _by_name(config.addresses)

    normal = addresses["normal-net"]
    assert normal.uuid == "11111111-1111-1111-1111-111111111111"
    assert normal.associated_interface == "port6"
    assert normal.allow_routing == "enable"
    assert normal.color == 9
    assert normal.cache_ttl == 300
    assert normal.extra_settings == {
        "password": "[REDACTED]",
    }

    assert addresses["geo"].country == "MY"

    ems = addresses["ems"]
    assert ems.sub_type == "ems-tag"
    assert ems.obj_tag == "Deleum_ADUser"
    assert ems.tag_type == "zero_trust"
    assert ems.obj_type == "mac"
    assert ems.dirty == "clean"

    ipv6 = addresses["ipv6-test"]
    assert ipv6.ip6 == "fdff:ffff::/120"
    assert ipv6.is_ipv6 is True
    assert ipv6.fabric_object == "enable"
    assert ipv6.extra_settings == {}

    sslvpn_ipv6 = addresses["SSLVPN_TUNNEL_IPv6_ADDR1"]
    assert sslvpn_ipv6.uuid == "17523864-65a4-51e9-c45e-65c6367ea4e3"
    assert sslvpn_ipv6.ip6 == "fdff:ffff::/120"
    assert sslvpn_ipv6.is_ipv6 is True

    multicast6 = [
        item
        for item in config.addresses
        if item.name == "all" and item.is_multicast
    ][0]
    assert multicast6.ip6 == "ff00::/8"
    assert multicast6.is_ipv6 is True
    assert multicast6.is_multicast is True
    assert multicast6.extra_settings == {"visibility": "enable"}

    wildcard = _by_name(config.wildcard_fqdns)
    assert wildcard["cdn-apple"].uuid == "cccccccc-cccc-cccc-cccc-cccccccccccc"
    assert wildcard["cdn-apple"].extra_settings == {"cache_ttl": "60"}

    multicast = addresses["multicast-test"]
    assert multicast.is_multicast is True
    assert multicast.extra_settings == {"visibility": "enable"}


def test_fortigate_address_transform_preserves_semantics_and_source_metadata():
    ir = FGToIRTransformer(
        parse_fortigate_config(ADDRESS_CONFIG)
    ).transform()
    addresses = _by_name(ir.addresses)
    groups = _by_name(ir.address_groups)

    normal = addresses["normal-net"]
    assert normal.type == AddressType.NETWORK
    assert normal.value == "192.168.10.0/24"
    assert normal.source_uuid == "11111111-1111-1111-1111-111111111111"
    assert normal.associated_interface == "port6"
    assert normal.allow_routing is True
    assert normal.source_color == 9
    assert normal.source_cache_ttl == 300
    assert normal.source_attributes == {
        "cache_ttl": 300,
        "password": "[REDACTED]",
        "subnet": "192.168.10.0 255.255.255.0",
    }

    geo = addresses["geo"]
    assert geo.type == AddressType.GEO
    assert geo.geo_code == "MY"

    missing_geo = addresses["geo-missing-country"]
    assert missing_geo.value == ""
    assert missing_geo.requires_manual_review is True
    assert missing_geo.requires_manual_review is True
    assert "unknown" not in (missing_geo.raw_value or "")

    ems = groups["ems"]
    assert ems.dynamic_filter == "'Deleum_ADUser'"
    assert ems.source_sub_type == "ems-tag"
    assert ems.source_obj_tag == "Deleum_ADUser"
    assert ems.source_tag_type == "zero_trust"
    assert ems.source_obj_type == "mac"
    assert ems.source_dirty == "clean"

    ipv6 = addresses["ipv6-test"]
    assert ipv6.type == AddressType.NETWORK
    assert ipv6.value == "fdff:ffff::/120"
    assert ipv6.is_ipv6 is True
    assert ipv6.source_uuid == "22222222-2222-2222-2222-222222222222"
    assert ipv6.source_fabric_object_setting == "enable"
    assert ipv6.source_attributes == {"fabric_object": "enable"}

    sslvpn_ipv6 = addresses["SSLVPN_TUNNEL_IPv6_ADDR1"]
    assert sslvpn_ipv6.value == "fdff:ffff::/120"
    assert sslvpn_ipv6.is_ipv6 is True
    assert sslvpn_ipv6.source_uuid == "17523864-65a4-51e9-c45e-65c6367ea4e3"

    ipv6_all = next(
        item for item in ir.addresses
        if item.name == "all" and item.is_ipv6 and not item.is_multicast
    )
    ipv6_none = next(
        item for item in ir.addresses
        if item.name == "none" and item.is_ipv6 and not item.is_multicast
    )
    multicast6_all = next(
        item for item in ir.addresses
        if item.name == "all" and item.is_ipv6 and item.is_multicast
    )
    assert ipv6_all.type == AddressType.SPECIAL
    assert ipv6_all.value == "all"
    assert ipv6_all.is_ipv6 is True
    assert ipv6_none.type == AddressType.SPECIAL
    assert ipv6_none.value == "none"
    assert ipv6_none.source_attributes == {"ip6": "::/128"}
    assert multicast6_all.type == AddressType.SPECIAL
    assert multicast6_all.value == "all"
    assert multicast6_all.is_multicast is True
    assert multicast6_all.source_attributes == {
        "visibility": "enable",
        "ip6": "ff00::/8",
    }

    cdn_apple = addresses["cdn-apple"]
    assert cdn_apple.value == "*.cdn-apple.com"
    assert cdn_apple.source_uuid == "cccccccc-cccc-cccc-cccc-cccccccccccc"
    assert cdn_apple.source_attributes == {
        "cache_ttl": "60",
        "wildcard_fqdn": "*.cdn-apple.com",
    }
    assert addresses["google-play"].value == "*play.google.com"

    multicast = addresses["multicast-test"]
    assert multicast.type == AddressType.RANGE
    assert multicast.is_multicast is True
    assert multicast.source_attributes == {"visibility": "enable"}

    mac = addresses["mac-source"]
    assert mac.type == AddressType.MAC
    assert mac.mac == "00:11:22:33:44:55"
    assert mac.value == "00:11:22:33:44:55"
    assert mac.requires_manual_review is False
    assert mac.source_uuid == "33333333-3333-3333-3333-333333333333"
    assert mac.associated_interface == "port7"


def test_fortigate_special_addresses_are_preserved_without_fabricated_values():
    config = parse_fortigate_config(SPECIAL_ADDRESS_CONFIG)
    assert len(config.addresses) == 8

    ir = FGToIRTransformer(config).transform()
    ipv4_unicast = {
        item.name: item
        for item in ir.addresses
        if not item.is_ipv6 and not item.is_multicast
    }

    assert set(ipv4_unicast) == {
        "all",
        "none",
        "FABRIC_DEVICE",
        "FIREWALL_AUTH_PORTAL_ADDRESS",
    }
    assert ipv4_unicast["all"].type == AddressType.SPECIAL
    assert ipv4_unicast["all"].value == "all"
    assert ipv4_unicast["all"].source_uuid.endswith("0001")
    assert ipv4_unicast["all"].source_attributes["subnet"] == (
        "0.0.0.0 0.0.0.0"
    )

    none = ipv4_unicast["none"]
    assert none.type == AddressType.SPECIAL
    assert none.value == "none"
    assert none.source_attributes["subnet"] == "0.0.0.0 255.255.255.255"
    assert none.requires_manual_review is True
    assert none.value not in {"any", "0.0.0.0/0", "0.0.0.0/32"}
    assert not none.value.startswith(("198.18.", "198.19."))

    for name in ("FABRIC_DEVICE", "FIREWALL_AUTH_PORTAL_ADDRESS"):
        special = ipv4_unicast[name]
        assert special.type == AddressType.SPECIAL
        assert special.value == name
        assert special.original_type == "fortigate_reserved"
        assert special.requires_manual_review is True

    ipv6 = [
        item for item in ir.addresses
        if item.is_ipv6 and not item.is_multicast
    ]
    assert [(item.name, item.value) for item in ipv6] == [
        ("all", "all"),
        ("none", "none"),
    ]
    assert all(item.type == AddressType.SPECIAL for item in ipv6)

    multicast4 = next(
        item for item in ir.addresses
        if item.name == "none" and item.is_multicast and not item.is_ipv6
    )
    assert multicast4.type == AddressType.SPECIAL
    assert multicast4.value == "none"

    multicast6 = next(
        item for item in ir.addresses
        if item.name == "FABRIC_DEVICE" and item.is_multicast and item.is_ipv6
    )
    assert multicast6.type == AddressType.SPECIAL
    assert multicast6.value == "FABRIC_DEVICE"


def test_special_address_value_uses_original_value_without_typed_network_fields():
    address = IRAddress(
        name="none",
        type=AddressType.SPECIAL,
        original_type="fortigate_reserved",
        original_value="none",
        requires_manual_review=True,
    )

    assert address.value == "none"


def test_fortigate_address_excel_exposes_source_metadata():
    ir = FGToIRTransformer(
        parse_fortigate_config(ADDRESS_CONFIG)
    ).transform()
    workbook = load_workbook(
        io.BytesIO(IRExcelExporter(ir).generate())
    )
    sheet = workbook["Addresses"]
    headers = {
        cell.value: cell.column
        for cell in sheet[3]
    }

    assert list(headers) == [
        "Name",
        "Source UUID",
            "Type",
            "Value",
            "MAC Entries",
            "MAC Count",
            "Source Section",
        "Address Family",
        "Source Type",
        "Original Type",
        "Original Value",
        "IPv6",
        "Multicast",
        "Associated Interface",
        "Allow Routing",
        "Source Color",
        "FSSO Group",
        "Hardware Vendor",
        "Hardware Model",
        "Source Interface",
        "Resolved Interface Subnet",
        "Interface Reference Resolved",
        "Cache TTL",
        "ClearPass SPT",
        "EPG Name",
        "Fabric Object",
        "Dynamic Filter",
        "SDN",
        "SDN Connector",
        "SDN Address Type",
        "SDN Tag",
        "Organization",
        "OS",
        "Policy Group",
        "Route Tag",
        "Subnet Name",
        "Software Version",
        "Tag Detection Level",
        "Tenant",
        "Node IP Only",
        "NSX Object ID",
        "EMS Sub-Type",
        "EMS Object Tag",
        "EMS Tag Type",
        "EMS Object Type",
        "EMS Dirty",
        "Tags",
        "IP List",
        "Object Tagging",
        "Migration Status",
        "Manual Review",
        "Audit Note",
        "Parse Error",
        "Additional Settings",
        "Description",
    ]

    row_by_name = {
        sheet.cell(row, headers["Name"]).value: row
        for row in range(4, sheet.max_row + 1)
    }
    normal_row = row_by_name["normal-net"]
    assert sheet.cell(normal_row, headers["Source UUID"]).value == (
        "11111111-1111-1111-1111-111111111111"
    )
    assert sheet.cell(normal_row, headers["Associated Interface"]).value == "port6"
    assert sheet.cell(normal_row, headers["Allow Routing"]).value == "TRUE"
    assert sheet.cell(normal_row, headers["Source Color"]).value == 9
    assert sheet.cell(normal_row, headers["Cache TTL"]).value == 300
    assert sheet.cell(normal_row, headers["ClearPass SPT"]).value is None
    assert sheet.cell(normal_row, headers["EPG Name"]).value is None
    assert sheet.cell(normal_row, headers["Fabric Object"]).value is None
    assert sheet.cell(normal_row, headers["Dynamic Filter"]).value is None
    assert sheet.cell(normal_row, headers["SDN Connector"]).value is None
    assert sheet.cell(normal_row, headers["Node IP Only"]).value is None
    assert sheet.cell(normal_row, headers["NSX Object ID"]).value is None
    assert sheet.cell(normal_row, headers["Source Section"]).value == "firewall address"
    assert sheet.cell(normal_row, headers["Address Family"]).value == "ipv4"
    assert sheet.cell(normal_row, headers["Source Type"]).value == "ipmask"
    assert sheet.cell(normal_row, headers["Additional Settings"]).value == (
        "password=[REDACTED]"
        "; subnet=192.168.10.0 255.255.255.0"
    )

    ipv6_row = row_by_name["ipv6-test"]
    assert sheet.cell(ipv6_row, headers["Value"]).value == "fdff:ffff::/120"
    assert sheet.cell(ipv6_row, headers["IPv6"]).value == "Yes"
    assert sheet.cell(ipv6_row, headers["Fabric Object"]).value == "enable"

    sslvpn_row = row_by_name["SSLVPN_TUNNEL_IPv6_ADDR1"]
    assert sheet.cell(sslvpn_row, headers["Value"]).value == "fdff:ffff::/120"
    assert sheet.cell(sslvpn_row, headers["Source UUID"]).value == (
        "17523864-65a4-51e9-c45e-65c6367ea4e3"
    )
    google_play_row = row_by_name["google-play"]
    assert sheet.cell(google_play_row, headers["Value"]).value == "*play.google.com"
    assert sheet.cell(google_play_row, headers["Source Section"]).value == (
        "firewall wildcard-fqdn custom"
    )

    mac_row = row_by_name["mac-source"]
    assert sheet.cell(mac_row, headers["Type"]).value == "mac"
    assert sheet.cell(mac_row, headers["Value"]).value == "00:11:22:33:44:55"

    groups_sheet = workbook["Address Groups"]
    group_headers = {
        cell.value: cell.column
        for cell in groups_sheet[3]
    }
    group_row_by_name = {
        groups_sheet.cell(row, group_headers["Name"]).value: row
        for row in range(4, groups_sheet.max_row + 1)
    }
    ems_row = group_row_by_name["ems"]
    assert groups_sheet.cell(
        ems_row,
        group_headers["EMS Sub-Type"],
    ).value == "ems-tag"
    assert groups_sheet.cell(
        ems_row,
        group_headers["EMS Object Tag"],
    ).value == "Deleum_ADUser"
    assert groups_sheet.cell(
        ems_row,
        group_headers["EMS Tag Type"],
    ).value == "zero_trust"
    assert groups_sheet.cell(
        ems_row,
        group_headers["EMS Object Type"],
    ).value == "mac"
    assert groups_sheet.cell(
        ems_row,
        group_headers["EMS Dirty"],
    ).value == "clean"


def test_fortigate_invalid_and_missing_mac_values_are_not_replaced():
    ir = FGToIRTransformer(
        parse_fortigate_config(
            """
config firewall address
    edit "invalid-mac"
        set type mac
        set macaddr "not-a-mac"
    next
    edit "missing-mac"
        set type mac
    next
end
"""
        )
    ).transform()
    addresses = _by_name(ir.addresses)

    invalid = addresses["invalid-mac"]
    assert invalid.type == AddressType.MAC
    assert invalid.value == "not-a-mac"
    assert invalid.requires_manual_review is True
    assert invalid.parse_error is not None

    missing = addresses["missing-mac"]
    assert missing.type == AddressType.MAC
    assert missing.value == ""
    assert missing.requires_manual_review is True
    assert missing.parse_error is not None

    assert all(
        not address.value.startswith(("198.18.", "198.19."))
        for address in (invalid, missing)
    )


def test_fortigate_macaddr_supports_ordered_singles_ranges_and_partial_invalid():
    ir = FGToIRTransformer(parse_fortigate_config('''
config firewall address
    edit "single"
        set type mac
        set macaddr "00:11:22:33:44:aa"
    next
    edit "mixed"
        set type mac
        set macaddr "00:11:22:33:44:01 00:11:22:33:44:10-00:11:22:33:44:20 00:11:22:33:44:30"
    next
    edit "reversed"
        set type mac
        set macaddr "00:11:22:33:44:ff-00:11:22:33:44:00"
    next
    edit "partial"
        set type mac
        set macaddr "00:11:22:33:44:55 BAD_MAC 00:11:22:33:44:66"
    next
end
''')).transform()
    addresses = _by_name(ir.addresses)
    assert addresses["single"].mac == "00:11:22:33:44:AA"
    assert addresses["mixed"].value == "00:11:22:33:44:01; 00:11:22:33:44:10-00:11:22:33:44:20; 00:11:22:33:44:30"
    assert addresses["mixed"].source_attributes["macaddr"].endswith("00:11:22:33:44:30")
    assert addresses["reversed"].requires_manual_review is True
    assert addresses["partial"].migration_status == "PARTIALLY_NORMALIZED"
    assert [entry.start for entry in addresses["partial"].mac_entries] == [
        "00:11:22:33:44:55", "00:11:22:33:44:66"
    ]
    assert addresses["partial"].requires_manual_review is True

    sheet = load_workbook(io.BytesIO(IRExcelExporter(ir).generate()))["Addresses"]
    headers = {cell.value: cell.column for cell in sheet[3]}
    row = {sheet.cell(i, headers["Name"]).value: i for i in range(4, sheet.max_row + 1)}
    assert sheet.cell(row["mixed"], headers["MAC Count"]).value == 3
    assert sheet.cell(row["mixed"], headers["MAC Entries"]).value.endswith("00:11:22:33:44:30")


ADVANCED_ADDRESS_CONFIG = """
config system interface
    edit "port1"
        set role lan
        set ip 192.168.100.1 255.255.255.0
    next
end
config firewall address
    edit "empty-vpn-helper"
        set allow-routing enable
    next
    edit "wildcard-address"
        set type wildcard
        set wildcard 192.168.0.0 0.0.255.255
    next
    edit "interface-subnet-address"
        set type interface-subnet
        set interface "port1"
    next
    edit "route-tag-address"
        set type route-tag
        set route-tag 12345
    next
    edit "sdn-dynamic"
        set type dynamic
        set sub-type sdn
        set sdn "aws"
        set filter "Tag.Name=production"
    next
    edit "ems-dynamic"
        set type dynamic
        set sub-type ems-tag
        set obj-tag "compliant"
    next
    edit "nested-address"
        set subnet 10.10.10.0 255.255.255.0
        config list
            edit "10.10.10.10"
            next
            edit "10.10.10.11"
            next
        end
        config tagging
            edit "owner"
                set category "department"
                set tags "ICT" "Security"
            next
        end
    next
end
config firewall wildcard-fqdn custom
    edit "wildcard-custom"
        set wildcard-fqdn "*.example.com"
    next
end
"""


def test_advanced_address_types_are_never_silently_lost():
    ir = FGToIRTransformer(parse_fortigate_config(ADVANCED_ADDRESS_CONFIG)).transform()
    addresses = _by_name(ir.addresses)
    groups = _by_name(ir.address_groups)
    empty = addresses["empty-vpn-helper"]
    assert empty.value == ""
    assert empty.value != "192.168.100.0/24"
    assert empty.source_section == "firewall address"
    assert empty.address_family == "ipv4"
    assert empty.source_type == "ipmask"
    assert empty.requires_manual_review
    assert addresses["wildcard-address"].type == AddressType.WILDCARD_MASK
    assert addresses["wildcard-address"].value == "192.168.0.0 0.0.255.255"
    assert addresses["interface-subnet-address"].value == "192.168.100.0/24"
    assert addresses["route-tag-address"].value == "12345"
    assert addresses["sdn-dynamic"].source_sub_type == "sdn"
    assert "sdn-dynamic" not in groups
    assert groups["ems-dynamic"].source_section == "firewall address"
    assert groups["ems-dynamic"].address_family == "ipv4"
    nested = addresses["nested-address"]
    assert nested.source_list_entries == ["10.10.10.10", "10.10.10.11"]
    assert nested.source_tagging_entries[0].tags == ["ICT", "Security"]
    assert addresses["wildcard-custom"].source_section == "firewall wildcard-fqdn custom"


def test_fortigate_address_cache_ttl_boundaries_and_source_only_preservation():
    config = parse_fortigate_config(
        """
config firewall address
    edit "unspecified"
        set type fqdn
        set fqdn "none.example"
    next
    edit "zero"
        set type fqdn
        set fqdn "zero.example"
        set cache-ttl 0
    next
    edit "maximum"
        set type fqdn
        set fqdn "max.example"
        set cache-ttl 86400
    next
    edit "invalid"
        set type fqdn
        set fqdn "invalid.example"
        set cache-ttl 86401
    next
    edit "non-fqdn"
        set subnet 10.0.0.0 255.255.255.0
        set cache-ttl 100
    next
    edit "source-only"
        set type dynamic
        set sub-type fabric
        set cache-ttl 200
    next
end
"""
    )
    parsed = _by_name(config.addresses)
    assert parsed["zero"].cache_ttl == 0
    assert parsed["maximum"].cache_ttl == 86400
    assert parsed["invalid"].cache_ttl == 86401

    addresses = _by_name(FGToIRTransformer(config).transform().addresses)
    assert addresses["unspecified"].source_cache_ttl is None
    assert addresses["zero"].source_cache_ttl == 0
    assert addresses["zero"].source_attributes["cache_ttl"] == 0
    assert addresses["maximum"].requires_manual_review is False
    assert addresses["invalid"].source_cache_ttl == 86401
    assert addresses["invalid"].source_attributes["cache_ttl"] == 86401
    assert addresses["invalid"].requires_manual_review is True
    assert "cache-ttl" in addresses["invalid"].audit_note
    assert addresses["non-fqdn"].source_cache_ttl == 100
    assert addresses["source-only"].source_cache_ttl == 200
    assert addresses["source-only"].source_attributes["cache_ttl"] == 200


def test_fortigate_address_cache_ttl_excel_values_preserve_none_and_zero():
    ir = FGToIRTransformer(
        parse_fortigate_config(
            """
config firewall address
    edit "blank"
        set type fqdn
        set fqdn "blank.example"
    next
    edit "zero"
        set type fqdn
        set fqdn "zero.example"
        set cache-ttl 0
    next
end
"""
        )
    ).transform()
    sheet = load_workbook(io.BytesIO(IRExcelExporter(ir).generate()))["Addresses"]
    headers = {cell.value: cell.column for cell in sheet[3]}
    rows = {sheet.cell(row, headers["Name"]).value: row for row in range(4, sheet.max_row + 1)}
    assert sheet.cell(rows["blank"], headers["Cache TTL"]).value is None
    assert sheet.cell(rows["zero"], headers["Cache TTL"]).value == 0


def test_fortigate_remaining_address_fields_are_typed_and_preserved():
    config = parse_fortigate_config("""
config firewall address
    edit "dynamic-metadata"
        set type dynamic
        set sw-version "10.2.3"
        set tag-detection-level "deep-match"
        set tenant "tenant-a"
    next
    edit "wildcard-address"
        set type fqdn
        set wildcard-fqdn "api*.example.com"
        set subnet-name "inside-subnet"
    next
    edit "overlong"
        set subnet-name "aaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa"
    next
end
""")
    parsed = _by_name(config.addresses)
    assert parsed["dynamic-metadata"].sw_version == "10.2.3"
    assert parsed["wildcard-address"].wildcard_fqdn == "api*.example.com"

    addresses = _by_name(FGToIRTransformer(config).transform().addresses)
    dynamic = addresses["dynamic-metadata"]
    assert dynamic.type == AddressType.DYNAMIC
    assert dynamic.source_sw_version == "10.2.3"
    assert dynamic.source_tag_detection_level == "deep-match"
    assert dynamic.source_tenant == "tenant-a"
    wildcard = addresses["wildcard-address"]
    assert wildcard.type == AddressType.WILDCARD_FQDN
    assert wildcard.fqdn == "api*.example.com"
    assert wildcard.source_subnet_name == "inside-subnet"
    assert addresses["overlong"].requires_manual_review

    sheet = load_workbook(io.BytesIO(IRExcelExporter(FGToIRTransformer(config).transform()).generate()))["Addresses"]
    headers = {cell.value: cell.column for cell in sheet[3]}
    rows = {sheet.cell(row, headers["Name"]).value: row for row in range(4, sheet.max_row + 1)}
    assert sheet.cell(rows["wildcard-address"], headers["Type"]).value == "wildcard"
    assert sheet.cell(rows["wildcard-address"], headers["Value"]).value == "api*.example.com"
    assert sheet.cell(rows["dynamic-metadata"], headers["Software Version"]).value == "10.2.3"


def test_fortigate_clearpass_spt_and_epg_name_are_typed_and_validated():
    values = ["unknown", "healthy", "quarantine", "checkup", "transient", "infected"]
    for value in values:
        ir = FGToIRTransformer(parse_fortigate_config(f"""
config firewall address
    edit "cp"
        set type dynamic
        set sub-type clearpass-spt
        set clearpass-spt {value}
    next
end
""")).transform()
        address = ir.addresses[0]
        assert address.source_clearpass_spt == value
        assert address.source_attributes["clearpass_spt"] == value
        assert address.requires_manual_review is True

    config = parse_fortigate_config("""
config firewall address
    edit "future"
        set type dynamic
        set sub-type ems-tag
        set clearpass-spt future-state
    next
    edit "epg4"
        set subnet 10.1.0.0 255.255.0.0
        set epg-name "Production"
    next
end
config firewall address6
    edit "epg6"
        set ip6 2001:db8::/64
        set epg-name "IPv6-Production"
    next
end
""")
    ir = FGToIRTransformer(config).transform()
    addresses = _by_name(ir.addresses)
    assert addresses["future"].source_clearpass_spt == "future-state"
    assert addresses["future"].requires_manual_review is True
    assert addresses["epg4"].source_epg_name == "Production"
    assert addresses["epg6"].source_epg_name == "IPv6-Production"

    for length, reviewed in ((255, False), (256, True)):
        address = _by_name(FGToIRTransformer(parse_fortigate_config(f"""
    config firewall address
    edit "epg"
        set subnet 10.0.0.0 255.255.255.0
        set epg-name "{"x" * length}"
    next
end
""")).transform().addresses)["epg"]
        assert len(address.source_epg_name) == length
        assert address.requires_manual_review is reviewed


def test_fortigate_address_fabric_object_is_typed_preserved_and_validated():
    config = parse_fortigate_config("""
config firewall address
    edit "enabled"
        set subnet 10.0.0.0 255.255.255.0
        set fabric-object enable
    next
    edit "disabled"
        set subnet 10.1.0.0 255.255.255.0
        set fabric-object disable
    next
    edit "unspecified"
        set subnet 10.2.0.0 255.255.255.0
    next
    edit "future"
        set subnet 10.3.0.0 255.255.255.0
        set fabric-object something-new
    next
    edit "source-only"
        set type dynamic
        set sub-type fabric
        set fabric-object enable
    next
end
""")
    parsed = _by_name(config.addresses)
    assert parsed["enabled"].fabric_object == "enable"
    assert parsed["disabled"].fabric_object == "disable"
    assert parsed["unspecified"].fabric_object is None
    assert parsed["future"].fabric_object == "something-new"

    addresses = _by_name(FGToIRTransformer(config).transform().addresses)
    assert addresses["enabled"].source_fabric_object_setting == "enable"
    assert addresses["enabled"].source_attributes["fabric_object"] == "enable"
    assert addresses["disabled"].source_fabric_object_setting == "disable"
    assert addresses["unspecified"].source_fabric_object_setting is None
    assert addresses["future"].source_fabric_object_setting == "something-new"
    assert addresses["future"].requires_manual_review is True
    assert addresses["source-only"].source_fabric_object_setting == "enable"
    assert addresses["source-only"].source_attributes["fabric_object"] == "enable"


def test_fortigate_dynamic_filter_is_typed_and_keeps_exact_source_expression():
    expression = "Vpc = prod-vpc && env in (prod, staging)"
    config = parse_fortigate_config(f'''
config firewall address
    edit "filtered"
        set type dynamic
        set sub-type sdn
        set sdn "ibm"
        set filter "{expression}"
    next
    edit "no-filter"
        set type dynamic
        set sdn "aws"
    next
    edit "wrong-context"
        set subnet 10.0.0.0 255.255.255.0
        set filter "Vpc=prod"
    next
end
''')
    parsed = _by_name(config.addresses)
    assert parsed["filtered"].filter == expression
    assert parsed["filtered"].sdn == "ibm"

    ir = FGToIRTransformer(config).transform()
    addresses = _by_name(ir.addresses)
    filtered = addresses["filtered"]
    assert filtered.type == AddressType.DYNAMIC
    assert filtered.dynamic_filter == expression
    assert filtered.source_sdn == "ibm"
    assert filtered.source_attributes["filter"] == expression
    assert filtered.source_attributes["sdn"] == "ibm"
    assert filtered.requires_manual_review is True
    assert addresses["no-filter"].requires_manual_review is True
    assert addresses["wrong-context"].source_attributes["filter"] == "Vpc=prod"
    assert addresses["wrong-context"].requires_manual_review is True

    workbook = load_workbook(io.BytesIO(IRExcelExporter(ir).generate()))
    sheet = workbook["Addresses"]
    headers = {cell.value: cell.column for cell in sheet[3]}
    row = next(
        row for row in range(4, sheet.max_row + 1)
        if sheet.cell(row, headers["Name"]).value == "filtered"
    )
    assert sheet.cell(row, headers["Dynamic Filter"]).value == expression
    assert sheet.cell(row, headers["SDN Connector"]).value == "ibm"


def test_fortigate_dynamic_filter_length_boundary_is_preserved_for_review():
    for length, reviewed in ((2047, False), (2048, True)):
        address = _by_name(FGToIRTransformer(parse_fortigate_config(f'''
config firewall address
    edit "filtered"
        set type dynamic
        set filter "{"x" * length}"
    next
end
''')).transform().addresses)["filtered"]
        assert address.dynamic_filter == "x" * length
        assert address.requires_manual_review is reviewed


def test_fortigate_node_ip_only_and_nsx_obj_id_are_typed_and_preserved():
    config = parse_fortigate_config('''
config firewall address
    edit "k8s"
        set type dynamic
        set node-ip-only enable
        set obj-id "nsx-12345"
    next
    edit "disabled"
        set type dynamic
        set node-ip-only disable
    next
    edit "wrong-context"
        set type ipmask
        set subnet 10.0.0.0 255.255.255.0
        set node-ip-only enable
    next
    edit "unknown"
        set type dynamic
        set node-ip-only something-new
    next
    edit "long-id"
        set type dynamic
        set obj-id "''' + "x" * 256 + '''"
    next
end
''')
    parsed = _by_name(config.addresses)
    assert parsed["k8s"].node_ip_only == "enable"
    assert parsed["k8s"].obj_id == "nsx-12345"

    addresses = _by_name(FGToIRTransformer(config).transform().addresses)
    assert addresses["k8s"].source_node_ip_only is True
    assert addresses["k8s"].source_obj_id == "nsx-12345"
    assert addresses["k8s"].source_attributes["node_ip_only"] == "enable"
    assert addresses["k8s"].source_attributes["obj_id"] == "nsx-12345"
    assert addresses["disabled"].source_node_ip_only is False
    assert addresses["disabled"].requires_manual_review is True
    assert addresses["wrong-context"].source_node_ip_only is True
    assert addresses["wrong-context"].requires_manual_review is True
    assert addresses["unknown"].source_node_ip_only is None
    assert addresses["unknown"].source_attributes["node_ip_only"] == "something-new"
    assert addresses["unknown"].requires_manual_review is True
    assert len(addresses["long-id"].source_obj_id) == 256
    assert addresses["long-id"].requires_manual_review is True

    sheet = load_workbook(io.BytesIO(IRExcelExporter(FGToIRTransformer(config).transform()).generate()))["Addresses"]
    headers = {cell.value: cell.column for cell in sheet[3]}
    rows = {sheet.cell(i, headers["Name"]).value: i for i in range(4, sheet.max_row + 1)}
    assert sheet.cell(rows["k8s"], headers["Node IP Only"]).value == "TRUE"
    assert sheet.cell(rows["k8s"], headers["NSX Object ID"]).value == "nsx-12345"


def test_empty_vpn_helper_addresses_are_not_inferred():
    config = """
    config system interface
        edit "port1"
            set role lan
            set ip 192.168.100.1 255.255.255.0
        next
    end
    config firewall address
        edit "to_TEST_local_subnet_1"
            set allow-routing enable
        next
        edit "to_TEST_remote_subnet_1"
            set allow-routing enable
        next
    end
    config router static
        edit 1
            set dst 10.20.30.0 255.255.255.0
            set device "to_TEST"
        next
    end
    """
    addresses = _by_name(
        FGToIRTransformer(parse_fortigate_config(config)).transform().addresses
    )
    for name in ("to_TEST_local_subnet_1", "to_TEST_remote_subnet_1"):
        item = addresses[name]
        assert item.value == ""
        assert item.requires_manual_review
        assert item.migration_status == "PARTIALLY_NORMALIZED"


def test_address_coverage_counts_exact_source_provenance():
    config = """
    config firewall address
        edit "net1"
            set subnet 10.0.0.0 255.255.255.0
        next
        edit "ems1"
            set type dynamic
            set sub-type ems-tag
            set obj-tag "tag1"
        next
    end
    config firewall wildcard-fqdn custom
        edit "wild1"
            set wildcard-fqdn "*.example.com"
        next
    end
    config firewall addrgrp
        edit "grp1"
            set member "net1"
        next
    end
    """
    coverage = {
        item.path: item for item in extract_fortigate_config(config).source_sections
    }
    for path, expected in (
        ("firewall address", 2),
        ("firewall wildcard-fqdn custom", 1),
        ("firewall addrgrp", 1),
    ):
        section = coverage[path]
        assert section.object_count_source == expected
        assert section.object_count_parsed == expected
        assert section.object_count_normalized == expected


def test_fortigate_address_source_matching_fields_are_typed_validated_and_exported():
    config = parse_fortigate_config('''
config firewall address
    edit "dynamic"
        set type dynamic
        set sub-type sdn
        set filter "env=prod"
        set organization "acme/finance"
        set os "Windows"
        set policy-group "group1"
        set sdn "aws-prod"
        set sdn-addr-type private
        set sdn-tag "production"
    next
    edit "route"
        set type route-tag
        set route-tag 4294967295
    next
    edit "source-only"
        set type ipmask
        set subnet 10.0.0.0 255.255.255.0
        set sdn "aws"
        set sdn-addr-type unknown
    next
end
''')
    parsed = _by_name(config.addresses)
    assert parsed["dynamic"].policy_group == "group1"
    assert parsed["route"].route_tag == 4294967295

    ir = FGToIRTransformer(config).transform()
    addresses = _by_name(ir.addresses)
    dynamic = addresses["dynamic"]
    assert {
        dynamic.source_organization,
        dynamic.source_os,
        dynamic.source_policy_group,
        dynamic.source_sdn,
        dynamic.source_sdn_addr_type,
        dynamic.source_sdn_tag,
    } == {"acme/finance", "Windows", "group1", "aws-prod", "private", "production"}
    assert dynamic.source_attributes["sdn_addr_type"] == "private"
    assert dynamic.requires_manual_review is True
    assert addresses["route"].source_type == "route-tag"
    assert addresses["route"].source_route_tag == 4294967295
    assert addresses["source-only"].source_sdn == "aws"
    assert addresses["source-only"].source_sdn_addr_type == "unknown"
    assert addresses["source-only"].requires_manual_review is True

    workbook = load_workbook(io.BytesIO(IRExcelExporter(ir).generate()))
    sheet = workbook["Addresses"]
    headers = {cell.value: cell.column for cell in sheet[3]}
    row = next(i for i in range(4, sheet.max_row + 1) if sheet.cell(i, headers["Name"]).value == "dynamic")
    assert sheet.cell(row, headers["Organization"]).value == "acme/finance"
    assert sheet.cell(row, headers["OS"]).value == "Windows"
    assert sheet.cell(row, headers["Policy Group"]).value == "group1"
    assert sheet.cell(row, headers["SDN"]).value == "aws-prod"
    assert sheet.cell(row, headers["SDN Address Type"]).value == "private"
    assert sheet.cell(row, headers["SDN Tag"]).value == "production"


def test_fortigate_address_source_matching_limits_and_context_require_review():
    ir = FGToIRTransformer(parse_fortigate_config('''
config firewall address
    edit "too-long"
        set type dynamic
        set filter "x"
        set organization "xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx"
        set os "xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx"
        set policy-group "xxxxxxxxxxxxxxxx"
        set sdn "xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx"
        set sdn-tag "xxxxxxxxxxxxxxxx"
    next
    edit "wrong"
        set type ipmask
        set subnet 10.0.0.0 255.255.255.0
        set organization "acme"
        set route-tag 0
    next
end
''')).transform()
    addresses = _by_name(ir.addresses)
    assert addresses["too-long"].source_organization == "x" * 36
    assert addresses["too-long"].source_sdn_tag == "x" * 16
    assert addresses["too-long"].requires_manual_review is True
    assert addresses["wrong"].source_organization == "acme"
    assert addresses["wrong"].source_route_tag == 0
    assert addresses["wrong"].requires_manual_review is True


def test_fortigate_dynamic_address_source_criteria_are_typed_and_preserved():
    config = parse_fortigate_config('''
config firewall address
    edit "fsso"
        set type dynamic
        set sub-type fsso
        set fsso-group "Finance Users"
    next
    edit "hardware"
        set type dynamic
        set sdn "aws"
        set filter "tag.Name=prod"
        set hw-vendor "Fortinet"
        set hw-model "FortiPhone"
    next
    edit "wrong-context"
        set type ipmask
        set subnet 10.0.0.0 255.255.255.0
        set hw-vendor "Fortinet"
    next
    edit "too-long"
        set type dynamic
        set fsso-group "{fsso}"
        set hw-model "{model}"
        set hw-vendor "{vendor}"
    next
end
'''.format(fsso="x" * 512, model="x" * 36, vendor="x" * 36))
    parsed = _by_name(config.addresses)
    assert parsed["fsso"].fsso_group == "Finance Users"

    addresses = _by_name(FGToIRTransformer(config).transform().addresses)
    fsso = addresses["fsso"]
    assert fsso.source_fsso_group == "Finance Users"
    assert fsso.source_sub_type == "fsso"
    assert fsso.source_attributes["fsso_group"] == "Finance Users"

    hardware = addresses["hardware"]
    assert hardware.type == AddressType.DYNAMIC
    assert hardware.dynamic_filter == "tag.Name=prod"
    assert hardware.source_sdn == "aws"
    assert hardware.source_hw_vendor == "Fortinet"
    assert hardware.source_hw_model == "FortiPhone"

    assert addresses["wrong-context"].source_hw_vendor == "Fortinet"
    assert addresses["wrong-context"].requires_manual_review is True
    assert addresses["too-long"].requires_manual_review is True

    sheet = load_workbook(io.BytesIO(IRExcelExporter(FGToIRTransformer(config).transform()).generate()))["Addresses"]
    headers = {cell.value: cell.column for cell in sheet[3]}
    row = {sheet.cell(row, headers["Name"]).value: row for row in range(4, sheet.max_row + 1)}
    assert sheet.cell(row["hardware"], headers["FSSO Group"]).value is None
    assert sheet.cell(row["hardware"], headers["Hardware Vendor"]).value == "Fortinet"
    assert sheet.cell(row["hardware"], headers["Hardware Model"]).value == "FortiPhone"


def test_fortigate_interface_subnet_resolves_in_context_and_preserves_relationship():
    ir = FGToIRTransformer(parse_fortigate_config("""
config system interface
    edit "port1"
        set ip 172.16.200.1 255.255.255.0
    next
    edit "dhcp1"
        set mode dhcp
    next
end
config firewall address
    edit "match"
        set type interface-subnet
        set interface "port1"
        set subnet 172.16.200.0 255.255.255.0
    next
    edit "mismatch"
        set type interface-subnet
        set interface "port1"
        set subnet 172.16.100.0 255.255.255.0
    next
    edit "missing"
        set type interface-subnet
        set interface "missing-port"
    next
    edit "dhcp"
        set type interface-subnet
        set interface "dhcp1"
    next
end
config vdom
    edit "VDOM-A"
        config system interface
            edit "port1"
                set ip 10.1.0.1 255.255.255.0
            next
        end
        config firewall address
            edit "a"
                set type interface-subnet
                set interface "port1"
            next
        end
    next
    edit "VDOM-B"
        config system interface
            edit "port1"
                set ip 10.2.0.1 255.255.255.0
            next
        end
        config firewall address
            edit "b"
                set type interface-subnet
                set interface "port1"
            next
        end
    next
end
""")).transform()
    addresses = _by_name(ir.addresses)

    assert addresses["match"].subnet == "172.16.200.0/24"
    assert addresses["match"].source_type == "interface-subnet"
    assert addresses["match"].source_interface == "port1"
    assert addresses["match"].resolved_interface_subnet == "172.16.200.0/24"
    assert addresses["match"].interface_reference_resolved is True
    assert addresses["match"].requires_manual_review is True
    assert addresses["match"].source_attributes["interface"] == "port1"
    assert addresses["match"].source_attributes["subnet"] == "172.16.200.0 255.255.255.0"

    assert addresses["mismatch"].resolved_interface_subnet == "172.16.200.0/24"
    assert addresses["mismatch"].requires_manual_review is True
    assert addresses["missing"].interface_reference_resolved is False
    assert addresses["missing"].resolved_interface_subnet is None
    assert addresses["missing"].migration_status == "SOURCE_ONLY"
    assert addresses["dhcp"].interface_reference_resolved is True
    assert addresses["dhcp"].resolved_interface_subnet is None
    assert addresses["dhcp"].migration_status == "SOURCE_ONLY"
    assert addresses["a"].resolved_interface_subnet == "10.1.0.0/24"
    assert addresses["b"].resolved_interface_subnet == "10.2.0.0/24"

    workbook = load_workbook(io.BytesIO(IRExcelExporter(ir).generate()))
    sheet = workbook["Addresses"]
    headers = {cell.value: cell.column for cell in sheet[3]}
    row = {sheet.cell(index, headers["Name"]).value: index for index in range(4, sheet.max_row + 1)}
    assert sheet.cell(row["match"], headers["Source Interface"]).value == "port1"
    assert sheet.cell(row["match"], headers["Resolved Interface Subnet"]).value == "172.16.200.0/24"
    assert sheet.cell(row["match"], headers["Interface Reference Resolved"]).value == "TRUE"
