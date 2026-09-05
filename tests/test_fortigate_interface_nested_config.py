import io

from openpyxl import load_workbook

from fwmigrate.extraction.models import ExtractionStatus
from fwmigrate.parsers.fortigate.extractor import extract_fortigate_config
from fwmigrate.parsers.fortigate.parser import parse_fortigate_config
from fwmigrate.parsers.fortigate.transformer import FGToIRTransformer
from fwmigrate.report.excel_exporter import IRExcelExporter


NESTED_INTERFACE_CONFIG = """
config system interface
    edit "port1"
        set vdom "root"
        set ip 10.0.0.1 255.255.255.0
        set type physical

        config client-options
            edit 1
                set code 60
                set type string
                set value "fortigate"
            next
        end

        config ipv6
            set ip6-address 2001:db8:1::1/64
            set ip6-allowaccess ping https

            config ip6-extra-addr
                edit 2001:db8:1::2/64
                next
            end

            config ip6-prefix-list
                edit 2001:db8:1::/64
                    set autonomous-flag enable
                    set onlink-flag enable
                    set valid-life-time 86400
                    set preferred-life-time 14400
                next
            end
        end

        config vrrp
            edit 1
                set version 3
                set vrip 10.0.0.254
                set priority 150

                config proxy-arp
                    edit 1
                        set ip 10.0.0.10
                    next
                end
            next
        end

        config tagging
            edit "site"
                set category "location"
                set tags "HQ" "WAN"
            next
        end

        config l2tp-client-settings
            set user "testuser"
            set password "secret-value"
            set peer-host "vpn.example.com"
        end
    next

    edit "port2"
        set type physical

        config ipv6
            set ip6-address 2001:db8:2::1/64
        end
    next
end
"""


def _interface(fg, name):
    return next(
        item
        for item in fg.interfaces
        if item.name == name
    )


def test_nested_config_stays_with_parent_interface():
    fg = parse_fortigate_config(NESTED_INTERFACE_CONFIG)

    port1 = _interface(fg, "port1")
    port2 = _interface(fg, "port2")

    # Use a list, not a set: source ordering is part of the preservation contract.
    assert [
        node.name
        for node in port1.nested_configs
    ] == [
        "client-options",
        "ipv6",
        "vrrp",
        "tagging",
        "l2tp-client-settings",
    ]

    assert [
        node.name
        for node in port2.nested_configs
    ] == [
        "ipv6",
    ]


def test_nested_ipv6_structure_is_recursive():
    fg = parse_fortigate_config(NESTED_INTERFACE_CONFIG)
    port1 = _interface(fg, "port1")

    ipv6 = next(
        node
        for node in port1.nested_configs
        if node.name == "ipv6"
    )

    ip6_address = next(
        command
        for command in ipv6.commands
        if command.key == "ip6-address"
    )

    assert ip6_address.values == [
        "2001:db8:1::1/64"
    ]

    prefix_config = next(
        child
        for child in ipv6.children
        if child.name == "ip6-prefix-list"
    )

    prefix = prefix_config.children[0]

    assert prefix.node_type == "edit"
    assert prefix.name == "2001:db8:1::/64"

    autonomous = next(
        command
        for command in prefix.commands
        if command.key == "autonomous-flag"
    )

    assert autonomous.values == ["enable"]


def test_nested_vrrp_proxy_arp_is_preserved():
    fg = parse_fortigate_config(NESTED_INTERFACE_CONFIG)
    port1 = _interface(fg, "port1")

    vrrp = next(
        node
        for node in port1.nested_configs
        if node.name == "vrrp"
    )

    vrrp_entry = vrrp.children[0]
    assert vrrp_entry.name == "1"

    proxy_arp = next(
        child
        for child in vrrp_entry.children
        if child.name == "proxy-arp"
    )

    proxy_entry = proxy_arp.children[0]

    ip_command = next(
        command
        for command in proxy_entry.commands
        if command.key == "ip"
    )

    assert ip_command.values == ["10.0.0.10"]


def test_nested_interface_secret_is_redacted():
    fg = parse_fortigate_config(NESTED_INTERFACE_CONFIG)
    port1 = _interface(fg, "port1")

    l2tp = next(
        node
        for node in port1.nested_configs
        if node.name == "l2tp-client-settings"
    )

    password = next(
        command
        for command in l2tp.commands
        if command.key == "password"
    )

    assert "secret-value" not in " ".join(password.values)
    assert password.values == ["[REDACTED]"]


def test_nested_configs_are_not_in_interface_source_attributes():
    fg = parse_fortigate_config(NESTED_INTERFACE_CONFIG)
    port1 = _interface(fg, "port1")

    assert "nested_configs" not in port1.source_attributes
    assert port1.source_attributes["vdom"] == "root"
    assert port1.source_attributes["type"] == "physical"


def test_nested_interface_config_reaches_ir():
    fg = parse_fortigate_config(NESTED_INTERFACE_CONFIG)
    ir = FGToIRTransformer(fg).transform()

    port1 = next(
        item
        for item in ir.interfaces
        if item.name == "port1"
    )

    assert len(port1.nested_source_configs) == 5
    assert port1.requires_manual_review is True

    assert [
        node.name
        for node in port1.nested_source_configs
    ] == [
        "client-options",
        "ipv6",
        "vrrp",
        "tagging",
        "l2tp-client-settings",
    ]


def test_phase1_interface_nested_settings_are_typed_and_preserved():
    config = """
config system interface
    edit "port1"
        set detectprotocol ping tcp
        set bfd enable
        set bfd-desired-min-tx 100
        set snmp-index 7

        config client-options
            edit 1
                set code 60
                set type string
                set value "fortigate client"
                set future-option preserved
            next
        end

        config dhcp-snooping-server-list
            edit "dhcp-primary"
                set server-ip 192.0.2.53
            next
        end

        config tagging
            edit "site"
                set category "location"
                set tags "HQ" "WAN"
            next
        end

        config vrrp
            edit 12
                set version 3
                set vrip 192.0.2.254
                set priority 200
                set vrdst-priority 10
                config proxy-arp
                    edit 1
                        set ip 192.0.2.100
                    next
                end
            next
        end

        config egress-queues
            set cos0 "voice"
            set cos7 "bulk"
        end
    next
end
"""

    interface = parse_fortigate_config(config).interfaces[0]

    assert interface.detectprotocol == ["ping", "tcp"]
    assert interface.bfd == "enable"
    assert interface.bfd_desired_min_tx == 100
    assert interface.snmp_index == 7
    assert interface.source_attributes["detectprotocol"] == ["ping", "tcp"]
    assert interface.source_attributes["bfd"] == "enable"

    assert interface.client_options[0].id == 1
    assert interface.client_options[0].value == "fortigate client"
    assert interface.client_options[0].extra_settings["future_option"] == "preserved"
    assert interface.dhcp_snooping_server_list[0].server_ip == "192.0.2.53"
    assert interface.tagging[0].tags == ["HQ", "WAN"]
    assert interface.vrrp[0].vrid == 12
    assert interface.vrrp[0].version == 3
    assert interface.vrrp[0].proxy_arp[0].ip == "192.0.2.100"
    assert interface.egress_queues.cos0 == "voice"
    assert interface.egress_queues.cos7 == "bulk"

    assert [node.name for node in interface.nested_configs] == [
        "client-options",
        "dhcp-snooping-server-list",
        "tagging",
        "vrrp",
        "egress-queues",
    ]
    vrrp_node = next(node for node in interface.nested_configs if node.name == "vrrp")
    assert vrrp_node.children[0].children[0].name == "proxy-arp"
    assert "future-option" in {
        command.key
        for command in interface.nested_configs[0].children[0].commands
    }


def test_secondaryip_remains_typed_not_generic():
    config = """
config system interface
    edit "port1"
        set secondary-IP enable

        config secondaryip
            edit 1
                set ip 10.0.0.2 255.255.255.0
                set allowaccess ping
            next
        end
    next
end
"""

    fg = parse_fortigate_config(config)
    interface = fg.interfaces[0]

    assert len(interface.secondary_ips) == 1
    assert interface.nested_configs == []


def test_nested_interface_coverage():
    result = extract_fortigate_config(NESTED_INTERFACE_CONFIG)

    interface_section = next(
        section
        for section in result.source_sections
        if section.path == "system interface"
    )

    assert (
        interface_section.status
        == ExtractionStatus.PARTIALLY_NORMALIZED
    )

    ipv6_section = next(
        section
        for section in result.source_sections
        if section.path == "system interface ipv6"
    )

    assert (
        ipv6_section.status
        == ExtractionStatus.EXTRACT_ONLY
    )

    vrrp_section = next(
        section
        for section in result.source_sections
        if section.path == "system interface vrrp"
    )

    assert (
        vrrp_section.status
        == ExtractionStatus.EXTRACT_ONLY
    )


def test_excel_contains_interface_nested_configuration():
    result = extract_fortigate_config(NESTED_INTERFACE_CONFIG)

    # ExtractionResult exposes canonical_ir, not result.ir.
    workbook_bytes = IRExcelExporter(
        result.canonical_ir,
        result,
    ).generate()

    workbook = load_workbook(
        io.BytesIO(workbook_bytes)
    )

    assert (
        "Interface Nested Configuration"
        in workbook.sheetnames
    )

    sheet = workbook[
        "Interface Nested Configuration"
    ]

    rows = list(
        sheet.iter_rows(values_only=True)
    )

    flattened = "\n".join(
        " | ".join(
            str(value or "")
            for value in row
        )
        for row in rows
    )

    assert "port1" in flattened
    assert "ipv6" in flattened
    assert "ip6-address" in flattened
    assert "2001:db8:1::1/64" in flattened

    assert "vrrp" in flattened
    assert "proxy-arp" in flattened
    assert "10.0.0.10" in flattened

    assert "secret-value" not in flattened
