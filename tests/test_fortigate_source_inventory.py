import io

from openpyxl import load_workbook

from fwmigrate.parsers.fortigate.extractor import extract_fortigate_config
from fwmigrate.extraction.models import ExtractionStatus
from fwmigrate.report.excel_exporter import IRExcelExporter


OPERATIONAL_CONFIG = '''config system ha
    set group-name "FGT-HA"
    set mode a-p
    set hbdev "ha1" 50
end
config system physical-switch
    edit "sw0"
        set age-val 300
    next
end
config system ike
    set dh-group 14
end
config system settings
    set opmode nat
end
config firewall ssh setting
    set caname "SSH-CA"
end
config log fortianalyzer setting
    set status enable
    set server "faz.example.test"
end
config system ntp
    set ntpsync enable
    set type custom
end
config system autoupdate schedule
    set frequency daily
end
config firewall address
    edit "dedicated-address"
        set subnet 192.0.2.0 255.255.255.0
    next
end
config application list
    edit "dedicated-application-list"
        set comment "Already shown in source security profiles"
    next
end
'''

AUTOMATION_CONFIG = '''config system automation-trigger
    edit "High CPU"
        set event-type event-log
        set logid 0100037904
        unset description
        append fields "cpu=high"
    next
end
config system automation-action
    edit "Backup Config"
        set action-type cli-script
        set script "execute backup config flash backup.conf"
    next
end
config system automation-stitch
    edit "backup-config"
        set trigger "High CPU"
        config actions
            edit 1
                set action "Backup Config"
                set required enable
            next
        end
    next
end
'''


def _source_sheet(config: str):
    result = extract_fortigate_config(config)
    workbook = load_workbook(
        io.BytesIO(
            IRExcelExporter(
                result.canonical_ir,
                extraction_result=result,
            ).generate()
        )
    )
    sheet = workbook["FortiGate Source Configuration"]
    headers = {cell.value: cell.column for cell in sheet[3]}
    rows = [
        {
            name: sheet.cell(row, column).value
            for name, column in headers.items()
        }
        for row in range(4, sheet.max_row + 1)
    ]
    return result, workbook, rows


def test_unmodeled_operational_commands_are_visible_without_ir_models() -> None:
    result, workbook, rows = _source_sheet(OPERATIONAL_CONFIG)

    assert "FortiGate Source Configuration" in workbook.sheetnames
    by_path = {}
    for row in rows:
        by_path.setdefault(row["Source Path"], []).append(row)

    assert {
        row["Setting"]: row["Value"]
        for row in by_path["system ha"]
    } == {
        "group-name": "FGT-HA",
        "mode": "a-p",
        "hbdev": "ha1\n50",
    }
    physical_switch = by_path["system physical-switch"][0]
    assert physical_switch["Object"] == "sw0"
    assert physical_switch["Setting"] == "age-val"

    assert by_path["system ike"][0]["Setting"] == "dh-group"
    assert by_path["system settings"][0]["Setting"] == "opmode"
    assert by_path["firewall ssh setting"][0]["Setting"] == "caname"
    assert by_path["log fortianalyzer setting"][0]["Setting"] == "status"
    assert {row["Setting"] for row in by_path["system ntp"]} == {"ntpsync", "type"}
    assert by_path["system autoupdate schedule"][0]["Setting"] == "frequency"

    inventory_paths = {item.source_path for item in result.inventory_items}
    assert "system ha" in inventory_paths
    assert "system physical-switch" in inventory_paths

    statuses = {section.path: section.status for section in result.source_sections}
    for path in (
        "system ha",
        "system physical-switch",
        "system ike",
        "system settings",
        "firewall ssh setting",
        "log fortianalyzer setting",
        "system ntp",
        "system autoupdate schedule",
    ):
        assert statuses[path] == ExtractionStatus.EXTRACT_ONLY

    assert {row["Category"] for row in by_path["system ha"]} == {"System Behaviour"}
    assert {row["Category"] for row in by_path["system ntp"]} == {"Management / Logging"}
    assert {row["Category"] for row in by_path["system autoupdate schedule"]} == {
        "Other Operational"
    }
    assert all(row["Manual Review"] == "Yes" for row in rows)
    assert not result.unsupported_items


def test_dedicated_inventory_is_not_duplicated_in_generic_sheet() -> None:
    _, _, rows = _source_sheet(OPERATIONAL_CONFIG)
    assert all(row["Source Path"] != "firewall address" for row in rows)
    assert all(row["Source Path"] != "application list" for row in rows)


def test_automation_hierarchy_and_operations_survive_source_tree_and_excel() -> None:
    result, _, rows = _source_sheet(AUTOMATION_CONFIG)
    items = {item.source_path: item for item in result.inventory_items}

    trigger = items["system automation-trigger"]
    assert trigger.name == "High CPU"
    assert "structured-operational-config" in trigger.notes
    assert [(command.operation, command.key) for command in trigger.commands] == [
        ("set", "event-type"),
        ("set", "logid"),
        ("unset", "description"),
        ("append", "fields"),
    ]

    stitch = items["system automation-stitch"]
    assert stitch.name == "backup-config"
    assert stitch.children[0].name == "actions"
    assert stitch.children[0].children[0].name == "1"
    nested_commands = stitch.children[0].children[0].commands
    assert [(command.key, command.values) for command in nested_commands] == [
        ("action", ["Backup Config"]),
        ("required", ["enable"]),
    ]

    trigger_rows = [
        row for row in rows
        if row["Source Path"] == "system automation-trigger"
    ]
    assert {row["Object"] for row in trigger_rows} == {"High CPU"}
    assert {row["Operation"] for row in trigger_rows} == {"set", "unset", "append"}

    nested_rows = [
        row for row in rows
        if row["Source Path"] == "system automation-stitch"
        and row["Parent / Subsection"] == "actions / 1"
    ]
    assert {row["Setting"] for row in nested_rows} == {"action", "required"}
    assert {row["Object"] for row in nested_rows} == {"backup-config"}
    assert {row["Category"] for row in rows} == {"Automation"}
    assert all(row["Migration Status"] == "EXTRACT_ONLY" for row in rows)
    assert all(row["Manual Review"] == "Yes" for row in rows)
    assert not result.unsupported_items


def test_management_secrets_are_field_redacted_across_extraction_and_excel() -> None:
    sentinels = {
        "community": "SNMP_COMMUNITY_SECRET",
        "auth-pwd": "SNMP_AUTH_SECRET",
        "priv-pwd": "SNMP_PRIV_SECRET",
        "password": "EMAIL_PASSWORD_SECRET",
        "api-key": "FORTIGUARD_API_SECRET",
        "token": "LOGGING_TOKEN_SECRET",
        "private-key": "PRIVATE_KEY_SECRET",
    }
    config = f'''config system snmp community
    edit 1
        set community "{sentinels["community"]}"
        set auth-pwd "{sentinels["auth-pwd"]}"
        set priv-pwd "{sentinels["priv-pwd"]}"
    next
end
config system email-server
    set password "{sentinels["password"]}"
end
config system fortiguard
    set api-key "{sentinels["api-key"]}"
end
config log syslogd setting
    set token "{sentinels["token"]}"
    set private-key "{sentinels["private-key"]}"
end
'''

    result, workbook, rows = _source_sheet(config)
    serialized = result.model_dump_json()
    workbook_text = "\n".join(
        str(cell.value)
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if cell.value is not None
    )
    for sentinel in sentinels.values():
        assert sentinel not in serialized
        assert sentinel not in workbook_text

    redacted_settings = {
        row["Setting"]: row["Value"]
        for row in rows
    }
    assert redacted_settings == {
        key: "[REDACTED]"
        for key in sentinels
    }
    assert {row["Category"] for row in rows} == {"Management / Logging"}
    assert all(row["Migration Status"] == "EXTRACT_ONLY" for row in rows)


def test_miscellaneous_operational_families_and_numeric_source_id_survive() -> None:
    config = '''config system autoupdate schedule
    set frequency weekly
end
config system search-engine
    edit "safe-token-name"
        set hostname "search.example.test"
    next
end
config system threat-weight
    edit 42
        set status enable
    next
end
'''

    result, _, rows = _source_sheet(config)
    assert {row["Source Path"] for row in rows} == {
        "system autoupdate schedule",
        "system search-engine",
        "system threat-weight",
    }
    search_row = next(row for row in rows if row["Source Path"] == "system search-engine")
    assert search_row["Object"] == "safe-token-name"
    threat_row = next(row for row in rows if row["Source Path"] == "system threat-weight")
    assert threat_row["Object"] == "42"
    assert threat_row["Source ID"] == "42"
    assert all(row["Migration Status"] == "EXTRACT_ONLY" for row in rows)
    assert not result.unsupported_items
