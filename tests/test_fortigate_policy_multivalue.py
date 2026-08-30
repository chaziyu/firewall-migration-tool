import io
import json

from openpyxl import load_workbook

from fwmigrate.parsers.fortigate.extractor import extract_fortigate_config
from fwmigrate.parsers.fortigate.parser import (
    SECTION_LIST_FIELDS,
    parse_fortigate_config,
)
from fwmigrate.report.excel_exporter import IRExcelExporter


def _policy_config(policy_settings: str) -> str:
    return f'''config firewall policy
    edit 42
        set name "Multi Value Policy"
{policy_settings}
    next
end
'''


def test_policy_multivalue_fields_preserve_normalized_names_boundaries_and_order():
    list_fields = SECTION_LIST_FIELDS["firewall policy"]
    assert {
        "custom_log_fields",
        "pcp_poolname",
        "ztna_ems_tag_secondary",
    } <= list_fields
    assert not any("-" in field for field in list_fields)

    policy = parse_fortigate_config(
        _policy_config(
            '''        set custom-log-fields "field1" "field with spaces" "field3"
        set pcp-poolname "pool1" "pool with spaces" "pool3"
        set ztna-ems-tag-secondary "tag1" "tag with spaces" "tag3"'''
        )
    ).policies[0]

    assert policy.extra_settings["custom_log_fields"] == [
        "field1", "field with spaces", "field3"
    ]
    assert policy.extra_settings["pcp_poolname"] == [
        "pool1", "pool with spaces", "pool3"
    ]
    assert policy.ztna_ems_tag_secondary == [
        "tag1", "tag with spaces", "tag3"
    ]
    assert "custom-log-fields" not in policy.extra_settings
    assert "pcp-poolname" not in policy.extra_settings
    assert "ztna-ems-tag-secondary" not in policy.model_dump()


def test_policy_multivalue_fields_keep_one_value_as_a_list_and_omit_absent_values():
    config = '''config firewall policy
    edit 1
        set custom-log-fields "field1"
        set pcp-poolname "pool1"
        set ztna-ems-tag-secondary "tag1"
    next
    edit 2
        set name "Omitted Values"
    next
end
'''

    policies = parse_fortigate_config(config).policies
    configured, omitted = policies

    assert configured.extra_settings["custom_log_fields"] == ["field1"]
    assert configured.extra_settings["pcp_poolname"] == ["pool1"]
    assert configured.ztna_ems_tag_secondary == ["tag1"]

    assert "custom_log_fields" not in omitted.extra_settings
    assert "pcp_poolname" not in omitted.extra_settings
    assert omitted.ztna_ems_tag_secondary == []


def test_policy_multivalue_fields_survive_extraction_ir_and_excel_source_settings():
    config = _policy_config(
        '''        set custom-log-fields "field1" "field two"
        set pcp-poolname "pool1" "pool two"
        set ztna-ems-tag-secondary "tag1" "tag two"'''
    )
    extraction = extract_fortigate_config(config)
    policy = extraction.canonical_ir.policies[0]

    expected = {
        "custom_log_fields": ["field1", "field two"],
        "pcp_poolname": ["pool1", "pool two"],
    }
    assert policy.source_extra_settings == expected
    assert policy.source_ztna_ems_tags_secondary == ["tag1", "tag two"]

    source_item = next(
        item
        for item in extraction.inventory_items
        if item.source_path == "firewall policy"
    )
    source_values = {
        command.key: command.values
        for command in source_item.commands
        if command.key in {
            "custom-log-fields",
            "pcp-poolname",
            "ztna-ems-tag-secondary",
        }
    }
    assert source_values == {
        "custom-log-fields": ["field1", "field two"],
        "pcp-poolname": ["pool1", "pool two"],
        "ztna-ems-tag-secondary": ["tag1", "tag two"],
    }

    workbook = load_workbook(
        io.BytesIO(
            IRExcelExporter(
                extraction.canonical_ir,
                extraction_result=extraction,
            ).generate()
        )
    )
    sheet = workbook["Firewall Policy Source Settings"]
    headers = {cell.value: cell.column for cell in sheet[3]}
    values = {
        sheet.cell(row, headers["Setting"]).value: sheet.cell(
            row, headers["Ordered Source Values"]
        ).value
        for row in range(4, sheet.max_row + 1)
        if sheet.cell(row, headers["Setting"]).value in source_values
    }
    assert values == {
        key: json.dumps(value, ensure_ascii=False)
        for key, value in source_values.items()
    }
