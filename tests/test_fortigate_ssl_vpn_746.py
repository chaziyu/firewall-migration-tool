import io
from pathlib import Path

from openpyxl import load_workbook

from fwmigrate.parsers.fortigate.extractor import extract_fortigate_config
from fwmigrate.parsers.fortigate.parser import parse_fortigate_config
from fwmigrate.report.excel_exporter import IRExcelExporter


FIXTURE = Path(__file__).parent / "fixtures" / "fortigate" / "fortios_7_4_6_ssl_vpn_full.conf"
SECRET = "ssl-vpn-secret-must-not-survive"


def test_fortios_746_ssl_vpn_survives_cli_fg_ir_excel_without_secrets():
    content = FIXTURE.read_text()
    fg = parse_fortigate_config(content)
    extraction = extract_fortigate_config(content)
    ir = extraction.canonical_ir

    assert fg.ssl_vpn_settings.source_address6 == ["SSLVPN_SOURCE6"]
    assert len(fg.ssl_vpn_settings.authentication_rules) == 2
    assert ir.ssl_vpn_settings.dtls_heartbeat_interval == 5
    assert ir.ssl_vpn_settings.source_address6 == ["SSLVPN_SOURCE6"]
    assert len(ir.ssl_vpn_portals) == 2

    portal = ir.ssl_vpn_portals[0]
    assert portal.default_window_width == 1280
    assert portal.ipv6_split_tunneling_routing_addresses == ["V6_ROUTE"]
    assert len(portal.split_dns) == 1
    assert len(portal.mac_address_check_rules) == 1
    assert len(portal.os_check_list) == 1
    assert len(portal.bookmark_groups[0].bookmarks) == 2
    assert portal.bookmark_groups[0].bookmarks[0].has_logon_password
    assert portal.bookmark_groups[0].bookmarks[0].form_data[0].value_configured
    assert len(portal.landing_pages[0].form_data) == 1

    assert any("SSLVPN_SOURCE" in str(item) for item in ir.audit_entries)
    assert SECRET not in extraction.model_dump_json()

    workbook = load_workbook(io.BytesIO(IRExcelExporter(ir, extraction).generate()))
    assert workbook["SSL VPN Portal Split DNS"].max_row == 4
    assert workbook["SSL VPN Bookmarks"].max_row == 5
    assert workbook["SSL VPN Bookmark Form Data"].max_row == 4
    assert SECRET not in "".join(
        str(cell.value)
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
    )


def test_fortios_746_ssl_vpn_global_settings_are_normalized():
    extraction = extract_fortigate_config(FIXTURE.read_text())
    paths = {section.path: section for section in extraction.source_sections}

    assert paths["vpn ssl settings"].status.value == "PARTIALLY_NORMALIZED"
    assert paths["vpn ssl settings authentication-rule"].status.value == "EXTRACT_ONLY"

    known_only = extract_fortigate_config("""
config vpn ssl settings
    set status enable
    set servercert "FortiWeb"
    set source-interface "wan1"
    set source-address "ssl-source"
    set tunnel-ip-pools "ssl-pool"
    set dns-server1 192.0.2.53
    set ipv6-dns-server1 2001:db8::53
    set login-timeout 120
end
""")
    assert {
        section.path: section.status.value
        for section in known_only.source_sections
    }["vpn ssl settings"] == "NORMALIZED"

    for path in (
        "vpn ssl web portal split-dns",
        "vpn ssl web portal mac-addr-check-rule",
        "vpn ssl web portal os-check-list",
        "vpn ssl web portal bookmark-group",
        "vpn ssl web portal bookmark-group bookmarks",
        "vpn ssl web portal bookmark-group bookmarks form-data",
        "vpn ssl web portal landing-page",
        "vpn ssl web portal landing-page form-data",
    ):
        assert paths[path].status.value == "EXTRACT_ONLY"


def test_fortios_746_typed_portal_inventory_is_not_forced_to_manual_review():
    from fwmigrate.parsers.fortigate.coverage import extract_only_requires_manual_review

    assert extract_only_requires_manual_review("vpn ssl web portal") is False
    assert extract_only_requires_manual_review(
        "vpn ssl web portal bookmark-group bookmarks form-data"
    ) is False
    extraction = extract_fortigate_config("""
config vpn ssl web portal
    edit "typed"
        set tunnel-mode enable
        config split-dns
            edit 1
                set domains "corp.example"
            next
        end
    next
end
""")
    portal_inventory = next(
        item for item in extraction.inventory_items
        if item.source_path == "vpn ssl web portal"
    )
    assert portal_inventory.requires_manual_review is False
