import io

from openpyxl import load_workbook

from fwmigrate.core.constants import IR_KEYWORD_ANY
from fwmigrate.ir.enums import ServiceProtocol
from fwmigrate.parsers.fortigate.parser import parse_fortigate_config
from fwmigrate.parsers.fortigate.extractor import extract_fortigate_config
from fwmigrate.parsers.fortigate.transformer import FGToIRTransformer
from fwmigrate.report.excel_exporter import IRExcelExporter


SERVICE_CONFIG = """
config firewall service category
    edit "General"
        set comment "General services."
        set color 1
    next
    edit "Web Access"
    next
    edit "File Access"
    next
    edit "Email"
    next
    edit "Network Services"
    next
    edit "Authentication"
    next
    edit "Remote Access"
    next
    edit "Tunneling"
    next
    edit "VoIP, Messaging & Other Applications"
    next
    edit "Web Proxy"
    next
end
config firewall service custom
    edit "ALL"
        set uuid 00000000-0000-0000-0000-000000000001
        set category "General"
        set protocol IP
    next
    edit "FTP"
        set uuid 00000000-0000-0000-0000-000000000002
        set category "File Access"
        set tcp-portrange 21
        set helper ftp
    next
    edit "DNS"
        set category "Network Services"
        set tcp-portrange 53
        set udp-portrange 53
    next
    edit "KERBEROS"
        set category "Authentication"
        set tcp-portrange 88 464
        set udp-portrange 88 464
    next
    edit "ALL_ICMP"
        set protocol ICMP
    next
    edit "ALL_ICMP6"
        set protocol ICMP6
    next
    edit "GRE"
        set protocol IP
        set protocol-number 47
    next
    edit "PING"
        set protocol ICMP
        set icmptype 8
    next
    edit "PING6"
        set protocol ICMP6
        set icmptype 128
    next
    edit "RLOGIN"
        set tcp-portrange 513:512-1023
    next
    edit "RSH"
        set tcp-portrange 514:512-1023
    next
    edit "NONE"
        set tcp-portrange 0
    next
    edit "webproxy"
        set uuid 00000000-0000-0000-0000-000000000013
        set proxy enable
        set category "Web Proxy"
        set protocol ALL
        set tcp-portrange 0-65535:0-65535
    next
    edit "DIAMETER-SCTP"
        set category "Authentication"
        set protocol TCP/UDP/SCTP
        set sctp-portrange 3868 3869
    next
    edit "SCTP-SOURCE-CONSTRAINT"
        set protocol TCP/UDP/SCTP
        set sctp-portrange 5000:1024-65535
    next
    edit "EXPLICIT-DEFAULT"
        set protocol TCP/UDP/SCTP
        set tcp-portrange 12345
    next
    edit "MY-ANY-IP"
        set protocol IP
        set protocol-number 0
    next
    edit "FULL-RANGE"
        set tcp-portrange 0-65535
    next
    edit "FQDN-HTTPS"
        set tcp-portrange 443
        set fqdn "service.example.com"
    next
    edit "SESSION-SERVICE"
        set tcp-portrange 444
        set session-ttl 300
    next
    edit "SOURCE-METADATA"
        set tcp-portrange 445
        set color 4
        set fabric-object enable
    next
end
config firewall service group
    edit "Web Access"
        set uuid 00000000-0000-0000-0000-000000000014
        set member "DNS" "FTP"
        set color 4
    next
    edit "Unsafe Child"
        set member "FQDN-HTTPS"
    next
    edit "Unsafe Parent"
        set member "Unsafe Child"
    next
    edit "Missing Member"
        set member "DOES-NOT-EXIST"
    next
    edit "Safe Group"
        set member "DNS"
    next
end
"""


def _by_name(items):
    return {item.name: item for item in items}


def test_service_parser_preserves_source_metadata_and_categories():
    parsed = parse_fortigate_config(SERVICE_CONFIG)
    categories = _by_name(parsed.service_categories)
    services = _by_name(parsed.services)
    groups = _by_name(parsed.service_groups)

    assert len(categories) == 10
    assert categories["General"].comment == "General services."
    assert categories["General"].extra_settings == {"color": "1"}
    assert services["FTP"].uuid == "00000000-0000-0000-0000-000000000002"
    assert services["FTP"].category == "File Access"
    assert services["FTP"].extra_settings == {"helper": "ftp"}
    assert services["FTP"].protocol == "tcp/udp/sctp"
    assert services["FTP"].source_protocol_configured is None
    assert services["EXPLICIT-DEFAULT"].source_protocol_configured == "TCP/UDP/SCTP"
    assert services["SOURCE-METADATA"].color == 4
    assert services["SOURCE-METADATA"].fabric_object == "enable"
    assert services["webproxy"].proxy == "enable"
    assert services["DIAMETER-SCTP"].sctp_portrange == "3868,3869"
    assert groups["Web Access"].uuid == "00000000-0000-0000-0000-000000000014"
    assert groups["Web Access"].color == 4
    assert groups["Web Access"].extra_settings == {}


def test_service_semantics_are_preserved_without_permissive_rewriting():
    ir = FGToIRTransformer(
        parse_fortigate_config(SERVICE_CONFIG)
    ).transform()
    services = _by_name(ir.services)

    assert services["ALL"].ports[0].protocol == ServiceProtocol.ANY
    assert services["ALL"].source_protocol == "IP"

    assert services["FTP"].source_uuid == "00000000-0000-0000-0000-000000000002"
    assert services["FTP"].source_category == "File Access"
    assert services["FTP"].source_protocol == "tcp/udp/sctp"
    assert services["FTP"].source_protocol_configured is None
    assert services["FTP"].ports[0].port == "21"
    assert services["FTP"].source_attributes == {"helper": "ftp"}
    assert services["FTP"].source_unmodeled_semantic_settings == ["helper"]
    assert services["FTP"].requires_manual_review is True
    assert services["FTP"].migration_status == "PARTIALLY_NORMALIZED"

    explicit = services["EXPLICIT-DEFAULT"]
    assert explicit.source_protocol_configured == "TCP/UDP/SCTP"
    assert explicit.source_protocol == "TCP/UDP/SCTP"

    any_ip = services["MY-ANY-IP"]
    assert any_ip.source_protocol_number == 0
    assert any_ip.ports[0].protocol == ServiceProtocol.ANY
    assert any_ip.ports[0].port == IR_KEYWORD_ANY

    full_range = services["FULL-RANGE"]
    assert full_range.ports[0].port == "0-65535"
    assert full_range.requires_manual_review is False
    assert "destination port 0" not in (full_range.audit_note or "")

    fqdn = services["FQDN-HTTPS"]
    assert fqdn.ports[0].port == "443"
    assert fqdn.source_attributes["fqdn"] == "service.example.com"
    assert fqdn.source_unmodeled_semantic_settings == ["fqdn"]
    assert fqdn.requires_manual_review is True
    assert fqdn.migration_status == "PARTIALLY_NORMALIZED"

    session_service = services["SESSION-SERVICE"]
    assert session_service.source_attributes["session_ttl"] == "300"
    assert session_service.source_unmodeled_semantic_settings == ["session_ttl"]
    assert session_service.requires_manual_review is True

    metadata = services["SOURCE-METADATA"]
    assert metadata.source_color == 4
    assert metadata.source_fabric_object == "enable"
    assert metadata.source_unmodeled_semantic_settings == []

    assert {port.protocol for port in services["DNS"].ports} == {
        ServiceProtocol.TCP,
        ServiceProtocol.UDP,
    }
    assert [port.port for port in services["KERBEROS"].ports] == [
        "88", "464", "88", "464"
    ]

    assert services["ALL_ICMP"].ports[0].protocol == ServiceProtocol.ICMP
    assert services["ALL_ICMP6"].ports[0].protocol == ServiceProtocol.ICMPV6
    assert services["GRE"].ports[0].protocol == ServiceProtocol.IP
    assert services["GRE"].ports[0].port == "47"
    assert services["PING"].ports[0].icmptype == 8
    assert services["PING6"].ports[0].protocol == ServiceProtocol.ICMPV6
    assert services["PING6"].ports[0].icmptype == 128

    rlogin = services["RLOGIN"].ports[0]
    assert rlogin.port == "513"
    assert rlogin.source_port == "512-1023"
    assert rlogin.raw_source_value == "513:512-1023"
    rsh = services["RSH"].ports[0]
    assert rsh.port == "514"
    assert rsh.source_port == "512-1023"

    none = services["NONE"]
    assert none.ports[0].port == "0"
    assert none.ports[0].port != "1-65535"
    assert none.requires_manual_review is True
    assert none.migration_status == "PARTIALLY_NORMALIZED"
    assert "destination port 0" in none.audit_note

    webproxy = services["webproxy"]
    assert webproxy.ports[0].port == "0-65535"
    assert webproxy.ports[0].source_port == "0-65535"
    assert webproxy.source_proxy is True
    assert webproxy.source_protocol == "ALL"
    assert webproxy.requires_manual_review is True
    assert webproxy.migration_status == "PARTIALLY_NORMALIZED"
    assert "proxy service semantics" in webproxy.audit_note
    assert "destination port 0" not in webproxy.audit_note

    diameter = services["DIAMETER-SCTP"]
    assert [
        port.port for port in diameter.ports
        if port.protocol == ServiceProtocol.SCTP
    ] == ["3868", "3869"]
    assert diameter.requires_manual_review
    source_constraint = services["SCTP-SOURCE-CONSTRAINT"].ports[0]
    assert source_constraint.protocol == ServiceProtocol.SCTP
    assert source_constraint.port == "5000"
    assert source_constraint.source_port == "1024-65535"
    assert source_constraint.raw_source_value == "5000:1024-65535"

    groups = _by_name(ir.service_groups)
    group = groups["Web Access"]
    assert group.members == ["DNS", "FTP"]
    assert group.source_uuid == "00000000-0000-0000-0000-000000000014"
    assert group.source_color == 4
    assert group.source_attributes == {}
    assert group.unsafe_members == ["FTP"]
    assert group.requires_manual_review is True
    assert groups["Unsafe Child"].unsafe_members == ["FQDN-HTTPS"]
    assert groups["Unsafe Parent"].unsafe_members == ["Unsafe Child"]
    assert groups["Missing Member"].unsafe_members == ["DOES-NOT-EXIST"]
    assert "unresolved service/service-group" in groups["Missing Member"].audit_note
    assert groups["Safe Group"].unsafe_members == []
    assert groups["Safe Group"].migration_status == "NORMALIZED"
    assert groups["Safe Group"].requires_manual_review is False


def test_service_inventory_reaches_excel():
    ir = FGToIRTransformer(
        parse_fortigate_config(SERVICE_CONFIG)
    ).transform()
    workbook = load_workbook(
        io.BytesIO(IRExcelExporter(ir).generate())
    )

    categories = workbook["Service Categories"]
    assert categories.max_row == 13
    assert categories["A4"].value == "General"
    category_headers = {cell.value: cell.column for cell in categories[3]}
    assert categories.cell(4, category_headers["Extraction Status"]).value == "EXTRACT_ONLY"
    assert categories.cell(4, category_headers["Additional Settings"]).value == "color=1"

    services = workbook["Services"]
    headers = {cell.value: cell.column for cell in services[3]}
    rows = {
        services.cell(row, headers["Name"]).value: row
        for row in range(4, services.max_row + 1)
    }

    ftp_row = rows["FTP"]
    assert services.cell(ftp_row, headers["Source UUID"]).value == (
        "00000000-0000-0000-0000-000000000002"
    )
    assert services.cell(ftp_row, headers["Category"]).value == "File Access"
    assert services.cell(ftp_row, headers["Configured Protocol"]).value is None
    assert services.cell(ftp_row, headers["Effective Protocol"]).value == "tcp/udp/sctp"
    assert services.cell(ftp_row, headers["Unmodeled Semantic Settings"]).value == "helper"
    assert services.cell(ftp_row, headers["Migration Status"]).value == "PARTIALLY_NORMALIZED"
    assert services.cell(ftp_row, headers["Protocol / Destination Port"]).value == "tcp/21"

    explicit_row = rows["EXPLICIT-DEFAULT"]
    assert services.cell(explicit_row, headers["Configured Protocol"]).value == "TCP/UDP/SCTP"
    assert services.cell(explicit_row, headers["Effective Protocol"]).value == "TCP/UDP/SCTP"

    any_ip_row = rows["MY-ANY-IP"]
    assert services.cell(any_ip_row, headers["Source Protocol Number"]).value == 0

    metadata_row = rows["SOURCE-METADATA"]
    assert services.cell(metadata_row, headers["Source Color"]).value == 4
    assert services.cell(metadata_row, headers["Fabric Object"]).value == "enable"

    rlogin_row = rows["RLOGIN"]
    assert services.cell(rlogin_row, headers["Protocol / Destination Port"]).value == "tcp/513"
    assert services.cell(rlogin_row, headers["Source Port Constraint"]).value == "512-1023"

    none_row = rows["NONE"]
    assert services.cell(none_row, headers["Protocol / Destination Port"]).value == "tcp/0"
    assert "1-65535" not in services.cell(none_row, headers["Protocol / Destination Port"]).value

    proxy_row = rows["webproxy"]
    assert services.cell(proxy_row, headers["Protocol / Destination Port"]).value == "tcp/0-65535"
    assert services.cell(proxy_row, headers["Source Port Constraint"]).value == "0-65535"
    assert services.cell(proxy_row, headers["Proxy"]).value == "TRUE"
    assert services.cell(proxy_row, headers["Manual Review"]).value == "TRUE"
    assert services.cell(proxy_row, headers["Migration Status"]).value == "PARTIALLY_NORMALIZED"
    assert "destination port 0" not in services.cell(proxy_row, headers["Audit Note"]).value
    diameter_row = rows["DIAMETER-SCTP"]
    assert "sctp/3868" in services.cell(
        diameter_row, headers["Protocol / Destination Port"]
    ).value

    groups = workbook["Service Groups"]
    group_headers = {cell.value: cell.column for cell in groups[3]}
    assert groups.cell(4, group_headers["Source UUID"]).value == (
        "00000000-0000-0000-0000-000000000014"
    )
    assert groups.cell(4, group_headers["Source Color"]).value == 4
    assert groups.cell(4, group_headers["Unsafe Members"]).value == "FTP"

    summary = {
        workbook["Summary"].cell(row, 1).value:
            workbook["Summary"].cell(row, 2).value
        for row in range(1, workbook["Summary"].max_row + 1)
    }
    assert summary["Service Categories"] == 10


def test_target_generators_do_not_flatten_source_port_or_proxy_semantics():
    from fwmigrate.generators.checkpoint.cli_generator import CheckPointCLIGenerator
    from fwmigrate.generators.cisco_asa.cli_generator import CiscoASACLIGenerator
    from fwmigrate.generators.fortigate.cli_generator import FortiGateCLIGenerator
    from fwmigrate.generators.juniper_srx.cli_generator import JuniperSRXCLIGenerator
    from fwmigrate.generators.palo_alto.terraform_generator import PANOSTerraformGenerator
    from fwmigrate.generators.palo_alto.xml_generator import PANOSXMLGenerator

    ir = FGToIRTransformer(
        parse_fortigate_config(SERVICE_CONFIG)
    ).transform()

    for svc in ir.services:
        svc.migration_status = "NORMALIZED"
        svc.requires_manual_review = False
        svc.parse_error = None

    fortigate_output = FortiGateCLIGenerator().generate(ir)[0].content
    assert "set tcp-portrange 513:512-1023" in fortigate_output
    assert "set tcp-portrange 0-65535:0-65535" in fortigate_output
    assert "set proxy enable" in fortigate_output
    assert "set sctp-portrange 3868" in fortigate_output
    assert "set sctp-portrange 5000:1024-65535" in fortigate_output
    assert "Service FTP withheld: unmodeled FortiGate service semantics" in fortigate_output
    assert 'edit "FQDN-HTTPS"' not in fortigate_output
    assert 'edit "NONE"' in fortigate_output
    assert "set tcp-portrange 0" in fortigate_output
    assert "set protocol-number 0" in fortigate_output
    assert "set protocol TCP/UDP/SCTP" in fortigate_output
    assert "set icmptype 8" in fortigate_output
    assert "set icmptype 128" in fortigate_output
    assert "Service group Unsafe Child withheld" in fortigate_output

    dns_block = fortigate_output.split('edit "DNS"', 1)[1].split("next", 1)[0]
    assert "set protocol" not in dns_block

    cisco_output = CiscoASACLIGenerator().generate(ir)
    checkpoint_output = CheckPointCLIGenerator().generate(ir)
    juniper_output = JuniperSRXCLIGenerator().generate(ir)
    for output in (cisco_output, checkpoint_output, juniper_output):
        assert "Service RLOGIN withheld" in output
        assert "Service webproxy withheld" in output
    assert "Service group Unsafe Child withheld" in cisco_output
    assert "Service group Unsafe Child withheld" in juniper_output

    panos_xml = PANOSXMLGenerator().generate(ir)[0].content
    assert '<entry name="RLOGIN">' not in panos_xml
    assert '<entry name="webproxy">' not in panos_xml
    assert '<entry name="Unsafe Child">' not in panos_xml

    panos_tf = "\n".join(
        artifact.content
        for artifact in PANOSTerraformGenerator().generate(ir)
    )
    assert "Service RLOGIN withheld" in panos_tf
    assert "Service webproxy withheld" in panos_tf
    assert "Service group Unsafe Child withheld" in panos_tf


def test_service_and_group_coverage_reflects_semantic_review_state():
    result = extract_fortigate_config(SERVICE_CONFIG)
    coverage = {item.path: item for item in result.source_sections}

    services = coverage["firewall service custom"]
    assert services.object_count_source == len(result.canonical_ir.services)
    assert services.object_count_parsed == services.object_count_source
    assert services.object_count_normalized == services.object_count_source
    assert services.status.value == "PARTIALLY_NORMALIZED"

    groups = coverage["firewall service group"]
    assert groups.object_count_source == len(result.canonical_ir.service_groups)
    assert groups.object_count_parsed == groups.object_count_source
    assert groups.object_count_normalized == groups.object_count_source
    assert groups.status.value == "PARTIALLY_NORMALIZED"

    safe = extract_fortigate_config("""
config firewall service custom
    edit "HTTPS"
        set tcp-portrange 443
    next
end
config firewall service group
    edit "Web"
        set member "HTTPS"
    next
end
""")
    safe_coverage = {item.path: item for item in safe.source_sections}
    assert safe_coverage["firewall service custom"].status.value == "NORMALIZED"
    assert safe_coverage["firewall service group"].status.value == "NORMALIZED"
