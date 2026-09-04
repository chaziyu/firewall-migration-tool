import io
import json
from datetime import datetime, timezone

from openpyxl import load_workbook

from fwmigrate.parsers.palo_alto.parser import PANOSSourceParser
from fwmigrate.report.excel_exporter import IRExcelExporter


PRIVATE_KEY_MARKER = "PHASE71_PRIVATE_KEY_SECRET_MARKER"


def _extract(certificates: str, profiles: str = ""):
    xml = f"<config><shared><certificate>{certificates}</certificate>{profiles}</shared></config>"
    return PANOSSourceParser().extract(xml)


def _certificate(name: str, fields: str = "") -> str:
    return f'<entry name="{name}">{fields}</entry>'


def test_pan_gmt_and_iso_certificate_dates_are_utc():
    result = _extract(
        _certificate(
            "pan-cert",
            "<not-valid-before>Mar 19 05:02:53 2021 GMT</not-valid-before>"
            "<not-valid-after>Mar 19 05:02:53 2022 GMT</not-valid-after>",
        )
        + _certificate("iso-z", "<not-valid-before>2024-01-15T10:20:30Z</not-valid-before>")
        + _certificate(
            "iso-offset",
            "<not-valid-before>2024-01-15T18:20:30+08:00</not-valid-before>",
        )
        + _certificate(
            "iso-naive",
            "<not-valid-before>2024-01-15T10:20:30</not-valid-before>",
        )
    )

    certificates = {item.name: item for item in result.canonical_ir.certificates}
    assert certificates["pan-cert"].valid_from == datetime(2021, 3, 19, 5, 2, 53, tzinfo=timezone.utc)
    assert certificates["pan-cert"].valid_until == datetime(2022, 3, 19, 5, 2, 53, tzinfo=timezone.utc)
    assert certificates["iso-z"].valid_from == datetime(2024, 1, 15, 10, 20, 30, tzinfo=timezone.utc)
    assert certificates["iso-offset"].valid_from == datetime(2024, 1, 15, 10, 20, 30, tzinfo=timezone.utc)
    assert certificates["iso-naive"].valid_from == datetime(2024, 1, 15, 10, 20, 30, tzinfo=timezone.utc)


def test_malformed_partial_and_missing_certificate_dates_are_distinct():
    result = _extract(
        _certificate("malformed", "<not-valid-before>not-a-certificate-date</not-valid-before>")
        + _certificate(
            "partial",
            "<not-valid-before>Mar 19 05:02:53 2021 GMT</not-valid-before>"
            "<not-valid-after>not-a-certificate-date</not-valid-after>",
        )
        + _certificate("missing")
    )
    certificates = {item.name: item for item in result.canonical_ir.certificates}

    malformed = certificates["malformed"]
    assert malformed.valid_from is None
    assert malformed.source_attributes["pan_malformed_certificate_dates"] == {
        "not-valid-before": "not-a-certificate-date"
    }
    assert malformed.source_attributes["review_reasons"] == ["unparsed-certificate-valid-from"]

    partial = certificates["partial"]
    assert partial.valid_from == datetime(2021, 3, 19, 5, 2, 53, tzinfo=timezone.utc)
    assert partial.valid_until is None
    assert partial.source_attributes["review_reasons"] == ["unparsed-certificate-valid-until"]

    missing = certificates["missing"]
    assert missing.valid_from is None and missing.valid_until is None
    assert missing.source_attributes.get("review_reasons", []) == []
    assert missing.source_attributes.get("pan_malformed_certificate_dates", {}) == {}

    inventory = next(item for item in result.inventory_items if item.name == "malformed")
    assert "unparsed-certificate-valid-from" in inventory.notes


def test_certificate_metadata_relationships_and_private_key_safety_remain_intact():
    result = _extract(
        _certificate(
            "metadata-cert",
            "<public-key>PUBLIC-CERTIFICATE-DATA</public-key>"
            "<algorithm>RSA</algorithm><ca>yes</ca>"
            "<subject>CN=Subject</subject><issuer>CN=Issuer</issuer>"
            f"<private-key><key>{PRIVATE_KEY_MARKER}</key></private-key>",
        )
        + _certificate("non-ca", "<ca>no</ca>")
        + _certificate("malformed-ca", "<ca>maybe</ca>"),
        "<trusted-root-CA><member>metadata-cert</member></trusted-root-CA>"
        '<ssl-tls-service-profile><entry name="tls-profile">'
        "<certificate>metadata-cert</certificate>"
        "<protocol-settings><min-version>TLSv1.2</min-version><max-version>TLSv1.3</max-version>"
        "</protocol-settings></entry></ssl-tls-service-profile>",
    )
    certificates = {item.name: item for item in result.canonical_ir.certificates}
    metadata = certificates["metadata-cert"]
    assert metadata.public_certificate_pem == "PUBLIC-CERTIFICATE-DATA"
    assert metadata.public_key_algorithm == "RSA"
    assert metadata.is_ca is True
    assert metadata.has_private_key is True
    assert metadata.subject == "CN=Subject"
    assert metadata.issuer == "CN=Issuer"
    assert metadata.source_attributes["pan_trusted_root_references"] == ["metadata-cert"]
    assert certificates["non-ca"].is_ca is False
    assert certificates["malformed-ca"].source_attributes["pan_malformed_ca"] == "maybe"

    profile = result.canonical_ir.ssl_tls_service_profiles[0]
    assert profile.certificate == "metadata-cert"
    assert profile.certificate_resolved is True
    assert profile.minimum_tls_version == "TLSv1.2"
    assert profile.maximum_tls_version == "TLSv1.3"

    serialized = json.dumps(result.model_dump(), default=str)
    assert PRIVATE_KEY_MARKER not in serialized


def test_pan_certificate_dates_flow_to_excel_and_expired_status():
    result = _extract(
        _certificate("expired", "<not-valid-after>Mar 19 05:02:53 2000 GMT</not-valid-after>")
        + _certificate("future", "<not-valid-after>Mar 19 05:02:53 2099 GMT</not-valid-after>")
    )
    workbook = load_workbook(io.BytesIO(IRExcelExporter(result.canonical_ir).generate()))
    sheet = workbook["Certificates"]
    headers = {cell.value: cell.column for cell in sheet[3]}
    rows = {sheet.cell(row, headers["Name"]).value: row for row in range(4, sheet.max_row + 1)}

    assert sheet.cell(rows["expired"], headers["Valid Until"]).value == "2000-03-19T05:02:53+00:00"
    assert sheet.cell(rows["expired"], headers["Expired"]).value == "Yes"
    assert sheet.cell(rows["future"], headers["Valid Until"]).value == "2099-03-19T05:02:53+00:00"
    assert sheet.cell(rows["future"], headers["Expired"]).value == "No"
