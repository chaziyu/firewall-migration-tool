import io

from openpyxl import load_workbook

from fwmigrate.parsers.fortigate.extractor import extract_fortigate_config
from fwmigrate.parsers.fortigate.model import FGSSLVPNHostCheckSoftware
from fwmigrate.parsers.fortigate.parser import parse_fortigate_config
from fwmigrate.report.excel_exporter import IRExcelExporter


SECRET = "identity-secret-must-not-survive"

FSSO_PHASE5_CONFIG = '''
config user fsso
    edit "agents"
        set server "192.0.2.10"
        set password "secret-1"
        set server2 "192.0.2.11"
        set password2 "secret-2"
        set port2 8001
        set sni "collector.example.test"
    next
end
config user fsso-polling
    edit 1
        set server "dc.example.test"
        set password "secret-3"
        set polling-frequency 5
        config adgrp
            edit "DOMAIN/Users"
            next
        end
    next
end
'''

FSSO_PHASE18_CONFIG = '''
config user fsso
    edit "collector"
        set server "collector.example.test"
        set port 8000
        set server2 "collector-2.example.test"
        set port2 8001
        set password "COLLECTOR_SECRET"
        set password2 "COLLECTOR_SECRET_2"
        set ldap-poll enable
        set ldap-poll-filter "(objectCategory=group)"
        set ldap-poll-interval 60
        set group-poll-interval 30
        set source-ip 192.0.2.20
        set source-ip6 2001:db8::20
        set ssl enable
        set ssl-server-host-ip-check enable
        set ssl-trusted-cert "collector-ca"
        set interface-select-method specify
        set interface "mgmt"
        set vrf-select 7
        set vendor-option first
        set vendor-option second
    next
end
config user fsso-polling
    edit 1
        set status enable
        set server "dc-1.example.test"
        set default-domain "EXAMPLE"
        set port 389
        set user "svc-fsso"
        set password "POLLING_SECRET"
        set ldap-server "ldap.example.test"
        set logon-history 24
        set polling-frequency 5
        set smbv1 disable
        set smb-ntlmv1-auth enable
        set vendor-option first
        set vendor-option second
        config adgrp
            edit "EXAMPLE/Domain Users"
                set vendor-option one
                set vendor-option two
            next
            edit "EXAMPLE/Network Admins"
            next
        end
    next
end
'''

LOCAL_GROUP_PHASE6_CONFIG = '''
config user local
    edit "alice"
        set id 7
        set type password
        set tacacs+-server "tacacs"
        set ppk-identity "alice-key"
        set ppk-secret "secret-ppk"
    next
end
config user group
    edit "vpn-users"
        set group-type firewall
        set auth-concurrent-override enable
        set auth-concurrent-value 3
        set member "alice" "missing-user"
        config guest
            edit 1
                set name "Guest"
                set password "secret-guest"
            next
        end
    next
end
config user tacacs+
    edit "tacacs"
    next
end
'''


ACCESS_CONFIG = f"""
config user ldap
    edit "corp-ldap"
        set server "ldap.example.test"
        set secondary-server "ldap2.example.test"
        set tertiary-server "ldap3.example.test"
        set cnid "sAMAccountName"
        set dn "dc=example,dc=test"
        set type regular
        set username "bind-user"
        set password "{SECRET}"
        set port 636
        set secure ldaps
        set server-identity-check enable
        set source-ip "192.0.2.10"
        set source-port 636
        set group-filter "(&(objectClass=group)(member={0}))"
        set group-search-base "ou=Groups,dc=example,dc=test"
        set group-member-check user-attr
        set group-object-filter "(objectClass=group)"
        set member-attr member
        set password-attr userPassword
        set obtain-user-info enable
        set search-type recursive nested
        set ssl-min-proto-version tls1-2
    next
end
config user saml
    edit "corp-saml"
        set entity-id "https://fw.example.test/saml"
        set single-sign-on-url "https://fw.example.test/login"
        set single-logout-url "https://fw.example.test/logout"
        set idp-entity-id "https://idp.example.test"
        set idp-single-sign-on-url "https://idp.example.test/sso"
        set idp-single-logout-url "https://idp.example.test/slo"
        set idp-cert "IDP_CERT"
        set user-name "username"
        set group-name "group"
        set digest-method sha256
    next
end
config user local
    edit "local-review"
        set status enable
        set type password
        set passwd "{SECRET}"
        set seed "{SECRET}"
        set activation-code "{SECRET}"
    next
end
config user fsso
    edit "corp-fsso"
        set server "10.10.10.10"
        set password "{SECRET}"
        set custom-option test
    next
end
config user adgrp
    edit "CORP/DOMAIN USERS"
        set server-name "corp-fsso"
    next
    edit "CORP/IT"
        set server-name "corp-fsso"
    next
end
config user group
    edit "remote-users"
        set type firewall
        set member "local-review"
        set authtimeout 30
        config match
            edit 1
                set server-name "corp-ldap"
                set group-name "CN=VPN,DC=example,DC=test"
            next
        end
    next
    edit "corp-fsso-users"
        set group-type fsso-service
        set member "CORP/DOMAIN USERS" "CORP/IT"
    next
end
config firewall policy
    edit 100
        set name "FSSO policy"
        set srcintf "any"
        set dstintf "any"
        set srcaddr "all"
        set dstaddr "all"
        set groups "corp-fsso-users"
        set action accept
        set schedule "always"
        set service "ALL"
    next
end
config vpn ssl web portal
    edit "full-access"
        set tunnel-mode enable
        set ipv6-tunnel-mode enable
        set ip-pools "SSLVPN_POOL"
        set ipv6-pools "SSLVPN_POOL6"
        set split-tunneling enable
        set limit-user-logins enable
        set forticlient-download disable
        set display-bookmark enable
        set host-check custom
        set host-check-policy "endpoint-agent"
    next
end
config vpn ssl web host-check-software
    edit "endpoint-agent"
        set type antivirus
        set guid "agent-guid"
        set version "1.2.3"
        set os-type windows
    next
end
config vpn ssl settings
    set status enable
    set ssl-min-proto-ver tls1-2
    set banned-cipher RSA 3DES
    set servercert "VPN_CERT"
    set source-interface "wan1" "wan2"
    set source-address "allowed-admins"
    set tunnel-ip-pools "SSLVPN_POOL"
    set default-portal "full-access"
    set login-attempt-limit 3
    config authentication-rule
        edit 1
            set groups "remote-users"
            set portal "full-access"
            set client-cert enable
        next
    end
end
config firewall DoS-policy
    edit 10
        set name "Protect published service"
        set status enable
        set interface "wan1"
        set srcaddr "all"
        set dstaddr "VIP_A"
        set service "ALL"
        set comments "Protect published service"
        set policyid 10
        config anomaly
            edit "tcp_syn_flood"
                set status enable
                set log enable
                set action block
                set threshold 2000
                set threshold(default) 1000
                set quarantine attacker
                set quarantine-expiry 10m
                set quarantine-log enable
            next
        end
    next
end
config firewall sniffer
    edit 5
        set uuid "sniffer-uuid"
        set logtraffic all
        set ipv6 enable
        set non-ip disable
        set application-list-status enable
        set application-list "monitor-apps"
        set ips-sensor-status enable
        set ips-sensor "strict-ips"
        set av-profile-status enable
        set av-profile "strict-av"
        set webfilter-profile-status enable
        set webfilter-profile "strict-web"
        set interface "port1"
    next
end
config authentication scheme
    edit "browser-auth"
        set method basic
        set user-database "remote-users"
        set require-tfa enable
    next
end
config authentication rule
    edit "admin-auth"
        set srcintf "wan1" "wan2"
        set srcaddr "allowed-admins"
        set active-auth-method "browser-auth"
        set protocol http
    next
end
"""


SSL_VPN_FIDELITY_CONFIG = """
config firewall address
    edit "Allowed-Source"
        set subnet 192.0.2.0 255.255.255.0
    next
    edit "SSL_POOL"
        set subnet 10.20.0.0 255.255.255.0
    next
end
config user group
    edit "VPN-Users"
    next
end
config vpn ssl web host-check-software
    edit "FortiClient-AV"
        set guid "11111111-2222-3333-4444-555555555555"
    next
    edit "Custom-FW"
        set type fw
        set os-type windows
        set version "10"
        set guid "AAAAAAAA-BBBB-CCCC-DDDD-EEEEEEEEEEEE"
        set custom-setting test
        config check-item-list
            edit 1
                set action require
                set type process
                set target "agent.exe"
                set version "1.2.3"
            next
            edit 2
                set action deny
                set type file
                set target "bad.exe"
                set md5s "abc" "def"
                set custom-nested foo
            next
        end
    next
end
config vpn ssl web portal
    edit "full-access"
        set tunnel-mode enable
        set ip-pools "SSL_POOL"
        set host-check custom
        set host-check-policy "FortiClient-AV"
        set host-check-interval 120
    next
    edit "portal-bad"
        set host-check custom
        set host-check-policy "Missing-A" "FortiClient-AV"
    next
end
config vpn ssl settings
    set status disable
    set ssl-min-proto-ver tls1-1
    set banned-cipher SHA1 SHA256 SHA384
    set servercert ''
    set login-block-time 300
    set tunnel-ip-pools "SSL_POOL"
    set dns-server1 192.0.2.53
    set dns-server2 198.51.100.53
    set source-interface "wan1" "wan2"
    set source-address "Allowed-Source"
    set default-portal "full-access"
    config authentication-rule
        edit 1
            set groups "VPN-Users"
            set portal "missing-portal"
            set users "alice" "bob"
            set source-interface "wan1" "wan2"
            set source-address "Allowed-Source"
            set client-cert enable
            set realm "employees"
            set custom-rule-value retained
        next
    end
end
"""


IPV6_DOS_CONFIG = """
config firewall DoS-policy6
    edit 20
        set name "IPv6 DoS protection"
        set status enable
        set interface "wan1"
        set srcaddr "IPv6_Source_A" "IPv6_Source_B"
        set dstaddr "IPv6_Server"
        set service "HTTPS" "SSH"
        config anomaly
            edit "tcp_syn_flood"
                set status enable
                set log enable
                set action block
                set quarantine attacker
                set quarantine-expiry 10m
                set quarantine-log enable
                set threshold 3000
                set threshold(default) 1000
            next
            edit "udp_flood"
                set status enable
                set threshold 5000
                set threshold(default) 2000
            next
        end
    next
end
"""


def test_identity_inventory_strips_credentials_and_preserves_safe_metadata():
    fg = parse_fortigate_config(ACCESS_CONFIG)
    result = extract_fortigate_config(ACCESS_CONFIG)
    ir = result.canonical_ir

    assert fg.user_ldap_servers[0].has_password is True
    assert fg.local_users[0].has_password is True
    assert len(fg.fsso_servers) == 1
    assert len(fg.ad_groups) == 2
    assert SECRET not in fg.model_dump_json()
    assert SECRET not in ir.model_dump_json()
    assert SECRET not in result.model_dump_json()

    ldap = ir.user_ldap_servers[0]
    assert (ldap.server, ldap.cnid, ldap.dn, ldap.username) == (
        "ldap.example.test", "sAMAccountName", "dc=example,dc=test", "bind-user"
    )
    assert ldap.has_password is True
    assert ldap.secure == "ldaps"
    assert (
        ldap.secondary_server, ldap.tertiary_server, ldap.port,
        ldap.server_identity_check, ldap.source_ip, ldap.source_port,
        ldap.group_filter, ldap.group_search_base, ldap.group_member_check,
        ldap.group_object_filter, ldap.member_attr, ldap.password_attr,
        ldap.obtain_user_info, ldap.search_type, ldap.ssl_min_proto_version,
    ) == (
        "ldap2.example.test", "ldap3.example.test", 636, "enable",
        "192.0.2.10", 636, "(&(objectClass=group)(member=0))",
        "ou=Groups,dc=example,dc=test", "user-attr", "(objectClass=group)",
        "member", "userPassword", "enable", ["recursive", "nested"], "tls1-2",
    )
    assert ldap.client_certificate_resolved is None
    assert ldap.source_attributes == {}
    saml = ir.user_saml_servers[0]
    assert saml.idp_single_sign_on_url == "https://idp.example.test/sso"
    assert saml.idp_cert == "IDP_CERT"
    group = ir.user_groups[0]
    assert group.group_type == "firewall"
    assert group.members == ["local-review"]
    assert group.matches[0].server_name == "corp-ldap"
    assert group.matches[0].group_name == "CN=VPN,DC=example,DC=test"
    assert group.authtimeout == 30
    assert group.source_attributes == {}

    provider = fg.fsso_servers[0]
    assert provider.name == "corp-fsso"
    assert provider.server == "10.10.10.10"
    assert provider.has_password is True
    assert provider.extra_settings == {"custom_option": "test"}
    assert fg.ad_groups[0].name == "CORP/DOMAIN USERS"
    assert fg.ad_groups[0].server_name == "corp-fsso"

    assert len(ir.fsso_providers) == 1
    assert len(ir.fsso_ad_groups) == 2
    ir_provider = ir.fsso_providers[0]
    assert ir_provider.name == "corp-fsso"
    assert ir_provider.server == "10.10.10.10"
    assert ir_provider.has_password is True
    ad_group = ir.fsso_ad_groups[0]
    assert ad_group.name == "CORP/DOMAIN USERS"
    assert ad_group.provider_name == "corp-fsso"
    assert ad_group.provider_resolved is True
    assert ad_group.migration_status == "EXTRACT_ONLY"
    fsso_user_group = next(
        item for item in ir.user_groups if item.name == "corp-fsso-users"
    )
    assert fsso_user_group.group_type == "fsso-service"
    assert fsso_user_group.members == ["CORP/DOMAIN USERS", "CORP/IT"]
    assert ir.policies[0].source_user_groups == ["corp-fsso-users"]

    coverage = {item.path: item for item in result.source_sections}
    for path, count in (("user fsso", 1), ("user adgrp", 2)):
        assert coverage[path].status.value == "EXTRACT_ONLY"
        assert coverage[path].object_count_source == count
        assert coverage[path].object_count_parsed == count
        assert coverage[path].object_count_normalized == count


def test_ssl_vpn_dos_sniffer_and_authentication_stay_separate_inventory():
    result = extract_fortigate_config(ACCESS_CONFIG)
    fg = parse_fortigate_config(ACCESS_CONFIG)
    assert fg.ssl_vpn_portals[0].host_checks == []
    assert isinstance(fg.ssl_vpn_host_check_software[0], FGSSLVPNHostCheckSoftware)
    assert fg.ssl_vpn_host_check_software[0].os_type == "windows"

    ir = result.canonical_ir
    portal = ir.ssl_vpn_portals[0]
    assert portal.ip_pools == ["SSLVPN_POOL"]
    assert portal.ipv6_pools == ["SSLVPN_POOL6"]
    assert portal.split_tunneling == "enable"
    assert portal.host_check == "custom"
    assert portal.host_check_policies == ["endpoint-agent"]
    assert portal.unresolved_host_check_policies == []
    host_check = ir.ssl_vpn_host_checks[0]
    assert host_check.name == "endpoint-agent"
    assert host_check.check_type == "antivirus"
    assert host_check.os_type == "windows"
    assert host_check.version == "1.2.3"
    assert host_check.guid == "agent-guid"
    assert host_check.migration_status == "EXTRACT_ONLY"
    assert host_check.requires_manual_review is True
    assert ir.vpn_tunnels == []

    settings = ir.ssl_vpn_settings
    assert settings is not None
    assert settings.status == "enable"
    assert settings.ssl_min_proto_ver == "tls1-2"
    assert settings.banned_cipher == ["RSA", "3DES"]
    assert settings.source_interfaces == ["wan1", "wan2"]
    assert settings.authentication_rules[0].groups == ["remote-users"]
    assert settings.authentication_rules[0].portal == "full-access"
    assert settings.authentication_rules[0].client_cert == "enable"
    assert settings.authentication_rules[0].source_attributes == {}

    assert fg.dos_policies[0].name == "Protect published service"
    dos = ir.dos_policies[0]
    assert dos.source_id == 10
    assert dos.name == "Protect published service"
    assert dos.source_addresses == ["all"]
    assert dos.destination_addresses == ["VIP_A"]
    assert dos.services == ["ALL"]
    assert dos.anomalies[0].name == "tcp_syn_flood"
    assert dos.anomalies[0].threshold == 2000
    assert dos.anomalies[0].threshold_default == 1000
    assert dos.anomalies[0].quarantine == "attacker"
    assert dos.anomalies[0].quarantine_expiry == "10m"
    assert dos.anomalies[0].quarantine_log == "enable"
    assert dos.anomalies[0].source_attributes == {}

    sniffer = ir.firewall_sniffers[0]
    assert sniffer.source_id == 5
    assert sniffer.application_list == "monitor-apps"
    assert sniffer.ips_sensor == "strict-ips"
    assert sniffer.av_profile == "strict-av"
    assert sniffer.webfilter_profile == "strict-web"
    assert sniffer.source_attributes == {"interface": "port1"}
    assert [policy.name for policy in ir.policies] == ["FSSO policy"]

    assert ir.authentication_schemes[0].user_database == "remote-users"
    assert ir.authentication_schemes[0].source_attributes == {"require_tfa": "enable"}
    rule = ir.authentication_rules[0]
    assert rule.source_interfaces == ["wan1", "wan2"]
    assert rule.source_addresses == ["allowed-admins"]
    assert rule.active_auth_method == "browser-auth"
    assert rule.source_attributes == {"protocol": "http"}


def test_access_inventory_excel_contains_no_credentials():
    result = extract_fortigate_config(ACCESS_CONFIG)
    workbook = load_workbook(
        io.BytesIO(
            IRExcelExporter(
                result.canonical_ir,
                extraction_result=result,
            ).generate()
        )
    )
    all_text = "\n".join(
        str(cell.value)
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if cell.value is not None
    )
    assert SECRET not in all_text
    for sheet_name in (
        "LDAP Servers",
        "SAML Servers",
        "FSSO Servers",
        "FSSO AD Groups",
        "Local Users",
        "User Groups",
        "User Group Matches",
        "SSL VPN Settings",
        "SSL VPN Portals",
        "SSL VPN Authentication Rules",
        "SSL VPN Host Checks",
        "SSL VPN Host Check Items",
        "DoS Policies",
        "DoS Anomalies",
        "Firewall Sniffer",
        "Authentication Schemes",
        "Authentication Rules",
    ):
        assert sheet_name in workbook.sheetnames

    host_checks = workbook["SSL VPN Host Checks"]
    host_check_headers = {cell.value: cell.column for cell in host_checks[3]}
    assert host_checks.max_row == 4
    assert host_checks.cell(4, host_check_headers["Type"]).value == "antivirus"
    assert host_checks.cell(4, host_check_headers["OS Type"]).value == "windows"
    assert host_checks.cell(4, host_check_headers["Version"]).value == "1.2.3"
    assert host_checks.cell(4, host_check_headers["GUID"]).value == "agent-guid"
    assert host_checks.cell(4, host_check_headers["Extraction Status"]).value == "EXTRACT_ONLY"
    assert host_checks.cell(4, host_check_headers["Manual Review"]).value == "Yes"

    fsso_servers = workbook["FSSO Servers"]
    server_headers = {cell.value: cell.column for cell in fsso_servers[3]}
    assert fsso_servers.cell(4, server_headers["Name"]).value == "corp-fsso"
    assert fsso_servers.cell(4, server_headers["Server"]).value == "10.10.10.10"
    assert fsso_servers.cell(4, server_headers["Password Configured"]).value == "Yes"

    fsso_groups = workbook["FSSO AD Groups"]
    group_headers = {cell.value: cell.column for cell in fsso_groups[3]}
    assert fsso_groups.cell(4, group_headers["Name"]).value == "CORP/DOMAIN USERS"
    assert fsso_groups.cell(4, group_headers["FSSO Server"]).value == "corp-fsso"
    assert fsso_groups.cell(4, group_headers["Server Resolved"]).value == "Yes"
    assert fsso_groups.cell(4, group_headers["Extraction Status"]).value == "EXTRACT_ONLY"

    user_groups = workbook["User Groups"]
    user_group_headers = {cell.value: cell.column for cell in user_groups[3]}
    fsso_group_row = next(
        row for row in range(4, user_groups.max_row + 1)
        if user_groups.cell(row, user_group_headers["Name"]).value == "corp-fsso-users"
    )
    assert user_groups.cell(fsso_group_row, user_group_headers["Members"]).value == (
        "CORP/DOMAIN USERS\nCORP/IT"
    )

    policies = workbook["Policies"]
    policy_headers = {cell.value: cell.column for cell in policies[3]}
    assert policies.cell(4, policy_headers["User Groups"]).value == "corp-fsso-users"

    dos_policies = workbook["DoS Policies"]
    dos_headers = {cell.value: cell.column for cell in dos_policies[3]}
    assert "Policy Name" in dos_headers
    assert dos_policies.cell(4, dos_headers["Policy Name"]).value == "Protect published service"

    anomalies = workbook["DoS Anomalies"]
    anomaly_headers = {cell.value: cell.column for cell in anomalies[3]}
    assert "Quarantine" in anomaly_headers
    assert "Quarantine Expiry" in anomaly_headers
    assert "Quarantine Log" in anomaly_headers
    assert anomalies.cell(4, anomaly_headers["Quarantine"]).value == "attacker"
    assert anomalies.cell(4, anomaly_headers["Quarantine Expiry"]).value == "10m"
    assert anomalies.cell(4, anomaly_headers["Quarantine Log"]).value == "enable"
    assert "Threshold" in anomaly_headers
    assert "Default Threshold" in anomaly_headers
    assert anomalies.cell(4, anomaly_headers["Threshold"]).value == 2000
    assert anomalies.cell(4, anomaly_headers["Default Threshold"]).value == 1000

    summary = {
        workbook["Summary"].cell(row, 1).value:
        workbook["Summary"].cell(row, 2).value
        for row in range(1, workbook["Summary"].max_row + 1)
    }
    assert summary["FSSO Servers"] == 1
    assert summary["FSSO AD Groups"] == 2
    assert summary["SSL VPN Host Checks"] == 1


def test_fortigate_ipv6_dos_nested_anomalies_preserve_typed_values_and_excel_rows():
    result = extract_fortigate_config(IPV6_DOS_CONFIG)
    fg = parse_fortigate_config(IPV6_DOS_CONFIG)
    policy = result.canonical_ir.dos_policies[0]

    assert fg.dos_policies[0].address_family == "ipv6"
    assert policy.source_id == 20
    assert policy.name == "IPv6 DoS protection"
    assert policy.address_family == "ipv6"
    assert policy.source_addresses == ["IPv6_Source_A", "IPv6_Source_B"]
    assert policy.destination_addresses == ["IPv6_Server"]
    assert policy.services == ["HTTPS", "SSH"]
    assert len(policy.anomalies) == 2

    syn, udp = policy.anomalies
    assert syn.name == "tcp_syn_flood"
    assert syn.status == "enable"
    assert syn.log == "enable"
    assert syn.action == "block"
    assert syn.quarantine == "attacker"
    assert syn.quarantine_expiry == "10m"
    assert syn.quarantine_log == "enable"
    assert syn.threshold == 3000
    assert syn.threshold_default == 1000
    assert "threshold(default)" not in syn.source_attributes
    assert "threshold_default" not in syn.source_attributes
    assert udp.threshold == 5000
    assert udp.threshold_default == 2000

    workbook = load_workbook(
        io.BytesIO(IRExcelExporter(result.canonical_ir, extraction_result=result).generate())
    )
    sheet = workbook["DoS Anomalies"]
    headers = {cell.value: cell.column for cell in sheet[3]}
    assert "Threshold" in headers
    assert "Default Threshold" in headers
    assert sheet.max_row == 5
    assert sheet.cell(4, headers["Threshold"]).value == 3000
    assert sheet.cell(4, headers["Default Threshold"]).value == 1000
    assert sheet.cell(5, headers["Threshold"]).value == 5000
    assert sheet.cell(5, headers["Default Threshold"]).value == 2000


def test_fortigate_dos_policy_name_is_optional():
    result = extract_fortigate_config(
        """config firewall DoS-policy6
    edit 10
        set status enable
        config anomaly
            edit "no-quarantine"
                set status enable
            next
        end
    next
end
"""
    )

    assert result.canonical_ir.dos_policies[0].name is None
    workbook = load_workbook(io.BytesIO(IRExcelExporter(result.canonical_ir).generate()))
    headers = {cell.value: cell.column for cell in workbook["DoS Policies"][3]}
    assert workbook["DoS Policies"].cell(4, headers["Policy Name"]).value is None
    anomaly_headers = {cell.value: cell.column for cell in workbook["DoS Anomalies"][3]}
    assert workbook["DoS Anomalies"].cell(4, anomaly_headers["Quarantine"]).value is None
    assert workbook["DoS Anomalies"].cell(4, anomaly_headers["Quarantine Expiry"]).value is None
    assert workbook["DoS Anomalies"].cell(4, anomaly_headers["Quarantine Log"]).value is None


def test_top_level_ssl_vpn_host_checks_and_nested_items_preserve_source_fidelity():
    fg = parse_fortigate_config(SSL_VPN_FIDELITY_CONFIG)
    assert len(fg.ssl_vpn_host_check_software) == 2
    first, custom = fg.ssl_vpn_host_check_software
    assert first.name == "FortiClient-AV"
    assert first.guid == "11111111-2222-3333-4444-555555555555"
    assert custom.name == "Custom-FW"
    assert custom.type == "fw"
    assert custom.os_type == "windows"
    assert custom.version == "10"
    assert custom.guid == "AAAAAAAA-BBBB-CCCC-DDDD-EEEEEEEEEEEE"
    assert custom.extra_settings == {"custom_setting": "test"}
    assert len(custom.check_items) == 2
    assert custom.check_items[0].action == "require"
    assert custom.check_items[0].type == "process"
    assert custom.check_items[0].target == "agent.exe"
    assert custom.check_items[0].version == "1.2.3"
    assert custom.check_items[1].md5s == ["abc", "def"]
    assert custom.check_items[1].extra_settings == {"custom_nested": "foo"}

    result = extract_fortigate_config(SSL_VPN_FIDELITY_CONFIG)
    ir = result.canonical_ir
    assert len(ir.ssl_vpn_host_checks) == 2
    ir_custom = ir.ssl_vpn_host_checks[1]
    assert ir_custom.check_type == "fw"
    assert ir_custom.os_type == "windows"
    assert ir_custom.migration_status == "EXTRACT_ONLY"
    assert ir_custom.requires_manual_review is True
    assert ir_custom.source_attributes == {"custom_setting": "test"}
    assert ir_custom.check_items[1].md5s == ["abc", "def"]
    assert ir_custom.check_items[1].source_attributes == {"custom_nested": "foo"}

    coverage = {section.path: section for section in result.source_sections}
    host_checks = coverage["vpn ssl web host-check-software"]
    assert (
        host_checks.object_count_source,
        host_checks.object_count_parsed,
        host_checks.object_count_normalized,
        host_checks.status.value,
    ) == (2, 2, 2, "EXTRACT_ONLY")
    items = coverage["vpn ssl web host-check-software check-item-list"]
    assert (items.object_count_source, items.object_count_parsed, items.object_count_normalized) == (2, 2, 2)
    settings = coverage["vpn ssl settings"]
    assert (settings.object_count_source, settings.object_count_parsed, settings.object_count_normalized) == (1, 1, 1)


def test_ssl_vpn_references_settings_and_auth_rules_are_preserved_and_audited():
    result = extract_fortigate_config(SSL_VPN_FIDELITY_CONFIG)
    ir = result.canonical_ir
    good, bad = ir.ssl_vpn_portals
    assert good.host_check == "custom"
    assert good.host_check_policies == ["FortiClient-AV"]
    assert good.host_check_interval == 120
    assert good.unresolved_host_check_policies == []
    assert bad.host_check_policies == ["Missing-A", "FortiClient-AV"]
    assert bad.unresolved_host_check_policies == ["Missing-A"]
    assert any(
        entry.id == "ssl-vpn-portal:portal-bad:host-check-policy"
        and "Missing-A" in entry.message
        for entry in ir.audit_entries
    )

    settings = ir.ssl_vpn_settings
    assert settings is not None
    assert settings.status == "disable"
    assert settings.ssl_min_proto_ver == "tls1-1"
    assert settings.banned_cipher == ["SHA1", "SHA256", "SHA384"]
    assert settings.server_certificate == ""
    assert settings.server_certificate_configured is True
    assert settings.login_block_time == 300
    assert settings.dns_server1 == "192.0.2.53"
    assert settings.dns_server2 == "198.51.100.53"
    assert settings.source_interfaces == ["wan1", "wan2"]
    assert settings.source_addresses == ["Allowed-Source"]
    assert settings.tunnel_ip_pools == ["SSL_POOL"]
    assert settings.default_portal == "full-access"
    rule = settings.authentication_rules[0]
    assert rule.groups == ["VPN-Users"]
    assert rule.portal == "missing-portal"
    assert rule.users == ["alice", "bob"]
    assert rule.source_interfaces == ["wan1", "wan2"]
    assert rule.source_addresses == ["Allowed-Source"]
    assert rule.client_cert == "enable"
    assert rule.realm == "employees"
    assert rule.source_attributes == {"custom_rule_value": "retained"}
    assert any(
        entry.id == "ssl-vpn-auth-rule:1:portal"
        and "missing-portal" in entry.message
        for entry in ir.audit_entries
    )


def test_ssl_vpn_fidelity_excel_has_top_level_host_checks_items_and_empty_certificate_state():
    result = extract_fortigate_config(SSL_VPN_FIDELITY_CONFIG)
    workbook = load_workbook(io.BytesIO(IRExcelExporter(result.canonical_ir, extraction_result=result).generate()))
    checks = workbook["SSL VPN Host Checks"]
    check_headers = {cell.value: cell.column for cell in checks[3]}
    assert checks.max_row == 5
    assert checks.cell(5, check_headers["Name"]).value == "Custom-FW"
    assert checks.cell(5, check_headers["GUID"]).value == "AAAAAAAA-BBBB-CCCC-DDDD-EEEEEEEEEEEE"
    assert checks.cell(5, check_headers["Check Item Count"]).value == 2
    items = workbook["SSL VPN Host Check Items"]
    item_headers = {cell.value: cell.column for cell in items[3]}
    assert items.max_row == 5
    assert items.cell(5, item_headers["MD5s"]).value == "abc\ndef"
    settings = workbook["SSL VPN Settings"]
    setting_headers = {cell.value: cell.column for cell in settings[3]}
    assert settings.cell(4, setting_headers["Server Certificate"]).value is None
    assert settings.cell(4, setting_headers["Server Certificate Configured"]).value == "TRUE"
    assert settings.cell(4, setting_headers["DNS Server 1"]).value == "192.0.2.53"
    assert settings.cell(4, setting_headers["Login Block Time"]).value == 300


def test_missing_fsso_identity_references_are_preserved_and_audited():
    config = """
config user adgrp
    edit "CORP/MISSING"
        set server-name "missing-fsso"
    next
end
config user group
    edit "broken-fsso-users"
        set group-type fsso-service
        set member "CORP/MISSING" "CORP/UNKNOWN"
    next
end
"""
    ir = extract_fortigate_config(config).canonical_ir

    ad_group = ir.fsso_ad_groups[0]
    assert ad_group.name == "CORP/MISSING"
    assert ad_group.provider_name == "missing-fsso"
    assert ad_group.provider_resolved is False
    assert ad_group.requires_manual_review is True
    assert ir.user_groups[0].members == ["CORP/MISSING", "CORP/UNKNOWN"]

    messages = [entry.message for entry in ir.audit_entries]
    assert (
        "FSSO AD group 'CORP/MISSING' references missing FSSO provider "
        "'missing-fsso'."
    ) in messages
    assert (
        "User group 'broken-fsso-users' references missing FSSO AD group "
        "'CORP/UNKNOWN'."
    ) in messages


def test_fsso_endpoints_and_polling_are_typed_and_redacted():
    result = extract_fortigate_config(FSSO_PHASE5_CONFIG)
    ir = result.canonical_ir

    assert [(item.index, item.server, item.port, item.has_password) for item in ir.fsso_providers[0].endpoints] == [
        (1, "192.0.2.10", None, True),
        (2, "192.0.2.11", 8001, True),
    ]
    assert ir.fsso_providers[0].sni == "collector.example.test"
    assert ir.fsso_polling[0].ad_groups[0].name == "DOMAIN/Users"
    assert ir.fsso_polling[0].polling_frequency == 5
    assert "secret-" not in result.model_dump_json()
    workbook = load_workbook(io.BytesIO(IRExcelExporter(ir, extraction_result=result).generate()))
    assert workbook["FSSO Polling"]["A4"].value == "1"
    assert workbook["FSSO Polling"]["N4"].value == "DOMAIN/Users"


def test_phase18_fsso_provider_directory_and_repeated_options_are_preserved():
    parsed = parse_fortigate_config(FSSO_PHASE18_CONFIG)

    provider = parsed.fsso_servers[0]
    assert [(item.server, item.port, item.has_password) for item in provider.endpoints] == [
        ("collector.example.test", 8000, True),
        ("collector-2.example.test", 8001, True),
    ]
    assert (provider.ldap_poll, provider.ldap_poll_filter) == (
        "enable", "(objectCategory=group)"
    )
    assert (provider.ldap_poll_interval, provider.group_poll_interval) == (60, 30)
    assert (provider.source_ip, provider.source_ip6, provider.interface) == (
        "192.0.2.20", "2001:db8::20", "mgmt"
    )
    assert provider.extra_settings["vrf_select"] == "7"
    assert provider.extra_settings["vendor_option"] == ["first", "second"]

    polling = parsed.fsso_polling[0]
    assert (polling.server, polling.default_domain, polling.ldap_server) == (
        "dc-1.example.test", "EXAMPLE", "ldap.example.test"
    )
    assert (polling.logon_history, polling.polling_frequency) == (24, 5)
    assert [group.name for group in polling.ad_groups] == [
        "EXAMPLE/Domain Users", "EXAMPLE/Network Admins"
    ]
    assert polling.extra_settings["vendor_option"] == ["first", "second"]
    assert polling.ad_groups[0].extra_settings["vendor_option"] == ["one", "two"]
    assert "SECRET" not in parsed.model_dump_json()


def test_local_user_and_group_semantics_preserve_typed_fields_and_unresolved_members():
    result = extract_fortigate_config(LOCAL_GROUP_PHASE6_CONFIG)
    local = result.canonical_ir.local_users[0]
    group = result.canonical_ir.user_groups[0]

    assert (local.id, local.tacacs_server, local.ppk_identity, local.has_ppk_secret) == (
        7, "tacacs", "alice-key", True
    )
    assert group.auth_concurrent_value == 3
    assert group.guests[0].name == "Guest"
    assert group.guests[0].has_password is True
    assert group.unresolved_members == ["missing-user"]
    assert "secret-ppk" not in result.model_dump_json()
    assert "secret-guest" not in result.model_dump_json()
