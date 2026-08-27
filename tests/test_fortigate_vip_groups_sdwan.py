import io

from openpyxl import load_workbook

from fwmigrate.extraction.models import ExtractionStatus
from fwmigrate.ir.enums import AddressType
from fwmigrate.parsers.fortigate.extractor import extract_fortigate_config
from fwmigrate.report.excel_exporter import IRExcelExporter


def test_vip_group_inventory_preserves_source_fields_without_creating_nat():
    result = extract_fortigate_config("""
config firewall vipgrp
    edit "published-services"
        set uuid "vip-group-uuid"
        set interface "wan1"
        set member "VIP_A" "VIP_B"
        set color 9
        set comments "Published services"
        set visibility enable
    next
end
""")
    group = result.canonical_ir.virtual_ip_groups[0]
    assert group.name == "published-services"
    assert group.source_uuid == "vip-group-uuid"
    assert group.interface == "wan1"
    assert group.members == ["VIP_A", "VIP_B"]
    assert group.source_color == 9
    assert group.description == "Published services"
    assert group.source_attributes == {"visibility": "enable"}
    assert group.migration_status == "EXTRACT_ONLY"
    assert result.canonical_ir.nat_rules == []


def test_ipv6_nat_resources_are_extraction_only_and_do_not_create_ipv4_nat():
    result = extract_fortigate_config("""
config firewall ippool6
    edit "POOL6"
        set startip 2001:db8:1::10
        set endip 2001:db8:1::20
        set nat46 enable
        set add-nat46-route enable
        set custom-pool6-setting "retained"
    next
end
config firewall vip6
    edit "VIP6"
        set uuid "vip6-uuid"
        set extip 2001:db8:2::10
        set mappedip 2001:db8:3::10 2001:db8:3::11
        set nat64 enable
        set nat66 enable
        set add-nat64-route enable
        set ndp-reply enable
        set ipv4-mappedip 192.0.2.10
        set ipv4-mappedport 8443
        set embedded-ipv4-address enable
        set custom-vip6-setting "retained"
    next
end
config firewall vipgrp6
    edit "VIP_GROUP6"
        set uuid "vipgrp6-uuid"
        set member "VIP6"
        set comments "IPv6 published services"
        set custom-group6-setting "retained"
    next
end
""")

    pool = result.canonical_ir.ip_pools[0]
    assert pool.address_family == "ipv6"
    assert (pool.start_ip, pool.end_ip) == ("2001:db8:1::10", "2001:db8:1::20")
    assert pool.nat46 is True
    assert pool.add_nat46_route is True
    assert pool.source_attributes == {"custom_pool6_setting": "retained"}
    assert pool.migration_status == "EXTRACT_ONLY"
    assert pool.requires_manual_review is True

    vip = result.canonical_ir.virtual_ips[0]
    assert vip.address_family == "ipv6"
    assert vip.external_ip == "2001:db8:2::10"
    assert vip.mapped_ips == ["2001:db8:3::10", "2001:db8:3::11"]
    assert (vip.nat64, vip.nat66, vip.add_nat64_route, vip.ndp_reply) == (
        True, True, True, True
    )
    assert vip.ipv4_mapped_ip == "192.0.2.10"
    assert vip.extra_settings == {"custom_vip6_setting": "retained"}
    assert vip.migration_status == "EXTRACT_ONLY"
    assert vip.requires_manual_review is True

    group = result.canonical_ir.virtual_ip_groups[0]
    assert group.address_family == "ipv6"
    assert group.members == ["VIP6"]
    assert group.description == "IPv6 published services"
    assert group.source_attributes == {"custom_group6_setting": "retained"}
    assert group.migration_status == "EXTRACT_ONLY"
    assert result.canonical_ir.nat_rules == []

    statuses = {section.path: section.status for section in result.source_sections}
    assert statuses["firewall ippool6"] == ExtractionStatus.EXTRACT_ONLY
    assert statuses["firewall vip6"] == ExtractionStatus.EXTRACT_ONLY
    assert statuses["firewall vipgrp6"] == ExtractionStatus.EXTRACT_ONLY
    inventory_paths = {item.source_path for item in result.inventory_items}
    assert {
        "firewall ippool6", "firewall vip6", "firewall vipgrp6"
    }.issubset(inventory_paths)

    workbook = load_workbook(
        io.BytesIO(IRExcelExporter(result.canonical_ir, result).generate())
    )
    for sheet_name in ("IP Pools", "Virtual IPs", "VIP Groups"):
        sheet = workbook[sheet_name]
        headers = {cell.value: cell.column for cell in sheet[3]}
        assert sheet.cell(4, headers["Address Family"]).value == "ipv6"
        assert sheet.cell(4, headers["Extraction Status"]).value == "EXTRACT_ONLY"
        assert sheet.cell(4, headers["Manual Review"]).value == "TRUE"


def test_named_multicast_ranges_and_special_names_survive():
    result = extract_fortigate_config("""
config firewall multicast-address
    edit "EIGRP"
        set start-ip 224.0.0.10
        set end-ip 224.0.0.10
    next
    edit "OSPF"
        set start-ip 224.0.0.5
        set end-ip 224.0.0.6
    next
    edit "all"
        set start-ip 224.0.0.0
        set end-ip 239.255.255.255
    next
    edit "none"
        set start-ip 0.0.0.0
        set end-ip 0.0.0.0
    next
end
""")
    addresses = {item.name: item for item in result.canonical_ir.addresses}
    assert addresses["EIGRP"].type == AddressType.HOST
    assert addresses["EIGRP"].subnet == "224.0.0.10/32"
    assert addresses["OSPF"].ip_range_start == "224.0.0.5"
    assert addresses["OSPF"].ip_range_end == "224.0.0.6"
    assert addresses["all"].type == AddressType.SPECIAL
    assert addresses["all"].value == "all"
    assert addresses["all"].is_multicast is True
    assert addresses["all"].source_attributes == {
        "start_ip": "224.0.0.0",
        "end_ip": "239.255.255.255",
    }
    assert addresses["none"].type == AddressType.SPECIAL
    assert addresses["none"].value == "none"
    assert addresses["none"].is_multicast is True
    assert addresses["none"].requires_manual_review is True


def test_full_sdwan_inventory_preserves_nested_source_semantics():
    config = """
config system sdwan
    set status enable
    set load-balance-mode source-ip-based
    set duplication-max-num 2
    config zone
        edit "Internet"
            set minimum-sla-meet-members 1
        next
    end
    config members
        edit 1
            set interface "wan1"
            set zone "Internet"
            set gateway 192.0.2.1
            set weight 20
            set priority 5
            set cost 10
        next
    end
    config health-check
        edit "internet-sla"
            set server "1.1.1.1"
            set members 1 bad-member
            set interval 500
            set failtime 3
            config sla
                edit 1
                    set latency-threshold 100
                    set jitter-threshold 20
                next
            end
        next
    end
    config service
        edit 10
            set name "preferred-web"
            set mode priority
            set src "LAN_NET"
            set dst "all"
            set health-check "internet-sla"
            set priority-members 1
            set internet-service enable
            set internet-service-name "Microsoft-Office365"
            set internet-service-app-ctrl 12345
            set use-shortcut-sla enable
            set tie-break fib-best-match
        next
    end
end
"""
    result = extract_fortigate_config(config)
    sdwan = result.canonical_ir.sdwan
    assert sdwan is not None
    assert sdwan.status == "enable"
    assert sdwan.load_balance_mode == "source-ip-based"
    assert sdwan.source_attributes == {"duplication_max_num": "2"}
    assert sdwan.migration_status == "EXTRACT_ONLY"
    assert sdwan.requires_manual_review is True

    assert sdwan.zones[0].name == "Internet"
    assert sdwan.zones[0].source_attributes == {"minimum_sla_meet_members": "1"}
    member = sdwan.members[0]
    assert (member.source_id, member.interface, member.zone) == (1, "wan1", "Internet")
    assert (member.gateway, member.weight, member.priority) == ("192.0.2.1", 20, 5)
    assert member.source_attributes == {"cost": "10"}

    check = sdwan.health_checks[0]
    assert check.server == "1.1.1.1"
    assert check.member_ids == [1]
    assert check.interval == 500
    assert check.source_attributes == {
        "failtime": "3",
        "unparsed_members": ["bad-member"],
    }
    assert check.sla[0].source_id == 1
    assert check.sla[0].source_attributes == {
        "jitter_threshold": "20",
        "latency_threshold": "100",
    }

    rule = sdwan.rules[0]
    assert rule.source_id == 10
    assert rule.name == "preferred-web"
    assert rule.mode == "priority"
    assert rule.source_addresses == ["LAN_NET"]
    assert rule.destination_addresses == ["all"]
    assert rule.health_check == "internet-sla"
    assert rule.priority_member_ids == [1]
    assert rule.internet_service == "enable"
    assert rule.internet_service_names == ["Microsoft-Office365"]
    assert rule.internet_service_app_ctrl == [12345]
    assert rule.use_shortcut_sla == "enable"
    assert rule.source_attributes == {"tie_break": "fib-best-match"}

    statuses = {item.path: item.status for item in result.source_sections}
    assert statuses["system sdwan"] == ExtractionStatus.EXTRACT_ONLY
    assert statuses["system sdwan health-check sla"] == ExtractionStatus.EXTRACT_ONLY

    workbook = load_workbook(
        io.BytesIO(IRExcelExporter(result.canonical_ir, result).generate())
    )
    for sheet_name in (
        "SD-WAN", "SD-WAN Members", "SD-WAN Health Checks", "SD-WAN SLAs", "SD-WAN Rules"
    ):
        assert sheet_name in workbook.sheetnames
    assert workbook["SD-WAN Members"]["D4"].value == "192.0.2.1"
    assert workbook["SD-WAN SLAs"]["B4"].value == 1


def test_vip_group_excel_membership_is_unchanged():
    result = extract_fortigate_config("""
config firewall vipgrp
    edit "published-services"
        set member "VIP_A" "VIP_B"
    next
end
""")
    workbook = load_workbook(
        io.BytesIO(IRExcelExporter(result.canonical_ir, result).generate())
    )
    sheet = workbook["VIP Groups"]
    headers = {cell.value: cell.column for cell in sheet[3]}
    assert sheet.cell(4, headers["Members"]).value == "VIP_A\nVIP_B"
