import io

from openpyxl import load_workbook

from fwmigrate.extraction.models import ExtractionStatus
from fwmigrate.parsers.fortigate.extractor import extract_fortigate_config
from fwmigrate.parsers.fortigate.parser import parse_fortigate_config
from fwmigrate.parsers.fortigate.transformer import FGToIRTransformer
from fwmigrate.report.excel_exporter import IRExcelExporter


PRIMARY_IPV6_CONFIG = """
config system interface
    edit "v6-only"
        config ipv6
            set ip6-address 2001:0DB8:0:0::1/64
            set ip6-allowaccess ping https ssh
            set ip6-mode static
            set ip6-send-adv enable
            set ip6-manage-flag enable
            set ip6-other-flag disable
        end
    next
end
"""

SCALAR_IPV6_CONFIG = """
config system interface
    edit "v6-scalars"
        config ipv6
            set autoconf enable
            set dhcp6-client-options rapid iapd iana
            set dhcp6-information-request enable
            set dhcp6-prefix-delegation enable
            set dhcp6-relay-ip 2001:db8::1 2001:db8::2
            set icmp6-send-redirect disable
            set interface-identifier 2001:db8::10
            set ip6-default-life 1800
            set ip6-hop-limit 64
            set ip6-link-mtu 1500
            set ip6-max-interval 600
            set ip6-min-interval 198
            set ip6-prefix-mode ra
            set ip6-reachable-time 30000
            set ip6-retrans-time 1000
            set ip6-subnet 2001:db8:1::/64
            set ip6-upstream-interface wan
        end
    next
end
"""

PREFIX_AND_IAPD_CONFIG = """
config system interface
    edit "v6-ra"
        config ipv6
            config ip6-prefix-list
                edit 2001:DB8::/64
                    set autonomous-flag enable
                    set dnssl corp.example lab.example
                    set onlink-flag disable
                    set preferred-life-time 100
                    set rdnss 2001:db8::53 2001:db8::54
                    set valid-life-time 200
                next
                edit not-an-ipv6/64
                next
            end
            config ip6-delegated-prefix-list
                edit 1
                    set autonomous-flag disable
                    set delegated-prefix-iaid 7
                    set onlink-flag enable
                    set rdnss 2001:db8::53
                    set rdnss-service delegated
                    set subnet 2001:db8:1::/64
                    set upstream-interface wan
                next
            end
            config dhcp6-iapd-list
                edit 9
                    set prefix-hint 2001:db8:2::/56
                    set prefix-hint-plt 0
                    set prefix-hint-vlt bad
                next
                edit bad-iaid
                    set prefix-hint 2001:db8:3::/56
                next
            end
        end
    next
end
"""


def _interface(config, name):
    return next(item for item in config.interfaces if item.name == name)


def test_primary_ipv6_values_are_typed_and_source_preserved():
    interface = _interface(parse_fortigate_config(PRIMARY_IPV6_CONFIG), "v6-only")

    assert interface.ip6_address == "2001:0DB8:0:0::1/64"
    assert interface.ip6_allowaccess == ["ping", "https", "ssh"]
    assert interface.ip6_mode == "static"
    assert interface.ip6_send_adv == "enable"
    assert interface.ip6_manage_flag == "enable"
    assert interface.ip6_other_flag == "disable"
    assert interface.ipv6_source_settings == {
        "ip6_address": "2001:0DB8:0:0::1/64",
        "ip6_allowaccess": ["ping", "https", "ssh"],
        "ip6_mode": "static",
        "ip6_send_adv": "enable",
        "ip6_manage_flag": "enable",
        "ip6_other_flag": "disable",
    }


def test_scalar_ipv6_values_are_typed_ordered_and_reviewed():
    fg_interface = _interface(parse_fortigate_config(SCALAR_IPV6_CONFIG), "v6-scalars")

    assert fg_interface.ipv6_autoconf == "enable"
    assert fg_interface.dhcp6_client_options == ["rapid", "iapd", "iana"]
    assert fg_interface.dhcp6_relay_ip == ["2001:db8::1", "2001:db8::2"]
    assert fg_interface.ip6_hop_limit == 64
    assert fg_interface.ip6_subnet == "2001:db8:1::/64"
    assert fg_interface.ipv6_source_settings["ip6_prefix_mode"] == "ra"

    result = extract_fortigate_config(SCALAR_IPV6_CONFIG)
    interface = result.canonical_ir.interfaces[0]
    assert interface.source_dhcp6_client_options == ["rapid", "iapd", "iana"]
    assert interface.source_dhcp6_relay_ip == ["2001:db8::1", "2001:db8::2"]
    assert interface.source_ip6_hop_limit == 64
    assert interface.migration_status == "PARTIALLY_NORMALIZED"
    assert interface.requires_manual_review is True

    workbook = load_workbook(
        io.BytesIO(IRExcelExporter(result.canonical_ir, result).generate())
    )
    values = dict(zip(
        [cell.value for cell in workbook["Interfaces"][3]],
        [cell.value for cell in workbook["Interfaces"][4]],
    ))
    assert values["DHCPv6 Client Options"] == "rapid\niapd\niana"
    assert values["IPv6 Hop Limit"] == 64


def test_ipv6_prefix_lists_and_iapd_are_separate_typed_ordered_collections():
    result = extract_fortigate_config(PREFIX_AND_IAPD_CONFIG)
    interface = result.canonical_ir.interfaces[0]

    assert [item.source_prefix for item in interface.ipv6_prefix_advertisements] == [
        "2001:DB8::/64", "not-an-ipv6/64"
    ]
    assert [item.prefix for item in interface.ipv6_prefix_advertisements] == [
        "2001:db8::/64", None
    ]
    assert interface.ipv6_prefix_advertisements[0].dnssl == [
        "corp.example", "lab.example"
    ]
    assert interface.ipv6_prefix_advertisements[0].rdnss == [
        "2001:db8::53", "2001:db8::54"
    ]

    delegated = interface.ipv6_delegated_prefixes[0]
    assert delegated.prefix_id == "1"
    assert delegated.subnet == "2001:db8:1::/64"
    assert delegated.upstream_interface == "wan"
    assert delegated.rdnss_service == "delegated"

    assert [item.source_iaid for item in interface.dhcp6_iapd] == ["9", "bad-iaid"]
    assert interface.dhcp6_iapd[0].prefix_hint_plt == 0
    assert interface.dhcp6_iapd[0].prefix_hint_vlt is None
    assert interface.dhcp6_iapd[1].iaid is None
    assert interface.migration_status == "PARTIALLY_NORMALIZED"

    workbook = load_workbook(
        io.BytesIO(IRExcelExporter(result.canonical_ir, result).generate())
    )
    values = dict(zip(
        [cell.value for cell in workbook["Interfaces"][3]],
        [cell.value for cell in workbook["Interfaces"][4]],
    ))
    assert "2001:db8::/64" in values["IPv6 Prefix Advertisements"]
    assert "2001:DB8::/64" in values["IPv6 Prefix Advertisements"]
    assert "wan" in values["IPv6 Delegated Prefixes"]
    assert "bad-iaid" in values["DHCPv6 IA-PD"]


def test_ipv6_only_interface_is_retained_and_normalized():
    result = extract_fortigate_config(PRIMARY_IPV6_CONFIG)
    interface = result.canonical_ir.interfaces[0]

    assert interface.ip is None
    assert interface.ipv6_address == "2001:db8::1/64"
    assert interface.source_ipv6_address == "2001:0DB8:0:0::1/64"
    assert interface.source_ipv6_management_access == ["ping", "https", "ssh"]
    assert interface.requires_manual_review is False
    assert interface.migration_status == "NORMALIZED"
    assert interface.review_reasons == []
    assert result.generation_safe is True


def test_simple_ipv6_interface_does_not_get_generic_nested_review():
    ir = FGToIRTransformer(parse_fortigate_config(PRIMARY_IPV6_CONFIG)).transform()
    interface = ir.interfaces[0]

    assert [node.name for node in interface.nested_source_configs] == ["ipv6"]
    assert not any("nested" in reason.lower() for reason in interface.review_reasons)
    assert not any(
        entry.category == "Interface Nested Configuration"
        for entry in ir.audit_entries
    )


def test_nested_ipv6_tree_preserves_additional_addresses_and_order():
    config = """
config system interface
    edit "port1"
        config ipv6
            set ip6-address 2001:db8:1::1/64
            config ip6-extra-addr
                edit 2001:db8:1::2/64
                next
                edit 2001:db8:1::3/64
                next
            end
            config ip6-prefix-list
                edit 2001:db8:1::/64
                    set autonomous-flag enable
                next
            end
        end
    next
end
"""
    interface = _interface(parse_fortigate_config(config), "port1")
    ipv6 = next(node for node in interface.nested_configs if node.name == "ipv6")

    extra_addresses = next(
        child for child in ipv6.children if child.name == "ip6-extra-addr"
    )
    assert [child.name for child in extra_addresses.children] == [
        "2001:db8:1::2/64",
        "2001:db8:1::3/64",
    ]
    assert [child.name for child in ipv6.children] == [
        "ip6-extra-addr",
        "ip6-prefix-list",
    ]
    assert [item.source_address for item in interface.ipv6_extra_addresses] == [
        "2001:db8:1::2/64", "2001:db8:1::3/64"
    ]


def test_extra_ipv6_addresses_are_normalized_in_order_and_invalid_values_preserved():
    config = """
config system interface
    edit "port1"
        config ipv6
            config ip6-extra-addr
                edit 2001:DB8::2/64
                next
                edit not-an-ipv6/64
                next
                edit 2001:db8::2/64
                next
            end
        end
    next
end
"""
    result = extract_fortigate_config(config)
    interface = result.canonical_ir.interfaces[0]

    assert [item.address for item in interface.additional_ipv6_addresses] == [
        "2001:db8::2/64", None, "2001:db8::2/64"
    ]
    assert [item.source_address for item in interface.additional_ipv6_addresses] == [
        "2001:DB8::2/64", "not-an-ipv6/64", "2001:db8::2/64"
    ]
    assert interface.migration_status == "PARTIALLY_NORMALIZED"
    assert any("ipv6-extra-addr" in error for error in interface.parse_errors)

    workbook = load_workbook(
        io.BytesIO(IRExcelExporter(result.canonical_ir, result).generate())
    )
    values = dict(zip(
        [cell.value for cell in workbook["Interfaces"][3]],
        [cell.value for cell in workbook["Interfaces"][4]],
    ))
    assert values["Additional IPv6 Source Addresses"] == (
        "2001:DB8::2/64\nnot-an-ipv6/64\n2001:db8::2/64"
    )


def test_complex_ipv6_interface_is_partially_normalized_and_reviewed():
    config = """
config system interface
    edit "port1"
        config ipv6
            set ip6-address 2001:db8::1/64
            config ip6-prefix-list
                edit 2001:db8::/64
                    set autonomous-flag enable
                next
            end
        end
    next
end
"""
    interface = FGToIRTransformer(parse_fortigate_config(config)).transform().interfaces[0]

    assert interface.ipv6_address == "2001:db8::1/64"
    assert interface.migration_status == "PARTIALLY_NORMALIZED"
    assert interface.requires_manual_review is True
    assert interface.review_reasons == [
        "FortiGate IPv6 interface contains source-specific behavior requiring target-platform review"
    ]


def test_vrrp6_is_preserved_and_reviewed():
    config = """
config system interface
    edit "port1"
        config ipv6
            set ip6-address 2001:db8::1/64
            config vrrp6
                edit 1
                    set accept-mode enable
                    set adv-interval 1
                    set priority 150
                    set preempt enable
                    set status enable
                    set vrdst6 2001:db8::fd
                    set vrip6 2001:db8::fe
                next
                edit 2
                    set status disable
                next
            end
        end
    next
end
"""
    ir = FGToIRTransformer(parse_fortigate_config(config)).transform()
    interface = ir.interfaces[0]

    assert any(
        child.name == "vrrp6"
        for node in interface.nested_source_configs
        for child in node.children
    )
    assert [item.source_vrid for item in interface.vrrp6] == ["1", "2"]
    assert interface.vrrp6[0].vrid == 1
    assert interface.vrrp6[0].vrip6 == "2001:db8::fe"
    assert interface.vrrp6[0].vrdst6 == "2001:db8::fd"
    assert interface.vrrp6[1].status == "disable"
    assert interface.migration_status == "PARTIALLY_NORMALIZED"
    assert interface.requires_manual_review is True
    assert any(
        "vrrp6" in reason.lower()
        for reason in interface.review_reasons
    )


def test_malformed_ipv6_address_is_retained_without_repair():
    config = """
config system interface
    edit "broken-v6"
        config ipv6
            set ip6-address not-an-ipv6/64
        end
    next
end
"""
    result = extract_fortigate_config(config)
    interface = result.canonical_ir.interfaces[0]

    assert interface.ipv6_address is None
    assert interface.source_ipv6_address == "not-an-ipv6/64"
    assert interface.ipv6_source_settings["ip6_address"] == "not-an-ipv6/64"
    assert interface.parse_errors
    assert "ipv6-address" in interface.parse_errors[0]
    assert interface.requires_manual_review is True
    assert result.generation_safe is False

    section = next(
        item for item in result.source_sections if item.path == "system interface"
    )
    assert section.status == ExtractionStatus.PARTIALLY_NORMALIZED


def test_ipv6_fields_are_visible_in_excel_interface_inventory():
    result = extract_fortigate_config(PRIMARY_IPV6_CONFIG)
    workbook = load_workbook(
        io.BytesIO(IRExcelExporter(result.canonical_ir, result).generate())
    )
    sheet = workbook["Interfaces"]
    headers = [cell.value for cell in sheet[3]]
    values = [cell.value for cell in sheet[4]]
    values_by_header = dict(zip(headers, values))

    assert {
        "IPv6 Address",
        "IPv6 Management Access",
        "Migration Status",
        "Manual Review",
        "Review Reasons",
    }.issubset(headers)
    assert values_by_header["IPv6 Address"] == "2001:db8::1/64"
    assert values_by_header["IPv6 Source Address"] == "2001:0DB8:0:0::1/64"
    assert values_by_header["IPv6 Management Access"] == "ping\nhttps\nssh"
    assert values_by_header["Migration Status"] == "NORMALIZED"
    assert values_by_header["Manual Review"] == "FALSE"
    assert values_by_header["Review Reasons"] is None
