import io
import zipfile
from pathlib import Path

import pytest
from openpyxl import load_workbook

from fwmigrate.report.excel_exporter import IRExcelExporter
from fwmigrate.web import _extract_source_config, create_app


@pytest.fixture
def client():
    return create_app({"TESTING": True}).test_client()


WEB_CONFIG = '''config firewall address
    edit "TEST"
        set subnet 10.0.0.0 255.255.255.0
    next
end
config unsupported-test-section
    edit "X"
        set foo bar
    next
end
'''


def _post_config(client, endpoint: str, content: bytes, **fields):
    return client.post(
        endpoint,
        data={
            "source_vendor": "fortigate",
            "file": (io.BytesIO(content), "test.conf"),
            **fields,
        },
        content_type="multipart/form-data",
    )


def _rows_by_first_column(sheet):
    return {
        sheet.cell(row, 1).value: row
        for row in range(4, sheet.max_row + 1)
    }


def test_public_excel_export_uses_authoritative_fortigate_extraction(client):
    response = _post_config(client, "/api/extract/excel", WEB_CONFIG.encode())
    assert response.status_code == 200
    workbook = load_workbook(io.BytesIO(response.data))

    coverage = workbook["Extraction Coverage"]
    headers = {cell.value: cell.column for cell in coverage[3]}
    rows = _rows_by_first_column(coverage)
    assert coverage.cell(rows["firewall address"], headers["Status"]).value == "NORMALIZED"
    assert coverage.cell(rows["unsupported-test-section"], headers["Status"]).value == "UNSUPPORTED"
    assert coverage.cell(rows["unsupported-test-section"], headers["Source Objects"]).value == 1

    unsupported = workbook["Unsupported"]
    unsupported_rows = _rows_by_first_column(unsupported)
    row = unsupported_rows["unsupported-test-section"]
    assert unsupported.cell(row, 3).value == "UNSUPPORTED"


def test_migration_zip_inventory_uses_authoritative_fortigate_extraction(client):
    response = _post_config(
        client,
        "/api/migrate",
        WEB_CONFIG.encode(),
        target_vendor="palo_alto",
    )
    assert response.status_code == 200
    with zipfile.ZipFile(io.BytesIO(response.data)) as archive:
        workbook = load_workbook(
            io.BytesIO(archive.read("source_inventory_fortigate.xlsx"))
        )
    assert "unsupported-test-section" in _rows_by_first_column(workbook["Unsupported"])


@pytest.mark.parametrize("endpoint", ["/api/extract/excel", "/api/migrate"])
def test_uploaded_invalid_utf8_is_rejected_without_silent_loss(client, endpoint):
    bad_content = (
        b"config system global\n"
        b'    set hostname "FW"\n'
        b"\xff\xfe"
        b"end\n"
    )
    response = _post_config(client, endpoint, bad_content)
    assert response.status_code == 400
    body = response.get_json()
    assert body["stage"] == "decode"
    assert "not valid UTF-8" in body["error"]


def test_non_fortigate_helper_preserves_ir_only_export_behavior():
    ir, extraction = _extract_source_config(
        "cisco_asa",
        "hostname asa-test\n",
    )
    assert ir.metadata.source_vendor == "cisco_asa"
    assert extraction is None
    assert IRExcelExporter(ir, extraction_result=None).generate()


def test_full_fixture_is_exposed_through_public_excel_path(client):
    fixture = Path(__file__).parent / "fixtures" / "example_fortigate.conf"
    response = _post_config(client, "/api/extract/excel", fixture.read_bytes())
    assert response.status_code == 200
    workbook = load_workbook(io.BytesIO(response.data), read_only=True)
    expected = {
        "Extraction Coverage",
        "Unsupported",
        "IPS Sensors",
        "IPS Sensor Entries",
        "Traffic Shapers",
        "Proxy Addresses",
        "Web Proxy Settings",
        "VIP Groups",
        "SD-WAN",
        "SD-WAN Members",
        "SD-WAN Health Checks",
        "SD-WAN Rules",
        "SSL VPN Settings",
        "SSL VPN Portals",
        "LDAP Servers",
        "SAML Servers",
        "DoS Policies",
        "Firewall Sniffer",
        "Routing Protocols",
        "Source Security Profiles",
    }
    assert expected <= set(workbook.sheetnames)
