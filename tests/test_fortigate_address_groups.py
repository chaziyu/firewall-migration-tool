import io

from openpyxl import load_workbook

from fwmigrate.parsers.fortigate.parser import parse_fortigate_config
from fwmigrate.parsers.fortigate.transformer import FGToIRTransformer
from fwmigrate.report.excel_exporter import IRExcelExporter


ADDRESS_GROUP_CONFIG = """
config firewall addrgrp
    edit "protection.outlook.com"
        set uuid ab64b5ce-6e6c-51e9-ffae-77f57a7e5008
        set allow-routing enable
        set member "microsoft1" "microsoft2" "microsoft3" "microsoft4" "microsoft5"
        set color 25
        set visibility enable
    next
    edit "Deleum_VPN"
        set category ztna-ems-tag
        set uuid 6d3d0d36-5881-51f0-94af-9d32beb6cb28
        set member "EMS1_ZTNA_all_registered_clients" "MAC_EMS1_ZTNA_all_registered_clients"
    next
    edit "Deleum_ICT"
        set category ztna-ems-tag
        set uuid b80a9cf4-58a7-51f0-3a26-8e6daa2814fb
        set member "EMS1_ZTNA_Deleum_ADUser" "EMS1_ZTNA_Deleum_AV" "EMS1_ZTNA_Deleum_CriticalVul" "MAC_EMS1_ZTNA_all_registered_clients" "MAC_EMS1_ZTNA_Deleum_ADUser" "MAC_EMS1_ZTNA_Deleum_AV" "MAC_EMS1_ZTNA_Deleum_CriticalVul" "EMS1_ZTNA_all_registered_clients"
    next
end
"""


def _by_name(items):
    return {item.name: item for item in items}


def test_address_group_metadata_survives_parser_and_ir():
    parsed = parse_fortigate_config(ADDRESS_GROUP_CONFIG)
    source_groups = _by_name(parsed.address_groups)
    protection = source_groups["protection.outlook.com"]
    assert protection.uuid == "ab64b5ce-6e6c-51e9-ffae-77f57a7e5008"
    assert protection.allow_routing == "enable"
    assert protection.color == 25
    assert protection.extra_settings == {"visibility": "enable"}

    ir = FGToIRTransformer(parsed).transform()
    groups = _by_name(ir.address_groups)
    protection_ir = groups["protection.outlook.com"]
    assert protection_ir.source_uuid == "ab64b5ce-6e6c-51e9-ffae-77f57a7e5008"
    assert protection_ir.allow_routing is True
    assert protection_ir.source_color == 25
    assert protection_ir.source_attributes == {"visibility": "enable"}

    deleum_ict = groups["Deleum_ICT"]
    assert deleum_ict.source_category == "ztna-ems-tag"
    assert deleum_ict.members == source_groups["Deleum_ICT"].member


def test_address_group_metadata_reaches_excel():
    ir = FGToIRTransformer(
        parse_fortigate_config(ADDRESS_GROUP_CONFIG)
    ).transform()
    workbook = load_workbook(
        io.BytesIO(IRExcelExporter(ir).generate())
    )
    sheet = workbook["Address Groups"]
    headers = {cell.value: cell.column for cell in sheet[3]}
    rows = {
        sheet.cell(row, headers["Name"]).value: row
        for row in range(4, sheet.max_row + 1)
    }

    protection_row = rows["protection.outlook.com"]
    assert sheet.cell(protection_row, headers["Allow Routing"]).value == "TRUE"
    assert sheet.cell(protection_row, headers["Source Color"]).value == 25
    assert sheet.cell(protection_row, headers["Additional Settings"]).value == "visibility=enable"

    ict_row = rows["Deleum_ICT"]
    assert sheet.cell(ict_row, headers["Source Category"]).value == "ztna-ems-tag"
    assert "EMS1_ZTNA_Deleum_ADUser" in sheet.cell(ict_row, headers["Members"]).value


def test_addrgrp6_exclude_members_remain_a_list():
    config = """
    config firewall addrgrp6
        edit "IPv6-Excluded"
            set uuid 11111111-2222-3333-4444-555555555555
            set exclude enable
            set member "IPv6-All"
            set exclude-member "IPv6-A" "IPv6-B"
            set fabric-object enable
            config tagging
                edit "classification"
                    set category "security"
                    set tags "internal" "restricted"
                next
            end
        next
    end
    """
    parsed = parse_fortigate_config(config)
    group = parsed.address_groups[0]
    assert group.is_ipv6 is True
    assert group.exclude_member == ["IPv6-A", "IPv6-B"]
    result = FGToIRTransformer(parsed).transform().address_groups[0]
    assert result.source_section == "firewall addrgrp6"
    assert result.address_family == "ipv6"
    assert result.exclusion_enabled is True
    assert result.exclude_members == ["IPv6-A", "IPv6-B"]
    assert result.requires_manual_review is True
