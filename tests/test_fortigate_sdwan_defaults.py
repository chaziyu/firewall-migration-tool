import io

from openpyxl import load_workbook

from fwmigrate.ir.io import load_ir_payload
from fwmigrate.parsers.fortigate.extractor import extract_fortigate_config
from fwmigrate.report.excel_exporter import IRExcelExporter


def _sdwan(config_text: str):
    result = extract_fortigate_config(config_text)
    assert result.canonical_ir.sdwan is not None
    return result.canonical_ir.sdwan


def _headers(sheet):
    return {cell.value: cell.column for cell in sheet[3]}


def test_sdwan_member_effective_defaults_and_explicit_values():
    sdwan = _sdwan(
        '''
config system sdwan
    set load-balance-mode weight-based
    config members
        edit 4
            set interface "wan4"
        next
        edit 5
            set interface "wan5"
            set zone "custom-zone"
            set weight 20
            set priority 7
            set status disable
        next
    end
end
'''
    )

    defaulted, explicit = sdwan.members
    assert (defaulted.weight, defaulted.priority, defaulted.status) == (
        1,
        1,
        "enable",
    )
    assert defaulted.zone == "virtual-wan-link"
    assert not {"weight", "priority", "status", "zone"}.intersection(
        defaulted.source_explicit_fields
    )
    assert "weight" not in defaulted.source_attributes
    assert "priority" not in defaulted.source_attributes
    assert "status" not in defaulted.source_attributes

    assert (explicit.weight, explicit.priority, explicit.status) == (
        20,
        7,
        "disable",
    )
    assert explicit.zone == "custom-zone"
    assert {"weight", "priority", "status", "zone"} <= set(
        explicit.source_explicit_fields
    )


def test_sdwan_health_check_effective_defaults_and_explicit_values():
    sdwan = _sdwan(
        '''
config system sdwan
    config health-check
        edit "google"
            set server "8.8.8.8"
        next
        edit "quad9"
            set protocol https
            set interval 1000
            set failtime 3
            set recoverytime 8
        next
    end
end
'''
    )

    defaulted, explicit = sdwan.health_checks
    assert (
        defaulted.protocol,
        defaulted.interval,
        defaulted.failtime,
        defaulted.recoverytime,
    ) == ("ping", 500, 5, 5)
    assert not {
        "protocol",
        "interval",
        "failtime",
        "recoverytime",
    }.intersection(defaulted.source_explicit_fields)
    assert "failtime" not in defaulted.source_attributes

    assert (
        explicit.protocol,
        explicit.interval,
        explicit.failtime,
        explicit.recoverytime,
    ) == ("https", 1000, 3, 8)
    assert {
        "protocol",
        "interval",
        "failtime",
        "recoverytime",
    } <= set(explicit.source_explicit_fields)
    assert explicit.source_attributes["failtime"] == "3"


def test_sdwan_service_effective_defaults_and_explicit_priority_mode():
    sdwan = _sdwan(
        '''
config system sdwan
    config service
        edit 7
            set name "rule-7"
        next
        edit 6
            set name "rule-6"
            set mode priority
            set status disable
        next
    end
end
'''
    )

    defaulted, explicit = sdwan.rules
    assert (defaulted.mode, defaulted.status) == ("manual", "enable")
    assert not {"mode", "status"}.intersection(defaulted.source_explicit_fields)

    assert (explicit.mode, explicit.status) == ("priority", "disable")
    assert {"mode", "status"} <= set(explicit.source_explicit_fields)


def test_sdwan_defaults_are_exported_as_effective_values():
    result = extract_fortigate_config(
        '''
config system sdwan
    set load-balance-mode weight-based
    config members
        edit 4
            set interface "wan4"
        next
    end
    config health-check
        edit "unifi2"
            set server "192.0.2.2"
        next
    end
    config service
        edit 7
            set name "rule-7"
        next
    end
end
'''
    )

    workbook = load_workbook(
        io.BytesIO(IRExcelExporter(result.canonical_ir, result).generate())
    )
    member_sheet = workbook["SD-WAN Members"]
    member_headers = _headers(member_sheet)
    assert member_sheet.cell(4, member_headers["Weight"]).value == 1
    assert member_sheet.cell(4, member_headers["Priority"]).value == 1
    assert member_sheet.cell(4, member_headers["Status"]).value == "enable"

    health_sheet = workbook["SD-WAN Health Checks"]
    health_headers = _headers(health_sheet)
    assert health_sheet.cell(4, health_headers["Protocol"]).value == "ping"
    assert health_sheet.cell(4, health_headers["Interval"]).value == 500
    assert health_sheet.cell(4, health_headers["Fail Time"]).value == 5
    assert health_sheet.cell(4, health_headers["Recovery Time"]).value == 5

    rule_sheet = workbook["SD-WAN Rules"]
    rule_headers = _headers(rule_sheet)
    assert rule_sheet.cell(4, rule_headers["Mode"]).value == "manual"
    assert rule_sheet.cell(4, rule_headers["Status"]).value == "enable"


def test_schema_1_17_sdwan_objects_receive_provenance_defaults():
    ir = load_ir_payload(
        {
            "schema_version": "1.17",
            "metadata": {
                "hostname": "Legacy-FW",
                "source_vendor": "fortigate",
            },
            "sdwan": {
                "members": [
                    {
                        "source_id": 4,
                        "interface": "wan4",
                        "zone": "virtual-wan-link",
                    }
                ],
                "health_checks": [{"name": "google"}],
                "rules": [{"source_id": 7}],
            },
        }
    )

    assert ir.sdwan is not None
    assert ir.sdwan.members[0].source_explicit_fields == []
    assert ir.sdwan.members[0].preferred_source is None
    assert ir.sdwan.members[0].transport_group is None
    assert ir.sdwan.members[0].review_reasons == []
    assert ir.sdwan.health_checks[0].source_explicit_fields == []
    assert ir.sdwan.rules[0].source_explicit_fields == []
