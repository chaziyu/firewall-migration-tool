import io

import pytest
from openpyxl import load_workbook

from fwmigrate.ir.enums import ServiceProtocol
from fwmigrate.parsers.fortigate.parser import parse_fortigate_config
from fwmigrate.parsers.fortigate.transformer import FGToIRTransformer
from fwmigrate.report.excel_exporter import IRExcelExporter


def _parse_and_transform(body: str, name: str = "VIP_Test"):
    config = f"""
config firewall vip
    edit "{name}"
{body}
    next
end
"""
    parsed = parse_fortigate_config(config)
    return parsed, FGToIRTransformer(parsed).transform()


def test_static_vip_preserves_core_scalar_and_multi_value_fields():
    parsed, ir = _parse_and_transform("""
        set id 42
        set uuid "vip-uuid"
        set type static-nat
        set status disable
        set extip 203.0.113.20
        set extaddr "PUBLIC_A" "PUBLIC_B"
        set mappedip "10.10.10.20" "10.10.10.21"
        set mapped-addr "WEB_SERVERS"
        set extintf "wan1"
        set arp-reply disable
        set nat-source-vip enable
        set src-filter "192.0.2.0/24" "198.51.100.4/32"
        set srcintf-filter "wan1" "wan2"
        set service "HTTPS" "HTTP"
        set gratuitous-arp-interval 30
        set comment "External web servers"
        set dns-mapping-ttl 300
        set private-key customer-private-material
""")

    source = parsed.vips[0]
    assert source.mappedip == ["10.10.10.20", "10.10.10.21"]
    assert source.extaddr == ["PUBLIC_A", "PUBLIC_B"]
    assert source.src_filter == ["192.0.2.0/24", "198.51.100.4/32"]
    assert source.srcintf_filter == ["wan1", "wan2"]
    assert source.portforward == "disable"
    assert source.extra_settings == {
        "dns_mapping_ttl": "300",
        "private_key": "[REDACTED]",
    }

    vip = ir.virtual_ips[0]
    assert vip.source_id == 42
    assert vip.source_uuid == "vip-uuid"
    assert vip.vip_type == "static-nat"
    assert vip.enabled is False
    assert vip.external_ip == "203.0.113.20"
    assert vip.external_addresses == ["PUBLIC_A", "PUBLIC_B"]
    assert vip.external_interface == "wan1"
    assert vip.mapped_ips == ["10.10.10.20", "10.10.10.21"]
    assert vip.mapped_address == "WEB_SERVERS"
    assert vip.arp_reply is False
    assert vip.port_forward is False
    assert vip.gratuitous_arp_interval == 30
    assert vip.nat_source_vip is True
    assert vip.services == ["HTTPS", "HTTP"]
    assert vip.extra_settings == {
        "dns_mapping_ttl": "300",
        "private_key": "[REDACTED]",
    }

    assert ir.nat_rules == []


@pytest.mark.parametrize(
    ("protocol", "expected_protocol"),
    [("tcp", ServiceProtocol.TCP), ("udp", ServiceProtocol.UDP)],
)
def test_unreferenced_port_forward_vip_stays_inventory_only(protocol, expected_protocol):
    parsed, ir = _parse_and_transform(f"""
        set extip 203.0.113.20
        set mappedip "10.10.10.20"
        set portforward enable
        set protocol {protocol}
        set extport 443
        set mappedport 8443
    """)

    assert parsed.vips[0].protocol == protocol
    assert ir.virtual_ips[0].port_forward is True
    assert ir.virtual_ips[0].protocol == protocol
    assert expected_protocol.value == protocol
    assert ir.services == []
    assert ir.nat_rules == []


def test_load_balance_vip_attaches_nested_real_servers_to_parent():
    parsed, ir = _parse_and_transform("""
        set type server-load-balance
        set extip 203.0.113.10
        set mappedip "10.10.10.10"
        set extintf "wan1"
        set ldb-method round-robin
        set server-type https
        set persistence http-cookie
        set http-redirect enable
        set monitor "https-health" "tcp-health"
        set max-embryonic-connections 1000
        config realservers
            edit 1
                set ip 10.10.10.10
                set port 443
                set status active
                set weight 10
                set holddown-interval 30
            next
            edit 2
                set ip 10.10.10.11
                set port 443
                set status standby
                set weight 20
            next
        end
    """, name="Web_LB")

    assert len(parsed.vips) == 1
    assert [server.id for server in parsed.vips[0].realservers] == [1, 2]
    assert parsed.vips[0].realservers[0].holddown_interval == 30

    vip = ir.virtual_ips[0]
    assert vip.load_balance_method == "round-robin"
    assert vip.server_type == "https"
    assert vip.persistence == "http-cookie"
    assert vip.http_redirect is True
    assert vip.monitors == ["https-health", "tcp-health"]
    assert vip.max_embryonic_connections == 1000
    assert [server.address for server in vip.real_servers] == ["10.10.10.10", "10.10.10.11"]
    assert vip.real_servers[0].holddown_interval == 30


def test_virtual_ip_excel_inventory_and_real_server_rows():
    _, ir = _parse_and_transform("""
        set type server-load-balance
        set extip 203.0.113.10
        set mappedip "10.10.10.10"
        set protocol udp
        set comment "Load-balanced VIP"
        set dns-mapping-ttl 300
        config realservers
            edit 1
                set ip 10.10.10.10
                set port 443
                set status active
                set weight 10
                set holddown-interval 30
            next
        end
    """, name="Web_LB")

    workbook = load_workbook(io.BytesIO(IRExcelExporter(ir).generate()))
    vips = workbook["Virtual IPs"]
    servers = workbook["VIP Real Servers"]

    assert vips["A4"].value == "Web_LB"
    headers = {cell.value: cell.column for cell in vips[3]}
    assert vips.cell(4, headers["Type"]).value == "server-load-balance"
    assert vips.cell(4, headers["External IP"]).value == "203.0.113.10"
    assert vips.cell(4, headers["Mapped IPs"]).value == "10.10.10.10"
    assert vips.cell(4, headers["Protocol"]).value == "udp"
    assert vips.cell(4, headers["Additional Settings"]).value == "dns-mapping-ttl=300"
    server_headers = {cell.value: cell.column for cell in servers[3]}
    assert servers.cell(4, server_headers["VIP Name"]).value == "Web_LB"
    assert servers.cell(4, server_headers["Server ID"]).value == 1
    assert servers.cell(4, server_headers["Address Type"]).value == "ip"
    assert servers.cell(4, server_headers["IP"]).value == "10.10.10.10"
    assert servers.cell(4, server_headers["Port"]).value == 443
    assert servers.cell(4, server_headers["Status"]).value == "active"
    assert servers.cell(4, server_headers["Weight"]).value == 10
    assert servers.cell(4, server_headers["Holddown Interval"]).value == 30


def test_real_server_address_reference_and_advanced_fields_survive():
    parsed, ir = _parse_and_transform("""
        set type server-load-balance
        set extip 203.0.113.90
        config realservers
            edit 1
                set type address
                set address "DYNAMIC_BACKEND"
                set port 8443
                set status active
                set weight 20
                set holddown-interval 45
                set healthcheck enable
                set http-host "backend.example.com"
                set translate-host "internal.example.com"
                set max-connections 500
                set monitor "HTTPS_MON" "TCP_MON"
                set client-ip "10.0.0.0/24"
                set custom-setting "retained"
            next
        end
    """, name="Address_Backend_VIP")

    source = parsed.vips[0].realservers[0]
    assert (source.type, source.address, source.ip, source.monitor) == (
        "address", "DYNAMIC_BACKEND", None, ["HTTPS_MON", "TCP_MON"]
    )

    server = ir.virtual_ips[0].real_servers[0]
    assert server.address_type == "address"
    assert server.ip_address is None
    assert server.address_reference == "DYNAMIC_BACKEND"
    assert server.address == "DYNAMIC_BACKEND"
    assert (server.port, server.status, server.weight) == (8443, "active", 20)
    assert server.holddown_interval == 45
    assert server.healthcheck == "enable"
    assert server.http_host == "backend.example.com"
    assert server.translate_host == "internal.example.com"
    assert server.max_connections == 500
    assert server.monitors == ["HTTPS_MON", "TCP_MON"]
    assert server.client_ip == "10.0.0.0/24"
    assert server.source_attributes == {"custom_setting": "retained"}
    assert server.migration_status == "PARTIALLY_NORMALIZED"
    assert server.requires_manual_review is True

    workbook = load_workbook(io.BytesIO(IRExcelExporter(ir).generate()))

    summary_counts = {
        workbook["Summary"].cell(row, 1).value: workbook["Summary"].cell(row, 2).value
        for row in range(1, workbook["Summary"].max_row + 1)
    }
    assert summary_counts["Virtual IPs"] == 1
    assert summary_counts["VIP Real Servers"] == 1

    coverage_counts = {
        workbook["Extraction Coverage"].cell(row, 1).value:
            workbook["Extraction Coverage"].cell(row, 3).value
        for row in range(4, workbook["Extraction Coverage"].max_row + 1)
    }
    assert coverage_counts["Virtual IPs"] == 1
    assert coverage_counts["VIP Real Servers"] == 1
