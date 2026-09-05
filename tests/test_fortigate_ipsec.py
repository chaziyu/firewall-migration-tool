import io

from openpyxl import load_workbook

from fwmigrate.extraction.models import ExtractionStatus
from fwmigrate.parsers.fortigate.extractor import extract_fortigate_config
from fwmigrate.parsers.fortigate.parser import parse_fortigate_config
from fwmigrate.parsers.fortigate.transformer import FGToIRTransformer
from fwmigrate.report.excel_exporter import IRExcelExporter


IPSEC_CONFIG = """
config vpn ipsec phase1-interface
    edit "phase1-a"
        set interface "port1"
        set remote-gw 203.0.113.10
        set proposal aes256-sha256
    next
    edit "actual-phase1"
        set interface "port2"
        set remote-gw 203.0.113.20
    next
end

config vpn ipsec phase2-interface
    edit "phase2-a"
        set phase1name "phase1-a"
        set proposal aes128-sha1 aes256-sha256
        set auto-negotiate enable
        set src-addr-type name
        set dst-addr-type name
        set src-name "local-net"
        set dst-name "remote-net"
        set src-subnet 10.0.0.0 255.255.255.0
        set dst-subnet 10.1.0.0 255.255.255.0
        set comments "Phase2 test"
        set custom-setting test-value
    next
    edit "phase2-different-name"
        set phase1name "actual-phase1"
        set proposal aes256gcm
        set dhgrp 20
        set keepalive enable
    next
    edit "orphan-phase2"
        set phase1name "missing-phase1"
        set proposal aes256-sha256
    next
end
"""


PHASE1_FIDELITY_CONFIG = """
config vpn ipsec phase1-interface
    edit "toMiriWH"
        set type static
        set interface "unifi_port1"
        set local-gw 60.53.219.65
        set remote-gw 219.93.103.225
        set mode aggressive
        set peertype any
        set net-device enable
        set proposal aes128-sha256 aes256-sha256 aes128-sha1 aes256-sha1
        set psksecret "PSK_SECRET_SENTINEL"
        set custom-setting test-value
        set nattraversal enable
    next
    edit "FortiClient"
        set type dynamic
        set interface "unifi_port1"
        set ike-version 2
        set mode-cfg enable
        set proposal aes128-sha256 aes256-sha256
        set eap enable
        set eap-identity send-request
        set authusrgrp "Deleum_IPSEC"
        set ipv4-start-ip 10.10.110.10
        set ipv4-end-ip 10.10.110.110
        set dns-mode auto
        set ipv4-split-include "trust_fixed_IP" "branch_fixed_IP"
        set dpd-retryinterval 60
    next
end
"""


def _by_name(items):
    return {item.name: item for item in items}


def test_fortigate_phase1_parser_and_transform_preserve_source_fidelity():
    config = parse_fortigate_config(PHASE1_FIDELITY_CONFIG)
    phase1 = _by_name(config.phase1_interfaces)

    site = phase1["toMiriWH"]
    assert site.type == "static"
    assert site.interface == "unifi_port1"
    assert site.local_gw == "60.53.219.65"
    assert site.remote_gw == "219.93.103.225"
    assert site.mode == "aggressive"
    assert site.peertype == "any"
    assert site.net_device == "enable"
    assert site.proposal == [
        "aes128-sha256",
        "aes256-sha256",
        "aes128-sha1",
        "aes256-sha1",
    ]
    assert site.has_psk is True
    assert site.extra_settings == {
        "custom_setting": "test-value",
        "nattraversal": "enable",
    }
    assert "PSK_SECRET_SENTINEL" not in config.model_dump_json()
    assert "psksecret" not in site.model_dump()

    remote = phase1["FortiClient"]
    assert remote.type == "dynamic"
    assert remote.ike_version == "2"
    assert remote.mode_cfg == "enable"
    assert remote.proposal == ["aes128-sha256", "aes256-sha256"]
    assert remote.eap == "enable"
    assert remote.eap_identity == "send-request"
    assert remote.authusrgrp == "Deleum_IPSEC"
    assert remote.ipv4_start_ip == "10.10.110.10"
    assert remote.ipv4_end_ip == "10.10.110.110"
    assert remote.dns_mode == "auto"
    assert remote.ipv4_split_include == [
        "trust_fixed_IP",
        "branch_fixed_IP",
    ]
    assert remote.dpd_retryinterval == 60

    ir = FGToIRTransformer(config).transform()
    tunnels = _by_name(ir.vpn_tunnels)
    site_ir = tunnels["toMiriWH"]
    assert site_ir.source_local_gateway == "60.53.219.65"
    assert site_ir.source_type == "static"
    assert site_ir.source_mode == "aggressive"
    assert site_ir.source_peer_type == "any"
    assert site_ir.source_net_device is True
    assert site_ir.source_proposals == site.proposal
    assert site_ir.has_psk is True
    assert site_ir.psk is None
    assert site_ir.ike_crypto_profile is None
    assert site_ir.ipsec_crypto_profile is None
    assert site_ir.migration_status == "PARTIALLY_NORMALIZED"
    assert site_ir.requires_manual_review is True
    assert site_ir.source_attributes == site.extra_settings

    remote_ir = tunnels["FortiClient"]
    assert remote_ir.peer_address == "dynamic"
    assert remote_ir.ike_version == "v2"
    assert remote_ir.source_type == "dynamic"
    assert remote_ir.source_mode_config is True
    assert remote_ir.source_eap is True
    assert remote_ir.source_eap_identity == "send-request"
    assert remote_ir.source_auth_user_group == "Deleum_IPSEC"
    assert remote_ir.source_client_ip_start == "10.10.110.10"
    assert remote_ir.source_client_ip_end == "10.10.110.110"
    assert remote_ir.source_dns_mode == "auto"
    assert remote_ir.source_split_include == [
        "trust_fixed_IP",
        "branch_fixed_IP",
    ]
    assert remote_ir.source_dpd_retry_interval == 60

    audit = {entry.id: entry.message for entry in ir.audit_entries}
    assert "Pre-Shared Key" in audit["toMiriWH"]
    assert "intentionally redacted" in audit["toMiriWH"]
    assert "Pre-Shared Key" not in audit["FortiClient"]


def test_fortigate_phase1_coverage_and_excel_preserve_partial_source_inventory():
    result = extract_fortigate_config(PHASE1_FIDELITY_CONFIG)
    section = next(
        item
        for item in result.source_sections
        if item.path == "vpn ipsec phase1-interface"
    )
    assert section.object_count_source == 2
    assert section.object_count_parsed == 2
    assert section.object_count_normalized == 2
    assert section.status == ExtractionStatus.PARTIALLY_NORMALIZED

    workbook = load_workbook(
        io.BytesIO(IRExcelExporter(result.canonical_ir, result).generate())
    )
    sheet = workbook["VPN Tunnels"]
    headers = {cell.value: cell.column for cell in sheet[3]}
    rows = {
        sheet.cell(row, headers["Name"]).value: row
        for row in range(4, sheet.max_row + 1)
    }

    site_row = rows["toMiriWH"]
    assert sheet.cell(site_row, headers["Type"]).value == "static"
    assert sheet.cell(site_row, headers["Peer Address"]).value == "219.93.103.225"
    assert sheet.cell(site_row, headers["Local Interface"]).value == "unifi_port1"
    assert sheet.cell(site_row, headers["Local Gateway"]).value == "60.53.219.65"
    assert sheet.cell(site_row, headers["Mode"]).value == "aggressive"
    assert sheet.cell(site_row, headers["Peer Type"]).value == "any"
    assert sheet.cell(site_row, headers["Net Device"]).value == "TRUE"
    assert sheet.cell(site_row, headers["IKE Proposal"]).value == (
        "aes128-sha256\naes256-sha256\naes128-sha1\naes256-sha1"
    )
    assert sheet.cell(site_row, headers["PSK"]).value == "Configured / Redacted"
    assert sheet.cell(site_row, headers["Extraction Status"]).value == (
        "PARTIALLY_NORMALIZED"
    )
    assert sheet.cell(site_row, headers["Manual Review"]).value == "TRUE"
    assert sheet.cell(site_row, headers["Additional Settings"]).value == (
        "custom-setting=test-value; nattraversal=enable"
    )
    assert sheet.cell(site_row, headers["IKE Crypto Profile"]).value is None
    assert sheet.cell(site_row, headers["IPsec Crypto Profile"]).value is None

    remote_row = rows["FortiClient"]
    assert sheet.cell(remote_row, headers["Type"]).value == "dynamic"
    assert sheet.cell(remote_row, headers["IKE Version"]).value == "v2"
    assert sheet.cell(remote_row, headers["Mode Config"]).value == "TRUE"
    assert sheet.cell(remote_row, headers["EAP"]).value == "TRUE"
    assert sheet.cell(remote_row, headers["EAP Identity"]).value == "send-request"
    assert sheet.cell(remote_row, headers["Auth User Group"]).value == "Deleum_IPSEC"
    assert sheet.cell(remote_row, headers["Client IP Start"]).value == "10.10.110.10"
    assert sheet.cell(remote_row, headers["Client IP End"]).value == "10.10.110.110"
    assert sheet.cell(remote_row, headers["Client IP Range"]).value == (
        "10.10.110.10 - 10.10.110.110"
    )
    assert sheet.cell(remote_row, headers["DNS Mode"]).value == "auto"
    assert sheet.cell(remote_row, headers["Split Include"]).value == (
        "trust_fixed_IP\nbranch_fixed_IP"
    )
    assert sheet.cell(remote_row, headers["DPD Retry Interval"]).value == 60

    excel_cells = "\n".join(
        str(cell.value)
        for worksheet in workbook.worksheets
        for row in worksheet.iter_rows()
        for cell in row
        if cell.value is not None
    )
    assert "PSK_SECRET_SENTINEL" not in result.model_dump_json()
    assert "PSK_SECRET_SENTINEL" not in excel_cells
    assert "IKE Crypto Profile = default" not in excel_cells


def test_fortigate_phase1_preserves_explicit_default_without_inventing_profiles():
    config = parse_fortigate_config("""
config vpn ipsec phase1-interface
    edit "explicit-default"
        set interface "wan1"
        set proposal default
    next
end
""")

    tunnel = FGToIRTransformer(config).transform().vpn_tunnels[0]
    assert tunnel.source_proposals == ["default"]
    assert tunnel.ike_crypto_profile is None
    assert tunnel.ipsec_crypto_profile is None


def test_fortigate_phase1_does_not_fabricate_missing_or_unknown_source_values():
    config = parse_fortigate_config("""
config vpn ipsec phase1-interface
    edit "missing-values"
        set interface "wan1"
    next
    edit "explicit-dynamic"
        set type dynamic
        set interface "wan1"
    next
    edit "future-values"
        set type static
        set interface "wan1"
        set ike-version 3
        set net-device future-state
        set mode-cfg future-state
        set eap future-state
    next
end
""")

    phase1 = _by_name(config.phase1_interfaces)
    missing = phase1["missing-values"]
    assert missing.type is None
    assert missing.ike_version is None
    assert missing.peertype is None
    assert missing.net_device is None

    ir = FGToIRTransformer(config).transform()
    tunnels = _by_name(ir.vpn_tunnels)

    missing_ir = tunnels["missing-values"]
    assert missing_ir.peer_address is None
    assert missing_ir.ike_version is None
    assert missing_ir.source_type is None
    assert missing_ir.source_peer_type is None
    assert missing_ir.source_net_device is None

    dynamic_ir = tunnels["explicit-dynamic"]
    assert dynamic_ir.peer_address == "dynamic"
    assert dynamic_ir.source_type == "dynamic"

    future_ir = tunnels["future-values"]
    assert future_ir.peer_address is None
    assert future_ir.ike_version is None
    assert future_ir.source_net_device is None
    assert future_ir.source_mode_config is None
    assert future_ir.source_eap is None
    assert future_ir.source_attributes == {
        "net_device": "future-state",
        "mode_cfg": "future-state",
        "eap": "future-state",
        "ike_version": "3",
    }

    audit = {entry.id: entry.message for entry in ir.audit_entries}
    assert "Pre-Shared Key" not in audit["missing-values"]
    assert "Pre-Shared Key" not in audit["explicit-dynamic"]
    assert "Pre-Shared Key" not in audit["future-values"]

    workbook = load_workbook(io.BytesIO(IRExcelExporter(ir).generate()))
    sheet = workbook["VPN Tunnels"]
    headers = {cell.value: cell.column for cell in sheet[3]}
    rows = {
        sheet.cell(row, headers["Name"]).value: row
        for row in range(4, sheet.max_row + 1)
    }
    missing_row = rows["missing-values"]
    assert sheet.cell(missing_row, headers["Type"]).value is None
    assert sheet.cell(missing_row, headers["Peer Address"]).value is None
    assert sheet.cell(missing_row, headers["IKE Version"]).value is None
    assert sheet.cell(missing_row, headers["Peer Type"]).value is None
    assert sheet.cell(missing_row, headers["Net Device"]).value is None


def test_fortigate_phase2_parser_preserves_typed_and_unknown_settings():
    config = parse_fortigate_config(IPSEC_CONFIG)
    phase2 = _by_name(config.phase2_interfaces)

    assert len(phase2) == 3
    first = phase2["phase2-a"]
    assert first.phase1name == "phase1-a"
    assert first.proposal == ["aes128-sha1", "aes256-sha256"]
    assert first.auto_negotiate == "enable"
    assert first.src_addr_type == "name"
    assert first.dst_addr_type == "name"
    assert first.src_name == ["local-net"]
    assert first.dst_name == ["remote-net"]
    assert first.src_subnet == "10.0.0.0 255.255.255.0"
    assert first.dst_subnet == "10.1.0.0 255.255.255.0"
    assert first.comments == "Phase2 test"
    assert first.extra_settings == {"custom_setting": "test-value"}

    second = phase2["phase2-different-name"]
    assert second.phase1name == "actual-phase1"
    assert second.dhgrp == [20]
    assert second.keepalive == "enable"


def test_fortigate_phase2_transform_preserves_semantics_and_references():
    ir = FGToIRTransformer(parse_fortigate_config(IPSEC_CONFIG)).transform()
    phase2 = _by_name(ir.vpn_phase2)

    assert len(phase2) == 3
    first = phase2["phase2-a"]
    assert first.phase1_name == "phase1-a"
    assert first.proposals == ["aes128-sha1", "aes256-sha256"]
    assert first.source_address_type == "name"
    assert first.destination_address_type == "name"
    assert first.source_names == ["local-net"]
    assert first.destination_names == ["remote-net"]
    assert first.source_subnet == "10.0.0.0 255.255.255.0"
    assert first.destination_subnet == "10.1.0.0 255.255.255.0"
    assert first.auto_negotiate is True
    assert first.source_attributes == {"custom_setting": "test-value"}

    second = phase2["phase2-different-name"]
    assert second.phase1_name == "actual-phase1"
    assert second.proposals == ["aes256gcm"]
    assert second.dh_groups == [20]
    assert second.keepalive is True
    assert second.requires_manual_review is True

    orphan = phase2["orphan-phase2"]
    assert orphan.phase1_name == "missing-phase1"
    assert orphan.requires_manual_review is True
    assert any(
        entry.id == "vpn-phase2:orphan-phase2:phase1"
        and "missing-phase1" in entry.message
        for entry in ir.audit_entries
    )


def test_fortigate_phase2_coverage_reports_structured_partial_inventory():
    result = extract_fortigate_config(IPSEC_CONFIG)
    section = next(
        item
        for item in result.source_sections
        if item.path == "vpn ipsec phase2-interface"
    )

    assert section.object_count_source == 3
    assert section.object_count_parsed == 3
    assert section.object_count_normalized == 3
    assert section.status == ExtractionStatus.EXTRACT_ONLY


def test_fortigate_phase2_excel_has_dedicated_inventory_and_summary():
    result = extract_fortigate_config(IPSEC_CONFIG)
    workbook = load_workbook(
        io.BytesIO(
            IRExcelExporter(
                result.canonical_ir,
                extraction_result=result,
            ).generate()
        )
    )

    assert "VPN Tunnels" in workbook.sheetnames
    assert workbook.sheetnames.index("VPN Phase 2") == (
        workbook.sheetnames.index("VPN Tunnels") + 1
    )

    sheet = workbook["VPN Phase 2"]
    headers = {cell.value: cell.column for cell in sheet[3]}
    assert list(headers) == [
        "Name",
        "Phase 1",
        "Proposal",
        "Source Address Type",
        "Source Selector",
        "Destination Address Type",
        "Destination Selector",
        "Source Subnet",
        "Destination Subnet",
        "Auto Negotiate",
        "DH / PFS Groups",
        "Keepalive",
        "Extraction Status",
        "Manual Review",
        "Additional Settings",
        "Description",
    ]
    rows = {
        sheet.cell(row, headers["Name"]).value: row
        for row in range(4, sheet.max_row + 1)
    }

    first_row = rows["phase2-a"]
    assert sheet.cell(first_row, headers["Phase 1"]).value == "phase1-a"
    assert sheet.cell(first_row, headers["Proposal"]).value == (
        "aes128-sha1\naes256-sha256"
    )
    assert sheet.cell(first_row, headers["Source Address Type"]).value == "name"
    assert sheet.cell(first_row, headers["Source Selector"]).value == "local-net"
    assert sheet.cell(first_row, headers["Destination Address Type"]).value == "name"
    assert sheet.cell(first_row, headers["Destination Selector"]).value == "remote-net"
    assert sheet.cell(first_row, headers["Auto Negotiate"]).value == "TRUE"
    assert sheet.cell(first_row, headers["Extraction Status"]).value == (
        "PARTIALLY_NORMALIZED"
    )
    assert sheet.cell(first_row, headers["Manual Review"]).value == "TRUE"
    assert sheet.cell(first_row, headers["Additional Settings"]).value == (
        "custom-setting=test-value"
    )
    assert sheet.cell(first_row, headers["Description"]).value == "Phase2 test"

    second_row = rows["phase2-different-name"]
    assert sheet.cell(second_row, headers["Phase 1"]).value == "actual-phase1"
    assert sheet.cell(second_row, headers["DH / PFS Groups"]).value == "20"
    assert sheet.cell(second_row, headers["Keepalive"]).value == "TRUE"
    assert sheet.cell(second_row, headers["Manual Review"]).value == "TRUE"

    summary = workbook["Summary"]
    counts = {
        summary.cell(row, 1).value: summary.cell(row, 2).value
        for row in range(1, summary.max_row + 1)
    }
    assert counts["VPN Tunnels"] == 2
    assert counts["VPN Phase 2"] == 3
