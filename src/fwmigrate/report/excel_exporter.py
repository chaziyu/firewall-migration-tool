"""Vendor-neutral Excel inventory export for :class:`IRConfig`."""

from __future__ import annotations

import io
import json
import re
from datetime import datetime, timezone
from enum import Enum
from typing import Any, Iterable, Sequence

from fwmigrate.ir.core import IRConfig
from fwmigrate.ir.enums import MigrationConfidence
from fwmigrate.parsers.fortigate.coverage import fortigate_source_category

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover - exercised only without the reports dependency
    Workbook = None


XLSX_MIMETYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_FORMULA_PREFIXES = ("=", "+", "-", "@")
_ILLEGAL_XML_CHARS = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")
_MAX_CELL_TEXT = 32767


class ExcelExportUnavailableError(RuntimeError):
    """Raised when the optional workbook dependency is unavailable."""


class IRExcelExporter:
    """Render a source firewall inventory directly from vendor-neutral IR."""

    OVERVIEW_SHEETS = (
        "Summary",
    )

    CORE_INVENTORY_SHEETS = (
        "System Settings",
        "DNS Settings",
        "Interfaces",
        "Interface Secondary IPs",
        "Zones",
        "Addresses",
        "Address Groups",
        "Address Group Tags",
        "Service Categories",
        "Services",
        "Service Groups",
        "Schedules",
        "Policies",
        "IP Pools",
        "Virtual IPs",
        "VIP Real Servers",
        "VIP Groups",
        "NAT Rules",
        "Routes",
        "VPN Tunnels",
        "VPN Phase 2",
    )

    NETWORK_ACCESS_SHEETS = (
        "DHCP Servers",
        "DHCP IP Ranges",
        "DHCP Reservations",
        "Traffic Shapers",
        "Session Helpers",
        "Session TTL Overrides",
        "SD-WAN",
        "SD-WAN Members",
        "SD-WAN Health Checks",
        "SD-WAN SLAs",
        "SD-WAN Rules",
        "SD-WAN Duplication",
        "SD-WAN Neighbors",
        "SD-WAN Rule SLAs",
        "Routing Protocols",
        "Routing Protocol Settings",
        "Routing Dependencies",
        "Routing Dependency Settings",
        "SSL VPN Settings",
        "SSL VPN Portals",
        "SSL VPN Authentication Rules",
        "SSL VPN Host Checks",
        "LDAP Servers",
        "SAML Servers",
        "FSSO Servers",
        "FSSO AD Groups",
        "Local Users",
        "User Groups",
        "User Group Matches",
        "Administrators",
        "Admin Profiles",
        "Admin Profile Permissions",
        "FortiTokens",
        "ZTNA Providers",
        "Authentication Schemes",
        "Authentication Rules",
        "Certificates",
    )

    SOURCE_DETAIL_SHEETS = (
        "FortiGate Source Configuration",
        "Firewall Policy Source Settings",
        "Interface Source Settings",
        "Interface Nested Configuration",
        "Proxy Addresses",
        "Web Proxy Settings",
        "SSH Keys",
        "Internet Services",
        "Internet Service Definitions",
        "Internet Service Def Entries",
        "Internet Service Def Ports",
        "IPS Sensors",
        "IPS Sensor Entries",
        "Security Profiles",
        "Source Security Profiles",
        "Source Security Profile Setting",
        "DoS Policies",
        "DoS Anomalies",
        "Firewall Sniffer",
    )

    AUDIT_SHEETS = (
        "Warnings",
        "Unsupported",
        "Extraction Coverage",
    )

    SHEET_ORDER = (
        OVERVIEW_SHEETS
        + CORE_INVENTORY_SHEETS
        + NETWORK_ACCESS_SHEETS
        + SOURCE_DETAIL_SHEETS
        + AUDIT_SHEETS
    )

    _NAVY = "17324D"
    _TEAL = "0F766E"
    _LIGHT_TEAL = "D7F0EC"
    _LIGHT_BLUE = "E8F0F7"
    _LIGHT_AMBER = "FEF3C7"
    _LIGHT_RED = "FEE2E2"
    _WHITE = "FFFFFF"
    _TEXT = "1F2937"
    _MUTED = "64748B"
    _BORDER = "CBD5E1"

    _FORTIGATE_DEDICATED_INVENTORY_PATHS = {
        "system global",
        "system dns",
        "system interface",
        "system zone",
        "system dhcp server",
        "system session-helper",
        "system session-ttl",
        "system sdwan",
        "endpoint-control fctems",
        "firewall address",
        "firewall address6",
        "firewall multicast-address",
        "firewall multicast-address6",
        "firewall addrgrp",
        "firewall wildcard-fqdn custom",
        "firewall service category",
        "firewall service custom",
        "firewall service group",
        "firewall schedule recurring",
        "firewall schedule onetime",
        "firewall shaper traffic-shaper",
        "firewall proxy-address",
        "web-proxy global",
        "firewall policy",
        "firewall ippool",
        "firewall vip",
        "firewall vipgrp",
        "firewall internet-service-name",
        "firewall internet-service-definition",
        "firewall DoS-policy",
        "firewall sniffer",
        "firewall ssh local-key",
        "firewall ssh local-ca",
        "router static",
        "router static6",
        "router rip",
        "router ripng",
        "router ospf",
        "router ospf6",
        "router bgp",
        "router isis",
        "router multicast",
        "vpn ipsec phase1-interface",
        "vpn ipsec phase2-interface",
        "vpn certificate remote",
        "vpn certificate local",
        "vpn certificate ca",
        "vpn ssl web portal",
        "vpn ssl settings",
        "ips sensor",
        "user ldap",
        "user fsso",
        "user adgrp",
        "user saml",
        "user local",
        "user group",
        "system admin",
        "system accprofile",
        "user fortitoken",
        "authentication scheme",
        "authentication rule",
    }

    def __init__(self, ir_config: IRConfig, extraction_result: Any = None):
        self.ir = ir_config
        self.extraction = extraction_result

    def generate(self) -> bytes:
        """Generate a complete ``.xlsx`` workbook and return its bytes."""
        if Workbook is None:
            raise ExcelExportUnavailableError(
                "Excel export requires openpyxl. Install the project with the reports extra."
            )

        workbook = Workbook()
        workbook.remove(workbook.active)

        workbook.properties.title = "Firewall Source Inventory"
        workbook.properties.subject = "Vendor-neutral firewall configuration extraction"
        workbook.properties.creator = "Firewall Migration Tool"

        # Build inventory sheets first.
        #
        # Summary is intentionally built last so its navigation section can derive
        # actual worksheet record counts and hyperlinks from the finished workbook.
        self._build_system_settings(workbook)

        self._build_interfaces(workbook)
        self._build_interface_secondary_ips(workbook)
        self._build_interface_source_settings(workbook)
        self._build_interface_nested_configuration(workbook)

        self._build_dhcp_servers(workbook)
        self._build_dhcp_ip_ranges(workbook)
        self._build_dhcp_reservations(workbook)

        self._build_zones(workbook)

        self._build_addresses(workbook)
        self._build_address_groups(workbook)
        self._build_address_group_tags(workbook)
        self._build_proxy_addresses(workbook)
        self._build_web_proxy_settings(workbook)

        self._build_service_categories(workbook)
        self._build_services(workbook)
        self._build_service_groups(workbook)
        self._build_session_helpers(workbook)
        self._build_session_ttl_overrides(workbook)

        self._build_schedules(workbook)
        self._build_traffic_shapers(workbook)
        self._build_policies(workbook)
        self._build_firewall_policy_source_settings(workbook)
        self._build_ztna_providers(workbook)

        self._build_ip_pools(workbook)
        self._build_virtual_ips(workbook)
        self._build_vip_real_servers(workbook)
        self._build_vip_groups(workbook)
        self._build_nat_rules(workbook)

        self._build_vpn_tunnels(workbook)
        self._build_vpn_phase2(workbook)
        self._build_ssl_vpn(workbook)
        self._build_certificates(workbook)
        self._build_ssh_keys(workbook)

        self._build_routes(workbook)
        self._build_routing_protocols(workbook)
        self._build_routing_dependencies(workbook)
        self._build_sdwan(workbook)

        self._build_internet_services(workbook)
        self._build_internet_service_definitions(workbook)
        self._build_ips_sensors(workbook)
        self._build_ips_sensor_entries(workbook)

        self._build_security_profiles(workbook)
        self._build_source_security_profiles(workbook)
        self._build_fortigate_source_configuration(workbook)

        self._build_identity_inventory(workbook)
        self._build_administrator_inventory(workbook)
        self._build_dos_inventory(workbook)
        self._build_firewall_sniffers(workbook)
        self._build_authentication_inventory(workbook)

        self._build_warnings(workbook)
        self._build_unsupported(workbook)
        self._build_extraction_coverage(workbook)

        # Summary is generated after all inventory sheets so navigation can use
        # actual generated worksheet counts.
        self._build_summary(workbook)

        self._order_sheets(workbook)

        output = io.BytesIO()
        workbook.save(output)
        return output.getvalue()

    @staticmethod
    def _format_source_command_values(
        values: Sequence[Any],
    ) -> str:
        if not values:
            return ""

        if len(values) == 1:
            return str(values[0])

        return json.dumps(
            list(values),
            ensure_ascii=False,
        )

    def _interface_nested_config_rows(
        self,
    ) -> Iterable[tuple[Any, ...]]:
        def walk(
            interface_name: str,
            node: Any,
            parent_path: list[str],
        ) -> Iterable[tuple[Any, ...]]:
            if node.node_type == "config":
                config_path = [
                    *parent_path,
                    str(node.name),
                ]
                object_name = None
            else:
                config_path = list(
                    parent_path
                )
                object_name = str(node.name)

            if node.commands:
                for command in node.commands:
                    yield (
                        interface_name,
                        " / ".join(
                            config_path
                        ),
                        node.node_type,
                        object_name,
                        command.operation,
                        command.key,
                        self._format_source_command_values(
                            command.values
                        ),
                        "EXTRACT_ONLY",
                        "Yes",
                    )

            elif not node.children:
                # Preserve the existence of an empty
                # nested config/edit block.
                yield (
                    interface_name,
                    " / ".join(
                        config_path
                    ),
                    node.node_type,
                    object_name,
                    None,
                    None,
                    None,
                    "EXTRACT_ONLY",
                    "Yes",
                )

            for child in node.children:
                child_parent = (
                    config_path
                    if node.node_type == "config"
                    else [
                        *config_path,
                        str(node.name),
                    ]
                )

                yield from walk(
                    interface_name,
                    child,
                    child_parent,
                )

        for interface in self.ir.interfaces:
            for root in (
                interface.nested_source_configs
            ):
                yield from walk(
                    interface.name,
                    root,
                    [],
                )

    def _build_interface_nested_configuration(
        self,
        workbook: Any,
    ) -> None:
        self._table_sheet(
            workbook,
            "Interface Nested Configuration",
            (
                "Interface",
                "Config Path",
                "Node Type",
                "Object / Edit",
                "Operation",
                "Setting",
                "Value",
                "Extraction Status",
                "Manual Review",
            ),
            self._interface_nested_config_rows(),
            empty_note=(
                "No nested interface configuration "
                "was extracted from the source firewall."
            ),
            subtitle=(
                "Nested FortiGate interface configuration "
                "retained as sanitized extraction-only "
                "source data. These settings are not "
                "consumed by target generators."
            ),
        )

    @classmethod
    def _has_dedicated_fortigate_inventory(cls, source_path: str) -> bool:
        if source_path.startswith("system sdwan"):
            return source_path in {
                "system sdwan",
                "system sdwan zone",
                "system sdwan members",
                "system sdwan health-check",
                "system sdwan health-check sla",
                "system sdwan service",
                "system sdwan service sla",
                "system sdwan duplication",
                "system sdwan neighbor",
            }
        return any(
            source_path == path or source_path.startswith(f"{path} ")
            for path in cls._FORTIGATE_DEDICATED_INVENTORY_PATHS
        )

    def _fortigate_source_inventory_items(self) -> list[Any]:
        if self.extraction is None:
            return []
        if str(self.ir.metadata.source_vendor).lower() not in {"fortigate", "fortinet"}:
            return []
        return [
            item
            for item in self.extraction.inventory_items
            if not self._has_dedicated_fortigate_inventory(item.source_path)
            and "structured-security-profile" not in item.notes
            and "structured-routing-protocol" not in item.notes
        ]

    @staticmethod
    def _flatten_fortigate_source_item(item: Any) -> list[tuple[Any, ...]]:
        rows: list[tuple[Any, ...]] = []

        def walk(node: Any, hierarchy: list[str]) -> None:
            for command in node.commands:
                rows.append((
                    fortigate_source_category(item.source_path),
                    item.source_path,
                    item.name,
                    item.source_id,
                    " / ".join(hierarchy),
                    command.operation,
                    command.key,
                    command.values,
                    item.status,
                    item.requires_manual_review,
                ))
            for child in node.children:
                walk(child, [*hierarchy, str(child.name)])

        walk(item, [])
        return rows

    def _build_fortigate_source_configuration(self, workbook: Any) -> None:
        items = self._fortigate_source_inventory_items()
        self._table_sheet(
            workbook,
            "FortiGate Source Configuration",
            (
                "Category", "Source Path", "Object", "Source ID",
                "Parent / Subsection", "Operation", "Setting", "Value",
                "Migration Status", "Manual Review",
            ),
            (
                row
                for item in items
                for row in self._flatten_fortigate_source_item(item)
            ),
            empty_note="No fallback FortiGate source configuration was retained.",
            subtitle=(
                "Sanitized source-only FortiGate configuration retained outside "
                "canonical migration IR. Dedicated inventory sections are omitted."
            ),
        )

    def _build_system_settings(self, workbook: Any) -> None:
        settings = self.ir.system_settings
        self._table_sheet(
            workbook,
            "System Settings",
            ("Hostname", "Timezone", "Admin HTTPS Port", "Additional Settings"),
            [] if settings is None else [(
                settings.hostname,
                settings.timezone,
                settings.admin_https_port,
                self._format_settings(settings.source_attributes),
            )],
        )

        dns = self.ir.dns_settings
        self._table_sheet(
            workbook,
            "DNS Settings",
            ("Primary DNS", "Secondary DNS", "Additional Settings"),
            [] if dns is None else [(
                dns.primary,
                dns.secondary,
                self._format_settings(dns.source_attributes),
            )],
        )

    def _build_summary(self, workbook: Any) -> None:
        sheet = workbook.create_sheet("Summary")
        sheet.sheet_view.showGridLines = False

        sheet.merge_cells("A1:E1")
        sheet["A1"] = "Firewall Source Inventory"
        sheet["A1"].font = Font(
            name="Aptos Display",
            size=20,
            bold=True,
            color=self._WHITE,
        )
        sheet["A1"].fill = PatternFill(
            "solid",
            fgColor=self._NAVY,
        )
        sheet["A1"].alignment = Alignment(
            vertical="center",
        )
        sheet.row_dimensions[1].height = 34

        sheet.merge_cells("A2:E2")
        sheet["A2"] = (
            "Vendor-neutral extraction generated before migration optimization"
        )
        sheet["A2"].font = Font(
            name="Aptos",
            size=10,
            italic=True,
            color=self._MUTED,
        )
        sheet["A2"].alignment = Alignment(
            vertical="center",
        )
        sheet.row_dimensions[2].height = 22

        if self.extraction is not None:
            unsupported_count = (
                len(self.extraction.unsupported_items)
                + sum(
                    1
                    for entry in self.ir.audit_entries
                    if entry.confidence == MigrationConfidence.UNSUPPORTED
                )
            )
        else:
            unsupported_count = sum(
                1
                for entry in self.ir.audit_entries
                if entry.confidence == MigrationConfidence.UNSUPPORTED
            )

        unresolved_count = sum(
            1
            for entry in self.ir.audit_entries
            if "unresolved" in entry.message.lower()
        )

        if unsupported_count:
            extraction_status = "COMPLETE_WITH_UNSUPPORTED_ITEMS"
        elif self.ir.audit_entries:
            extraction_status = "COMPLETE_WITH_WARNINGS"
        else:
            extraction_status = "COMPLETE"

        metadata_rows = [
            ("Source Vendor", self.ir.metadata.source_vendor),
            ("Hostname", self.ir.metadata.hostname),
            ("Input Type", self.ir.metadata.input_type),
            ("Source Version", self.ir.metadata.source_version),
            ("IR Schema Version", self.ir.schema_version),
            ("Source Context", self.ir.metadata.source_context),
            ("Extracted At (UTC)", self.ir.metadata.migration_timestamp),
            ("Extraction Status", extraction_status),
        ]

        self._summary_section(
            sheet,
            4,
            "Extraction Metadata",
            metadata_rows,
        )

        navigation_end_row = self._build_summary_navigation(
            sheet,
            workbook,
            start_row=14,
        )

        inventory_rows = [
            ("Interfaces", len(self.ir.interfaces)),
            (
                "Interface Secondary IPs",
                sum(
                    len(intf.secondary_ips)
                    for intf in self.ir.interfaces
                ),
            ),
            ("DHCP Servers", len(self.ir.dhcp_servers)),
            (
                "DHCP IP Ranges",
                sum(
                    len(server.ip_ranges)
                    for server in self.ir.dhcp_servers
                ),
            ),
            (
                "DHCP Reservations",
                sum(
                    len(server.reservations)
                    for server in self.ir.dhcp_servers
                ),
            ),
            ("Zones", len(self.ir.zones)),
            ("Addresses", len(self.ir.addresses)),
            ("Address Groups", len(self.ir.address_groups)),
            ("Proxy Addresses", len(self.ir.proxy_addresses)),
            (
                "Web Proxy Settings",
                1 if self.ir.web_proxy_settings is not None else 0,
            ),
            ("Service Categories", len(self.ir.service_categories)),
            ("Services", len(self.ir.services)),
            ("Service Groups", len(self.ir.service_groups)),
            ("Session Helpers", len(self.ir.session_helpers)),
            (
                "Session TTL Overrides",
                len(self.ir.session_ttl_overrides),
            ),
            ("Schedules", len(self.ir.schedules)),
            ("Traffic Shapers", len(self.ir.traffic_shapers)),
            ("Policies", len(self.ir.policies)),
            ("ZTNA Providers", len(self.ir.ztna_providers)),
            ("IP Pools", len(self.ir.ip_pools)),
            ("Virtual IPs", len(self.ir.virtual_ips)),
            (
                "VIP Real Servers",
                sum(
                    len(vip.real_servers)
                    for vip in self.ir.virtual_ips
                ),
            ),
            ("VIP Groups", len(self.ir.virtual_ip_groups)),
            ("NAT Rules", len(self.ir.nat_rules)),
            ("VPN Tunnels", len(self.ir.vpn_tunnels)),
            ("VPN Phase 2", len(self.ir.vpn_phase2)),
            ("SSL VPN Portals", len(self.ir.ssl_vpn_portals)),
            (
                "SD-WAN Rules",
                len(self.ir.sdwan.rules)
                if self.ir.sdwan
                else 0,
            ),
            ("LDAP Servers", len(self.ir.user_ldap_servers)),
            ("SAML Servers", len(self.ir.user_saml_servers)),
            ("FSSO Servers", len(self.ir.fsso_providers)),
            ("FSSO AD Groups", len(self.ir.fsso_ad_groups)),
            ("Local Users", len(self.ir.local_users)),
            ("User Groups", len(self.ir.user_groups)),
            ("DoS Policies", len(self.ir.dos_policies)),
            ("Firewall Sniffers", len(self.ir.firewall_sniffers)),
            ("Certificates", len(self.ir.certificates)),
            ("Routes", len(self.ir.routes)),
            ("Internet Services", len(self.ir.internet_services)),
            ("Internet Service Definitions", len(self.ir.internet_service_definitions)),
            (
                "Internet Service Def Entries",
                sum(len(definition.entries) for definition in self.ir.internet_service_definitions),
            ),
            (
                "Internet Service Def Ports",
                sum(
                    len(entry.port_ranges)
                    for definition in self.ir.internet_service_definitions
                    for entry in definition.entries
                ),
            ),
            ("IPS Sensors", len(self.ir.ips_sensors)),
            (
                "IPS Sensor Entries",
                sum(
                    len(sensor.entries)
                    for sensor in self.ir.ips_sensors
                ),
            ),
            (
                "Security Profiles",
                len(self.ir.security_profile_groups),
            ),
            ("Warnings", len(self.ir.audit_entries)),
            ("Unsupported Items", unsupported_count),
            ("Unresolved References", unresolved_count),
        ]

        self._summary_section(
            sheet,
            navigation_end_row + 2,
            "Inventory Counts",
            inventory_rows,
        )

        sheet.column_dimensions["A"].width = 22
        sheet.column_dimensions["B"].width = 34
        sheet.column_dimensions["C"].width = 12
        sheet.column_dimensions["D"].width = 48
        sheet.column_dimensions["E"].width = 16

        sheet.freeze_panes = "A4"

    @staticmethod
    def _set_internal_link(cell: Any, sheet_name: str) -> None:
        escaped = sheet_name.replace("'", "''")
        cell.hyperlink = f"#'{escaped}'!A1"
        cell.style = "Hyperlink"

    def _summary_section(
        self, sheet: Any, start_row: int, title: str, rows: Sequence[tuple[str, Any]]
    ) -> None:
        sheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=2)
        title_cell = sheet.cell(start_row, 1, title)
        title_cell.font = Font(name="Aptos", size=11, bold=True, color=self._WHITE)
        title_cell.fill = PatternFill("solid", fgColor=self._TEAL)
        title_cell.alignment = Alignment(vertical="center")
        sheet.row_dimensions[start_row].height = 23
        for row_index, (label, value) in enumerate(rows, start_row + 1):
            sheet.cell(row_index, 1, self._safe_value(label))
            sheet.cell(row_index, 2, self._safe_value(value))
            sheet.cell(row_index, 1).font = Font(name="Aptos", bold=True, color=self._TEXT)
            sheet.cell(row_index, 1).fill = PatternFill("solid", fgColor=self._LIGHT_BLUE)
            sheet.cell(row_index, 2).font = Font(name="Aptos", color=self._TEXT)
            sheet.cell(row_index, 2).alignment = Alignment(wrap_text=True, vertical="top")

    def _build_summary_navigation(
        self,
        sheet: Any,
        workbook: Any,
        start_row: int,
    ) -> int:
        """Build workbook navigation using the already-generated worksheets."""
        sheet.merge_cells(
            start_row=start_row,
            start_column=1,
            end_row=start_row,
            end_column=5,
        )

        title_cell = sheet.cell(
            start_row,
            1,
            "Workbook Navigation",
        )
        title_cell.font = Font(
            name="Aptos",
            size=11,
            bold=True,
            color=self._WHITE,
        )
        title_cell.fill = PatternFill(
            "solid",
            fgColor=self._TEAL,
        )
        title_cell.alignment = Alignment(
            vertical="center",
        )
        sheet.row_dimensions[start_row].height = 23

        header_row = start_row + 1

        headers = (
            "Category",
            "Sheet",
            "Records",
            "Purpose",
            "Manual Review",
        )

        for column, header in enumerate(headers, 1):
            cell = sheet.cell(
                header_row,
                column,
                header,
            )
            cell.font = Font(
                name="Aptos",
                bold=True,
                color=self._WHITE,
            )
            cell.fill = PatternFill(
                "solid",
                fgColor=self._NAVY,
            )
            cell.alignment = Alignment(
                wrap_text=True,
                vertical="center",
            )

        generated_sheets = {
            worksheet.title: worksheet
            for worksheet in workbook.worksheets
            if worksheet.title != "Summary"
        }

        row_index = header_row

        for sheet_name in self.SHEET_ORDER:
            if sheet_name == "Summary":
                continue

            target_sheet = generated_sheets.get(sheet_name)
            if target_sheet is None:
                continue

            row_index += 1

            category = self._sheet_category(sheet_name)

            # All normal table worksheets use:
            # row 1 = title
            # row 2 = note
            # row 3 = headers
            #
            # Therefore max_row - 3 is the actual record count.
            record_count = max(
                target_sheet.max_row - 3,
                0,
            )

            values = (
                category,
                sheet_name,
                record_count,
                self._sheet_purpose(
                    sheet_name,
                    category,
                ),
                (
                    "Yes"
                    if self._sheet_requires_review(
                        sheet_name,
                        category,
                    )
                    else "No"
                ),
            )

            for column, value in enumerate(values, 1):
                cell = sheet.cell(
                    row_index,
                    column,
                    self._safe_value(value),
                )
                cell.font = Font(
                    name="Aptos",
                    size=10,
                    color=self._TEXT,
                )
                cell.alignment = Alignment(
                    wrap_text=True,
                    vertical="top",
                )

            self._set_internal_link(
                sheet.cell(row_index, 2),
                sheet_name,
            )

            if (row_index - header_row) % 2 == 0:
                for column in range(1, 6):
                    sheet.cell(
                        row_index,
                        column,
                    ).fill = PatternFill(
                        "solid",
                        fgColor="F8FAFC",
                    )

        sheet.auto_filter.ref = (
            f"A{header_row}:E{row_index}"
        )

        return row_index


    def _sheet_category(
        self,
        sheet_name: str,
    ) -> str:
        if sheet_name in self.CORE_INVENTORY_SHEETS:
            return "Core Inventory"

        if sheet_name in self.NETWORK_ACCESS_SHEETS:
            return "Network / Access"

        if sheet_name in self.SOURCE_DETAIL_SHEETS:
            return "Source Detail"

        if sheet_name in self.AUDIT_SHEETS:
            return "Audit"

        return "Overview"


    @staticmethod
    def _sheet_purpose(
        sheet_name: str,
        category: str,
    ) -> str:
        purposes = {
            "System Settings": "System-level firewall settings",
            "DNS Settings": "Configured DNS settings",
            "Interfaces": "Interface inventory",
            "Zones": "Security/interface zones",
            "Addresses": "Address objects",
            "Address Groups": "Address object groups",
            "Services": "Service and protocol objects",
            "Service Groups": "Service object groups",
            "Schedules": "Policy schedule objects",
            "Policies": "Firewall security policies",
            "IP Pools": "Source NAT pools",
            "Virtual IPs": "Destination NAT/VIP objects",
            "VIP Real Servers": "VIP backend servers",
            "VIP Groups": "FortiGate VIP groups",
            "NAT Rules": "Normalized NAT inventory",
            "Routes": "Static route inventory",
            "VPN Tunnels": "IPsec Phase 1 / tunnel inventory",
            "VPN Phase 2": "IPsec Phase 2 selectors and settings",
            "Interface Source Settings": (
                "Explicit FortiGate/source interface settings"
            ),
            "Source Security Profiles": (
                "Source security-profile inventory"
            ),
            "Source Security Profile Setting": (
                "Detailed source security-profile settings"
            ),
            "Warnings": (
                "Extraction and migration review warnings"
            ),
            "Unsupported": (
                "Unsupported source sections/items"
            ),
            "Extraction Coverage": (
                "Source-to-parser extraction coverage"
            ),
        }

        return purposes.get(
            sheet_name,
            f"{category}: {sheet_name}",
        )


    @staticmethod
    def _sheet_requires_review(
        sheet_name: str,
        category: str,
    ) -> bool:
        if category in {
            "Source Detail",
            "Audit",
        }:
            return True

        return sheet_name in {
            "Addresses",
            "Interface Secondary IPs",
            "Policies",
            "NAT Rules",
            "Routes",
            "VPN Tunnels",
            "VPN Phase 2",
        }


    def _order_sheets(
        self,
        workbook: Any,
    ) -> None:
        """Validate and apply deterministic logical workbook ordering."""
        if len(self.SHEET_ORDER) != len(
            set(self.SHEET_ORDER)
        ):
            duplicates = sorted(
                {
                    sheet_name
                    for sheet_name in self.SHEET_ORDER
                    if self.SHEET_ORDER.count(
                        sheet_name
                    ) > 1
                }
            )

            raise ValueError(
                "Excel sheet order contains duplicate entries: "
                f"{duplicates}"
            )

        expected = set(self.SHEET_ORDER)
        actual = {
            worksheet.title
            for worksheet in workbook.worksheets
        }

        unknown = actual - expected

        if unknown:
            raise ValueError(
                "Excel sheet order missing entries for: "
                f"{sorted(unknown)}"
            )

        workbook._sheets.sort(
            key=lambda worksheet: (
                self.SHEET_ORDER.index(
                    worksheet.title
                )
            )
        )

        for worksheet in workbook.worksheets:
            worksheet.sheet_properties.tabColor = (
                self._sheet_tab_color(
                    worksheet.title
                )
            )


    def _sheet_tab_color(
        self,
        sheet_name: str,
    ) -> str:
        if sheet_name == "Summary":
            return self._NAVY

        if sheet_name in self.CORE_INVENTORY_SHEETS:
            return self._TEAL

        if sheet_name in self.NETWORK_ACCESS_SHEETS:
            return self._LIGHT_BLUE

        if sheet_name in self.SOURCE_DETAIL_SHEETS:
            return self._MUTED

        if sheet_name == "Unsupported":
            return self._LIGHT_RED

        if sheet_name in self.AUDIT_SHEETS:
            return self._LIGHT_AMBER

        return self._LIGHT_BLUE
    def _build_interfaces(self, workbook: Any) -> None:
        headers = (
            "Name", "Source VDOM", "Zone", "IP / Prefix", "Remote IP / Prefix",
            "Enabled", "Interface Type",
            "Role", "Addressing Mode", "DHCP Client", "Management Access", "Alias",
            "Parent / Underlay Interface",
            "Tag", "VLAN ID", "Management Profile", "PPPoE Mode", "PPPoE Username",
            "Description",
        )
        rows = [
            (
                item.name, item.source_vdom, item.zone, item.ip, item.remote_ip, item.status,
                item.interface_type,
                item.role, item.addressing_mode, item.dhcp_client, item.management_access,
                item.alias, item.parent, item.tag, item.vlanid, item.management_profile,
                item.pppoe_mode, item.pppoe_username, item.description,
            )
            for item in self.ir.interfaces
        ]
        self._table_sheet(workbook, "Interfaces", headers, rows)

    def _build_interface_secondary_ips(self, workbook: Any) -> None:
        headers = (
            "Interface",
            "Source ID",
            "Source IP",
            "IP / Prefix",
            "Management Access",
            "Extraction Status",
            "Manual Review",
            "Parse Error",
            "Additional Settings",
        )
        rows = []
        for intf in self.ir.interfaces:
            for sec in getattr(intf, "secondary_ips", []):
                if sec.parse_error:
                    status = "PARSE_ERROR"
                elif sec.requires_manual_review:
                    status = "PARTIALLY_NORMALIZED"
                else:
                    status = "NORMALIZED"

                rows.append(
                    (
                        intf.name,
                        sec.source_id,
                        sec.source_ip,
                        sec.ip,
                        sec.management_access,
                        status,
                        self._optional_bool_literal(sec.requires_manual_review),
                        sec.parse_error,
                        self._format_settings(sec.source_attributes),
                    )
                )

        self._table_sheet(
            workbook,
            "Interface Secondary IPs",
            headers,
            rows,
            empty_note="No secondary interface IP addresses were extracted from the source firewall.",
            subtitle="Secondary interface IP configuration extracted for migration and inventory review.",
        )

    def _build_interface_source_settings(self, workbook: Any) -> None:
        """Expose every explicitly configured interface setting without reinterpreting it."""
        rows = []
        for item in self.ir.interfaces:
            for setting, value in item.source_attributes.items():
                display_setting = str(setting).replace("_", "-")
                if isinstance(value, set):
                    display_value = json.dumps(sorted(value), ensure_ascii=False, default=str)
                elif isinstance(value, (dict, list, tuple)):
                    display_value = json.dumps(
                        value, ensure_ascii=False, sort_keys=True, default=str
                    )
                else:
                    display_value = value
                rows.append(
                    (
                        item.name,
                        self.ir.metadata.source_vendor,
                        display_setting,
                        display_value,
                        "EXTRACT_ONLY",
                    )
                )

        self._table_sheet(
            workbook,
            "Interface Source Settings",
            ("Interface", "Source Vendor", "Setting", "Value", "Extraction Status"),
            rows,
            empty_note="No explicit source-interface settings were retained by the parser.",
            subtitle=(
                "Explicit source settings retained for inventory. These values are extraction-only "
                "and are not consumed by target generators."
            ),
        )

    def _build_dhcp_servers(
        self,
        workbook: Any,
    ) -> None:
        rows = [
        (
            item.source_id,
            item.interface,
            item.enabled,
            item.default_gateway,
            item.netmask,
            item.lease_time_seconds,
            item.dns_service,
            item.dns_servers,
            item.timezone_option,
            item.migration_status,
            item.requires_manual_review,
            self._format_settings(
                item.source_attributes
            ),
        )
        for item in self.ir.dhcp_servers
        ]

        self._table_sheet(
            workbook,
            "DHCP Servers",
        (
            "Server ID",
            "Interface",
            "Enabled",
            "Default Gateway",
            "Netmask",
            "Lease Time (Seconds)",
            "DNS Service",
            "DNS Servers",
            "Timezone Option",
            "Extraction Status",
            "Manual Review",
            "Additional Settings",
        ),
            rows,
            empty_note=(
            "No DHCP server configuration was "
            "extracted from the source firewall."
        ),
            subtitle=(
            "DHCP server configuration retained for "
            "migration review."
        ),
        )


    def _build_dhcp_ip_ranges(
        self,
        workbook: Any,
    ) -> None:
        rows = [
        (
            server.source_id,
            server.interface,
            item.source_id,
            item.start_ip,
            item.end_ip,
            "EXTRACT_ONLY",
            True,
            self._format_settings(
                item.source_attributes
            ),
        )
        for server in self.ir.dhcp_servers
        for item in server.ip_ranges
        ]

        self._table_sheet(
            workbook,
            "DHCP IP Ranges",
        (
            "Server ID",
            "Interface",
            "Range ID",
            "Start IP",
            "End IP",
            "Extraction Status",
            "Manual Review",
            "Additional Settings",
        ),
            rows,
            empty_note=(
            "No DHCP IP ranges were extracted "
            "from the source firewall."
        ),
        )


    def _build_dhcp_reservations(
        self,
        workbook: Any,
    ) -> None:
        rows = [
        (
            server.source_id,
            server.interface,
            item.source_id,
            item.ip_address,
            item.mac_address,
            "EXTRACT_ONLY",
            True,
            self._format_settings(
                item.source_attributes
            ),
        )
        for server in self.ir.dhcp_servers
        for item in server.reservations
        ]

        self._table_sheet(
            workbook,
            "DHCP Reservations",
        (
            "Server ID",
            "Interface",
            "Reservation ID",
            "IP Address",
            "MAC Address",
            "Extraction Status",
            "Manual Review",
            "Additional Settings",
        ),
            rows,
            empty_note=(
            "No DHCP reservations were extracted "
            "from the source firewall."
        ),
        )

    def _build_zones(self, workbook: Any) -> None:
        rows = [(item.name, item.interfaces, item.description) for item in self.ir.zones]
        self._table_sheet(workbook, "Zones", ("Name", "Interfaces", "Description"), rows)

    def _build_addresses(self, workbook: Any) -> None:
        rows = [
            (
                item.name,
                item.source_uuid,
                item.type,
                item.value,
                item.source_section,
                item.address_family,
                item.source_type,
                item.original_type,
                item.original_value,
                item.is_ipv6,
                item.is_multicast,
                item.associated_interface,
                self._optional_bool_literal(
                    item.allow_routing
                ),
                item.source_color,
                item.source_sub_type,
                item.source_obj_tag,
                item.source_tag_type,
                item.source_obj_type,
                item.source_dirty,
                item.tags,
                item.source_list_entries,
                [
                    {"name": entry.name, "category": entry.category, "tags": entry.tags}
                    for entry in item.source_tagging_entries
                ],
                item.migration_status,
                item.requires_manual_review,
                item.audit_note,
                item.parse_error,
                self._format_settings(
                    item.source_attributes
                ),
                item.description,
            )
            for item in self.ir.addresses
        ]
        self._table_sheet(
            workbook,
            "Addresses",
            (
                "Name",
                "Source UUID",
                "Type",
                "Value",
                "Source Section",
                "Address Family",
                "Source Type",
                "Original Type",
                "Original Value",
                "IPv6",
                "Multicast",
                "Associated Interface",
                "Allow Routing",
                "Source Color",
                "EMS Sub-Type",
                "EMS Object Tag",
                "EMS Tag Type",
                "EMS Object Type",
                "EMS Dirty",
                "Tags",
                "IP List",
                "Object Tagging",
                "Migration Status",
                "Manual Review",
                "Audit Note",
                "Parse Error",
                "Additional Settings",
                "Description",
            ),
            rows,
        )

    def _build_address_groups(self, workbook: Any) -> None:
        rows = [
            (
                item.name,
                item.source_uuid,
                item.members,
                item.is_dynamic,
                item.dynamic_filter,
                self._optional_bool_literal(
                    item.allow_routing
                ),
                item.source_color,
                item.source_category,
                item.source_sub_type,
                item.source_obj_tag,
                item.source_tag_type,
                item.source_obj_type,
                item.source_dirty,
                item.tags,
                item.source_section,
                item.address_family,
                item.exclusion_enabled,
                item.exclude_members,
                item.source_exclude_setting,
                item.source_group_type,
                item.source_fabric_object_setting,
                item.migration_status,
                self._optional_bool_literal(item.requires_manual_review),
                item.audit_note,
                self._format_settings(
                    item.source_attributes
                ),
                item.description,
            )
            for item in self.ir.address_groups
        ]
        self._table_sheet(
            workbook,
            "Address Groups",
            (
                "Name",
                "Source UUID",
                "Members",
                "Dynamic",
                "Dynamic Filter",
                "Allow Routing",
                "Source Color",
                "Source Category",
                "EMS Sub-Type",
                "EMS Object Tag",
                "EMS Tag Type",
                "EMS Object Type",
                "EMS Dirty",
                "Tags",
                "Source Section",
                "Address Family",
                "Exclusion Enabled",
                "Exclude Members",
                "Source Exclude Setting",
                "Group Type",
                "Fabric Object",
                "Migration Status",
                "Manual Review",
                "Audit Note",
                "Additional Settings",
                "Description",
            ),
            rows,
        )

    def _build_service_categories(self, workbook: Any) -> None:
        rows = [
            (
                item.name,
                item.description,
                item.source_fabric_object,
                item.migration_status,
                self._format_settings(
                    item.source_attributes
                ),
            )
            for item in self.ir.service_categories
        ]
        self._table_sheet(
            workbook,
            "Service Categories",
            (
                "Name",
                "Description",
                "Fabric Object",
                "Extraction Status",
                "Additional Settings",
            ),
            rows,
        )

    def _build_services(self, workbook: Any) -> None:
        rows = [
            (
                item.name,
                item.source_uuid,
                item.source_category,
                item.source_protocol,
                [
                    self._format_port(port)
                    for port in item.ports
                ],
                [
                    port.source_port
                    for port in item.ports
                    if port.source_port is not None
                ],
                self._optional_bool_literal(
                    item.source_proxy
                ),
                self._optional_bool_literal(
                    item.requires_manual_review
                ),
                item.audit_note,
                self._format_settings(
                    item.source_attributes
                ),
                item.description,
            )
            for item in self.ir.services
        ]
        self._table_sheet(
            workbook,
            "Services",
            (
                "Name",
                "Source UUID",
                "Category",
                "Source Protocol",
                "Protocol / Destination Port",
                "Source Port Constraint",
                "Proxy",
                "Manual Review",
                "Audit Note",
                "Additional Settings",
                "Description",
            ),
            rows,
        )

    def _build_service_groups(
    self,
    workbook: Any,
    ) -> None:
        rows = [
            (
                item.name,
                item.source_uuid,
                item.members,
                self._optional_bool_literal(item.source_proxy),
                item.source_color,
                item.source_fabric_object,
                item.migration_status,
                self._optional_bool_literal(item.requires_manual_review),
                item.audit_note,
                self._format_settings(
                    item.source_attributes
                ),
                item.description,
            )
            for item in self.ir.service_groups
        ]

        self._table_sheet(
            workbook,
            "Service Groups",
            (
                "Name",
                "Source UUID",
                "Members",
                "Proxy",
                "Source Color",
                "Fabric Object",
                "Migration Status",
                "Manual Review",
                "Audit Note",
                "Additional Settings",
                "Description",
            ),
            rows,
        )

    def _build_session_helpers(
        self,
        workbook: Any,
    ) -> None:
        """
        Export FortiGate session-helper / ALG inventory.

        Session helpers influence protocol handling but should not be
        converted into normal firewall service objects.
        """

        rows = [
            (
                item.source_id,
                item.name,
                item.protocol_name,
                item.protocol_number,
                item.port,
                item.classification,
                item.migration_status,
                item.requires_manual_review,
                self._format_settings(
                    item.source_attributes
                ),
            )
            for item in self.ir.session_helpers
        ]

        sheet = self._table_sheet(
            workbook,
            "Session Helpers",
            (
                "Source ID",
                "Name",
                "Protocol",
                "Protocol Number",
                "Port",
                "Classification",
                "Extraction Status",
                "Manual Review",
                "Additional Settings",
            ),
            rows,
            empty_note=(
                "No FortiGate session-helper entries were "
                "extracted from the source configuration."
            ),
            subtitle=(
                "FortiGate protocol/session helpers retained for "
                "traffic-behavior inventory. DEFAULT entries match "
                "the known FortiOS baseline. CUSTOM, CUSTOMIZED, "
                "or UNKNOWN entries require target-platform review. "
                "Session helpers are not converted into service objects."
            ),
        )

        for row in range(4, sheet.max_row + 1):
            classification = str(
                sheet.cell(row, 6).value or ""
            ).upper()

            if classification in {
                "CUSTOM",
                "CUSTOMIZED",
                "UNKNOWN",
            }:
                for column in range(1, 10):
                    sheet.cell(
                        row,
                        column,
                    ).fill = PatternFill(
                        "solid",
                        fgColor=self._LIGHT_AMBER,
                    )

    def _build_session_ttl_overrides(
        self,
        workbook: Any,
    ) -> None:
        """
        Export explicit FortiGate session timeout overrides.

        These settings affect actual session behaviour and require
        target-platform review.
        """

        rows = [
            (
                item.source_id,
                item.protocol_name,
                item.protocol_number,
                item.start_port,
                item.end_port,
                item.timeout_seconds,
                item.migration_status,
                item.requires_manual_review,
                self._format_settings(
                    item.source_attributes
                ),
            )
            for item in self.ir.session_ttl_overrides
        ]

        sheet = self._table_sheet(
            workbook,
            "Session TTL Overrides",
            (
                "Source ID",
                "Protocol",
                "Protocol Number",
                "Start Port",
                "End Port",
                "Timeout (Seconds)",
                "Extraction Status",
                "Manual Review",
                "Additional Settings",
            ),
            rows,
            empty_note=(
                "No explicit session TTL port overrides were "
                "extracted from the source configuration."
            ),
            subtitle=(
                "Explicit source session timeout overrides retained "
                "for traffic-behavior migration review. These settings "
                "are target-platform dependent and are not automatically "
                "converted into service or policy objects."
            ),
        )

        for row in range(4, sheet.max_row + 1):
            for column in range(1, 10):
                sheet.cell(
                    row,
                    column,
                ).fill = PatternFill(
                    "solid",
                    fgColor=self._LIGHT_AMBER,
                )

    def _build_schedules(
        self,
        workbook: Any,
    ) -> None:
        rows = [
            (
                item.name,
                item.schedule_type,
                item.start,
                item.end,
                item.days,
                item.source_color,
                item.expiration_days,
                self._format_settings(item.source_attributes),
            )
            for item in self.ir.schedules
        ]

        self._table_sheet(
            workbook,
            "Schedules",
            (
                "Name",
                "Type",
                "Start",
                "End",
                "Days",
                "Color",
                "Expiration Days",
                "Additional Settings",
            ),
            rows,
        )

    def _build_traffic_shapers(self, workbook: Any) -> None:
        rows = [
            (
                item.name,
                item.guaranteed_bandwidth,
                item.maximum_bandwidth,
                item.source_bandwidth_unit,
                item.priority,
                self._optional_bool_literal(item.per_policy),
                item.migration_status,
                self._optional_bool_literal(item.requires_manual_review),
                self._format_settings(item.source_attributes),
            )
            for item in self.ir.traffic_shapers
        ]
        self._table_sheet(
            workbook,
            "Traffic Shapers",
            (
                "Name",
                "Guaranteed Bandwidth",
                "Maximum Bandwidth",
                "Source Bandwidth Unit",
                "Priority",
                "Per Policy",
                "Extraction Status",
                "Manual Review",
                "Additional Settings",
            ),
            rows,
            subtitle=(
                "FortiGate shaping inventory; exact target QoS behavior requires manual review."
            ),
        )

    def _build_proxy_addresses(self, workbook: Any) -> None:
        rows = [
            (
                item.name,
                item.source_uuid,
                item.proxy_address_type,
                item.host,
                item.host_regex,
                item.path,
                item.query,
                item.migration_status,
                self._optional_bool_literal(item.requires_manual_review),
                self._format_settings(item.source_attributes),
            )
            for item in self.ir.proxy_addresses
        ]
        self._table_sheet(
            workbook,
            "Proxy Addresses",
            (
                "Name",
                "Source UUID",
                "Type",
                "Host",
                "Host Regex",
                "Path",
                "Query",
                "Extraction Status",
                "Manual Review",
                "Additional Settings",
            ),
            rows,
            subtitle=(
                "Source proxy-address inventory retained without conversion to firewall addresses."
            ),
        )

    def _build_web_proxy_settings(self, workbook: Any) -> None:
        settings = self.ir.web_proxy_settings
        rows = [] if settings is None else [
            (
                settings.proxy_fqdn,
                settings.migration_status,
                self._optional_bool_literal(settings.requires_manual_review),
                self._format_settings(settings.source_attributes),
            )
        ]
        self._table_sheet(
            workbook,
            "Web Proxy Settings",
            (
                "Proxy FQDN",
                "Extraction Status",
                "Manual Review",
                "Additional Settings",
            ),
            rows,
            empty_note="No web-proxy global settings were extracted.",
            subtitle="FortiGate global web-proxy settings retained as source inventory.",
        )

    def _build_policies(self, workbook: Any) -> None:
        rows = [
            (
                index, item.source_rule_id, item.source_uuid, item.name, item.source_from_interfaces,
                item.from_zone, item.source_to_interfaces, item.to_zone,
                item.source_address_references, item.source,
                item.source_address_negate_setting,
                item.source_ipv6_address_references,
                item.source_ipv6_address_negate_setting,
                item.destination_address_references, item.destination,
                item.destination_address_negate_setting,
                item.destination_ipv6_address_references,
                item.destination_ipv6_address_negate_setting,
                item.source_user_groups, item.source_users,
                item.source_service_references, item.service,
                item.source_service_negate_setting,
                item.source_action, item.action, item.source_schedule, item.schedule,
                item.disabled,
                item.source_vpn_tunnel,
                item.source_log_setting, item.source_log_start_setting,
                item.source_utm_status, item.log_start, item.log_end,
                self._optional_bool_literal(item.nat_enabled),
                self._optional_bool_literal(item.nat_pool_enabled), item.nat_pool_names,
                item.nat_pool_names6,
                item.applications, item.source_internet_service_status,
                item.internet_service, item.security_profile_group,
                item.antivirus, item.ips_sensor, item.webfilter, item.application_list,
                item.ssl_ssh_profile, item.source_profile_type,
                item.source_profile_group, item.source_profile_protocol_options,
                item.source_inspection_mode, item.source_ztna_status,
                item.source_ztna_ems_tags, self._format_settings(item.source_extra_settings),
                item.migration_status,
                self._optional_bool_literal(item.requires_manual_review),
                item.description,
            )
            for index, item in enumerate(self.ir.policies, 1)
        ]
        sheet = self._table_sheet(
            workbook,
            "Policies",
            (
                "Rule #",
                "Source Policy ID",
                "Source UUID",
                "Name",
                "Source Interface",
                "From Zone",
                "Destination Interface",
                "To Zone",
                "Source Address (FortiGate)",
                "Source Address (Normalized)",
                "Source Address Negate",
                "Source IPv6 Address",
                "Source IPv6 Address Negate",
                "Destination Address (FortiGate)",
                "Destination Address (Normalized)",
                "Destination Address Negate",
                "Destination IPv6 Address",
                "Destination IPv6 Address Negate",
                "User Groups",
                "Users",
                "Service (FortiGate)",
                "Service (Normalized)",
                "Service Negate",
                "Action (FortiGate)",
                "Action (Normalized)",
                "Schedule (FortiGate)",
                "Schedule (Normalized)",
                "Disabled",
                "VPN Tunnel",
                "Log Setting",
                "Log Start Setting",
                "UTM Status",
                "Log Start",
                "Log End",
                "NAT Enabled",
                "IP Pool Enabled",
                "NAT Pool",
                "NAT Pool IPv6",
                "Applications",
                "Internet Service Status",
                "Internet Services",
                "Security Profile Group",
                "Antivirus",
                "IPS Sensor",
                "Web Filter",
                "Application List",
                "SSL/SSH Profile",
                "Source Profile Type",
                "Source Profile Group",
                "Profile Protocol Options",
                "Inspection Mode",
                "ZTNA Status",
                "ZTNA EMS Tags",
                "Additional Settings",
                "Extraction Status",
                "Manual Review",
                "Description",
            ),
            rows,
        )

        # Keep title/note/header visible and retain the four most useful
        # identifier columns while scrolling horizontally.
        sheet.freeze_panes = "E4"

    def _build_firewall_policy_source_settings(self, workbook: Any) -> None:
        items = (
            []
            if self.extraction is None
            else [
                item
                for item in self.extraction.inventory_items
                if item.source_path == "firewall policy"
            ]
        )

        self._table_sheet(
            workbook,
            "Firewall Policy Source Settings",
            (
                "Source Policy ID",
                "Policy Name",
                "Operation",
                "Setting",
                "Ordered Source Values",
            ),
            (
                (
                    item.source_id,
                    item.name,
                    command.operation,
                    command.key,
                    json.dumps(list(command.values), ensure_ascii=False),
                )
                for item in items
                for command in item.commands
            ),
            empty_note="No FortiGate firewall policy source commands were retained.",
            subtitle=(
                "Sanitized, ordered FortiGate policy commands retained for audit. "
                "This extraction-only detail is not consumed by target generators."
            ),
        )

    def _build_address_group_tags(self, workbook: Any) -> None:
        rows = [
            (group.name, group.source_section, group.address_family, entry.name,
             entry.category, entry.tags, entry.migration_status,
             self._optional_bool_literal(entry.requires_manual_review),
             self._format_settings(entry.source_attributes))
            for group in self.ir.address_groups
            for entry in group.source_tagging_entries
        ]
        self._table_sheet(workbook, "Address Group Tags", (
            "Group Name", "Source Section", "Address Family", "Tag Entry",
            "Category", "Tags", "Extraction Status", "Manual Review",
            "Additional Settings",
        ), rows)

    def _build_ztna_providers(self, workbook: Any) -> None:
        """
        Export ZTNA / endpoint-posture provider dependencies.

        Policy IDs and EMS tags shown here are observed elsewhere in the
        same source configuration. They are not asserted to belong to a
        specific provider unless explicit source correlation exists.
        """

        # Collect all policies that contain ZTNA intent or EMS tag references.
        ztna_policies = [
            policy
            for policy in self.ir.policies
            if (
                policy.source_ztna_status == "enable"
                or policy.source_ztna_ems_tags
            )
        ]

        # Preserve policy order while removing duplicates.
        observed_policy_ids = list(
            dict.fromkeys(
                policy.source_rule_id or policy.name
                for policy in ztna_policies
                if policy.source_rule_id or policy.name
            )
        )

        # Collect all ZTNA EMS tags referenced by those policies.
        observed_ems_tags = list(
            dict.fromkeys(
                tag
                for policy in ztna_policies
                for tag in policy.source_ztna_ems_tags
                if tag
            )
        )

        rows = [
            (
                item.name,
                item.source_vendor or self.ir.metadata.source_vendor,
                item.source_id,
                item.provider_type,
                self._optional_bool_literal(item.enabled),
                item.source_serial,
                item.source_tenant_id,
                self._optional_bool_literal(
                    item.source_cloud_authentication
                ),
                item.verifying_ca,
                item.verified_cn,
                item.capabilities,
                observed_policy_ids,
                observed_ems_tags,
                item.migration_status,
                self._optional_bool_literal(
                    item.requires_manual_review
                ),
                item.migration_instruction,
                self._format_settings(
                    item.source_attributes
                ),
            )
            for item in self.ir.ztna_providers
        ]

        self._table_sheet(
            workbook,
            "ZTNA Providers",
            (
                "Name",
                "Source Vendor",
                "Source ID",
                "Provider Type",
                "Enabled",
                "Source Serial",
                "Tenant ID",
                "Cloud Authentication",
                "Verifying CA",
                "Verified CN",
                "Capabilities",
                "ZTNA Policy IDs (Observed)",
                "ZTNA EMS Tags (Observed)",
                "Extraction Status",
                "Manual Review",
                "Migration Instruction",
                "Additional Settings",
            ),
            rows,
            empty_note=(
                "No meaningful ZTNA / endpoint-posture providers "
                "were extracted from the source configuration."
            ),
            subtitle=(
                "Source ZTNA and endpoint-posture dependencies retained "
                "for migration review. Policy IDs and EMS tags are observed "
                "in the same source configuration and are not automatically "
                "claimed as an exact mapping to an individual provider. "
                "Provider-specific configuration is not consumed by target "
                "generators."
            ),
        )

    def _build_ip_pools(self, workbook: Any) -> None:
        rows = [
            (
                item.name,
                item.address_family,
                item.pool_type,
                item.start_ip,
                item.end_ip,
                item.source_start_ip,
                item.source_end_ip,
                item.source_prefix6,
                item.start_port,
                item.end_port,
                item.associated_interface,
                self._optional_bool_literal(item.arp_reply),
                item.arp_interface,
                self._optional_bool_literal(item.permit_any_host),
                item.excluded_ips,
                item.block_size,
                item.blocks_per_user,
                item.pba_timeout,
                item.pba_interim_log,
                item.ports_per_user,
                self._optional_bool_literal(item.privileged_port_use_pba),
                self._optional_bool_literal(item.nat64),
                self._optional_bool_literal(item.add_nat64_route),
                self._optional_bool_literal(item.nat46),
                self._optional_bool_literal(item.add_nat46_route),
                item.client_prefix_length,
                self._optional_bool_literal(item.include_subnet_broadcast),
                item.cgn_block_size,
                item.cgn_client_start_ip,
                item.cgn_client_end_ip,
                item.cgn_client_ipv6_shift,
                self._optional_bool_literal(item.cgn_fixed_allocation),
                self._optional_bool_literal(item.cgn_overload),
                item.cgn_port_start,
                item.cgn_port_end,
                self._optional_bool_literal(item.cgn_spa),
                item.utilization_alarm_clear,
                item.utilization_alarm_raise,
                item.tcp_session_quota,
                item.udp_session_quota,
                item.icmp_session_quota,
                item.migration_status,
                self._optional_bool_literal(item.requires_manual_review),
                item.audit_note,
                self._format_settings(item.source_attributes),
                item.description,
            )
            for item in self.ir.ip_pools
        ]
        self._table_sheet(
            workbook,
            "IP Pools",
            (
                "Name", "Address Family", "Type", "Start IP", "End IP", "Source Start IP",
                "Source End IP", "Source Prefix6", "Start Port", "End Port", "Associated Interface",
                "ARP Reply", "ARP Interface", "Permit Any Host", "Excluded IPs",
                "Block Size", "Blocks Per User", "PBA Timeout", "PBA Interim Log",
                "Ports Per User", "Privileged Port Uses PBA", "NAT64", "Add NAT64 Route",
                "NAT46", "Add NAT46 Route", "Client Prefix Length",
                "Include Subnet/Broadcast", "CGN Block Size", "CGN Client Start IP",
                "CGN Client End IP", "CGN Client IPv6 Shift", "CGN Fixed Allocation",
                "CGN Overload", "CGN Port Start", "CGN Port End", "CGN SPA",
                "Utilization Alarm Clear", "Utilization Alarm Raise", "TCP Session Quota",
                "UDP Session Quota", "ICMP Session Quota", "Extraction Status",
                "Manual Review", "Review Reason", "Additional Settings", "Description",
            ),
            rows,
        )

    def _build_nat_rules(self, workbook: Any) -> None:
        rows = [
            (
                index, item.name, item.type, item.source_policy_reference,
                item.source_policy_uuid, self._optional_bool_literal(item.enabled),
                item.source_from_interfaces, item.from_zone, item.source_to_interfaces,
                item.to_zone, item.source, item.destination, item.services,
                item.internet_services, item.source_translation_mode,
                item.source_pool_references, item.translated_sources,
                item.source_pool_type, item.source_pool_excluded_ips,
                self._optional_bool_literal(item.source_pool_permit_any_host),
                item.source_pool_original_start_ip, item.source_pool_original_end_ip,
                item.source_vip_reference, item.source_vip_group_reference,
                item.source_vip_type, self._optional_bool_literal(item.source_vip_enabled),
                self._optional_bool_literal(item.source_vip_nat_source_vip),
                item.source_vip_filters, item.source_vip_interface_filters,
                item.source_vip_services, item.source_vip_port_mapping_type,
                item.translated_destinations, item.original_destination_port,
                item.destination_protocol, item.translated_port,
                item.source_policy_fixed_port, item.source_policy_nat46,
                item.source_policy_nat64, item.source_policy_nat_inbound,
                item.source_policy_nat_outbound, item.source_policy_nat_ip,
                item.source_policy_match_vip, item.source_policy_match_vip_only,
                item.migration_status,
                self._optional_bool_literal(item.requires_manual_review),
                item.review_reasons,
                item.description,
            )
            for index, item in enumerate(self.ir.nat_rules, 1)
        ]
        self._table_sheet(
            workbook,
            "NAT Rules",
            (
                "Rule #", "Name", "Type", "Source Policy ID", "Source Policy UUID",
                "Enabled", "Source Interface", "From Zone", "Destination Interface",
                "To Zone", "Original Source", "Original Destination", "Services",
                "Internet Services", "Source Translation Mode", "IP Pool",
                "Translated Source", "IP Pool Type", "Pool Excluded IPs", "Pool Full Cone",
                "Pool Source Start IP", "Pool Source End IP", "VIP", "VIP Group",
                "VIP Type", "VIP Enabled", "VIP NAT Source VIP", "VIP Source Filters",
                "VIP Interface Filters", "VIP Services", "VIP Port Mapping Type",
                "Translated Destination", "Original Destination Port", "Destination Protocol",
                "Translated Port", "Policy Fixed Port", "Policy NAT46", "Policy NAT64",
                "Policy NAT Inbound", "Policy NAT Outbound", "Policy NAT IP",
                "Policy Match VIP", "Policy Match VIP Only", "Migration Status",
                "Manual Review", "Review Reasons", "Description",
            ),
            rows,
        )

    def _build_virtual_ips(self, workbook: Any) -> None:
        rows = [
            (
                item.name,
                item.address_family,
                item.source_id,
                item.source_uuid,
                item.vip_type,
                self._optional_bool_literal(item.enabled),
                item.external_ip,
                item.external_addresses,
                item.external_interface,
                item.mapped_ips,
                item.mapped_address,
                self._optional_bool_literal(item.port_forward),
                item.protocol,
                item.external_port,
                item.mapped_port,
                item.port_mapping_type,
                self._optional_bool_literal(item.arp_reply),
                item.gratuitous_arp_interval,
                self._optional_bool_literal(item.nat_source_vip),
                self._optional_bool_literal(item.nat44),
                self._optional_bool_literal(item.nat46),
                self._optional_bool_literal(item.nat64),
                self._optional_bool_literal(item.nat66),
                self._optional_bool_literal(item.add_nat46_route),
                self._optional_bool_literal(item.add_nat64_route),
                self._optional_bool_literal(item.ndp_reply),
                item.ipv6_mapped_ip,
                item.ipv6_mapped_port,
                item.ipv4_mapped_ip,
                item.ipv4_mapped_port,
                item.embedded_ipv4_address,
                item.source_filters,
                item.source_interface_filters,
                item.services,
                item.load_balance_method,
                item.server_type,
                item.persistence,
                self._optional_bool_literal(item.http_redirect),
                item.monitors,
                item.max_embryonic_connections,
                item.color,
                item.description,
                self._format_settings(item.extra_settings),
                item.migration_status,
                self._optional_bool_literal(item.requires_manual_review),
                item.audit_note,
            )
            for item in self.ir.virtual_ips
        ]
        self._table_sheet(
            workbook,
            "Virtual IPs",
            (
                "Name", "Address Family", "Source ID", "Source UUID", "Type", "Enabled", "External IP",
                "External Address Objects", "External Interface", "Mapped IPs",
                "Mapped Address", "Port Forward", "Protocol", "External Port",
                "Mapped Port", "Port Mapping Type", "ARP Reply", "Gratuitous ARP Interval",
                "NAT Source VIP",
                "NAT44", "NAT46", "NAT64", "NAT66", "Add NAT46 Route",
                "Add NAT64 Route", "NDP Reply", "IPv6 Mapped IP", "IPv6 Mapped Port",
                "IPv4 Mapped IP", "IPv4 Mapped Port", "Embedded IPv4 Address",
                "Source Filters", "Source Interface Filters", "Services",
                "Load Balance Method", "Server Type", "Persistence", "HTTP Redirect",
                "Monitors", "Max Embryonic Connections", "Color", "Description",
                "Additional Settings", "Extraction Status", "Manual Review", "Review Reason",
            ),
            rows,
        )

    def _build_vip_real_servers(self, workbook: Any) -> None:
        rows = [
            (
                vip.name,
                server.id,
                server.address_type,
                server.ip_address,
                server.address_reference,
                server.port,
                server.status,
                server.weight,
                server.holddown_interval,
                server.healthcheck,
                server.http_host,
                server.translate_host,
                server.max_connections,
                server.monitors,
                server.client_ip,
                server.migration_status,
                self._optional_bool_literal(server.requires_manual_review),
                server.audit_note,
                self._format_settings(server.source_attributes),
            )
            for vip in self.ir.virtual_ips
            for server in vip.real_servers
        ]
        self._table_sheet(
            workbook,
            "VIP Real Servers",
            (
                "VIP Name", "Server ID", "Address Type", "IP", "Address Object",
                "Port", "Status", "Weight", "Holddown Interval", "Health Check",
                "HTTP Host", "Translate Host", "Max Connections", "Monitors", "Client IP",
                "Extraction Status", "Manual Review", "Review Reason", "Additional Settings",
            ),
            rows,
        )

    def _build_vpn_tunnels(self, workbook: Any) -> None:
        rows = [
            (
                item.name,
                item.source_type,
                item.peer_address,
                item.local_interface,
                item.source_local_gateway,
                item.ike_version,
                item.source_mode,
                item.source_peer_type,
                self._optional_bool_literal(item.source_net_device),
                item.source_proposals,
                self._optional_bool_literal(item.source_mode_config),
                self._optional_bool_literal(item.source_eap),
                item.source_eap_identity,
                item.source_auth_user_group,
                item.source_client_ip_start,
                item.source_client_ip_end,
                (
                    f"{item.source_client_ip_start} - {item.source_client_ip_end}"
                    if item.source_client_ip_start and item.source_client_ip_end
                    else None
                ),
                item.source_dns_mode,
                item.source_split_include,
                item.source_dpd_retry_interval,
                (
                    "Configured / Redacted"
                    if item.has_psk or item.psk
                    else "Not configured"
                ),
                item.migration_status,
                self._optional_bool_literal(item.requires_manual_review),
                self._format_settings(item.source_attributes),
                item.ike_crypto_profile,
                item.ipsec_crypto_profile,
                item.description,
            )
            for item in self.ir.vpn_tunnels
        ]
        self._table_sheet(
            workbook,
            "VPN Tunnels",
            (
                "Name",
                "Type",
                "Peer Address",
                "Local Interface",
                "Local Gateway",
                "IKE Version",
                "Mode",
                "Peer Type",
                "Net Device",
                "IKE Proposal",
                "Mode Config",
                "EAP",
                "EAP Identity",
                "Auth User Group",
                "Client IP Start",
                "Client IP End",
                "Client IP Range",
                "DNS Mode",
                "Split Include",
                "DPD Retry Interval",
                "PSK",
                "Extraction Status",
                "Manual Review",
                "Additional Settings",
                "IKE Crypto Profile",
                "IPsec Crypto Profile",
                "Description",
            ),
            rows,
        )

    def _build_certificates(self, workbook: Any) -> None:
        extraction_timestamp = self.ir.metadata.migration_timestamp
        if extraction_timestamp.tzinfo is None:
            extraction_timestamp = extraction_timestamp.replace(
                tzinfo=timezone.utc
            )

        rows = []
        for item in self.ir.certificates:
            expired = None
            if item.valid_until is not None:
                valid_until = item.valid_until
                if valid_until.tzinfo is None:
                    valid_until = valid_until.replace(tzinfo=timezone.utc)
                expired = valid_until < extraction_timestamp

            rows.append(
                (
                    item.name,
                    item.certificate_type,
                    item.source_range,
                    item.source_origin,
                    item.subject,
                    item.issuer,
                    item.serial_number,
                    item.valid_from,
                    item.valid_until,
                    expired,
                    item.public_key_algorithm,
                    item.public_key_size,
                    item.signature_algorithm,
                    item.sha256_fingerprint,
                    item.is_ca,
                    item.is_self_signed,
                    item.has_certificate,
                    item.has_private_key,
                    item.private_key_encrypted,
                    item.has_password,
                    item.source_last_updated,
                    item.migration_status,
                    item.requires_manual_review,
                    item.parse_error,
                    self._format_settings(item.source_attributes),
                    item.description,
                )
            )

        self._table_sheet(
            workbook,
            "Certificates",
            (
                "Name",
                "Certificate Type",
                "Range",
                "Source",
                "Subject",
                "Issuer",
                "Serial Number",
                "Valid From",
                "Valid Until",
                "Expired",
                "Public Key Algorithm",
                "Key Size",
                "Signature Algorithm",
                "SHA-256 Fingerprint",
                "CA Certificate",
                "Self Signed",
                "Has Certificate",
                "Has Private Key",
                "Private Key Encrypted",
                "Has Password",
                "Last Updated",
                "Extraction Status",
                "Manual Review",
                "Parse Error",
                "Additional Settings",
                "Description",
            ),
            rows,
            empty_note="No remote, local, or CA certificates were extracted.",
            subtitle=(
                "Non-secret certificate inventory. Public certificate PEM, "
                "private keys, and passwords are intentionally excluded."
            ),
        )

    def _build_ssh_keys(self, workbook: Any) -> None:
        self._table_sheet(
            workbook,
            "SSH Keys",
            (
                "Name", "Type", "Source", "Has Public Key",
                "Has Private Key", "Has Password", "Extraction Status",
                "Manual Review", "Additional Settings",
            ),
            (
                (
                    item.name,
                    item.key_type,
                    item.source_origin,
                    bool(item.public_key),
                    item.has_private_key,
                    item.has_password,
                    item.migration_status,
                    item.requires_manual_review,
                    self._format_settings(item.source_attributes),
                )
                for item in self.ir.ssh_keys
            ),
            subtitle="Public-key presence is shown; private-key and password contents are never exported.",
        )

    def _build_routes(self, workbook: Any) -> None:
        rows = [
            (
                item.name,
                item.source_route_id,
                item.address_family,
                item.destination,
                item.destination,
                item.source_destination,
                item.source_destination_reference,
                item.source_prefix,
                item.interface,
                item.next_hop,
                item.administrative_distance,
                item.metric,
                item.priority,
                item.weight,
                item.blackhole,
                item.enabled,
                item.sdwan_zone,
                item.sdwan_zones,
                item.dynamic_gateway,
                item.link_monitor_exempt,
                item.bfd,
                item.vrf,
                item.route_tag,
                item.internet_service,
                item.internet_service_custom,
                item.migration_status,
                item.requires_manual_review,
                item.review_reasons,
                item.description,
                self._format_settings(item.source_attributes),
                item.parse_error,
            )
            for item in self.ir.routes
        ]
        self._table_sheet(
            workbook,
            "Routes",
            (
                "Name",
                "Source Route ID",
                "Address Family",
                "Destination",
                "Destination Prefix (Normalized)",
                "Source Destination",
                "Destination Object / Group",
                "Source Prefix",
                "Interface",
                "Next Hop",
                "Administrative Distance",
                "Metric",
                "Priority",
                "Weight",
                "Blackhole",
                "Enabled",
                "SD-WAN Zone",
                "SD-WAN Zones",
                "Dynamic Gateway",
                "Link Monitor Exempt",
                "BFD",
                "VRF",
                "Route Tag",
                "Internet Service",
                "Internet Service Custom",
                "Migration Status",
                "Manual Review",
                "Review Reasons",
                "Description",
                "Additional Settings",
                "Parse Error",
            ),
            rows,
        )
    def _build_vpn_phase2(self, workbook: Any) -> None:
        rows = [
            (
                item.name,
                item.phase1_name,
                item.proposals,
                item.source_address_type,
                item.source_names,
                item.destination_address_type,
                item.destination_names,
                item.source_subnet,
                item.destination_subnet,
                self._optional_bool_literal(item.auto_negotiate),
                item.dh_groups,
                self._optional_bool_literal(item.keepalive),
                item.migration_status,
                self._optional_bool_literal(item.requires_manual_review),
                self._format_settings(item.source_attributes),
                item.description,
            )
            for item in self.ir.vpn_phase2
        ]
        self._table_sheet(
            workbook,
            "VPN Phase 2",
            (
                "Name",
                "Phase 1",
                "Proposal",
                "Source Address Type",
                "Source Selector",
                "Destination Address Type",
                "Destination Selector",
                "Source Subnet",
                "Destination Subnet",
                "Auto Negotiate",
                "DH / PFS Groups",
                "Keepalive",
                "Extraction Status",
                "Manual Review",
                "Additional Settings",
                "Description",
            ),
            rows,
        )

    @staticmethod
    def _routing_protocol_label(source_path: str) -> str:
        return {
            "router rip": "RIP",
            "router ripng": "RIPng",
            "router ospf": "OSPF",
            "router ospf6": "OSPFv3",
            "router bgp": "BGP",
            "router isis": "ISIS",
            "router multicast": "Multicast Routing",
        }.get(source_path, source_path)

    def _structured_routing_items(self) -> list[Any]:
        if self.extraction is None:
            return []
        return [
            item for item in self.extraction.inventory_items
            if "structured-routing-protocol" in item.notes
        ]

    def _build_routing_protocols(self, workbook: Any) -> None:
        items = self._structured_routing_items()
        self._table_sheet(
            workbook,
            "Routing Protocols",
            (
                "Protocol", "Name or Instance", "Source Block Present",
                "Configured", "Extraction Status", "Manual Review",
            ),
            (
                (
                    self._routing_protocol_label(item.source_path),
                    item.name,
                    "Yes",
                    (
                        "Yes"
                        if self._flatten_source_profile_settings(item)
                        else "No"
                    ),
                    item.status,
                    self._optional_bool_literal(item.requires_manual_review),
                )
                for item in items
            ),
        )
        self._table_sheet(
            workbook,
            "Routing Protocol Settings",
            ("Protocol", "Object / Instance", "Subsection", "Entry", "Operation", "Setting", "Value"),
            (
                (self._routing_protocol_label(row[0]), *row[1:])
                for item in items
                for row in self._flatten_source_profile_settings(item)
            ),
        )

    def _build_internet_services(self, workbook: Any) -> None:
        rows = [
            (
                item.name,
                self.ir.metadata.source_vendor,
                item.source_id,
                item.description,
                self._format_settings(item.source_attributes),
            )
            for item in self.ir.internet_services
        ]

        self._table_sheet(
            workbook,
            "Internet Services",
            (
                "Name",
                "Source Vendor",
                "Source ID",
                "Description",
                "Additional Settings",
            ),
            rows,
            empty_note="No Internet Service objects were extracted from the source configuration.",
            subtitle=(
                "Internet Service objects referenced by source firewall policy. "
                "Source IDs are retained for traceability and must not be assumed "
                "equivalent to target-vendor identifiers."
            ),
        )

    def _build_security_profiles(self, workbook: Any) -> None:
        rows = [
            (
                item.name, item.antivirus, item.vulnerability, item.anti_spyware,
                item.url_filtering, item.file_blocking, item.wildfire,
                item.ssl_decryption, item.description,
            )
            for item in self.ir.security_profile_groups
        ]
        self._table_sheet(
            workbook,
            "Security Profiles",
            (
                "Name", "Antivirus", "Vulnerability", "Anti-Spyware", "URL Filtering",
                "File Blocking", "WildFire", "SSL Decryption", "Description",
            ),
            rows,
        )

    def _build_ips_sensors(self, workbook: Any) -> None:
        rows = [
            (
                sensor.name,
                sensor.description,
                sensor.block_malicious_url,
                sensor.scan_botnet_connections,
                len(sensor.entries),
                sensor.migration_status,
                sensor.requires_manual_review,
                self._format_settings(sensor.source_attributes),
            )
            for sensor in self.ir.ips_sensors
        ]
        self._table_sheet(
            workbook,
            "IPS Sensors",
            (
                "Name",
                "Description",
                "Block Malicious URL",
                "Scan Botnet Connections",
                "Entry Count",
                "Extraction Status",
                "Manual Review",
                "Additional Settings",
            ),
            rows,
            empty_note="No IPS sensors were extracted.",
            subtitle=(
                "FortiGate IPS sensor inventory retained as EXTRACT_ONLY; "
                "source signature IDs are not translated."
            ),
        )

    def _build_ips_sensor_entries(self, workbook: Any) -> None:
        rows = [
            (
                sensor.name,
                entry.source_id,
                ", ".join(str(value) for value in entry.source_signature_ids),
                ", ".join(entry.severities),
                entry.location,
                ", ".join(entry.protocols),
                entry.enabled,
                entry.action,
                entry.rate_count,
                entry.rate_duration,
                entry.quarantine,
                entry.quarantine_expiry,
                self._format_settings(entry.source_attributes),
            )
            for sensor in self.ir.ips_sensors
            for entry in sensor.entries
        ]
        self._table_sheet(
            workbook,
            "IPS Sensor Entries",
            (
                "Sensor",
                "Entry ID",
                "Signature IDs",
                "Severities",
                "Location",
                "Protocols",
                "Enabled",
                "Action",
                "Rate Count",
                "Rate Duration",
                "Quarantine",
                "Quarantine Expiry",
                "Additional Settings",
            ),
            rows,
            empty_note="No nested IPS sensor entries were extracted.",
            subtitle=(
                "One row per source entry. Signature IDs and filter values "
                "retain their FortiGate source meaning."
            ),
        )

    def _build_warnings(
        self,
        workbook: Any,
    ) -> None:
        rows = [
            (
                item.id,
                item.category,
                item.confidence,
                item.message,
            )
            for item in self.ir.audit_entries
        ]

        sheet = self._table_sheet(
            workbook,
            "Warnings",
            (
                "ID",
                "Category",
                "Confidence",
                "Message",
            ),
            rows,
            empty_note=(
                "No audit warnings were reported. "
                "See Extraction Coverage before assuming extraction is complete."
            ),
            subtitle=(
                "Audit and manual-review entries emitted while "
                "normalizing the source configuration."
            ),
        )

        # Highlight only the confidence/status cell rather than painting the
        # entire row. This preserves readability for large warning inventories.
        for row in range(
            4,
            sheet.max_row + 1,
        ):
            confidence_cell = sheet.cell(
                row,
                3,
            )

            confidence = str(
                confidence_cell.value or ""
            ).lower()

            if confidence in {
                "manual",
                "unsupported",
            }:
                fill = self._LIGHT_RED

            elif confidence == "partial":
                fill = self._LIGHT_AMBER

            else:
                continue

            confidence_cell.fill = PatternFill(
                "solid",
                fgColor=fill,
            )

    def _build_vip_groups(self, workbook: Any) -> None:
        self._table_sheet(
            workbook,
            "VIP Groups",
            (
                "Name", "Address Family", "Source UUID", "Interface", "Members", "Source Color",
                "Extraction Status", "Manual Review", "Review Reason",
                "Additional Settings", "Description",
            ),
            (
                (
                    item.name, item.address_family, item.source_uuid, item.interface, item.members,
                    item.source_color, item.migration_status,
                    self._optional_bool_literal(item.requires_manual_review),
                    item.audit_note, self._format_settings(item.source_attributes), item.description,
                )
                for item in self.ir.virtual_ip_groups
            ),
        )

    def _build_sdwan(self, workbook: Any) -> None:
        sdwan = self.ir.sdwan
        self._table_sheet(
            workbook,
            "SD-WAN",
            ("Status", "Load Balance Mode", "Extraction Status", "Manual Review", "Additional Settings"),
            [] if sdwan is None else [(
                sdwan.status, sdwan.load_balance_mode, sdwan.migration_status,
                self._optional_bool_literal(sdwan.requires_manual_review),
                self._format_settings(sdwan.source_attributes),
            )],
        )
        self._table_sheet(
            workbook,
            "SD-WAN Members",
            (
                "ID", "Interface", "Zone", "Gateway", "Source", "IPv6 Gateway",
                "IPv6 Source", "Cost", "Weight", "Priority", "IPv6 Priority",
                "Spillover Threshold", "Ingress Spillover Threshold", "Volume Ratio",
                "Status", "Description", "Additional Settings",
            ),
            [] if sdwan is None else (
                (
                    item.source_id, item.interface, item.zone, item.gateway,
                    item.source, item.gateway6, item.source6, item.cost, item.weight,
                    item.priority, item.priority6, item.spillover_threshold,
                    item.ingress_spillover_threshold, item.volume_ratio, item.status,
                    item.description, self._format_settings(item.source_attributes),
                )
                for item in sdwan.members
            ),
        )
        self._table_sheet(
            workbook,
            "SD-WAN Health Checks",
            (
                "Name", "Server", "Members", "Protocol", "Port", "Interval",
                "Probe Timeout", "Fail Time", "Recovery Time", "Update Static Route",
                "VRF", "Source Address", "SLA Count", "Additional Settings",
            ),
            [] if sdwan is None else (
                (
                    item.name, item.server, item.member_ids, item.protocol, item.port,
                    item.interval, item.probe_timeout, item.failtime, item.recoverytime,
                    item.update_static_route, item.vrf, item.source, len(item.sla),
                    self._format_settings(item.source_attributes),
                )
                for item in sdwan.health_checks
            ),
        )
        self._table_sheet(
            workbook,
            "SD-WAN SLAs",
            ("Health Check", "SLA ID", "Additional Settings"),
            [] if sdwan is None else (
                (check.name, sla.source_id, self._format_settings(sla.source_attributes))
                for check in sdwan.health_checks
                for sla in check.sla
            ),
        )
        self._table_sheet(
            workbook,
            "SD-WAN Rules",
            (
                "ID", "Name", "Mode", "Status", "Source", "Destination",
                "Health Checks", "Priority Members", "Priority Zones",
                "Internet Service", "Internet Service Names",
                "Internet Service App Control", "SLA Compare Method", "Tie Break",
                "Use Shortcut SLA", "Additional Settings",
            ),
            [] if sdwan is None else (
                (
                    item.source_id, item.name, item.mode, item.status, item.source_addresses,
                    item.destination_addresses, item.health_checks, item.priority_member_ids,
                    item.priority_zones,
                    item.internet_service, item.internet_service_names,
                    item.internet_service_app_ctrl, item.sla_compare_method, item.tie_break,
                    item.use_shortcut_sla,
                    self._format_settings(item.source_attributes),
                )
                for item in sdwan.rules
            ),
        )
        self._build_sdwan_source_details(workbook, sdwan)

    def _build_sdwan_source_details(self, workbook: Any, sdwan: Any) -> None:
        self._table_sheet(
            workbook,
            "SD-WAN Duplication",
            (
                "ID", "Service ID", "Source Addresses", "Destination Addresses",
                "IPv6 Source Addresses", "IPv6 Destination Addresses",
                "Source Interfaces", "Destination Interfaces", "Services",
                "Packet Duplication", "SLA Match Service", "Packet De-duplication",
                "Extraction Status", "Manual Review", "Additional Settings",
            ),
            [] if sdwan is None else (
                (
                    item.source_id, item.service_id, item.source_addresses,
                    item.destination_addresses, item.source_addresses6,
                    item.destination_addresses6, item.source_interfaces,
                    item.destination_interfaces, item.services, item.packet_duplication,
                    item.sla_match_service, item.packet_de_duplication,
                    item.migration_status,
                    self._optional_bool_literal(item.requires_manual_review),
                    self._format_settings(item.source_attributes),
                )
                for item in sdwan.duplication_rules
            ),
        )
        self._table_sheet(
            workbook,
            "SD-WAN Neighbors",
            ("Name", "Extraction Status", "Manual Review", "Additional Settings"),
            [] if sdwan is None else (
                (
                    item.name, item.migration_status,
                    self._optional_bool_literal(item.requires_manual_review),
                    self._format_settings(item.source_attributes),
                )
                for item in sdwan.neighbors
            ),
        )
        self._table_sheet(
            workbook,
            "SD-WAN Rule SLAs",
            ("Rule ID", "Rule Name", "SLA", "SLA ID", "Additional Settings"),
            [] if sdwan is None else (
                (
                    rule.source_id, rule.name, sla.name, sla.source_id,
                    self._format_settings(sla.source_attributes),
                )
                for rule in sdwan.rules
                for sla in rule.sla
            ),
        )

    def _structured_security_items(self) -> list[Any]:
        if self.extraction is None:
            return []
        return [
            item for item in self.extraction.inventory_items
            if "structured-security-profile" in item.notes
        ]

    def _profile_policy_references(self, name: Any) -> list[str]:
        if not name:
            return []
        references = []
        for policy in self.ir.policies:
            known = {
                policy.antivirus,
                policy.ips_sensor,
                policy.webfilter,
                policy.application_list,
                policy.ssl_ssh_profile,
            }
            if name in known:
                references.append(policy.source_rule_id or policy.name)
        return references

    def _flatten_source_profile_settings(self, item: Any) -> list[tuple[Any, ...]]:
        rows = []

        def walk(node: Any, subsections: list[str], entry: Any) -> None:
            node_type = next(
                (note.split(":", 1)[1] for note in node.notes if note.startswith("source-node:")),
                "edit",
            )
            current_subsections = subsections
            current_entry = entry
            if node is not item:
                if node_type == "config":
                    current_subsections = [*subsections, node.name]
                elif node_type == "edit":
                    current_entry = node.name
            for command in node.commands:
                rows.append((
                    item.source_path,
                    item.name,
                    " / ".join(current_subsections),
                    current_entry,
                    command.operation,
                    command.key,
                    command.values,
                ))
            for child in node.children:
                walk(child, current_subsections, current_entry)

        walk(item, [], item.name)
        return rows

    def _build_source_security_profiles(self, workbook: Any) -> None:
        items = self._structured_security_items()
        self._table_sheet(
            workbook,
            "Source Security Profiles",
            ("Profile Type", "Name", "Extraction Status", "Manual Review", "Referenced By Policy"),
            (
                (
                    item.source_path, item.name, item.status,
                    self._optional_bool_literal(item.requires_manual_review),
                    self._profile_policy_references(item.name),
                )
                for item in items
            ),
        )
        self._table_sheet(
            workbook,
            "Source Security Profile Setting",
            ("Profile Type", "Profile Name", "Subsection", "Entry", "Operation", "Setting", "Value"),
            (row for item in items for row in self._flatten_source_profile_settings(item)),
        )

    def _build_identity_inventory(self, workbook: Any) -> None:
        self._table_sheet(
            workbook, "LDAP Servers",
            ("Name", "Server", "CNID", "DN", "Type", "Username", "Password Configured", "Extraction Status", "Manual Review", "Additional Settings"),
            (
                (
                    item.name, item.server, item.cnid, item.dn, item.source_type,
                    item.username, item.has_password, item.migration_status,
                    item.requires_manual_review, self._format_settings(item.source_attributes),
                ) for item in self.ir.user_ldap_servers
            ),
        )
        self._table_sheet(
            workbook, "SAML Servers",
            ("Name", "Entity ID", "SSO URL", "SLO URL", "IdP Entity ID", "IdP SSO URL", "IdP SLO URL", "IdP Certificate", "User Name", "Group Name", "Digest Method", "Extraction Status", "Manual Review", "Additional Settings"),
            (
                (
                    item.name, item.entity_id, item.single_sign_on_url, item.single_logout_url,
                    item.idp_entity_id, item.idp_single_sign_on_url,
                    item.idp_single_logout_url, item.idp_cert, item.user_name,
                    item.group_name, item.digest_method, item.migration_status,
                    item.requires_manual_review, self._format_settings(item.source_attributes),
                ) for item in self.ir.user_saml_servers
            ),
        )
        self._table_sheet(
            workbook, "FSSO Servers",
            ("Name", "Server", "Password Configured", "Extraction Status", "Manual Review", "Additional Settings"),
            (
                (
                    item.name, item.server, item.has_password,
                    item.migration_status, item.requires_manual_review,
                    self._format_settings(item.source_attributes),
                ) for item in self.ir.fsso_providers
            ),
        )
        self._table_sheet(
            workbook, "FSSO AD Groups",
            ("Name", "FSSO Server", "Server Resolved", "Extraction Status", "Manual Review", "Additional Settings"),
            (
                (
                    item.name, item.provider_name, item.provider_resolved,
                    item.migration_status, item.requires_manual_review,
                    self._format_settings(item.source_attributes),
                ) for item in self.ir.fsso_ad_groups
            ),
        )
        self._table_sheet(
            workbook, "Local Users",
            ("Name", "Status", "Type", "Password Configured", "Extraction Status", "Manual Review", "Additional Settings"),
            (
                (
                    item.name, item.status, item.source_type, item.has_password,
                    item.migration_status, item.requires_manual_review,
                    self._format_settings(item.source_attributes),
                ) for item in self.ir.local_users
            ),
        )
        self._table_sheet(
            workbook, "User Groups",
            ("Name", "Type", "Members", "Match Count", "Extraction Status", "Manual Review", "Additional Settings"),
            (
                (
                    item.name, item.group_type, item.members, len(item.matches),
                    item.migration_status, item.requires_manual_review,
                    self._format_settings(item.source_attributes),
                ) for item in self.ir.user_groups
            ),
        )
        self._table_sheet(
            workbook, "User Group Matches",
            ("User Group", "ID", "Server Name", "Group Name"),
            (
                (group.name, match.source_id, match.server_name, match.group_name)
                for group in self.ir.user_groups for match in group.matches
            ),
        )

    def _build_ssl_vpn(self, workbook: Any) -> None:
        settings = self.ir.ssl_vpn_settings
        self._table_sheet(
            workbook, "SSL VPN Settings",
            ("Status", "Minimum Protocol", "Banned Ciphers", "Server Certificate", "Source Interfaces", "Source Addresses", "Tunnel IP Pools", "Default Portal", "Extraction Status", "Manual Review", "Additional Settings"),
            [] if settings is None else [(
                settings.status, settings.ssl_min_proto_ver, settings.banned_cipher,
                settings.server_certificate, settings.source_interfaces,
                settings.source_addresses, settings.tunnel_ip_pools,
                settings.default_portal, settings.migration_status,
                settings.requires_manual_review, self._format_settings(settings.source_attributes),
            )],
        )
        self._table_sheet(
            workbook, "SSL VPN Portals",
            ("Name", "Tunnel Mode", "IPv6 Tunnel Mode", "IP Pools", "IPv6 Pools", "Split Tunneling", "Limit User Logins", "FortiClient Download", "Extraction Status", "Manual Review", "Additional Settings"),
            (
                (
                    item.name, item.tunnel_mode, item.ipv6_tunnel_mode, item.ip_pools,
                    item.ipv6_pools, item.split_tunneling, item.limit_user_logins,
                    item.forticlient_download, item.migration_status,
                    item.requires_manual_review, self._format_settings(item.source_attributes),
                ) for item in self.ir.ssl_vpn_portals
            ),
        )
        self._table_sheet(
            workbook, "SSL VPN Authentication Rules",
            ("ID", "Groups", "Portal", "Additional Settings"),
            [] if settings is None else (
                (item.source_id, item.groups, item.portal, self._format_settings(item.source_attributes))
                for item in settings.authentication_rules
            ),
        )
        self._table_sheet(
            workbook, "SSL VPN Host Checks",
            ("Portal", "Name", "Type", "Version", "GUID", "Migration Status", "Manual Review", "Additional Settings"),
            (
                (
                    portal.name, item.name, item.source_type, item.version, item.guid,
                    item.migration_status, item.requires_manual_review,
                    self._format_settings(item.source_attributes),
                ) for portal in self.ir.ssl_vpn_portals for item in portal.host_checks
            ),
        )

    def _build_internet_service_definitions(self, workbook: Any) -> None:
        definitions = self.ir.internet_service_definitions
        self._table_sheet(
            workbook,
            "Internet Service Definitions",
            (
                "Definition ID", "Entry Count", "Migration Status",
                "Requires Manual Review", "Additional Settings",
            ),
            (
                (
                    definition.source_id,
                    len(definition.entries),
                    definition.migration_status,
                    definition.requires_manual_review,
                    self._format_settings(definition.source_attributes),
                )
                for definition in definitions
            ),
            empty_note="No Internet Service Definitions were extracted from the source configuration.",
            subtitle="FortiGate Internet Service Definitions are EXTRACT_ONLY and require manual review.",
        )
        self._table_sheet(
            workbook,
            "Internet Service Def Entries",
            (
                "Definition ID", "Sequence #", "Category ID", "Name", "Protocol #",
                "Protocol Name", "Port Range Count", "Additional Settings",
            ),
            (
                (
                    definition.source_id,
                    entry.source_sequence,
                    entry.category_id,
                    entry.name,
                    entry.protocol_number,
                    self._protocol_name(entry.protocol_number),
                    len(entry.port_ranges),
                    self._format_settings(entry.source_attributes),
                )
                for definition in definitions
                for entry in definition.entries
            ),
            empty_note="No Internet Service Definition entries were extracted from the source configuration.",
        )
        self._table_sheet(
            workbook,
            "Internet Service Def Ports",
            (
                "Definition ID", "Entry Sequence #", "Range ID", "Start Port",
                "End Port", "Additional Settings",
            ),
            (
                (
                    definition.source_id,
                    entry.source_sequence,
                    port_range.source_id,
                    port_range.start_port,
                    port_range.end_port,
                    self._format_settings(port_range.source_attributes),
                )
                for definition in definitions
                for entry in definition.entries
                for port_range in entry.port_ranges
            ),
            empty_note="No Internet Service Definition port ranges were extracted from the source configuration.",
        )

    @staticmethod
    def _protocol_name(protocol_number: int | None) -> str | None:
        return {1: "ICMP", 6: "TCP", 17: "UDP"}.get(protocol_number)

    def _build_administrator_inventory(self, workbook: Any) -> None:
        self._table_sheet(
            workbook,
            "Administrators",
            (
                "Name", "Access Profile", "VDOMs", "IPv4 Trusted Hosts", "IPv6 Trusted Hosts",
                "Two Factor", "FortiToken", "Guest User Groups", "Remote Auth", "Remote Group",
                "Schedule", "Peer Auth", "Peer Group", "SSH Certificate", "SSH Public Keys",
                "Credential Configured", "Migration Status", "Manual Review",
                "Additional Settings",
            ),
            (
                (
                    item.name, item.access_profile, item.vdoms, item.trusted_hosts_ipv4,
                    item.trusted_hosts_ipv6, item.two_factor, item.token_reference,
                    item.guest_user_groups, item.remote_auth, item.remote_group,
                    item.schedule, item.peer_auth, item.peer_group, item.ssh_certificate,
                    item.ssh_public_keys,
                    item.credential_configured, item.migration_status,
                    item.requires_manual_review,
                    self._format_settings(item.source_attributes),
                )
                for item in self.ir.administrators
            ),
        )
        self._table_sheet(
            workbook,
            "Admin Profiles",
            ("Name", "Migration Status", "Manual Review", "Additional Settings"),
            (
                (
                    item.name, item.migration_status, item.requires_manual_review,
                    self._format_settings(item.source_attributes),
                )
                for item in self.ir.admin_profiles
            ),
        )
    def _structured_routing_dependency_items(self) -> list[Any]:
        if self.extraction is None:
            return []
        return [
            item
            for item in self.extraction.inventory_items
            if "structured-routing-dependency" in item.notes
        ]

    def _build_routing_dependencies(self, workbook: Any) -> None:
        items = self._structured_routing_dependency_items()
        self._table_sheet(
            workbook,
            "Routing Dependencies",
            (
                "Type",
                "Name / Source ID",
                "Source Path",
                "Source Block Present",
                "Configured",
                "Extraction Status",
                "Manual Review",
            ),
            (
                (
                    item.source_path.removeprefix("router "),
                    item.name or item.source_id,
                    item.source_path,
                    "Yes",
                    (
                        "Yes"
                        if self._flatten_source_profile_settings(item)
                        else "No"
                    ),
                    item.status,
                    self._optional_bool_literal(item.requires_manual_review),
                )
                for item in items
            ),
        )
        self._table_sheet(
            workbook,
            "Routing Dependency Settings",
            (
                "Type",
                "Object",
                "Parent / Subsection",
                "Entry",
                "Operation",
                "Setting",
                "Value",
            ),
            (
                (row[0].removeprefix("router "), *row[1:])
                for item in items
                for row in self._flatten_source_profile_settings(item)
            ),
        )
        self._table_sheet(
            workbook,
            "Admin Profile Permissions",
            ("Profile", "Permission Group", "Setting", "Value", "Extraction Status", "Additional Settings"),
            (
                (
                    profile.name, block.name, setting, value, "EXTRACT_ONLY",
                    self._format_settings(block.source_attributes),
                )
                for profile in self.ir.admin_profiles
                for block in profile.permission_blocks
                for setting, value in {**block.settings, **block.source_attributes}.items()
            ),
            empty_note="No admin profile permissions were extracted from the source configuration.",
        )
        self._table_sheet(
            workbook,
            "FortiTokens",
            (
                "Serial / Name", "Status", "Assigned User", "Description",
                "Migration Status", "Manual Review", "Additional Settings",
            ),
            (
                (
                    item.serial, item.status, item.assigned_user, item.description,
                    item.migration_status, item.requires_manual_review,
                    self._format_settings(item.source_attributes),
                )
                for item in self.ir.fortitokens
            ),
        )

    def _build_dos_inventory(self, workbook: Any) -> None:
        self._table_sheet(
            workbook, "DoS Policies",
            ("Policy ID", "Status", "Interface", "Source Addresses", "Destination Addresses", "Services", "Anomaly Count", "Description", "Extraction Status", "Manual Review", "Additional Settings"),
            (
                (
                    item.source_id, item.status, item.interface, item.source_addresses,
                    item.destination_addresses, item.services, len(item.anomalies),
                    item.description, item.migration_status, item.requires_manual_review,
                    self._format_settings(item.source_attributes),
                ) for item in self.ir.dos_policies
            ),
        )
        self._table_sheet(
            workbook, "DoS Anomalies",
            ("Policy ID", "Name", "Status", "Log", "Action", "Threshold", "Additional Settings"),
            (
                (
                    policy.source_id, item.name, item.status, item.log, item.action,
                    item.threshold, self._format_settings(item.source_attributes),
                ) for policy in self.ir.dos_policies for item in policy.anomalies
            ),
        )

    def _build_firewall_sniffers(self, workbook: Any) -> None:
        self._table_sheet(
            workbook, "Firewall Sniffer",
            ("ID", "Source UUID", "Log Traffic", "IPv6", "Non-IP", "Application List Status", "Application List", "IPS Sensor Status", "IPS Sensor", "AV Profile Status", "AV Profile", "Web Filter Status", "Web Filter Profile", "Extraction Status", "Manual Review", "Additional Settings"),
            (
                (
                    item.source_id, item.source_uuid, item.logtraffic, item.ipv6,
                    item.non_ip, item.application_list_status, item.application_list,
                    item.ips_sensor_status, item.ips_sensor, item.av_profile_status,
                    item.av_profile, item.webfilter_profile_status,
                    item.webfilter_profile, item.migration_status,
                    item.requires_manual_review, self._format_settings(item.source_attributes),
                ) for item in self.ir.firewall_sniffers
            ),
        )

    def _build_authentication_inventory(self, workbook: Any) -> None:
        self._table_sheet(
            workbook, "Authentication Schemes",
            ("Name", "Method", "User Database", "Extraction Status", "Manual Review", "Additional Settings"),
            (
                (
                    item.name, item.method, item.user_database, item.migration_status,
                    item.requires_manual_review, self._format_settings(item.source_attributes),
                ) for item in self.ir.authentication_schemes
            ),
        )
        self._table_sheet(
            workbook, "Authentication Rules",
            ("Name", "Source Interfaces", "Source Addresses", "Active Auth Method", "Extraction Status", "Manual Review", "Additional Settings"),
            (
                (
                    item.name, item.source_interfaces, item.source_addresses,
                    item.active_auth_method, item.migration_status,
                    item.requires_manual_review, self._format_settings(item.source_attributes),
                ) for item in self.ir.authentication_rules
            ),
        )

    def _build_unsupported(self, workbook: Any) -> None:
        if self.extraction is not None:
            rows = [
                (
                    item.source_path,
                    item.source_name or "",
                    "UNSUPPORTED",
                    item.reason,
                    item.requires_manual_review,
                    item.raw_capture or "",
                )
                for item in self.extraction.unsupported_items
            ]
            sheet = self._table_sheet(
                workbook,
                "Unsupported",
                ("Section", "Item", "Status", "Reason", "Manual Review", "Raw Capture"),
                rows,
                empty_note="No unsupported source sections or objects were reported.",
                subtitle="Unsupported source evidence is shown without secret-bearing raw configuration.",
            )
            for row in range(4, sheet.max_row + 1):
                for column in range(1, 7):
                    sheet.cell(row, column).fill = PatternFill("solid", fgColor=self._LIGHT_RED)
            return

        rows = [
            (
                item.category, item.id, "Unsupported", item.message, True,
                "Audit entry; raw source intentionally excluded from workbook",
            )
            for item in self.ir.audit_entries
            if item.confidence == MigrationConfidence.UNSUPPORTED
        ]
        sheet = self._table_sheet(
            workbook,
            "Unsupported",
            ("Section", "Item", "Status", "Reason", "Manual Review", "Raw Capture"),
            rows,
            empty_note="No unsupported items were reported. See Extraction Coverage before assuming extraction is complete.",
            subtitle="Unsupported audit entries are listed without raw source configuration or secrets.",
        )
        for row in range(4, sheet.max_row + 1):
            for column in range(1, 7):
                sheet.cell(row, column).fill = PatternFill("solid", fgColor=self._LIGHT_RED)

    def _build_extraction_coverage(self, workbook: Any) -> None:
        if self.extraction is not None:
            rows = [
                (
                    section.path,
                    section.present,
                    section.object_count_source,
                    section.object_count_parsed,
                    section.object_count_normalized,
                    section.status.value,
                    section.parser_handler,
                    section.line_start,
                    section.line_end,
                    "; ".join(section.notes),
                )
                for section in self.extraction.source_sections
            ]
            sheet = self._table_sheet(
                workbook,
                "Extraction Coverage",
                (
                    "Source Section",
                    "Found",
                    "Source Objects",
                    "Parsed Objects",
                    "Normalized Objects",
                    "Status",
                    "Parser Handler",
                    "Line Start",
                    "Line End",
                    "Notes",
                ),
                rows,
                empty_note="No FortiGate config sections were discovered in the source.",
                subtitle="Coverage correlates independent source discovery with typed parsing and canonical IR.",
            )
            for row in range(4, sheet.max_row + 1):
                status = str(sheet.cell(row, 6).value or "").lower()
                if "unsupported" in status or "parse_error" in status:
                    fill = self._LIGHT_RED
                elif "partial" in status or "extract_only" in status:
                    fill = self._LIGHT_AMBER
                else:
                    continue
                for column in range(1, 11):
                    sheet.cell(row, column).fill = PatternFill("solid", fgColor=fill)
            return

        collections = (
            ("Interfaces", self.ir.interfaces),
            (
                "Interface Secondary IPs",
                [
                    sec
                    for intf in self.ir.interfaces
                    for sec in getattr(intf, "secondary_ips", [])
                ],
            ),
            (
                "DHCP Servers",
                self.ir.dhcp_servers,
            ),
            (
                "DHCP IP Ranges",
                [
                    ip_range
                    for server in self.ir.dhcp_servers
                    for ip_range in server.ip_ranges
                ],
            ),
            (
                "DHCP Reservations",
                [
                    reservation
                    for server in self.ir.dhcp_servers
                    for reservation in server.reservations
                ],
            ),
            ("Zones", self.ir.zones),
            ("Addresses", self.ir.addresses),
            (
                "Address Groups",
                self.ir.address_groups,
            ),
            ("Proxy Addresses", self.ir.proxy_addresses),
            (
                "Web Proxy Settings",
                [] if self.ir.web_proxy_settings is None else [self.ir.web_proxy_settings],
            ),
            (
                "Service Categories",
                self.ir.service_categories,
            ),
            ("Services", self.ir.services),
            (
                "Service Groups",
                self.ir.service_groups,
            ),
            (
                "Session Helpers",
                self.ir.session_helpers,
            ),
            (
                "Session TTL Overrides",
                self.ir.session_ttl_overrides,
            ),
            ("Schedules", self.ir.schedules),
            ("Traffic Shapers", self.ir.traffic_shapers),
            ("Policies", self.ir.policies),
            (
                "ZTNA Providers",
                self.ir.ztna_providers,
            ),
            ("IP Pools", self.ir.ip_pools),
            ("Virtual IPs", self.ir.virtual_ips),
            ("VIP Groups", self.ir.virtual_ip_groups),
            (
                "VIP Real Servers",
                [
                    server
                    for vip in self.ir.virtual_ips
                    for server in vip.real_servers
                ],
            ),
            ("NAT Rules", self.ir.nat_rules),
            ("VPN Tunnels", self.ir.vpn_tunnels),
            ("VPN Phase 2", self.ir.vpn_phase2),
            ("SSL VPN Portals", self.ir.ssl_vpn_portals),
            (
                "SSL VPN Settings",
                [] if self.ir.ssl_vpn_settings is None else [self.ir.ssl_vpn_settings],
            ),
            ("SD-WAN", [] if self.ir.sdwan is None else [self.ir.sdwan]),
            ("LDAP Servers", self.ir.user_ldap_servers),
            ("SAML Servers", self.ir.user_saml_servers),
            ("FSSO Servers", self.ir.fsso_providers),
            ("FSSO AD Groups", self.ir.fsso_ad_groups),
            ("Local Users", self.ir.local_users),
            ("User Groups", self.ir.user_groups),
            ("DoS Policies", self.ir.dos_policies),
            ("Firewall Sniffer", self.ir.firewall_sniffers),
            ("Authentication Schemes", self.ir.authentication_schemes),
            ("Authentication Rules", self.ir.authentication_rules),
            ("Certificates", self.ir.certificates),
            ("Routes", self.ir.routes),
            (
                "Internet Services",
                self.ir.internet_services,
            ),
            ("Internet Service Definitions", self.ir.internet_service_definitions),
            (
                "Internet Service Def Entries",
                [
                    entry
                    for definition in self.ir.internet_service_definitions
                    for entry in definition.entries
                ],
            ),
            (
                "Internet Service Def Ports",
                [
                    port_range
                    for definition in self.ir.internet_service_definitions
                    for entry in definition.entries
                    for port_range in entry.port_ranges
                ],
            ),
            ("IPS Sensors", self.ir.ips_sensors),
            (
                "IPS Sensor Entries",
                [
                    entry
                    for sensor in self.ir.ips_sensors
                    for entry in sensor.entries
                ],
            ),
            (
                "Security Profiles",
                self.ir.security_profile_groups,
            ),
        )
        rows = [
            (
                name, None, len(items), "Not reported",
                "Populated" if items else "Empty / unknown",
                "Phase 1 exports IR only; source-section coverage awaits ExtractionResult in Phase 2.",
            )
            for name, items in collections
        ]
        note = "Source-section evidence is unavailable in IR; inferred counts are not proof of parser completeness."
        sheet = self._table_sheet(
            workbook,
            "Extraction Coverage",
            ("Source Section", "Found", "Objects", "Parser Status", "IR Status", "Notes"),
            rows,
            subtitle=note,
        )
        for row in range(4, sheet.max_row + 1):
            status = str(sheet.cell(row, 5).value or "").lower()
            if "unsupported" in status or "not transformed" in status:
                fill = self._LIGHT_RED
            elif "partial" in status or "unknown" in status:
                fill = self._LIGHT_AMBER
            else:
                continue
            for column in range(1, 7):
                sheet.cell(row, column).fill = PatternFill("solid", fgColor=fill)

    def _table_sheet(
        self,
        workbook: Any,
        title: str,
        headers: Sequence[str],
        rows: Iterable[Sequence[Any]],
        empty_note: str = (
            "No objects were represented in this IR collection."
        ),
        subtitle: str = (
            "Vendor-neutral IR inventory exported before migration optimization."
        ),
    ) -> Any:
        rows = list(rows)

        sheet = workbook.create_sheet(title)
        sheet.sheet_view.showGridLines = False

        last_column = get_column_letter(
            len(headers)
        )

        sheet.merge_cells(
            f"A1:{last_column}1"
        )

        sheet["A1"] = title

        sheet["A1"].font = Font(
            name="Aptos Display",
            size=16,
            bold=True,
            color=self._WHITE,
        )

        sheet["A1"].fill = PatternFill(
            "solid",
            fgColor=self._NAVY,
        )

        sheet["A1"].alignment = Alignment(
            vertical="center",
        )

        sheet.row_dimensions[1].height = 30

        sheet.merge_cells(
            f"A2:{last_column}2"
        )

        note = (
            subtitle
            if rows
            else empty_note
        )

        sheet["A2"] = self._safe_value(
            f"Back to Summary  |  {note}"
        )

        self._set_internal_link(
            sheet["A2"],
            "Summary",
        )

        # Keep subtitle visually subtle even though A2 is also a hyperlink.
        sheet["A2"].font = Font(
            name="Aptos",
            size=9,
            italic=True,
            underline="single",
            color="0563C1",
        )

        sheet["A2"].alignment = Alignment(
            wrap_text=True,
            vertical="center",
        )

        sheet.row_dimensions[2].height = 26

        for column, header in enumerate(
            headers,
            1,
        ):
            # Headers are application-owned constants, not extracted values.
            cell = sheet.cell(
                3,
                column,
                str(header),
            )

            cell.font = Font(
                name="Aptos",
                bold=True,
                color=self._WHITE,
            )

            cell.fill = PatternFill(
                "solid",
                fgColor=self._TEAL,
            )

            cell.alignment = Alignment(
                wrap_text=True,
                vertical="center",
            )

        sheet.row_dimensions[3].height = 28

        row_count = 0

        for row_count, values in enumerate(
            rows,
            1,
        ):
            worksheet_row = (
                row_count + 3
            )

            for column, value in enumerate(
                values,
                1,
            ):
                cell = sheet.cell(
                    worksheet_row,
                    column,
                    self._safe_value(value),
                )

                cell.font = Font(
                    name="Aptos",
                    size=10,
                    color=self._TEXT,
                )

                cell.alignment = Alignment(
                    wrap_text=True,
                    vertical="top",
                )

            if row_count % 2 == 0:
                for column in range(
                    1,
                    len(headers) + 1,
                ):
                    sheet.cell(
                        worksheet_row,
                        column,
                    ).fill = PatternFill(
                        "solid",
                        fgColor="F8FAFC",
                    )

        sheet.auto_filter.ref = (
            f"A3:{last_column}"
            f"{max(3, row_count + 3)}"
        )

        sheet.freeze_panes = "A4"

        self._size_table(
            sheet,
            len(headers),
            row_count,
        )

        return sheet

    def _size_table(
        self,
        sheet: Any,
        column_count: int,
        row_count: int,
    ) -> None:
        thin = Side(
            style="thin",
            color=self._BORDER,
        )

        verbose_tokens = (
            "description",
            "message",
            "note",
            "reason",
            "additional settings",
            "audit",
            "migration instruction",
            "raw capture",
            "parse error",
        )

        for column in range(
            1,
            column_count + 1,
        ):
            header = str(
                sheet.cell(
                    3,
                    column,
                ).value
                or ""
            )

            max_length = len(header)

            # Inspect only a bounded number of records. Large source inventories
            # must not make workbook generation progressively more expensive.
            for row in range(
                4,
                min(
                    row_count + 4,
                    204,
                ),
            ):
                value = str(
                    sheet.cell(
                        row,
                        column,
                    ).value
                    or ""
                )

                max_length = max(
                    max_length,
                    max(
                        (
                            len(line)
                            for line
                            in value.splitlines()
                        ),
                        default=0,
                    ),
                )

            header_lower = header.lower()

            width_cap = (
                48
                if any(
                    token in header_lower
                    for token
                    in verbose_tokens
                )
                else 32
            )

            sheet.column_dimensions[
                get_column_letter(column)
            ].width = min(
                max(
                    max_length + 2,
                    11,
                ),
                width_cap,
            )

            sheet.cell(
                3,
                column,
            ).border = Border(
                bottom=thin,
            )

        for row in range(
            4,
            row_count + 4,
        ):
            max_lines = max(
                (
                    str(
                        sheet.cell(
                            row,
                            column,
                        ).value
                        or ""
                    ).count("\n")
                    + 1
                    for column
                    in range(
                        1,
                        column_count + 1,
                    )
                ),
                default=1,
            )

            sheet.row_dimensions[row].height = min(
                15 * max_lines + 5,
                90,
            )

    def _format_port(self, port: Any) -> str:
        protocol = self._enum_value(port.protocol)
        result = f"{protocol}/{port.port}"
        details = []
        if port.icmptype is not None:
            details.append(f"type={port.icmptype}")
        if port.icmpcode is not None:
            details.append(f"code={port.icmpcode}")
        if details:
            result += f" ({', '.join(details)})"
        return result

    @staticmethod
    def _format_settings(settings: dict[str, Any]) -> str:
        def format_value(value: Any) -> str:
            if isinstance(value, (list, tuple, set)):
                return " ".join(str(item) for item in value)
            return str(value)

        return "; ".join(
            f"{key.replace('_', '-')}={format_value(value)}"
            for key, value in sorted(settings.items())
        )

    def _safe_value(self, value: Any) -> Any:
        if value is None:
            return ""
        if isinstance(value, bool):
            return "Yes" if value else "No"
        if isinstance(value, (int, float)):
            return value
        if isinstance(value, datetime):
            value = value.isoformat()
        elif isinstance(value, Enum):
            value = value.value
        elif isinstance(value, (list, tuple, set)):
            value = "\n".join(str(self._enum_value(item)) for item in value)
        else:
            value = str(value)

        value = _ILLEGAL_XML_CHARS.sub(" ", value)
        if len(value) > _MAX_CELL_TEXT:
            suffix = "\n[truncated for Excel cell limit]"
            value = value[: _MAX_CELL_TEXT - len(suffix)] + suffix
        if value.lstrip().startswith(_FORMULA_PREFIXES):
            value = "'" + value
        return value

    @staticmethod
    def _enum_value(value: Any) -> Any:
        return value.value if isinstance(value, Enum) else value

    @staticmethod
    def _optional_bool_literal(value: bool | None) -> str | None:
        if value is None:
            return None
        return "TRUE" if value else "FALSE"
